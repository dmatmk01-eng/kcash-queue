"""
Utility functions: CSV export, statement parser, Line message, date helpers.
"""
import csv
import io
import re
from datetime import date, datetime, timedelta as _td
from difflib import SequenceMatcher


# ───────────────────── date helpers ─────────────────────

def today_str() -> str:
    return date.today().strftime("%Y-%m-%d")


def fmt_date(d: str) -> str:
    """แปลง ISO date เป็น DD/MM/YYYY."""
    if not d:
        return "-"
    try:
        return datetime.strptime(d[:10], "%Y-%m-%d").strftime("%d/%m/%Y")
    except ValueError:
        return d


def fmt_amount(amount) -> str:
    try:
        return f"{float(amount):,.2f}"
    except (TypeError, ValueError):
        return str(amount)


# ───────────────────── bank CSV ─────────────────────

def build_bank_csv(expenses: list, cfg: dict) -> str:
    """
    สร้าง CSV โอนเงินแบบ SCB/Krungthai generic format:
    Ref, Beneficiary Name, Account No, Amount, Date, Note
    """
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\r\n")

    writer.writerow(["ลำดับ", "ชื่อผู้รับ", "เลขบัญชี", "จำนวนเงิน (บาท)", "วันครบกำหนด", "เลขที่เอกสาร", "หมายเหตุ"])

    for i, exp in enumerate(expenses, 1):
        writer.writerow([
            i,
            _vendor_name(exp),
            _vendor_account(exp),
            fmt_amount(_amount(exp)),
            fmt_date(exp.get("dueDate") or exp.get("due_date") or ""),
            _doc_serial(exp),
            exp.get("remarks") or exp.get("projectName") or exp.get("note") or "",
        ])

    return output.getvalue()


def build_bank_excel(expenses: list, cfg: dict, path: str) -> None:
    """
    สร้างไฟล์ Excel (.xlsx) จัดรูปแบบสวยงาม อ่านง่าย:
    - หัวตารางสีเขียว ตัวหนา
    - ความกว้างคอลัมน์พอดี
    - จำนวนเงินจัดชิดขวา + คั่นหลักพัน
    - แถวรวมยอดท้ายตาราง
    - เส้นขอบทุกช่อง + สลับสีแถว
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Payment"

    headers = ["ลำดับ", "ชื่อผู้รับ", "แบรนด์", "เลขบัญชี",
               "จำนวนเงิน (บาท)", "วันครบกำหนด", "เลขที่เอกสาร", "หมายเหตุ"]

    # ── สไตล์ ──
    green       = "16A34A"
    green_light = "DCFCE7"
    thin        = Side(style="thin", color="CBD5E1")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_font   = Font(name="Tahoma", size=11, bold=True, color="FFFFFF")
    cell_font   = Font(name="Tahoma", size=10, color="0F172A")
    head_fill   = PatternFill("solid", fgColor=green)
    alt_fill    = PatternFill("solid", fgColor="F8FAFC")
    total_fill  = PatternFill("solid", fgColor=green_light)
    center      = Alignment(horizontal="center", vertical="center")
    left        = Alignment(horizontal="left",   vertical="center")
    right       = Alignment(horizontal="right",  vertical="center")

    # ── หัวข้อบริษัท ──
    company = cfg.get("company_name", "")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1, value=f"รายการโอนเงิน — {company}")
    c.font = Font(name="Tahoma", size=14, bold=True, color="15803D")
    c.alignment = left
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    c = ws.cell(row=2, column=1, value=f"วันที่ออกเอกสาร: {fmt_date(today_str())}")
    c.font = Font(name="Tahoma", size=10, color="64748B")
    c.alignment = left

    # ── หัวตาราง ──
    head_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=head_row, column=col, value=h)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = center
        cell.border = border

    # ── ข้อมูล ──
    total = 0.0
    r = head_row
    for i, exp in enumerate(expenses, 1):
        r += 1
        amt = _amount(exp)
        total += amt
        row_vals = [
            i,
            _vendor_name(exp),
            exp.get("_brand_name") or exp.get("brandCode") or "",
            _vendor_account(exp),
            amt,
            fmt_date(exp.get("dueDate") or exp.get("due_date") or ""),
            _doc_serial(exp),
            exp.get("remarks") or exp.get("projectName") or exp.get("note") or "",
        ]
        for col, val in enumerate(row_vals, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = cell_font
            cell.border = border
            if col == 1:
                cell.alignment = center
            elif col == 5:
                cell.alignment = right
                cell.number_format = "#,##0.00"
            elif col == 6:
                cell.alignment = center
            else:
                cell.alignment = left
            if i % 2 == 0:
                cell.fill = alt_fill

    # ── แถวรวมยอด ──
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(row=r, column=1, value=f"รวม {len(expenses)} รายการ")
    c.font = Font(name="Tahoma", size=11, bold=True, color="15803D")
    c.alignment = right
    c.fill = total_fill
    c.border = border
    c = ws.cell(row=r, column=5, value=total)
    c.font = Font(name="Tahoma", size=11, bold=True, color="15803D")
    c.alignment = right
    c.number_format = "#,##0.00"
    c.fill = total_fill
    c.border = border
    for col in range(6, len(headers) + 1):
        cell = ws.cell(row=r, column=col)
        cell.fill = total_fill
        cell.border = border

    # ── ความกว้างคอลัมน์ ──
    widths = [7, 28, 14, 18, 16, 14, 16, 24]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[head_row].height = 22
    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)

    wb.save(path)


def _due_iso(e: dict) -> str:
    return (e.get("dueDate") or e.get("due_date") or "")[:10]


def _holiday_config():
    """อ่านวันหยุดจาก config (วันหยุดประจำสัปดาห์ + วันหยุดพิเศษ)"""
    try:
        from config import load_config
        c = load_config()
        weekly = set(c.get("weekly_off_days") or [6])
        hol = set(c.get("holidays") or [])
        return weekly, hol
    except Exception:
        return {6}, set()


def _is_day_off(d, weekly=None, hol=None) -> bool:
    if weekly is None or hol is None:
        weekly, hol = _holiday_config()
    return d.weekday() in weekly or d.isoformat() in hol


def _next_business_day(d):
    weekly, hol = _holiday_config()
    d = d + _td(days=1)
    while _is_day_off(d, weekly, hol):   # ข้ามวันหยุดประจำสัปดาห์ + วันหยุดพิเศษ
        d = d + _td(days=1)
    return d


def distribute_into_days(expenses: list, limit: float, start=None) -> list:
    """
    กระจาย expense ลงแต่ละวันแบบ best-fit (เต็มวงเงิน/วัน, ข้ามเสาร์-อาทิตย์,
    เกินกำหนดก่อน) — คืน list ของ {'date': date, 'items': [exp]}
    """
    limit = float(limit or 0)
    today_iso = date.today().isoformat()

    def priority(e):
        name = (_vendor_name(e) or "").replace(" ", "")
        if "พชร" in name and "รัชนาทสกุล" in name:
            return (-1, "")          # นายพชร รัชนาทสกุล → จัดคิวก่อนเสมอ
        due = _due_iso(e)
        if due and due < today_iso:
            return (0, due)
        elif due:
            return (1, due)
        return (2, "9999-12-31")

    remaining = sorted([e for e in expenses if _amount(e) > 0], key=priority)
    groups = []
    while remaining:
        day, total, rest = [], 0.0, []
        for e in remaining:
            amt = _amount(e)
            if limit > 0 and day and (total + amt) > limit:
                rest.append(e)
                continue
            day.append(e)
            total += amt
        groups.append(day)
        remaining = rest

    return assign_dates(groups, start)


def assign_dates(groups: list, start=None) -> list:
    """ใส่วันที่ให้แต่ละกลุ่ม (ข้ามวันหยุด) — คืน [{'date': date, 'items': [...]}]"""
    result, d = [], (start or date.today())
    while _is_day_off(d):              # วันแรกถ้าตรงวันหยุด → เลื่อนเป็นวันทำการ
        d = _next_business_day(d)
    for grp in (groups or [[]]):
        result.append({"date": d, "items": grp})
        d = _next_business_day(d)
    return result


def _exp_brand(e: dict) -> str:
    return e.get("_brand_name") or e.get("brandCode") or ""


def _exp_remark(e: dict) -> str:
    # หมายเหตุที่ผู้ใช้พิมพ์เอง (ลูกค้าขอเพิ่ม) มาก่อนเสมอ
    return (e.get("_custom_remark") or e.get("remarks")
            or e.get("projectName") or e.get("note") or "")


def _exp_link(e: dict) -> str:
    """ลิงก์เอกสารสำหรับ Export (ข้อ 1)
    ใช้ลิงก์ "แชร์/เปิดดูเอกสาร" ที่ผู้ใช้วางเองก่อน (ถ้ามี) ไม่งั้นใช้ลิงก์หน้าแก้ไข"""
    shared = (e.get("_share_url") or "").strip()
    if shared:
        return shared
    sc  = e.get("_support_code") or ""
    rid = e.get("recordId") or e.get("documentId") or ""
    if sc and rid:
        return f"https://advance.flowaccount.com/{sc}/business/expenses/{rid}"
    return ""


def _exp_edit_link(e: dict) -> str:
    """ลิงก์หน้า 'แก้ไขเอกสาร' ใน FlowAccount (เปิดเพื่อแก้ไขใบ) — ข้อ 5"""
    sc  = e.get("_support_code") or ""
    rid = e.get("recordId") or e.get("documentId") or ""
    if not (sc and rid):
        return ""
    seg = {"po": "purchase-orders", "gr": "purchases"}.get(
        e.get("_doctype", "expense"), "expenses")
    return f"https://advance.flowaccount.com/{sc}/business/{seg}/{rid}"


def build_bank_excel_multiday(days: list, cfg: dict, path: str) -> None:
    """สร้าง Excel จัดกลุ่มตามวันจ่าย (หลายวัน) — มีหัววัน + ยอดรวมแต่ละวัน + ยอดรวมใหญ่"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook(); ws = wb.active; ws.title = "Payment"
    try:
        ws.page_setup.orientation = "landscape"     # พิมพ์แนวนอน
        ws.page_setup.fitToWidth = 1
    except Exception:
        pass
    headers = ["ลำดับ", "วันครบกำหนด", "แบรนด์", "เลขที่เอกสาร", "ชื่อผู้รับ",
               "ชื่อโปรเจ็ค/รายละเอียด", "จำนวนเงิน (บาท)", "เลขบัญชี", "หมายเหตุ",
               "ลิงก์", "ลิงก์แก้ไขใบ"]
    AMT_COL = 7                                  # คอลัมน์ 'จำนวนเงิน'
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    thai = ["จันทร์", "อังคาร", "พุธ", "พฤหัสบดี", "ศุกร์", "เสาร์", "อาทิตย์"]

    company = cfg.get("company_name", "")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(1, 1, f"คิวจ่ายเงิน (หลายวัน) — {company}")
    c.font = Font(name="Tahoma", size=14, bold=True, color="15803D")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    c = ws.cell(2, 1, f"วันที่ออกเอกสาร: {fmt_date(today_str())}")
    c.font = Font(name="Tahoma", size=10, color="64748B")

    r = 3
    grand = 0.0
    for di, day in enumerate(days, 1):
        items = day["items"]
        if not items:
            continue
        day_total = sum(_amount(e) for e in items)
        grand += day_total
        d = day["date"]
        # หัววัน
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
        cell = ws.cell(r, 1, f"📅 วันที่ {di}  ({thai[d.weekday()]} {fmt_date(d.isoformat())})"
                             f"   —   {len(items)} รายการ   รวม {fmt_amount(day_total)} บาท")
        cell.font = Font(name="Tahoma", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="16A34A")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        # หัวตาราง
        r += 1
        for col, h in enumerate(headers, 1):
            hc = ws.cell(r, col, h)
            hc.font = Font(name="Tahoma", size=10, bold=True, color="15803D")
            hc.fill = PatternFill("solid", fgColor="DCFCE7")
            hc.alignment = Alignment(horizontal="center", vertical="center")
            hc.border = border
        # รายการ
        for i, e in enumerate(items, 1):
            r += 1
            row_vals = [i,
                        fmt_date(e.get("dueDate") or e.get("due_date") or ""),
                        _exp_brand(e), _doc_serial(e), _vendor_name(e),
                        _exp_project(e), _amount(e), _vendor_account(e),
                        _exp_remark(e), "", ""]
            for col, val in enumerate(row_vals, 1):
                cc = ws.cell(r, col, val)
                cc.font = Font(name="Tahoma", size=10, color="0F172A")
                cc.border = border
                if col == 1:
                    cc.alignment = Alignment(horizontal="center", vertical="center")
                elif col == AMT_COL:
                    cc.alignment = Alignment(horizontal="right", vertical="center")
                    cc.number_format = "#,##0.00"
                else:
                    cc.alignment = Alignment(horizontal="left", vertical="center")
            # คอลัมน์ลิงก์ — เปิดดูเอกสาร (ข้อ 1) + ลิงก์แก้ไขใบ (ข้อ 5)
            link = _exp_link(e)
            if link:
                lc = ws.cell(r, len(headers) - 1)
                lc.value = "เปิดดูเอกสาร"
                lc.hyperlink = link
                lc.font = Font(name="Tahoma", size=10, color="2563EB", underline="single")
            edit = _exp_edit_link(e)
            if edit:
                ec = ws.cell(r, len(headers))
                ec.value = "แก้ไขใบ"
                ec.hyperlink = edit
                ec.font = Font(name="Tahoma", size=10, color="DC2626", underline="single")
        # ยอดรวมวัน
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=AMT_COL - 1)
        cc = ws.cell(r, 1, f"รวมวันที่ {di}")
        cc.font = Font(name="Tahoma", size=10, bold=True, color="15803D")
        cc.alignment = Alignment(horizontal="right", vertical="center")
        cc = ws.cell(r, AMT_COL, day_total)
        cc.font = Font(name="Tahoma", size=10, bold=True, color="15803D")
        cc.number_format = "#,##0.00"
        cc.alignment = Alignment(horizontal="right", vertical="center")
        r += 1   # เว้นบรรทัด

    # ยอดรวมใหญ่
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=AMT_COL - 1)
    cc = ws.cell(r, 1, f"รวมทั้งหมด {sum(len(d['items']) for d in days)} รายการ ({len([d for d in days if d['items']])} วัน)")
    cc.font = Font(name="Tahoma", size=12, bold=True, color="15803D")
    cc.fill = PatternFill("solid", fgColor="DCFCE7")
    cc.alignment = Alignment(horizontal="right", vertical="center")
    cc = ws.cell(r, AMT_COL, grand)
    cc.font = Font(name="Tahoma", size=12, bold=True, color="15803D")
    cc.fill = PatternFill("solid", fgColor="DCFCE7")
    cc.number_format = "#,##0.00"
    cc.alignment = Alignment(horizontal="right", vertical="center")

    # ความกว้างคอลัมน์ (ตามลำดับใหม่ 11 คอลัมน์)
    for col, w in enumerate([7, 14, 14, 18, 30, 30, 16, 20, 30, 14, 14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    wb.save(path)


def build_bank_pdf_multiday(days: list, cfg: dict, path: str) -> None:
    """สร้าง PDF จัดกลุ่มตามวันจ่าย แนวนอน (ภาษาไทย ใช้ฟอนต์ Tahoma จากระบบ)"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    import os as _os

    # ลงทะเบียนฟอนต์ไทยจากระบบ
    fname = "THSarabun"
    for fp in [r"C:\Windows\Fonts\tahoma.ttf", r"C:\Windows\Fonts\leelawui.ttf",
               r"C:\Windows\Fonts\angsau.ttf"]:
        if _os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont(fname, fp)); break
            except Exception:
                pass
    else:
        fname = "Helvetica"

    thai = ["จันทร์", "อังคาร", "พุธ", "พฤหัสบดี", "ศุกร์", "เสาร์", "อาทิตย์"]
    styles = getSampleStyleSheet()
    title_st = ParagraphStyle("t", parent=styles["Title"], fontName=fname,
                              fontSize=15, textColor=colors.HexColor("#15803d"))
    sub_st   = ParagraphStyle("s", parent=styles["Normal"], fontName=fname,
                              fontSize=9, textColor=colors.HexColor("#64748b"))
    day_st   = ParagraphStyle("d", parent=styles["Normal"], fontName=fname,
                              fontSize=11, textColor=colors.white)
    cell_st  = ParagraphStyle("c", parent=styles["Normal"], fontName=fname, fontSize=8)
    cellr_st = ParagraphStyle("cr", parent=styles["Normal"], fontName=fname,
                              fontSize=8, alignment=2)
    link_st  = ParagraphStyle("lk", parent=styles["Normal"], fontName=fname,
                              fontSize=8, alignment=1)   # กึ่งกลาง

    # แนวนอน (landscape A4) — ความกว้างใช้งานจริง ~277mm
    doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                            leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=10*mm, bottomMargin=10*mm)
    elems = [Paragraph(f"คิวจ่ายเงิน (หลายวัน) — {cfg.get('company_name','')}", title_st),
             Paragraph(f"วันที่ออกเอกสาร: {fmt_date(today_str())}", sub_st),
             Spacer(1, 6)]

    headers = ["#", "วันครบกำหนด", "แบรนด์", "เลขที่เอกสาร", "ชื่อผู้รับ",
               "ชื่อโปรเจ็ค/รายละเอียด", "จำนวนเงิน", "เลขบัญชี", "หมายเหตุ",
               "ลิงก์", "แก้ไขใบ"]
    col_w = [8*mm, 20*mm, 20*mm, 24*mm, 44*mm, 44*mm, 22*mm, 26*mm, 33*mm, 15*mm, 15*mm]
    # ให้เลขลำดับอยู่กึ่งกลาง ไม่ตัดบรรทัด (เลข 2 หลักขึ้นไปอยู่บรรทัดเดียว)
    num_st = ParagraphStyle("num", parent=cell_st, alignment=1)
    grand = 0.0

    for di, day in enumerate(days, 1):
        items = day["items"]
        if not items:
            continue
        day_total = sum(_amount(e) for e in items)
        grand += day_total
        d = day["date"]
        hdr = Table([[Paragraph(
            f"วันที่ {di}  ({thai[d.weekday()]} {fmt_date(d.isoformat())})  —  "
            f"{len(items)} รายการ   รวม {fmt_amount(day_total)} บาท", day_st)]],
            colWidths=[sum(col_w)])
        hdr.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#16a34a")),
                                 ("LEFTPADDING", (0, 0), (-1, -1), 6),
                                 ("TOPPADDING", (0, 0), (-1, -1), 4),
                                 ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
        elems.append(hdr)

        data = [[Paragraph(h, link_st if h in ("ลิงก์", "แก้ไขใบ") else cell_st)
                 for h in headers]]
        for i, e in enumerate(items, 1):
            link = _exp_link(e)
            link_para = (Paragraph(f'<link href="{link}"><font color="#2563eb"><u>เปิดดู</u></font></link>',
                                   link_st)
                         if link else Paragraph("-", link_st))
            edit = _exp_edit_link(e)
            edit_para = (Paragraph(f'<link href="{edit}"><font color="#dc2626"><u>แก้ไข</u></font></link>',
                                   link_st)
                         if edit else Paragraph("-", link_st))
            data.append([
                Paragraph(str(i), num_st),
                Paragraph(fmt_date(e.get("dueDate") or e.get("due_date") or ""), cell_st),
                Paragraph(_exp_brand(e), cell_st),
                Paragraph(_doc_serial(e), cell_st),
                Paragraph(_vendor_name(e), cell_st),
                Paragraph(_exp_project(e), cell_st),
                Paragraph(fmt_amount(_amount(e)), cellr_st),
                Paragraph(_vendor_account(e), cell_st),
                Paragraph(_exp_remark(e), cell_st),
                link_para,
                edit_para,
            ])
        data.append([Paragraph("รวมวันนี้", cellr_st), "", "", "", "", "",
                     Paragraph(fmt_amount(day_total), cellr_st), "", "", "", ""])
        t = Table(data, colWidths=col_w, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dcfce7")),
            ("SPAN", (0, -1), (5, -1)),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#f1f5f9")),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            # ลดระยะขอบในเซลล์ → มีที่พอ เลขลำดับ/ข้อความไม่ตัดบรรทัดง่าย
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ]))
        elems.append(t)
        elems.append(Spacer(1, 8))

    elems.append(Paragraph(
        f"รวมทั้งหมด {sum(len(d['items']) for d in days)} รายการ  •  "
        f"{fmt_amount(grand)} บาท", title_st))
    doc.build(elems)


_GH_HEADERS = {"Accept": "application/vnd.github+json", "User-Agent": "KCash-Queue"}


def upload_html_github(html: str, token: str) -> str:
    """อัป HTML เป็น secret gist บน GitHub → คืนลิงก์ถาวร (ไม่หมดอายุ) เปิดบนมือถือได้
    ต้องมี token ที่มีสิทธิ์ 'gist'. render ผ่าน githack CDN (เร็ว)
    """
    import requests
    h = dict(_GH_HEADERS); h["Authorization"] = f"token {token}"
    r = requests.post(
        "https://api.github.com/gists",
        headers=h,
        json={"public": False,
              "description": "KCash Queue",
              "files": {"kcash_queue.html": {"content": html}}},
        timeout=40,
    )
    if r.status_code not in (200, 201):
        raise RuntimeError(f"GitHub {r.status_code}: {r.text[:160]}")
    data = r.json()
    owner = (data.get("owner") or {}).get("login", "")
    gid = data["id"]
    # statically.io = CDN render gist เร็ว ไม่มีหน้าเตือน
    sta = f"https://cdn.statically.io/gist/{owner}/{gid}/raw/kcash_queue.html"
    try:
        t = requests.get(sta, timeout=8, headers={"User-Agent": "KCash"})
        if t.ok and "<html" in t.text.lower():
            return sta
    except Exception:
        pass
    # statically ล่ม → ใช้ githack แทน (ยังเปิดได้ แต่มีหน้าเตือนรอบแรก)
    return f"https://gistcdn.githack.com/{owner}/{gid}/raw/kcash_queue.html"


def delete_all_kcash_gists(token: str) -> int:
    """ลบ gist ของ KCash (description == 'KCash Queue') ทั้งหมด → คืนจำนวนที่ลบ
    ลิงก์ Export Link เก่าทั้งหมดจะใช้งานไม่ได้หลังลบ
    """
    import requests
    h = dict(_GH_HEADERS); h["Authorization"] = f"token {token}"
    deleted = 0
    page = 1
    while True:
        r = requests.get(f"https://api.github.com/gists?per_page=100&page={page}",
                         headers=h, timeout=30)
        if not r.ok:
            break
        items = r.json()
        if not items:
            break
        for g in items:
            if (g.get("description") or "") == "KCash Queue":
                d = requests.delete(f"https://api.github.com/gists/{g['id']}",
                                    headers=h, timeout=20)
                if d.status_code == 204:
                    deleted += 1
        if len(items) < 100:
            break
        page += 1
    return deleted


def upload_html_temp(html: str, hours: str = "72h") -> str:
    """อัป HTML ขึ้น litterbox (ฝากชั่วคราว ฟรี) → คืนลิงก์สุ่มที่หมดอายุเอง
    hours: '1h' | '12h' | '24h' | '72h'  (สูงสุด 72 ชม.)
    เปิดบนมือถือได้ (render เป็นหน้าเว็บ) ลิงก์สุ่มเดาไม่ได้
    """
    import requests
    r = requests.post(
        "https://litterbox.catbox.moe/resources/internals/api.php",
        data={"reqtype": "fileupload", "time": hours},
        files={"fileToUpload": ("kcash_queue.html", html.encode("utf-8"), "text/html")},
        headers={"User-Agent": "Mozilla/5.0"},
        timeout=60,
    )
    r.raise_for_status()
    url = r.text.strip()
    if not url.startswith("http"):
        raise RuntimeError(url or "อัปโหลดไม่สำเร็จ")
    return url


def build_simple_table_pdf(path: str, title: str, headers: list, rows: list) -> None:
    """สร้าง PDF ตารางง่าย ๆ (ภาษาไทย) จาก headers + rows — ใช้ทั่วไป"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import ParagraphStyle
    import os as _os
    fname = "THSarabun"
    for fp in [r"C:\Windows\Fonts\tahoma.ttf", r"C:\Windows\Fonts\leelawui.ttf"]:
        if _os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont(fname, fp)); break
            except Exception:
                pass
    else:
        fname = "Helvetica"
    title_st = ParagraphStyle("t", fontName=fname, fontSize=14, spaceAfter=6)
    cell_st = ParagraphStyle("c", fontName=fname, fontSize=8)
    doc = SimpleDocTemplate(path, pagesize=landscape(A4),
                            leftMargin=10*mm, rightMargin=10*mm,
                            topMargin=10*mm, bottomMargin=10*mm)
    data = [[Paragraph(str(h), cell_st) for h in headers]]
    for r in rows:
        data.append([Paragraph(str(c), cell_st) for c in r])
    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dcfce7")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTNAME", (0, 0), (-1, -1), fname),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    doc.build([Paragraph(title, title_st), Spacer(1, 4), t])


def build_bank_html_multiday(days: list, cfg: dict) -> str:
    """สร้างหน้า HTML (เหมือน PDF) สำหรับเปิดในเบราว์เซอร์ — คืน HTML string"""
    import html as _html

    thai = ["จันทร์", "อังคาร", "พุธ", "พฤหัสบดี", "ศุกร์", "เสาร์", "อาทิตย์"]
    company = _html.escape(cfg.get("company_name", ""))
    headers = ["#", "วันครบกำหนด", "แบรนด์", "เลขที่เอกสาร", "ชื่อผู้รับ",
               "ชื่อโปรเจ็ค/รายละเอียด", "จำนวนเงิน", "เลขบัญชี", "หมายเหตุ",
               "ลิงก์", "ลิงก์แก้ไขใบ"]

    def esc(x):
        return _html.escape(str(x or ""))

    parts = ["""<!DOCTYPE html><html lang="th"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>คิวจ่ายเงิน — KCash</title>
<style>
*{box-sizing:border-box}
body{font-family:'Segoe UI','Tahoma','Noto Sans Thai',sans-serif;margin:0;padding:24px;
     background:#f1f5f9;color:#0f172a;}
.wrap{max-width:1400px;margin:0 auto;background:white;padding:24px 28px;border-radius:10px;
      box-shadow:0 1px 6px rgba(0,0,0,.08);}
h1{color:#15803d;font-size:22px;margin:0 0 4px;}
.sub{color:#64748b;font-size:13px;margin-bottom:18px;}
.daybar{background:#16a34a;color:white;font-weight:700;font-size:14px;
        padding:8px 14px;border-radius:6px 6px 0 0;margin-top:18px;}
table{width:100%;border-collapse:collapse;font-size:13px;}
th{background:#dcfce7;color:#15803d;font-weight:700;padding:7px 8px;border:1px solid #cbd5e1;}
td{padding:6px 8px;border:1px solid #e2e8f0;vertical-align:top;}
td.num,td.amt{text-align:right;white-space:nowrap;}
td.ctr{text-align:center;white-space:nowrap;}
tr:nth-child(even) td{background:#f8fafc;}
.daytotal td{background:#f1f5f9;font-weight:700;color:#15803d;}
.grand{margin-top:20px;text-align:right;font-size:16px;font-weight:800;color:#15803d;
       background:#dcfce7;padding:12px 16px;border-radius:8px;}
a.lk{color:#2563eb;text-decoration:none;font-weight:600;}
a.lk:hover{text-decoration:underline;}
@media print{body{background:white;padding:0}.wrap{box-shadow:none}}
</style></head><body><div class="wrap">"""]
    parts.append(f"<h1>คิวจ่ายเงิน (หลายวัน) — {company}</h1>")
    parts.append(f"<div class='sub'>วันที่ออกเอกสาร: {esc(fmt_date(today_str()))}</div>")

    grand = 0.0
    for di, day in enumerate(days, 1):
        items = day["items"]
        if not items:
            continue
        day_total = sum(_amount(e) for e in items)
        grand += day_total
        d = day["date"]
        parts.append(
            f"<div class='daybar'>วันที่ {di} ({thai[d.weekday()]} {esc(fmt_date(d.isoformat()))}) "
            f"— {len(items)} รายการ &nbsp;รวม {esc(fmt_amount(day_total))} บาท</div>")
        parts.append("<table><thead><tr>"
                     + "".join(f"<th>{esc(h)}</th>" for h in headers)
                     + "</tr></thead><tbody>")
        for i, e in enumerate(items, 1):
            link = _exp_link(e)
            link_html = (f"<a class='lk' href='{esc(link)}' target='_blank'>เปิดดู</a>"
                         if link else "-")
            edit = _exp_edit_link(e)
            edit_html = (f"<a class='lk' style='color:#dc2626' href='{esc(edit)}' "
                         f"target='_blank'>แก้ไขใบ</a>" if edit else "-")
            parts.append(
                "<tr>"
                f"<td class='ctr'>{i}</td>"
                f"<td class='ctr'>{esc(fmt_date(e.get('dueDate') or e.get('due_date') or ''))}</td>"
                f"<td>{esc(_exp_brand(e))}</td>"
                f"<td>{esc(_doc_serial(e))}</td>"
                f"<td>{esc(_vendor_name(e))}</td>"
                f"<td>{esc(_exp_project(e))}</td>"
                f"<td class='amt'>{esc(fmt_amount(_amount(e)))}</td>"
                f"<td>{esc(_vendor_account(e))}</td>"
                f"<td>{esc(_exp_remark(e))}</td>"
                f"<td class='ctr'>{link_html}</td>"
                f"<td class='ctr'>{edit_html}</td>"
                "</tr>")
        parts.append(
            f"<tr class='daytotal'><td colspan='6' style='text-align:right'>รวมวันที่ {di}</td>"
            f"<td class='amt'>{esc(fmt_amount(day_total))}</td><td colspan='4'></td></tr>")
        parts.append("</tbody></table>")

    n_items = sum(len(d["items"]) for d in days)
    n_days = len([d for d in days if d["items"]])
    parts.append(f"<div class='grand'>รวมทั้งหมด {n_items} รายการ ({n_days} วัน) &nbsp;•&nbsp; "
                 f"{esc(fmt_amount(grand))} บาท</div>")
    parts.append("</div></body></html>")
    return "".join(parts)


def build_log_pdf(rows: list, cols: list, title: str, path: str) -> None:
    """สร้าง PDF ของ Log (แก้ไขไม่ได้) ไว้ดูย้อนหลัง — แนวตั้ง ภาษาไทย
    rows: list of dict, cols: [(key, header), ...]
    """
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    import os as _os

    fname = "THSarabun"
    for fp in [r"C:\Windows\Fonts\tahoma.ttf", r"C:\Windows\Fonts\leelawui.ttf"]:
        if _os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont(fname, fp)); break
            except Exception:
                pass
    else:
        fname = "Helvetica"

    styles = getSampleStyleSheet()
    title_st = ParagraphStyle("t", parent=styles["Title"], fontName=fname,
                              fontSize=15, textColor=colors.HexColor("#15803d"))
    sub_st = ParagraphStyle("s", parent=styles["Normal"], fontName=fname,
                            fontSize=9, textColor=colors.HexColor("#64748b"))
    cell_st = ParagraphStyle("c", parent=styles["Normal"], fontName=fname, fontSize=8)
    head_st = ParagraphStyle("h", parent=styles["Normal"], fontName=fname,
                             fontSize=8, textColor=colors.HexColor("#15803d"))

    doc = SimpleDocTemplate(path, pagesize=A4,
                            leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=10 * mm, bottomMargin=10 * mm)
    elems = [Paragraph(title, title_st),
             Paragraph(f"พิมพ์เมื่อ {fmt_date(today_str())} • {len(rows)} รายการ", sub_st),
             Spacer(1, 6)]

    headers = [c[1] for c in cols]
    keys = [c[0] for c in cols]
    # กว้างคอลัมน์: คอลัมน์สุดท้าย (รายละเอียด) กว้างสุด
    n = len(cols)
    avail = 190
    base_w = 26
    last_w = avail - base_w * (n - 1)
    col_w = [base_w * mm] * (n - 1) + [max(40, last_w) * mm]

    data = [[Paragraph(h, head_st) for h in headers]]
    for r in rows:
        line = []
        for k in keys:
            v = r.get(k, "")
            if k == "time" and v:
                v = str(v).replace("T", " ")
            line.append(Paragraph(str(v), cell_st))
        data.append(line)

    t = Table(data, colWidths=col_w, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dcfce7")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
    ]))
    elems.append(t)
    doc.build(elems)


def _vendor_name(exp: dict) -> str:
    return (
        exp.get("contactName")
        or exp.get("vendorName")
        or exp.get("contactCode")
        or exp.get("supplierName")
        or ""
    )


def _vendor_account(exp: dict) -> str:
    """ดึงเลขบัญชีธนาคารผู้รับ — ไล่หาหลายทาง:
    1) field ตรง (contactBankAccount ฯลฯ)
    2) เลขบัญชีที่ระบบเคยจำจากสลิป (account_memory) ตามชื่อผู้จำหน่าย
    3) ดึงเลขบัญชีจากข้อความ 'หมายเหตุ' (บัญชีมักพิมพ์ไว้ที่นี่)"""
    direct = (exp.get("contactBankAccount") or exp.get("bankAccount")
              or exp.get("bankAccountNumber") or "")
    if direct:
        return str(direct)
    # 2) reverse lookup จากที่ระบบเรียนรู้เลขบัญชี (จากการตัดบิล/จับคู่สลิป)
    try:
        import account_memory
        acct = account_memory.lookup_account_by_name(_vendor_name(exp))
        if acct:
            return acct
    except Exception:
        pass
    # 3) ดึงจากหมายเหตุ/โน้ต — เลขบัญชีมักถูกพิมพ์ไว้ (เช่น '035-110-2438', '409 052 0705')
    text = " ".join(str(exp.get(k) or "") for k in
                    ("_custom_remark", "remarks", "internalNotes", "note"))
    for m in re.findall(r"\d[\d\s\-]{7,}\d", text):
        digits = re.sub(r"\D", "", m)
        if 9 <= len(digits) <= 13:            # เลขบัญชีไทยมัก 10 หลัก (กันเลขภาษี 13/โทร)
            return re.sub(r"\s+", "", m).strip("-")
    return ""


def _exp_project(exp: dict) -> str:
    """ชื่อโปรเจ็ค + รายละเอียด (สำหรับคอลัมน์ 'ชื่อโปรเจ็ค/รายละเอียด')"""
    proj = str(exp.get("projectName") or "").strip()
    detail = ""
    items = exp.get("items") or []
    if items:
        it = items[0]
        detail = str(it.get("description") or it.get("nameLocal")
                     or it.get("nameForeign") or "").strip()
    if not detail:
        detail = str(exp.get("_detail") or exp.get("detail") or "").strip()
    parts = [p for p in (proj, detail) if p]
    return " / ".join(parts)


def _amount(exp: dict) -> float:
    for key in ("grandTotal", "totalAmount", "total", "amount", "amountDue"):
        val = exp.get(key)
        if val is not None:
            try:
                return float(val)
            except (TypeError, ValueError):
                pass
    return 0.0


def _doc_serial(exp: dict) -> str:
    return (exp.get("documentSerial") or exp.get("documentNumber")
            or exp.get("referenceNumber") or str(exp.get("recordId") or exp.get("documentId") or ""))


# ───────────────────── Line message ─────────────────────

def build_line_message(paid_expenses: list, payment_date: str, cfg: dict) -> str:
    company = cfg.get("company_name", "บริษัท")
    lines = [
        f"✅ สรุปการจ่ายเงิน — {company}",
        f"📅 วันที่จ่าย: {fmt_date(payment_date)}",
        f"📦 จำนวนรายการ: {len(paid_expenses)} รายการ",
        "",
    ]

    total = 0.0
    for exp in paid_expenses:
        amt = _amount(exp)
        total += amt
        vendor = _vendor_name(exp) or "(ไม่ระบุ)"
        doc_no = _doc_serial(exp)
        due = fmt_date(exp.get("dueDate") or exp.get("due_date") or "")
        lines.append(f"• {vendor}  |  {fmt_amount(amt)} บาท  |  ครบ {due}  |  #{doc_no}")

    lines += [
        "",
        f"💰 รวมทั้งสิ้น: {fmt_amount(total)} บาท",
        "—" * 30,
    ]
    return "\n".join(lines)


# ───────────────────── Statement parser ─────────────────────

def parse_statement_csv(content: str) -> list[dict]:
    """
    รับ CSV statement จากธนาคาร (SCB / KTB / BBL / Kbank).
    ลอง detect column ให้อัตโนมัติ.
    Returns list of { date, description, debit, credit, balance }
    """
    reader = csv.reader(io.StringIO(content))
    rows = list(reader)

    # หา header row
    header_idx = 0
    for i, row in enumerate(rows):
        joined = " ".join(row).lower()
        if any(k in joined for k in ["date", "วัน", "debit", "credit", "withdraw", "deposit", "รายการ"]):
            header_idx = i
            break

    headers = [h.strip().lower() for h in rows[header_idx]]

    col = {
        "date": _find_col(headers, ["date", "วันที่", "transaction date", "txn date"]),
        "desc": _find_col(headers, ["description", "รายการ", "detail", "particulars", "narrative"]),
        "debit": _find_col(headers, ["debit", "withdraw", "ถอน", "จ่าย", "payment"]),
        "credit": _find_col(headers, ["credit", "deposit", "ฝาก", "รับ", "receipt"]),
        "balance": _find_col(headers, ["balance", "ยอดคงเหลือ", "คงเหลือ"]),
        "amount": _find_col(headers, ["amount", "จำนวน", "ยอดเงิน"]),
    }

    result = []
    for row in rows[header_idx + 1:]:
        if len(row) < 2:
            continue
        get = lambda c: row[c].strip() if c is not None and c < len(row) else ""
        debit = _parse_number(get(col["debit"]) or get(col["amount"]))
        credit = _parse_number(get(col["credit"]))
        result.append({
            "date": get(col["date"]),
            "description": get(col["desc"]),
            "debit": debit,
            "credit": credit,
            "balance": _parse_number(get(col["balance"])),
            "raw": row,
        })

    return [r for r in result if r["description"] or r["debit"] or r["credit"]]


def _find_col(headers: list, candidates: list) -> int | None:
    for c in candidates:
        for i, h in enumerate(headers):
            if c in h:
                return i
    return None


def _parse_number(s: str) -> float:
    s = re.sub(r"[,\s฿$]", "", s or "")
    try:
        return float(s)
    except ValueError:
        return 0.0


# ───────────────────── Auto-match ─────────────────────

def match_statements(statements: list[dict], expenses: list[dict]) -> dict:
    """
    จับคู่ statement กับ expense โดยเปรียบเทียบ:
    1. ยอดเงินตรง (debit = totalAmount)
    2. ชื่อ vendor คล้ายกัน
    Returns: { expense_id: statement_row }
    """
    matched = {}
    used_stmt_indices = set()

    for exp in expenses:
        exp_id = str(exp.get("recordId") or exp.get("documentId") or exp.get("id") or "")
        exp_amount = _amount(exp)
        vendor = _vendor_name(exp).lower()

        best_score = 0.0
        best_idx = None

        for i, stmt in enumerate(statements):
            if i in used_stmt_indices:
                continue
            stmt_amount = stmt["debit"] or stmt["credit"]
            if stmt_amount == 0:
                continue

            # score: amount match (0-0.6) + name similarity (0-0.4)
            amount_score = 0.6 if abs(stmt_amount - exp_amount) < 1.0 else 0.0
            name_score = 0.0
            if vendor and stmt["description"]:
                name_score = 0.4 * SequenceMatcher(
                    None, vendor, stmt["description"].lower()
                ).ratio()

            score = amount_score + name_score
            if score > best_score and score >= 0.5:
                best_score = score
                best_idx = i

        if best_idx is not None:
            matched[exp_id] = statements[best_idx]
            used_stmt_indices.add(best_idx)

    return matched
