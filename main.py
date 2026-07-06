"""
KCash Queue System — PyQt6 Desktop App
รันด้วย: python main.py
Build .exe: pyinstaller kcash.spec
"""
import sys
import os
import re
import csv
import io
import json
import hashlib
import subprocess
from datetime import date, datetime

try:
    import psutil
except Exception:
    psutil = None

from PyQt6.QtCore import (
    Qt, QThread, pyqtSignal, QDate, QLocale, QTimer, QSortFilterProxyModel, QUrl,
)
from PyQt6.QtGui import (
    QColor, QFont, QAction, QBrush, QPalette, QIcon, QPixmap, QPainter, QPen,
    QPainterPath, QDesktopServices,
)
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QTableWidget, QTableWidgetItem, QCheckBox, QPushButton,
    QLabel, QComboBox, QLineEdit, QTextEdit, QFileDialog, QMessageBox,
    QDialog, QDialogButtonBox, QDateEdit, QGroupBox, QFormLayout,
    QHeaderView, QAbstractItemView, QStatusBar, QToolBar, QSplitter,
    QProgressBar, QFrame, QScrollArea, QSizePolicy,
    QTreeWidget, QTreeWidgetItem, QSpinBox, QPlainTextEdit,
    QListWidget, QListWidgetItem, QInputDialog,
)

from config import load_config, save_config, set_active_company, CONFIG_FILE
from flowaccount import (FlowAccountError, get_all_expenses, mark_expense_paid,
                          get_all_purchase_orders, convert_po_to_expense, upload_po_attachment,
                          get_all_purchases, get_share_link,
                          delete_expense, delete_document, set_purchase_order_status,
                          find_linked_po, PO_STATUS_APPROVED)
from utils import (build_bank_csv, build_bank_excel, build_line_message,
                   parse_statement_csv, match_statements, fmt_date, fmt_amount, today_str,
                   distribute_into_days, build_bank_excel_multiday, build_bank_pdf_multiday,
                   build_bank_html_multiday, upload_html_temp, upload_html_github,
                   delete_all_kcash_gists, assign_dates, _exp_edit_link)
from manual_expenses import load_manual, add_manual, delete_manual, update_manual, is_manual
from brand_assignments import load_assignments as load_brands, save_assignments as save_brands, get_brand, set_brand
from bank_slip import parse_payment_report, match_slip_to_expenses, _exp_serial as _slip_exp_serial
import activity_log
import paid_pending
import queue_plan
import rejected
import users
import remarks
import share_links


def _resource_path(rel_path: str) -> str:
    """หา path ของไฟล์ resource — ใช้ได้ทั้ง .py และ .exe ที่ build ด้วย PyInstaller"""
    import sys, os
    if getattr(sys, "frozen", False):
        base = getattr(sys, "_MEIPASS", os.path.dirname(sys.executable))
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, rel_path)

# ─────────────────────────── colours ────────────────────────────
C_PRIMARY   = "#2563eb"
C_SUCCESS   = "#16a34a"
C_DANGER    = "#dc2626"
C_WARNING   = "#d97706"
C_PURPLE    = "#7c3aed"
C_BG        = "#f8fafc"
C_SURFACE   = "#ffffff"
C_BORDER    = "#e2e8f0"
C_MUTED     = "#64748b"
C_OVERDUE   = QColor("#fee2e2")
C_DUE_SOON  = QColor("#fef9c3")
C_MATCHED   = QColor("#dcfce7")

# ─────────────── Dark Mode (อ่านตอนเริ่มโปรแกรม) ───────────────
# โปรแกรมกำหนดสีอ่อน inline ไว้หลายร้อยจุด → ดัก setStyleSheet แล้วแปลงสีอ่อน→เข้ม
# อัตโนมัติ ทำให้เป็นโหมดมืดทั้งโปรแกรมโดยไม่ต้องแก้ทีละจุด (ใช้ตอน build UI)
try:
    IS_DARK = bool(load_config().get("dark_mode", False))
except Exception:
    IS_DARK = False

# สถานะธีมปัจจุบัน (สลับสดได้ ไม่ต้องรีโปรแกรม)
_THEME = {"dark": IS_DARK}
_LIGHT_ROWS = {"o": "#fee2e2", "d": "#fef9c3", "m": "#dcfce7"}
_DARK_ROWS  = {"o": "#3a2c2e", "d": "#37331f", "m": "#26332b"}   # จางมาก ไม่เลอะ

def _apply_row_colors(dark: bool):
    """สีพื้นแถวตาราง (QColor ที่ setBackground ใช้) — สลับตามธีม"""
    global C_OVERDUE, C_DUE_SOON, C_MATCHED
    r = _DARK_ROWS if dark else _LIGHT_ROWS
    C_OVERDUE = QColor(r["o"]); C_DUE_SOON = QColor(r["d"]); C_MATCHED = QColor(r["m"])

_apply_row_colors(IS_DARK)

# แผนที่แปลงสีอ่อน→เข้ม (โทนเทาแบบ Google/Windows) ใช้กับ stylesheet ทุกชิ้น
_DARK_MAP = {
    "#ffffff": "#2a2b2e",
    "#f8fafc": "#202124", "#f1f5f9": "#303134", "#f9fafb": "#2a2b2e",
    "#f0fdf4": "#26332b", "#ecfdf5": "#26332b",
    "#fef9c3": "#37331f", "#fef3c7": "#37331f", "#fee2e2": "#3a2c2e",
    "#dcfce7": "#26332b", "#e7f5ed": "#26332b", "#dbeafe": "#1e3a5f",
    "#fef2f2": "#3a2c2e", "#fecaca": "#5c2a2a", "#eef2ff": "#2a2b2e",
    "#ccfbf1": "#143b38", "#bbf7d0": "#1d3a26",
    "#e2e8f0": "#3c4043", "#cbd5e1": "#5f6368", "#e5e7eb": "#3c4043",
    # ตัวอักษรเข้ม (บนพื้นอ่อนเดิม) → สว่างขึ้นให้อ่านออกบนพื้นดำ
    "#0f172a": "#e8eaed", "#334155": "#bdc1c6", "#475569": "#bdc1c6",
    "#64748b": "#9aa0a6",
    "#15803d": "#4ade80", "#92400e": "#fbbf24", "#991b1b": "#f87171",
    "#4338ca": "#a5b4fc", "#0c4a6e": "#7dd3fc",
}
import re as _re_theme
_dark_pat = _re_theme.compile(
    "|".join(_re_theme.escape(k) for k in sorted(_DARK_MAP, key=len, reverse=True)),
    _re_theme.IGNORECASE)
_orig_setss = QWidget.setStyleSheet

def _dark_setss(self, ss):
    """ดักทุก setStyleSheet: เก็บต้นฉบับ(สีอ่อน) แล้วแปลงเป็นสีเข้มถ้าโหมดมืดเปิด
    → สลับธีมสดได้ โดยเรียกใหม่ด้วยต้นฉบับที่เก็บไว้"""
    try:
        self._orig_ss = ss
    except Exception:
        pass
    out = ss
    if ss and _THEME["dark"]:
        for w in ("background:white", "background: white",
                  "background-color:white", "background-color: white"):
            out = out.replace(w, w.split(":")[0] + ":#2a2b2e")
        out = _dark_pat.sub(lambda m: _DARK_MAP[m.group(0).lower()], out)
    _orig_setss(self, out)
QWidget.setStyleSheet = _dark_setss


def _amount(exp: dict) -> float:
    for k in ("grandTotal", "totalAmount", "total", "amount", "amountDue"):
        v = exp.get(k)
        if v is not None:
            try:
                return float(v)
            except (TypeError, ValueError):
                pass
    return 0.0


def _vendor(exp: dict) -> str:
    return (exp.get("contactName") or exp.get("vendorName")
            or exp.get("contactCode") or exp.get("supplierName") or "")


def _is_vip_vendor(exp: dict) -> bool:
    """ผู้รับเงินที่ต้องจัดคิวก่อนเสมอ (นายพชร รัชนาทสกุล)"""
    name = (_vendor(exp) or "").replace(" ", "")
    return "พชร" in name and "รัชนาทสกุล" in name


def _doc_no(exp: dict) -> str:
    return (exp.get("documentSerial") or exp.get("documentNumber")
            or exp.get("referenceNumber")
            or str(exp.get("recordId") or exp.get("documentId") or ""))


def _exp_id(exp: dict) -> str:
    return str(exp.get("recordId") or exp.get("documentId") or exp.get("id") or "")


def _due(exp: dict) -> str:
    return exp.get("dueDate") or exp.get("due_date") or ""


def _status(exp: dict) -> str:
    return (exp.get("statusString") or exp.get("paymentStatus")
            or str(exp.get("status") or "")).lower()


def _brand_code(exp: dict) -> str:
    """หารหัสลูกค้า (เช่น N609850) จากหลาย field ที่ FlowAccount ใช้"""
    for k in ("contactCode", "customerCode", "vendorCode",
              "contactId", "contact_id", "customerId"):
        v = exp.get(k)
        if v: return str(v).strip()
    return ""


def _brand_name(exp: dict, assignments: dict = None) -> str:
    """หาแบรนด์ของ expense — มาจากบริษัทที่ดึงมา (อัตโนมัติ)"""
    # ป้ายแบรนด์อัตโนมัติจากบริษัทต้นทาง (โหมดหลายบริษัท)
    if exp.get("_brand_name"):
        return exp["_brand_name"]
    # manual entry → ใช้ค่าที่กรอกตอนเพิ่ม
    if exp.get("brandCode"):
        return exp["brandCode"]
    # ใช้ assignment dict ที่ส่งมาก่อน (สำหรับ batch load)
    if assignments is not None:
        return assignments.get(_exp_id(exp), "")
    # หรือโหลดจากไฟล์
    return get_brand(_exp_id(exp))


# ──────────────────── Worker threads ────────────────────

def _tag_brand(items, company):
    """ติดป้ายแบรนด์/บริษัทให้แต่ละ expense (ใช้แสดงในคอลัมน์แบรนด์ + จ่ายตามบริษัท)"""
    for e in items:
        e["_brand_name"]    = company.get("label", "")
        e["_brand_color"]   = company.get("color", "#334155")
        e["_co_id"]         = company.get("client_id", "")
        e["_co_secret"]     = company.get("client_secret", "")
        e["_support_code"]  = company.get("support_code", "")
    return items


def _payment_date(e: dict) -> str:
    """วันที่จ่ายเงิน — จาก payments ของ FlowAccount (ถ้ามีบิล/จ่ายแล้ว)"""
    p = e.get("payments")
    if isinstance(p, list) and p:
        p = p[0]
    if isinstance(p, dict):
        return (p.get("paymentDate") or "")[:10]
    return ""


def _share_link(e: dict) -> str:
    """สร้างลิงก์ไปหน้าเอกสารใน FlowAccount (ข้อมูลเพิ่มเติมของรายการ)"""
    sc  = e.get("_support_code") or ""
    rid = e.get("recordId") or e.get("documentId") or ""
    if sc and rid:
        return f"https://advance.flowaccount.com/{sc}/business/expenses/{rid}"
    return ""


def _exp_creds(e: dict) -> tuple:
    """คืน (client_id, client_secret) ของบริษัทที่เป็นเจ้าของรายการ"""
    cfg = load_config()
    cid  = e.get("_co_id")     or cfg.get("flowaccount_api_key", "")
    csec = e.get("_co_secret") or cfg.get("flowaccount_secret_key", "")
    return cid, csec


def _doc_converted_to_expense(doc: dict) -> bool:
    """เอกสาร PO/ใบรับสินค้า ที่ถูกแปลงเป็นค่าใช้จ่าย (EXP) แล้ว
    → referencedToMe มี documentType '13' (=expense)
    เอกสารพวกนี้ไม่ควรโชว์ในคิว เพราะ EXP เป็นตัวแทนสถานะจริง (เช่นจ่ายแล้ว) อยู่แล้ว
    กันโชว์ซ้ำ/สถานะดูไม่ตรง"""
    for ref in (doc.get("referencedToMe") or []):
        if str(ref.get("documentType")) == "13":
            return True
    return False


def _item_details(e: dict) -> list:
    """คืน line items (รายละเอียดเบื้องหลังยอด) — [{desc, qty, price, total}]"""
    out = []
    for it in (e.get("items") or []):
        out.append({
            "desc":  (it.get("description") or it.get("nameLocal")
                      or it.get("nameForeign") or "").strip(),
            "qty":   it.get("quantity", ""),
            "unit":  it.get("unitName", ""),
            "price": it.get("pricePerUnit", ""),
            "total": it.get("total", ""),
        })
    return out


class ShareLinkWorker(QThread):
    """ดึงลิงก์แชร์อัตโนมัติทีละรายการ (background) — requirement ลิงก์แชร์"""
    progress = pyqtSignal(int, int)   # done, total
    finished_all = pyqtSignal(int, int)   # success, fail

    def __init__(self, expenses: list, parent=None):
        super().__init__(parent)
        self.expenses = expenses
        self._stop = False

    def stop(self):
        self._stop = True

    def run(self):
        ok = fail = 0
        total = len(self.expenses)
        for i, e in enumerate(self.expenses, 1):
            if self._stop:
                break
            try:
                did = e.get("documentId") or e.get("recordId")
                cid, csec = _exp_creds(e)
                link = get_share_link(cid, csec, did, doctype=e.get("_doctype", "expense"))
                if link:
                    share_links.set(_exp_id(e), link)
                    ok += 1
                else:
                    fail += 1
            except Exception:
                fail += 1
            self.progress.emit(i, total)
        self.finished_all.emit(ok, fail)


class FetchWorker(QThread):
    done    = pyqtSignal(list)
    error   = pyqtSignal(str)

    def __init__(self, cfg: dict, company: dict = None):
        super().__init__()
        self.cfg = cfg
        self.company = company or {}

    def run(self):
        try:
            items = get_all_expenses(self.cfg["flowaccount_api_key"],
                                     self.cfg["flowaccount_secret_key"])
            if self.company:
                _tag_brand(items, self.company)
            self.done.emit(items)
        except FlowAccountError as exc:
            self.error.emit(str(exc))
        except Exception as exc:
            self.error.emit(f"เกิดข้อผิดพลาด: {exc}")


class FetchAllWorker(QThread):
    """ดึง expense จากทุกบริษัทมารวมกัน + ติดป้ายแบรนด์ตามต้นทาง"""
    done     = pyqtSignal(list)
    error    = pyqtSignal(str)
    progress = pyqtSignal(int, int, str)   # done, total, label

    def __init__(self, companies: list):
        super().__init__()
        self.companies = companies

    def run(self):
        from concurrent.futures import ThreadPoolExecutor, as_completed
        valid = [c for c in self.companies if c.get("client_id")]
        total = len(valid)
        all_items, fails = [], []
        self.progress.emit(0, total, "กำลังดึงข้อมูลทุกบริษัทพร้อมกัน...")

        def _fetch(c):
            items = get_all_expenses(c.get("client_id"), c.get("client_secret"))
            _tag_brand(items, c)
            return items

        done_n = 0
        # ดึงทุกบริษัท "พร้อมกัน" (แต่ละบริษัทก็ดึงหลายหน้าพร้อมกันอีกชั้น)
        with ThreadPoolExecutor(max_workers=min(4, max(1, total))) as ex:
            futs = {ex.submit(_fetch, c): c for c in valid}
            for f in as_completed(futs):
                c = futs[f]
                try:
                    all_items.extend(f.result())
                except Exception as exc:
                    fails.append(f"{c.get('label','')}: {exc}")
                done_n += 1
                self.progress.emit(done_n, total,
                                   f"ดึงเสร็จ {done_n}/{total} บริษัท ({len(all_items)} รายการ)")
        if fails and not all_items:
            self.error.emit("ดึงข้อมูลไม่สำเร็จ:\n" + "\n".join(fails))
        else:
            self.done.emit(all_items)


class DocFetchWorker(QThread):
    """ดึงใบสั่งซื้อ/ใบรับสินค้า (on-demand) สำหรับหน้าคิวรวม"""
    done  = pyqtSignal(list, list, bool, bool)   # pos, grs, did_po, did_gr
    error = pyqtSignal(str)

    def __init__(self, cfg, comps, combined, need_po, need_gr):
        super().__init__()
        self.cfg = cfg
        self.comps = comps or []
        self.combined = combined
        self.need_po = need_po
        self.need_gr = need_gr

    def _targets(self):
        if self.combined and self.comps:
            return [(c.get("client_id"), c.get("client_secret"), c)
                    for c in self.comps if c.get("client_id")]
        active = self.cfg.get("active_company", 0)
        company = self.comps[active] if (0 <= active < len(self.comps)) else {}
        return [(self.cfg.get("flowaccount_api_key"),
                 self.cfg.get("flowaccount_secret_key"), company)]

    def run(self):
        from concurrent.futures import ThreadPoolExecutor
        try:
            pos, grs = [], []
            # สร้าง "งานย่อย" ทุก (บริษัท × ประเภท) แล้วยิงพร้อมกันหมด
            # → PO และ GR โหลดพร้อมกัน เร็วพอ ๆ กับโหลดคิว
            tasks = []
            for cid, csec, company in self._targets():
                if not cid:
                    continue
                if self.need_po:
                    tasks.append(("po", cid, csec, company))
                if self.need_gr:
                    tasks.append(("gr", cid, csec, company))

            def _one(task):
                kind, cid, csec, company = task
                items = (get_all_purchase_orders(cid, csec) if kind == "po"
                         else get_all_purchases(cid, csec))
                if company:
                    _tag_brand(items, company)
                return kind, items

            with ThreadPoolExecutor(max_workers=min(6, max(1, len(tasks)))) as ex:
                for kind, items in ex.map(_one, tasks):
                    (pos if kind == "po" else grs).extend(items)
            self.done.emit(pos, grs, self.need_po, self.need_gr)
        except Exception as exc:
            self.error.emit(f"ดึงเอกสารไม่สำเร็จ: {exc}")


class MarkPaidWorker(QThread):
    progress = pyqtSignal(int, int)
    done     = pyqtSignal(list, list)  # ok_list, error_list

    def __init__(self, cfg: dict, expenses: list, payment_date: str, method: str):
        super().__init__()
        self.cfg          = cfg
        self.expenses     = expenses
        self.payment_date = payment_date
        self.method       = method

    def run(self):
        ok, errs = [], []
        for i, exp in enumerate(self.expenses):
            self.progress.emit(i + 1, len(self.expenses))
            # ใช้ key ของบริษัทที่ expense นั้นมาจาก (โหมดรวมทุกแบรนด์)
            cid  = exp.get("_co_id")     or self.cfg["flowaccount_api_key"]
            csec = exp.get("_co_secret") or self.cfg["flowaccount_secret_key"]
            try:
                mark_expense_paid(
                    cid, csec, _exp_id(exp),
                    {"paymentDate": self.payment_date, "paymentMethod": self.method,
                     "amount": _amount(exp)},
                )
                ok.append(exp)
            except FlowAccountError as exc:
                errs.append({"id": _exp_id(exp), "error": str(exc)})
        self.done.emit(ok, errs)


class RejectFlowAccountWorker(QThread):
    """ลบเอกสารใน FlowAccount จริง (EXP/PO/GR).
    ถ้าเป็น EXP → ลบแล้วคืนสถานะ PO ที่ผูกกัน เป็น 'อนุมัติ' ด้วย"""
    progress = pyqtSignal(int, int)
    finished_all = pyqtSignal(list, list)   # ok_ids, problems[{doc, err}]

    def __init__(self, exps, parent=None):
        super().__init__(parent)
        self.exps = exps

    def run(self):
        ok_ids, problems = [], []
        for i, e in enumerate(self.exps, 1):
            doc = _doc_no(e)
            dt = e.get("_doctype", "expense")
            cid, csec = _exp_creds(e)
            did = e.get("documentId") or e.get("recordId")
            try:
                delete_document(cid, csec, dt, did)     # 1) ลบเอกสาร
                ok_ids.append(_exp_id(e))
                # 2) ถ้าเป็น EXP → คืนสถานะ PO ที่ผูกกัน → อนุมัติ (ถ้ามี)
                if dt == "expense":
                    po_id, po_ser = find_linked_po(e)
                    if po_id:
                        try:
                            set_purchase_order_status(cid, csec, po_id, PO_STATUS_APPROVED)
                        except Exception as ex2:
                            problems.append({"doc": doc,
                                             "err": f"ลบ EXP แล้ว แต่แก้สถานะ {po_ser} ไม่ได้: {ex2}"})
            except Exception as ex:
                problems.append({"doc": doc, "err": f"ลบไม่ได้: {ex}"})
            self.progress.emit(i, len(self.exps))
        self.finished_all.emit(ok_ids, problems)


# ──────────────────── Dialogs ────────────────────

class SettingsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("⚙️ ตั้งค่า")
        self.setMinimumWidth(520)
        self.cfg = load_config()
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        # FlowAccount section
        fa_box = QGroupBox("🔑 FlowAccount API")
        fa_form = QFormLayout(fa_box)
        self.api_key    = QLineEdit(self.cfg.get("flowaccount_api_key", ""))
        self.secret_key = QLineEdit(self.cfg.get("flowaccount_secret_key", ""))
        self.secret_key.setEchoMode(QLineEdit.EchoMode.Password)
        self.ed_fetch_limit = QLineEdit(str(self.cfg.get("fetch_limit", 1000)))
        self.ed_fetch_limit.setFixedWidth(100)
        fa_form.addRow("Client ID (API Key):", self.api_key)
        fa_form.addRow("Client Secret:", self.secret_key)
        fa_form.addRow("จำนวนรายการที่ดึง (ล่าสุด):", self.ed_fetch_limit)
        fl_hint = QLabel("ดึงรายการใหม่สุดตามจำนวนนี้ (ดึงพร้อมกันหลายหน้า เร็วขึ้น) — มากไป=ช้าลง แนะนำ 1000")
        fl_hint.setStyleSheet(f"color:{C_MUTED};font-size:11px;")
        fl_hint.setWordWrap(True)
        fa_form.addRow("", fl_hint)
        layout.addWidget(fa_box)

        # Company section
        co_box = QGroupBox("🏢 ข้อมูลบริษัท")
        co_form = QFormLayout(co_box)
        self.company    = QLineEdit(self.cfg.get("company_name", ""))
        self.bank_acct  = QLineEdit(self.cfg.get("bank_account", ""))
        co_form.addRow("ชื่อบริษัท:", self.company)
        co_form.addRow("เลขบัญชีธนาคาร:", self.bank_acct)
        layout.addWidget(co_box)

        # monday.com section
        mn_box = QGroupBox("📋 monday.com Integration")
        mn_form = QFormLayout(mn_box)
        self.monday_token   = QLineEdit(self.cfg.get("monday_api_token",""))
        self.monday_token.setEchoMode(QLineEdit.EchoMode.Password)
        self.monday_token.setPlaceholderText("eyJhbGciOiJIUzI1NiJ9...")
        self.monday_board   = QLineEdit(self.cfg.get("monday_board_id",""))
        self.monday_board.setPlaceholderText("เลข Board ID จาก URL")
        sync_row = QHBoxLayout()
        self.monday_auto    = QCheckBox("Auto-sync ทุกวัน เวลา")
        self.monday_auto.setChecked(bool(self.cfg.get("monday_auto_sync")))
        self.monday_h       = QLineEdit(str(self.cfg.get("monday_sync_hour",15)))
        self.monday_h.setFixedWidth(45); self.monday_h.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.monday_m       = QLineEdit(str(self.cfg.get("monday_sync_minute",0)))
        self.monday_m.setFixedWidth(45); self.monday_m.setAlignment(Qt.AlignmentFlag.AlignCenter)
        sync_row.addWidget(self.monday_auto)
        sync_row.addWidget(self.monday_h)
        sync_row.addWidget(QLabel(":"))
        sync_row.addWidget(self.monday_m)
        sync_row.addWidget(QLabel("น."))
        sync_row.addStretch()
        mn_form.addRow("API Token:", self.monday_token)
        mn_form.addRow("Board ID:", self.monday_board)
        mn_form.addRow("", sync_row)
        layout.addWidget(mn_box)

        # Daily payment limit
        dl_box  = QGroupBox("💰 วงเงินจ่ายต่อวัน (default)")
        dl_form = QFormLayout(dl_box)
        cur = self.cfg.get("daily_payment_limit", 150000) or 0
        self.ed_daily_limit = QLineEdit(
            "ไม่จำกัด" if cur <= 0 else f"{float(cur):,.0f}"
        )
        self.ed_daily_limit.setAlignment(Qt.AlignmentFlag.AlignRight)
        dl_form.addRow("วงเงิน (บาท):", self.ed_daily_limit)
        dl_hint = QLabel("ใส่ 0 หรือ 'ไม่จำกัด' = ไม่จำกัดวงเงิน")
        dl_hint.setStyleSheet(f"color:{C_MUTED};font-size:11px;")
        dl_form.addRow("", dl_hint)
        layout.addWidget(dl_box)

        # วันหยุด (สำหรับคำนวณวันจ่าย)
        hd_box  = QGroupBox("📅 วันหยุด (ใช้คำนวณวันจ่าย)")
        hd_form = QFormLayout(hd_box)
        self.cmb_weekoff = QComboBox()
        self._weekoff_map = [
            ("หยุดอาทิตย์ (ทำงาน จ.–ส.)", [6]),
            ("หยุดเสาร์–อาทิตย์ (ทำงาน จ.–ศ.)", [5, 6]),
            ("ไม่หยุดประจำสัปดาห์", []),
        ]
        for label, _ in self._weekoff_map:
            self.cmb_weekoff.addItem(label)
        cur_off = sorted(self.cfg.get("weekly_off_days") or [6])
        for i, (_, val) in enumerate(self._weekoff_map):
            if sorted(val) == cur_off:
                self.cmb_weekoff.setCurrentIndex(i); break
        hd_form.addRow("วันหยุดประจำสัปดาห์:", self.cmb_weekoff)

        self.holiday_text = QTextEdit()
        self.holiday_text.setFixedHeight(80)
        self.holiday_text.setPlaceholderText("วันหยุดพิเศษ/นักขัตฤกษ์ — บรรทัดละ 1 วัน\n2026-04-13\n2026-04-14")
        self.holiday_text.setPlainText("\n".join(self.cfg.get("holidays") or []))
        hd_form.addRow("วันหยุดพิเศษ:", self.holiday_text)
        hd_hint = QLabel("รูปแบบวันที่: ปปปป-ดด-วว (ค.ศ.) เช่น 2026-04-13 — ระบบจะข้ามวันเหล่านี้ตอนจัดคิวจ่าย")
        hd_hint.setStyleSheet(f"color:{C_MUTED};font-size:11px;")
        hd_hint.setWordWrap(True)
        hd_form.addRow("", hd_hint)
        layout.addWidget(hd_box)

        # Brand list section
        br_box  = QGroupBox("🏷️ รายชื่อแบรนด์ (ตัวเลือกใน dropdown)")
        br_form = QVBoxLayout(br_box)
        br_help = QLabel(
            "ใส่ <b>ชื่อแบรนด์</b> ที่ต้องการให้เลือกได้ในตาราง<br>"
            "พิมพ์บรรทัดละ 1 แบรนด์ — บัญชีจะเลือกแบรนด์เองในแต่ละรายการ"
        )
        br_help.setStyleSheet(f"color:{C_MUTED}; font-size:11px;")
        br_help.setWordWrap(True)
        br_form.addWidget(br_help)
        self.brand_text = QTextEdit()
        self.brand_text.setFixedHeight(110)
        self.brand_text.setPlaceholderText("Hollywood\nD-MAT CNC\nDEKO\nYellow kitchen\nDDCUT")
        # โหลด brand_list ปัจจุบันมาแสดง
        blist = self.cfg.get("brand_list") or []
        self.brand_text.setPlainText("\n".join(blist))
        br_form.addWidget(self.brand_text)
        layout.addWidget(br_box)

        # Line section
        ln_box = QGroupBox("📱 Line Notify (ไม่บังคับ)")
        ln_form = QFormLayout(ln_box)
        self.line_token = QLineEdit(self.cfg.get("line_token", ""))
        self.line_token.setEchoMode(QLineEdit.EchoMode.Password)
        ln_form.addRow("Line Notify Token:", self.line_token)
        layout.addWidget(ln_box)

        # GitHub token — สำหรับ Export Link แบบลิงก์ถาวร (ไม่หมดอายุ)
        gh_box = QGroupBox("🔗 Export Link ถาวร (GitHub – ไม่บังคับ)")
        gh_form = QFormLayout(gh_box)
        self.github_token = QLineEdit(self.cfg.get("github_token", ""))
        self.github_token.setEchoMode(QLineEdit.EchoMode.Password)
        self.github_token.setPlaceholderText("ghp_... (สิทธิ์ gist)")
        gh_form.addRow("GitHub Token:", self.github_token)
        gh_hint = QLabel(
            "ใส่ token แล้ว ปุ่ม 'Export Link' จะได้ลิงก์ <b>ถาวร ไม่หมดอายุ</b> "
            "(เปิดบนมือถือได้)<br>ถ้าเว้นว่าง = ใช้ลิงก์ชั่วคราว 72 ชม. แทน")
        gh_hint.setStyleSheet(f"color:{C_MUTED};font-size:11px;")
        gh_hint.setWordWrap(True)
        gh_form.addRow("", gh_hint)
        layout.addWidget(gh_box)

        # Buttons
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(self._save)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _save(self):
        self.cfg["flowaccount_api_key"]    = self.api_key.text().strip()
        self.cfg["flowaccount_secret_key"] = self.secret_key.text().strip()
        try:
            self.cfg["fetch_limit"] = max(200, int(self.ed_fetch_limit.text().replace(",", "").strip()))
        except Exception:
            self.cfg["fetch_limit"] = 1000
        self.cfg["company_name"]           = self.company.text().strip()
        self.cfg["bank_account"]           = self.bank_acct.text().strip()
        self.cfg["line_token"]             = self.line_token.text().strip()
        self.cfg["github_token"]           = self.github_token.text().strip()
        self.cfg["monday_api_token"]       = self.monday_token.text().strip()
        self.cfg["monday_board_id"]        = self.monday_board.text().strip()
        self.cfg["monday_auto_sync"]       = self.monday_auto.isChecked()
        try: self.cfg["monday_sync_hour"]   = int(self.monday_h.text())
        except: self.cfg["monday_sync_hour"] = 15
        try: self.cfg["monday_sync_minute"] = int(self.monday_m.text())
        except: self.cfg["monday_sync_minute"] = 0

        # parse brand_list จาก textarea (บรรทัดละ 1 แบรนด์)
        blist = []
        for line in self.brand_text.toPlainText().splitlines():
            name = line.strip()
            if name and name not in blist:
                blist.append(name)
        self.cfg["brand_list"] = blist

        # วันหยุด
        self.cfg["weekly_off_days"] = self._weekoff_map[self.cmb_weekoff.currentIndex()][1]
        hols = []
        import re as _re
        for line in self.holiday_text.toPlainText().splitlines():
            s = line.strip()
            if not s:
                continue
            # รองรับ 2026-04-13 หรือ 13/04/2026
            m = _re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})$", s)
            if not m:
                m2 = _re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
                if m2:
                    s = f"{m2.group(3)}-{int(m2.group(2)):02d}-{int(m2.group(1)):02d}"
                    m = True
            else:
                s = f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
            if m and s not in hols:
                hols.append(s)
        self.cfg["holidays"] = hols

        # daily payment limit
        dl_txt = self.ed_daily_limit.text().strip().lower()
        if "ไม่จำกัด" in dl_txt or dl_txt in ("0", "∞", ""):
            self.cfg["daily_payment_limit"] = 0
        else:
            try:
                self.cfg["daily_payment_limit"] = max(0.0,
                    float(dl_txt.replace(",", "").replace("บาท", "").strip()))
            except Exception:
                self.cfg["daily_payment_limit"] = 150000

        save_config(self.cfg)
        self.accept()


class QueuePlanDialog(QDialog):
    """
    ตารางคิวจ่ายหลายวัน
    ─────────────────────
    กระจายรายการที่ยังไม่จ่ายลงแต่ละวันตามวงเงิน/วัน (เกินกำหนดก่อน)
    • ดูได้ว่าวันไหนจ่ายรายการไหนบ้าง ไล่ลงมาทั้งหมด
    • เลือกรายการแล้วกดปุ่มโยกขึ้น/ลง เพื่อสลับวันได้
    • กด "เลือกวันนี้ไปจ่าย" เพื่อ tick รายการของวันที่เลือกกลับไปหน้าหลัก
    """

    def __init__(self, candidates, daily_limit, assignments=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📋 ตารางคิวจ่าย (หลายวัน)")
        self.resize(940, 640)
        # ให้ย่อหน้าต่าง (minimize) ได้ — requirement ข้อ 5
        self.setWindowFlags(self.windowFlags()
                            | Qt.WindowType.WindowMinimizeButtonHint
                            | Qt.WindowType.WindowMaximizeButtonHint)
        self._limit = float(daily_limit or 0)
        self._assignments = assignments or {}
        self._selected_items = []   # ผลลัพธ์เมื่อกด "เลือกวันนี้ไปจ่าย"
        self._marked_pending = []   # exp ที่กด Mark จ่ายแล้วรออัพเดต
        self._saved_at = ""

        # โหลดการจัดเรียงที่บันทึกไว้ (ถ้ามี) ไม่งั้นกระจายอัตโนมัติ
        self._days = self._load_or_distribute(candidates)

        lay = QVBoxLayout(self)

        head = QLabel(
            f"กระจายรายการที่ยังไม่จ่ายลงแต่ละวัน — วงเงิน "
            f"{('฿{:,.0f}/วัน'.format(self._limit)) if self._limit > 0 else 'ไม่จำกัด'}\n"
            "เลือกรายการแล้วกดปุ่มลูกศรเพื่อย้ายไปวันอื่น • ดับเบิลคลิกหัววันเพื่อเลือกทั้งวันไปจ่าย")
        head.setStyleSheet("color:#15803d;font-size:13px;font-weight:600;padding:2px 0;")
        lay.addWidget(head)

        # ── toolbar ย้ายรายการ ──
        tb = QHBoxLayout()
        self.btn_up = QPushButton("⬆️ ย้ายไปวันก่อนหน้า")
        self.btn_down = QPushButton("⬇️ ย้ายไปวันถัดไป")
        for b in (self.btn_up, self.btn_down):
            b.setStyleSheet(
                "QPushButton{padding:5px 12px;border:1px solid #16a34a;border-radius:4px;"
                "background:white;color:#15803d;font-size:12px;font-weight:600;}"
                "QPushButton:hover{background:#e7f5ed;}"
                "QPushButton:disabled{color:#94a3b8;border-color:#cbd5e1;}")
        self.btn_up.setText("⬆️ ย้ายที่ติ๊กไปวันก่อนหน้า")
        self.btn_down.setText("⬇️ ย้ายที่ติ๊กไปวันถัดไป")
        self.btn_up.clicked.connect(lambda: self._move_selected(-1))
        self.btn_down.clicked.connect(lambda: self._move_selected(+1))
        tb.addWidget(self.btn_up)
        tb.addWidget(self.btn_down)
        tb.addStretch()
        self.lbl_hint = QLabel("")
        self.lbl_hint.setStyleSheet("color:#64748b;font-size:11px;")
        tb.addWidget(self.lbl_hint)
        lay.addLayout(tb)

        # ── toolbar 2: บันทึกคิว + Mark จ่ายแล้วรออัพเดต ──
        tb2 = QHBoxLayout()
        self.btn_save_queue = QPushButton("💾 บันทึกคิว")
        self.btn_save_queue.setStyleSheet(
            "QPushButton{padding:5px 14px;border:none;border-radius:4px;"
            "background:#2563eb;color:white;font-size:12px;font-weight:700;}"
            "QPushButton:hover{background:#1d4ed8;}")
        self.btn_save_queue.clicked.connect(self._save_queue)
        tb2.addWidget(self.btn_save_queue)

        self.btn_mark_pending = QPushButton("✅ Mark จ่ายแล้ว (รออัพเดต FlowAccount)")
        self.btn_mark_pending.setStyleSheet(
            "QPushButton{padding:5px 14px;border:none;border-radius:4px;"
            "background:#7c3aed;color:white;font-size:12px;font-weight:700;}"
            "QPushButton:hover{background:#6d28d9;}")
        self.btn_mark_pending.clicked.connect(self._mark_pending)
        tb2.addWidget(self.btn_mark_pending)

        # ไม่อนุมัติ (ที่ติ๊ก) + ดูรายการไม่อนุมัติ (ข้อ 2)
        self.btn_reject_p = QPushButton("🚫 ไม่อนุมัติ (ที่ติ๊ก)")
        self.btn_reject_p.setStyleSheet(
            "QPushButton{padding:5px 14px;border:none;border-radius:4px;"
            "background:#dc2626;color:white;font-size:12px;font-weight:700;}"
            "QPushButton:hover{background:#b91c1c;}")
        self.btn_reject_p.clicked.connect(self._reject_checked_plan)
        tb2.addWidget(self.btn_reject_p)

        self.lbl_saved = QLabel("")
        self.lbl_saved.setStyleSheet("color:#64748b;font-size:11px;")
        tb2.addSpacing(12)
        tb2.addWidget(self.lbl_saved)
        tb2.addStretch()
        chk_hint = QLabel("☑️ ติ๊กหน้ารายการ แล้วกด Mark")
        chk_hint.setStyleSheet("color:#7c3aed;font-size:11px;")
        tb2.addWidget(chk_hint)
        lay.addLayout(tb2)

        # ── toolbar 3: จัดคิว manual โดยระบุเลขเอกสาร (ลูกค้าขอเพิ่ม) ──
        tb3 = QHBoxLayout()
        tb3.addWidget(QLabel("🔢 จัดคิวเอง:"))
        self.ed_doc = QLineEdit()
        self.ed_doc.setPlaceholderText("พิมพ์เลขเอกสาร EXP/PO เช่น EXP038498")
        self.ed_doc.setFixedWidth(240)
        self.ed_doc.returnPressed.connect(self._manual_move_by_doc)
        tb3.addWidget(self.ed_doc)
        tb3.addWidget(QLabel("→ ใส่ไว้วันที่"))
        self.spin_day = QSpinBox()
        self.spin_day.setMinimum(1)
        self.spin_day.setMaximum(max(1, len(self._days)))
        self.spin_day.setFixedWidth(60)
        tb3.addWidget(self.spin_day)
        btn_manual = QPushButton("➕ ย้าย/ใส่ไปวันนี้")
        btn_manual.setStyleSheet(
            "QPushButton{padding:5px 14px;border:none;border-radius:4px;"
            "background:#0d9488;color:white;font-size:12px;font-weight:700;}"
            "QPushButton:hover{background:#0f766e;}")
        btn_manual.clicked.connect(self._manual_move_by_doc)
        tb3.addWidget(btn_manual)
        self.lbl_manual = QLabel("")
        self.lbl_manual.setStyleSheet("color:#64748b;font-size:11px;")
        tb3.addSpacing(8)
        tb3.addWidget(self.lbl_manual)
        tb3.addStretch()
        lay.addLayout(tb3)

        # ── แถบค้นหาในคิว (ค้นชื่อ/เลขเอกสาร แบบบางส่วน ไฮไลต์ให้เห็น) ──
        tb_s = QHBoxLayout()
        tb_s.addWidget(QLabel("🔍 ค้นหาในคิว:"))
        self.ed_search = QLineEdit()
        self.ed_search.setPlaceholderText("พิมพ์ชื่อผู้รับ หรือเลขเอกสาร (บางส่วนก็ได้ / หลายคำเว้นวรรค)")
        self.ed_search.setClearButtonEnabled(True)
        self.ed_search.setFixedWidth(360)
        self.ed_search.textChanged.connect(self._search_in_plan)
        tb_s.addWidget(self.ed_search)
        self.lbl_search = QLabel("")
        self.lbl_search.setStyleSheet("color:#2563eb;font-size:12px;font-weight:600;")
        tb_s.addSpacing(8)
        tb_s.addWidget(self.lbl_search)
        tb_s.addStretch()
        # ปุ่มรายการไม่อนุมัติ (ย้ายลงมาอยู่แถวนี้)
        self.btn_view_rejected = QPushButton("📋 รายการไม่อนุมัติ")
        self.btn_view_rejected.setStyleSheet(
            "QPushButton{padding:5px 12px;border:1px solid #dc2626;border-radius:4px;"
            "background:white;color:#dc2626;font-size:12px;font-weight:600;}"
            "QPushButton:hover{background:#fef2f2;}")
        self.btn_view_rejected.clicked.connect(self._view_rejected_plan)
        tb_s.addWidget(self.btn_view_rejected)
        lay.addLayout(tb_s)

        # ── tree ──
        self.tree = QTreeWidget()
        self.tree.setColumnCount(7)
        self.tree.setHeaderLabels(["รายการ / วันจ่าย", "เลขที่เอกสาร", "แบรนด์",
                                   "ครบกำหนด", "จำนวนเงิน (บาท)", "หมายเหตุ", "ลิงก์แชร์"])
        self.tree.setColumnWidth(0, 280)
        self.tree.setColumnWidth(1, 115)
        self.tree.setColumnWidth(2, 100)
        self.tree.setColumnWidth(3, 95)
        self.tree.setColumnWidth(4, 115)
        self.tree.setColumnWidth(5, 150)
        self.tree.header().setStretchLastSection(True)
        self.tree.setAlternatingRowColors(True)
        self.tree.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tree.setStyleSheet(
            "QTreeWidget{border:1px solid #e2e8f0;font-size:13px;}"
            "QTreeWidget::item{padding:3px 0;}")
        self.tree.itemSelectionChanged.connect(self._update_buttons)
        self.tree.itemDoubleClicked.connect(self._on_double_click)
        self.tree.itemClicked.connect(self._on_tree_clicked)        # คลิกเลขเอกสาร=copy
        self.tree.itemChanged.connect(self._on_tree_changed)        # บันทึกหมายเหตุ
        lay.addWidget(self.tree, 1)

        # ── ปุ่มล่าง ──
        bottom = QHBoxLayout()
        self.lbl_total = QLabel("")
        self.lbl_total.setStyleSheet("color:#15803d;font-size:12px;font-weight:600;")
        bottom.addWidget(self.lbl_total)
        bottom.addStretch()
        btn_reset = QPushButton("↺ จัดใหม่อัตโนมัติ")
        btn_reset.setToolTip("ลบการจัดเรียงที่บันทึกไว้ แล้วกระจายใหม่ตามวงเงิน")
        btn_reset.setStyleSheet(
            "QPushButton{padding:6px 14px;border:1px solid #cbd5e1;border-radius:4px;"
            "background:white;}QPushButton:hover{background:#fef3c7;}")
        btn_reset.clicked.connect(self._reset_plan)
        bottom.addWidget(btn_reset)
        btn_close = QPushButton("ปิด")
        btn_close.setStyleSheet(
            "QPushButton{padding:6px 18px;border:1px solid #cbd5e1;border-radius:4px;"
            "background:white;}QPushButton:hover{background:#f1f5f9;}")
        btn_close.clicked.connect(self.close)   # ผ่าน closeEvent (บังคับบันทึก)
        bottom.addWidget(btn_close)
        lay.addLayout(bottom)

        # โชว์เวลาบันทึกล่าสุด (ถ้ามีการบันทึกไว้)
        if self._saved_at:
            try:
                ts = self._saved_at.replace("T", " ")[:16]
                self.lbl_saved.setText(f"💾 ใช้การจัดเรียงที่บันทึกไว้ ({ts})")
            except Exception:
                pass

        self._render()
        self._update_buttons()

    def _load_or_distribute(self, candidates):
        """ถ้ามีการจัดเรียงที่บันทึกไว้ → ใช้อันนั้น (รายการใหม่ที่ยังไม่จัด ค่อยกระจายเพิ่ม)"""
        saved = queue_plan.load_plan()
        # พ้น 3 วันแล้ว → ดีดทิ้ง จัดใหม่ (ข้อ 10)
        if saved and saved.get("saved_at"):
            try:
                age = (datetime.now()
                       - datetime.fromisoformat(saved["saved_at"][:19])).days
                if age >= 3:
                    queue_plan.clear_plan()
                    saved = {}
            except Exception:
                pass
        if not (saved and saved.get("days")):
            return self._distribute(candidates)

        by_id = {_exp_id(e): e for e in candidates}
        used = set()
        days = []
        for day_ids in saved["days"]:
            grp = [by_id[i] for i in day_ids if i in by_id and i not in used]
            for i in day_ids:
                used.add(i)
            days.append(grp)
        # รายการใหม่ที่ยังไม่เคยจัด → กระจายเพิ่มต่อท้าย
        leftover = [e for e in candidates if _exp_id(e) not in used]
        if leftover:
            days.extend(self._distribute(leftover))
        days = [g for g in days if g] or [[]]
        self._saved_at = saved.get("saved_at", "")
        return days

    # ── การกระจายรายการ ──
    def _distribute(self, candidates):
        """
        เรียงตามความสำคัญ แล้วบรรจุลงแต่ละวันแบบ best-fit
        (ข้ามรายการที่ใส่ไม่พอ ลองตัวถัดไปให้เต็มวงเงินที่สุด)
        — ใช้ตรรกะเดียวกับปุ่ม 🤖 จัดคิวอัตโนมัติ เพื่อให้ยอดวันที่ 1 ตรงกัน
        """
        today_iso = date.today().isoformat()

        def priority(e):
            if _is_vip_vendor(e):
                return (-1, "")     # นายพชร รัชนาทสกุล → จัดคิวก่อนเสมอ
            due = (_due(e) or "")[:10]
            if due and due < today_iso:
                return (0, due)
            elif due:
                return (1, due)
            else:
                return (2, "9999-12-31")

        remaining = sorted(candidates, key=priority)

        days = []
        while remaining:
            day, total, rest = [], 0.0, []
            for e in remaining:
                amt = _amount(e)
                if self._limit > 0 and day and (total + amt) > self._limit:
                    rest.append(e)   # ใส่ไม่พอ → เก็บไว้วันถัดไป ลองตัวเล็กกว่าต่อ
                    continue
                day.append(e)
                total += amt
            days.append(day)
            remaining = rest
        return days or [[]]

    @staticmethod
    def _next_business_day(d):
        from utils import _next_business_day as _nbd
        return _nbd(d)

    def _day_dates(self):
        """คืน list ของวันที่ (date) ตามจำนวนวันใน self._days (ข้ามวันหยุด)"""
        from utils import _is_day_off
        dates = []
        d = date.today()
        while _is_day_off(d):           # วันแรกถ้าตรงวันหยุด → เลื่อนเป็นวันทำการถัดไป
            d = self._next_business_day(d)
        for _ in self._days:
            dates.append(d)
            d = self._next_business_day(d)
        return dates

    # ── render ──
    def _render(self):
        # จำ id ที่เลือกไว้ เพื่อ select กลับหลัง re-render
        keep = self._selected_exp_id()
        keep_checked = getattr(self, "_keep_checked", set())
        self.tree.blockSignals(True)
        self.tree.clear()
        dates = self._day_dates()
        grand = 0.0
        target_item = None

        thai_days = ["จันทร์", "อังคาร", "พุธ", "พฤหัสบดี", "ศุกร์", "เสาร์", "อาทิตย์"]
        for di, (grp, d) in enumerate(zip(self._days, dates)):
            day_total = sum(_amount(e) for e in grp)
            grand += day_total
            over = self._limit > 0 and day_total > self._limit
            label = (f"📅 วันที่ {di + 1}  ({thai_days[d.weekday()]} {fmt_date(d.isoformat())})"
                     f"  —  {len(grp)} รายการ")
            top = QTreeWidgetItem([label, "", "", "", fmt_amount(day_total), "", ""])
            top.setData(0, Qt.ItemDataRole.UserRole, ("day", di))
            # เช็คบ็อกซ์หัววัน — ติ๊ก=เลือกทั้งวัน / ปลด=เอาเครื่องหมายออกทั้งวัน
            top.setFlags(top.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            top.setCheckState(0, Qt.CheckState.Unchecked)
            f = top.font(0); f.setBold(True); top.setFont(0, f); top.setFont(4, f)
            bg = QColor("#fee2e2") if over else QColor("#dcfce7")
            for c in range(7):
                top.setBackground(c, QBrush(bg))
            top.setForeground(0, QBrush(QColor("#991b1b" if over else "#15803d")))
            top.setForeground(4, QBrush(QColor("#991b1b" if over else "#15803d")))
            top.setTextAlignment(4, Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            if over:
                top.setToolTip(0, "ยอดวันนี้เกินวงเงิน — ลองย้ายบางรายการไปวันอื่น")
            self.tree.addTopLevelItem(top)

            for e in grp:
                vendor = _vendor(e) or "—"
                doc = _doc_no(e)
                brand = _brand_name(e, self._assignments) or "—"
                due = fmt_date(_due(e)) if _due(e) else "—"
                eid = _exp_id(e)
                child = QTreeWidgetItem([vendor, doc, brand, due,
                                         fmt_amount(_amount(e)), remarks.get(eid), ""])
                child.setData(0, Qt.ItemDataRole.UserRole, ("exp", eid))
                child.setTextAlignment(4, Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                # ☑️ checkbox หน้าแต่ละรายการ (selector) + หมายเหตุ(col5) แก้ไขได้
                child.setFlags(child.flags() | Qt.ItemFlag.ItemIsUserCheckable
                               | Qt.ItemFlag.ItemIsEditable)
                child.setCheckState(0, Qt.CheckState.Checked
                                    if eid in keep_checked
                                    else Qt.CheckState.Unchecked)
                child.setToolTip(5, "ดับเบิลคลิกช่องนี้เพื่อพิมพ์หมายเหตุ")
                # เน้นสีรายการเกินกำหนด
                if _due(e) and (_due(e) or "")[:10] < date.today().isoformat():
                    child.setForeground(0, QBrush(QColor("#dc2626")))
                top.addChild(child)
                # ลิงก์แชร์ (col 6) — มีลิงก์แล้วโชว์ลิงก์ / ยังไม่มี = ปุ่มกดดึง+เปิด
                url = share_links.get(eid)
                if url:
                    lk = QLabel(f"🔗 <a href='{url}'>เปิดดู</a>")
                    lk.setTextFormat(Qt.TextFormat.RichText)
                    lk.setOpenExternalLinks(True)
                    lk.setStyleSheet("font-size:12px;padding-left:4px;background:transparent;")
                    self.tree.setItemWidget(child, 6, lk)
                else:
                    b = QPushButton("🔗 ดูใบ")
                    b.setCursor(Qt.CursorShape.PointingHandCursor)
                    b.setStyleSheet(
                        "QPushButton{padding:1px 6px;border:1px solid #2563eb;border-radius:4px;"
                        "background:white;color:#2563eb;font-size:11px;}"
                        "QPushButton:hover{background:#dbeafe;}")
                    b.clicked.connect(lambda _=False, ex=e: self._open_share_doc(ex))
                    bw = QWidget(); bl = QHBoxLayout(bw)
                    bl.addWidget(b); bl.setContentsMargins(2, 0, 2, 0)
                    bl.setAlignment(Qt.AlignmentFlag.AlignLeft)
                    self.tree.setItemWidget(child, 6, bw)
                if keep and eid == keep:
                    target_item = child

            top.setExpanded(True)

        self.lbl_total.setText(
            f"รวมทั้งหมด {sum(len(g) for g in self._days)} รายการ • "
            f"{fmt_amount(grand)} บาท • กระจาย {len(self._days)} วัน")

        if target_item:
            self.tree.setCurrentItem(target_item)
        self.tree.blockSignals(False)
        self._keep_checked = set()   # ใช้ครั้งเดียวแล้วล้าง
        # อัปเดตช่วงวันที่ของช่องจัดคิวเอง (เผื่อจำนวนวันเปลี่ยน) +1 เพื่อสร้างวันใหม่ได้
        if hasattr(self, "spin_day"):
            self.spin_day.setMaximum(max(1, len(self._days) + 1))

    # ── ตัวช่วยหา selection ──
    def _selected_node(self):
        items = self.tree.selectedItems()
        return items[0] if items else None

    def _selected_exp_id(self):
        node = self._selected_node()
        if node:
            data = node.data(0, Qt.ItemDataRole.UserRole)
            if data and data[0] == "exp":
                return data[1]
        return None

    def _locate_exp(self, exp_id):
        """คืน (day_index, item_index) ของ exp_id ใน self._days"""
        for di, grp in enumerate(self._days):
            for ii, e in enumerate(grp):
                if _exp_id(e) == exp_id:
                    return di, ii
        return None, None

    def _current_day_index(self):
        """วันของ node ที่เลือกอยู่ (ทั้งหัววันและรายการ)"""
        node = self._selected_node()
        if not node:
            return None
        data = node.data(0, Qt.ItemDataRole.UserRole)
        if not data:
            return None
        if data[0] == "day":
            return data[1]
        if data[0] == "exp":
            di, _ = self._locate_exp(data[1])
            return di
        return None

    # ── ปุ่ม ──
    def _update_buttons(self):
        has = bool(self._checked_exp_ids()) or bool(self._selected_exp_id())
        self.btn_up.setEnabled(has)
        self.btn_down.setEnabled(has)

    def _move_selected(self, direction):
        # ย้ายรายการที่ "ติ๊ก" (หลายรายการได้) ถ้าไม่ติ๊กใช้แถวที่ไฮไลต์
        ids = self._checked_exp_ids()
        if not ids:
            eid = self._selected_exp_id()
            ids = [eid] if eid else []
        if not ids:
            QMessageBox.information(self, "ยังไม่ได้เลือก",
                "ติ๊ก ☑️ หน้ารายการที่ต้องการย้ายก่อนครับ")
            return

        moved = 0
        for eid in ids:
            di, ii = self._locate_exp(eid)
            if di is None:
                continue
            target = di + direction
            if target < 0:
                continue
            e = self._days[di].pop(ii)
            while target >= len(self._days):
                self._days.append([])     # สร้างวันใหม่ต่อท้าย
            self._days[target].append(e)
            moved += 1
        # ลบวันที่ว่างกลางทาง
        self._days = [g for g in self._days if g] or [[]]
        self._keep_checked = set(ids)     # คงเครื่องหมายติ๊กไว้หลัง re-render
        self._render()
        self._update_buttons()
        if moved:
            self.lbl_hint.setText(f"ย้าย {moved} รายการแล้ว")

    def _manual_move_by_doc(self):
        """จัดคิวเอง: พิมพ์เลขเอกสาร EXP/PO แล้วย้าย/ใส่ไปวันที่ที่ระบุ"""
        q = (self.ed_doc.text() or "").strip().upper().replace(" ", "")
        if not q:
            self.lbl_manual.setText("⚠️ พิมพ์เลขเอกสารก่อน")
            return
        # หาในรายการทั้งหมด (ตรงเป๊ะก่อน ไม่งั้นค่อยจับบางส่วน)
        found_id = None
        exact = None
        partial = []
        for grp in self._days:
            for e in grp:
                doc = (_doc_no(e) or "").upper().replace(" ", "")
                if not doc:
                    continue
                if doc == q:
                    exact = _exp_id(e)
                elif q in doc:
                    partial.append(_exp_id(e))
        found_id = exact or (partial[0] if len(partial) == 1 else None)
        if not found_id:
            if len(partial) > 1:
                self.lbl_manual.setText(f"⚠️ เจอหลายรายการที่ตรงกับ '{q}' — พิมพ์ให้ครบ")
            else:
                self.lbl_manual.setText(f"❌ ไม่พบเอกสาร '{q}' ในคิว")
            return

        target = self.spin_day.value() - 1
        di, ii = self._locate_exp(found_id)
        if di is None:
            self.lbl_manual.setText("❌ ไม่พบรายการ")
            return
        e = self._days[di].pop(ii)
        while target >= len(self._days):
            self._days.append([])
        self._days[target].append(e)
        self._days = [g for g in self._days if g] or [[]]
        self._keep_checked = self._checked_exp_ids()
        self._render()
        self._update_buttons()
        self.lbl_manual.setText(
            f"✅ ย้าย {_doc_no(e)} ({_vendor(e)}) ไปวันที่ {target + 1} แล้ว")
        self.ed_doc.clear()
        self.ed_doc.setFocus()

    def _search_in_plan(self, text):
        """ค้นหารายการในคิว (ชื่อ/เลขเอกสาร บางส่วน หลายคำ) → ไฮไลต์ + เลื่อนไปหา"""
        q = (text or "").strip().lower()
        self.tree.clearSelection()
        if not q:
            self.lbl_search.setText("")
            return
        terms = [t for t in re.split(r"[,\s]+", q) if t]
        root = self.tree.invisibleRootItem()
        first = None
        n = 0
        for i in range(root.childCount()):
            day = root.child(i)
            day_hit = False
            for j in range(day.childCount()):
                ch = day.child(j)
                blob = (ch.text(0) + " " + ch.text(1)).lower()
                if any(t in blob for t in terms):
                    ch.setSelected(True)
                    day_hit = True
                    n += 1
                    if first is None:
                        first = ch
            if day_hit:
                day.setExpanded(True)
        if first:
            self.tree.scrollToItem(first)
            self.lbl_search.setText(f"🔍 เจอ {n} รายการ (ไฮไลต์สีน้ำเงิน)")
        else:
            self.lbl_search.setText("❌ ไม่เจอในคิว — อาจจ่ายแล้ว / ไม่อนุมัติ / คนละบริษัท")

    def _on_double_click(self, item, _col):
        data = item.data(0, Qt.ItemDataRole.UserRole)
        if not (data and data[0] == "exp"):
            return
        if _col == 5:        # ช่องหมายเหตุ → แก้ไข
            self.tree.editItem(item, 5)
            return
        # ดับเบิลคลิกช่องอื่น = ติ๊ก/ยกเลิกติ๊ก
        new = (Qt.CheckState.Unchecked if item.checkState(0) == Qt.CheckState.Checked
               else Qt.CheckState.Checked)
        item.setCheckState(0, new)

    def _on_tree_clicked(self, item, col):
        """คลิกคอลัมน์เลขเอกสาร (1) → คัดลอกรหัส EXP/PO (ข้อ 4)"""
        data = item.data(0, Qt.ItemDataRole.UserRole)
        if col == 1 and data and data[0] == "exp":
            code = item.text(1)
            if code:
                QApplication.clipboard().setText(code)
                self.lbl_hint.setText(f"📋 คัดลอกเลขเอกสาร: {code}")

    def _on_tree_changed(self, item, column):
        data = item.data(0, Qt.ItemDataRole.UserRole)
        # ติ๊ก/ปลดหัววัน → ตั้งค่าให้ทุกรายการในวันนั้น (ข้อ feedback)
        if data and data[0] == "day" and column == 0:
            state = item.checkState(0)
            self.tree.blockSignals(True)
            for j in range(item.childCount()):
                ch = item.child(j)
                d = ch.data(0, Qt.ItemDataRole.UserRole)
                if d and d[0] == "exp":
                    ch.setCheckState(0, state)
            self.tree.blockSignals(False)
        elif column == 5 and data and data[0] == "exp":
            remarks.set(data[1], item.text(5))
            self.lbl_hint.setText("💾 บันทึกหมายเหตุแล้ว")
        self._update_buttons()

    def _open_share_doc(self, e):
        """ดึงลิงก์แชร์ของเอกสารนี้ (ถ้ายังไม่มี) แล้วเปิดดู"""
        eid = _exp_id(e)
        url = share_links.get(eid)
        if not url:
            self.lbl_hint.setText("กำลังดึงลิงก์แชร์เอกสาร…")
            QApplication.processEvents()
            try:
                did = e.get("documentId") or e.get("recordId")
                cid, csec = _exp_creds(e)
                url = get_share_link(cid, csec, did, doctype=e.get("_doctype", "expense"))
                if url:
                    share_links.set(eid, url)
                    self._render()
            except Exception as ex:
                QMessageBox.warning(self, "เปิดไม่ได้", f"ดึงลิงก์แชร์ไม่สำเร็จ:\n{ex}")
                return
        if url:
            QDesktopServices.openUrl(QUrl(url))
            self.lbl_hint.setText("เปิดลิงก์แชร์แล้ว")
        else:
            QMessageBox.information(self, "ไม่มีลิงก์แชร์", "ดึงลิงก์แชร์ไม่สำเร็จ")

    def selected_day_items(self):
        return self._selected_items

    # ── ข้อ 5: checkbox + บันทึกคิว + Mark จ่ายแล้วรออัพเดต ──
    def _checked_exp_ids(self):
        ids = []
        root = self.tree.invisibleRootItem()
        for i in range(root.childCount()):
            day = root.child(i)
            for j in range(day.childCount()):
                ch = day.child(j)
                if ch.checkState(0) == Qt.CheckState.Checked:
                    data = ch.data(0, Qt.ItemDataRole.UserRole)
                    if data and data[0] == "exp":
                        ids.append(data[1])
        return ids

    def _all_exp_by_id(self):
        out = {}
        for grp in self._days:
            for e in grp:
                out[_exp_id(e)] = e
        return out

    def _check_all_plan(self, checked: bool):
        """ติ๊ก/ปลดติ๊ก ทุกรายการในตารางคิว (ข้อ 3)"""
        state = Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked
        self.tree.blockSignals(True)
        root = self.tree.invisibleRootItem()
        for i in range(root.childCount()):
            day = root.child(i)
            for j in range(day.childCount()):
                ch = day.child(j)
                data = ch.data(0, Qt.ItemDataRole.UserRole)
                if data and data[0] == "exp":
                    ch.setCheckState(0, state)
        self.tree.blockSignals(False)
        self._update_buttons()

    def _reject_checked_plan(self):
        """ไม่อนุมัติรายการที่ติ๊ก (ข้อ 2) — เอาออกจากคิว ไม่นำมาจัดอีกจนกว่าจะอนุมัติ"""
        ids = self._checked_exp_ids()
        if not ids:
            QMessageBox.information(self, "ยังไม่ได้เลือก",
                "ติ๊ก ☑️ หน้ารายการที่จะไม่อนุมัติก่อนครับ")
            return
        if QMessageBox.question(self, "🚫 ไม่อนุมัติ",
                f"ทำเครื่องหมาย 'ไม่อนุมัติ' {len(ids)} รายการ?\n\n"
                "• จะถูกเอาออกจากตารางคิว และไม่นำมาจัดคิวอีก\n"
                "• กด '📋 รายการไม่อนุมัติ' เพื่อดู/อนุมัติกลับ",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel
                ) != QMessageBox.StandardButton.Yes:
            return
        by_id = self._all_exp_by_id()
        for eid in ids:
            e = by_id.get(eid, {})
            rejected.reject(eid, {"vendor": _vendor(e), "amount": _amount(e),
                                  "doc": _doc_no(e), "user": activity_log.current_user()})
        self._days = [[e for e in grp if _exp_id(e) not in ids] for grp in self._days]
        self._days = [g for g in self._days if g] or [[]]
        activity_log.log("ไม่อนุมัติ (จากตารางคิว)", f"{len(ids)} รายการ")
        self._render()
        self._update_buttons()
        QMessageBox.information(self, "✅ สำเร็จ",
            f"ไม่อนุมัติ {len(ids)} รายการแล้ว\n"
            "ดู/อนุมัติกลับได้ที่ปุ่ม '📋 รายการไม่อนุมัติ'")

    def _view_rejected_plan(self):
        """หน้ารายการไม่อนุมัติ (ใช้ dialog กลาง) — อนุมัติกลับแล้วนำกลับเข้าคิวทันที"""
        dlg = RejectedListDialog(self)
        dlg.exec()
        if not dlg.approved_ids:
            return
        # นำที่อนุมัติกลับ ใส่กลับเข้าคิว (วันที่ 1) ทันที
        qtab = self.parent()
        pool = {_exp_id(e): e for e in getattr(qtab, "_expenses", [])}
        readded = 0
        for eid in dlg.approved_ids:
            e = pool.get(eid)
            if e:
                self._days[0].insert(0, e)
                readded += 1
        self._render()
        self._update_buttons()
        if readded:
            self.lbl_hint.setText(f"นำกลับเข้าคิว {readded} รายการ (วันที่ 1)")

    def _persist_plan(self) -> int:
        """บันทึกการจัดเรียงคิว + log ประวัติ (เงียบ ไม่มี popup) — คืนจำนวนรายการ"""
        dates = self._day_dates()
        items = []
        for di, (grp, d) in enumerate(zip(self._days, dates), 1):
            for e in grp:
                items.append({"doc": _doc_no(e), "vendor": _vendor(e),
                              "amount": _amount(e), "brand": _brand_name(e, self._assignments),
                              "day": di, "date": d.isoformat()})
        n_days = len([g for g in self._days if g])
        now_iso = datetime.now().isoformat(timespec="seconds")
        days_ids = [[_exp_id(e) for e in grp] for grp in self._days]
        queue_plan.save_plan(days_ids, now_iso)
        self._saved_at = now_iso
        activity_log.log_queue(
            f"บันทึกคิว {n_days} วัน {len(items)} รายการ", items=items)
        activity_log.log("บันทึกคิว", f"{n_days} วัน {len(items)} รายการ")
        return len(items)

    def _save_queue(self):
        n_items = self._persist_plan()
        n_days = len([g for g in self._days if g])
        ts = datetime.now().strftime("%d/%m %H:%M:%S")
        self.lbl_saved.setText(f"💾 บันทึกล่าสุด {ts}")
        QMessageBox.information(self, "✅ บันทึกคิวแล้ว",
            f"บันทึกการจัดเรียงคิว {n_days} วัน • {n_items} รายการ\n"
            f"เวลา {ts}\n\n"
            "ปิดแล้วเปิดใหม่ การจัดเรียงนี้จะยังอยู่ครับ\n"
            "(ดูประวัติได้ที่ปุ่ม 📜 ประวัติการจัดคิว)")

    def _mark_pending(self):
        ids = self._checked_exp_ids()
        if not ids:
            QMessageBox.information(self, "ยังไม่ได้เลือก",
                "ติ๊ก ☑️ หน้ารายการที่ต้องการก่อน แล้วค่อยกด Mark ครับ")
            return
        by_id = self._all_exp_by_id()
        for eid in ids:
            e = by_id.get(eid, {})
            paid_pending.mark(eid, {
                "amount": _amount(e), "vendor": _vendor(e),
                "doc": _doc_no(e), "user": activity_log.current_user(),
            })
        activity_log.log("Mark จ่ายแล้วรออัพเดต", f"{len(ids)} รายการ")
        # เอาออกจากคิว (ไม่นำไปคิดยอดต่อ)
        self._days = [[e for e in grp if _exp_id(e) not in ids] for grp in self._days]
        self._days = [g for g in self._days if g] or [[]]
        self._marked_pending.extend(ids)
        self._render()
        self._update_buttons()
        QMessageBox.information(self, "✅ สำเร็จ",
            f"Mark จ่ายแล้ว (รออัพเดต FlowAccount) {len(ids)} รายการ\n\n"
            "• รายการนี้จะไม่ถูกนำไปคิดยอดแล้ว\n"
            "• สถานะในหน้าหลักจะเป็น 'จ่ายแล้วรออัพเดต'")

    def closeEvent(self, ev):
        """บังคับบันทึกคิวก่อนปิดหน้าต่าง (ทั้งปุ่มปิดและกากบาท)"""
        r = QMessageBox.question(self, "💾 บันทึกคิวก่อนปิด",
            "ต้องการบันทึกการจัดคิวก่อนปิดหรือไม่?\n\n"
            "• บันทึก = เก็บการจัดเรียงไว้ (เปิดใหม่ยังอยู่)\n"
            "• ไม่บันทึก = ปิดโดยไม่เก็บการเปลี่ยนแปลงล่าสุด",
            QMessageBox.StandardButton.Save | QMessageBox.StandardButton.Discard
            | QMessageBox.StandardButton.Cancel,
            QMessageBox.StandardButton.Save)
        if r == QMessageBox.StandardButton.Cancel:
            ev.ignore()
            return
        if r == QMessageBox.StandardButton.Save:
            self._persist_plan()
            self.lbl_saved.setText("💾 บันทึกแล้ว")
        ev.accept()

    def _reset_plan(self):
        if QMessageBox.question(self, "จัดใหม่อัตโนมัติ",
                "ลบการจัดเรียงที่บันทึกไว้ แล้วกระจายใหม่ตามวงเงิน?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel
                ) != QMessageBox.StandardButton.Yes:
            return
        queue_plan.clear_plan()
        allitems = [e for grp in self._days for e in grp]
        self._days = self._distribute(allitems)
        self._saved_at = ""
        self.lbl_saved.setText("")
        self._render()
        self._update_buttons()


class QueueHistoryDialog(QDialog):
    """ประวัติการจัดคิว (ข้อ 9) — เก็บทุกครั้งที่กดบันทึกคิว ดูรายละเอียด/ค้นหา/ลบได้"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📜 ประวัติการจัดคิว")
        self.resize(820, 560)
        self.setWindowFlags(self.windowFlags()
                            | Qt.WindowType.WindowMinimizeButtonHint
                            | Qt.WindowType.WindowMaximizeButtonHint)
        lay = QVBoxLayout(self)

        head = QLabel("ประวัติการบันทึกคิวจ่ายทั้งหมด — ดับเบิลคลิกเพื่อดูรายละเอียดรายวัน")
        head.setStyleSheet("color:#15803d;font-size:13px;font-weight:600;")
        lay.addWidget(head)

        bar = QHBoxLayout()
        self.search = QLineEdit()
        self.search.setPlaceholderText("🔍 ค้นหาวัน (เช่น 2026-06-25) / รายละเอียด")
        self.search.setClearButtonEnabled(True)
        self.search.textChanged.connect(self._reload)
        bar.addWidget(self.search, 1)
        self.btn_del = QPushButton("🗑️ ลบที่เลือก")
        self.btn_del.setStyleSheet(
            "QPushButton{padding:5px 12px;border:none;border-radius:4px;"
            "background:#dc2626;color:white;font-weight:600;}"
            "QPushButton:hover{background:#b91c1c;}")
        self.btn_del.clicked.connect(self._delete_selected)
        bar.addWidget(self.btn_del)
        lay.addLayout(bar)

        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["เวลาบันทึก", "รายละเอียด", "จำนวนรายการ", "ยอดรวม (บาท)"])
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.table.setColumnWidth(0, 150)
        self.table.setColumnWidth(2, 110)
        self.table.setColumnWidth(3, 140)
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.cellDoubleClicked.connect(self._show_detail)
        lay.addWidget(self.table, 1)

        self.lbl_sum = QLabel("")
        self.lbl_sum.setStyleSheet("color:#64748b;font-size:12px;")
        lay.addWidget(self.lbl_sum)

        self._rows = []
        self._reload()

    def _reload(self):
        q = (self.search.text() or "").strip().lower()
        rows = activity_log.load_queue_log()
        if q:
            rows = [r for r in rows
                    if q in str(r.get("time", "")).lower()
                    or q in str(r.get("detail", "")).lower()]
        self._rows = rows
        self.table.setRowCount(0)
        for r in rows:
            total = sum(float(it.get("amount", 0) or 0) for it in (r.get("items") or []))
            row = self.table.rowCount()
            self.table.insertRow(row)
            vals = [str(r.get("time", "")).replace("T", " ")[:19],
                    r.get("detail", ""),
                    str(r.get("count", len(r.get("items") or []))),
                    fmt_amount(total)]
            for c, v in enumerate(vals):
                item = QTableWidgetItem(v)
                if c in (2, 3):
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.table.setItem(row, c, item)
        self.lbl_sum.setText(f"ทั้งหมด {len(rows)} ครั้งที่บันทึกคิว")

    def _show_detail(self, row, col):
        if not (0 <= row < len(self._rows)):
            return
        r = self._rows[row]
        items = r.get("items") or []
        dlg = QDialog(self)
        dlg.setWindowTitle(f"รายละเอียดคิว — {str(r.get('time','')).replace('T',' ')[:19]}")
        dlg.resize(680, 480)
        dl = QVBoxLayout(dlg)
        dl.addWidget(QLabel(r.get("detail", "")))
        t = QTableWidget(0, 5)
        t.setHorizontalHeaderLabels(["วันที่", "วัน", "เลขเอกสาร", "ผู้รับเงิน", "ยอด (บาท)"])
        t.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        t.setColumnWidth(0, 100); t.setColumnWidth(1, 50); t.setColumnWidth(2, 120)
        t.setColumnWidth(4, 110)
        t.verticalHeader().setVisible(False)
        t.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        for it in items:
            rr = t.rowCount(); t.insertRow(rr)
            cells = [fmt_date(it.get("date", "")) if it.get("date") else "—",
                     str(it.get("day", "")),
                     it.get("doc", ""), it.get("vendor", ""),
                     fmt_amount(it.get("amount", 0))]
            for c, v in enumerate(cells):
                cell = QTableWidgetItem(v)
                if c == 4:
                    cell.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                t.setItem(rr, c, cell)
        dl.addWidget(t, 1)
        total = sum(float(it.get("amount", 0) or 0) for it in items)
        dl.addWidget(QLabel(f"รวม {len(items)} รายการ • {fmt_amount(total)} บาท"))
        btn = QPushButton("ปิด"); btn.clicked.connect(dlg.accept)
        dl.addWidget(btn)
        dlg.exec()

    def _delete_selected(self):
        row = self.table.currentRow()
        if not (0 <= row < len(self._rows)):
            QMessageBox.information(self, "ยังไม่ได้เลือก", "คลิกเลือกแถวที่จะลบก่อนครับ")
            return
        r = self._rows[row]
        if QMessageBox.question(self, "ลบประวัติ",
                f"ลบประวัติการจัดคิวของ {str(r.get('time','')).replace('T',' ')[:19]} ?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel
                ) != QMessageBox.StandardButton.Yes:
            return
        activity_log.delete_queue_log(r.get("time", ""))
        self._reload()


class AdjustLimitDialog(QDialog):
    """ปรับวงเงินจ่ายวันนี้"""

    PRESETS = [
        ("100K",     100000),
        ("150K",     150000),
        ("200K",     200000),
        ("300K",     300000),
        ("500K",     500000),
        ("♾️ ไม่จำกัด", 0),
    ]

    def __init__(self, current_limit: float, parent=None):
        super().__init__(parent)
        self.setWindowTitle("🔧 ปรับวงเงินจ่าย")
        self.setMinimumWidth(380)
        self._build_ui(current_limit)

    def _build_ui(self, current_limit: float):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        # ช่องวงเงิน
        form = QFormLayout()
        self.ed_limit = QLineEdit("ไม่จำกัด" if current_limit <= 0 else f"{current_limit:,.0f}")
        self.ed_limit.setAlignment(Qt.AlignmentFlag.AlignRight)
        self.ed_limit.setStyleSheet("padding:6px;font-size:14px;font-weight:600;")
        form.addRow("วงเงินใหม่ (บาท):", self.ed_limit)
        layout.addLayout(form)

        # Preset buttons
        preset_row = QHBoxLayout()
        preset_row.addWidget(QLabel("เลือกด่วน:"))
        for label, val in self.PRESETS:
            b = QPushButton(label)
            b.setStyleSheet(
                "QPushButton{padding:4px 10px;border:1px solid #cbd5e1;"
                "border-radius:4px;background:white;font-size:11px;}"
                "QPushButton:hover{background:#dcfce7;border-color:#16a34a;}")
            b.clicked.connect(lambda _, v=val: self._set_preset(v))
            preset_row.addWidget(b)
        preset_row.addStretch()
        layout.addLayout(preset_row)

        # Save scope
        scope_box = QGroupBox("บังคับใช้กับ:")
        scope_v = QVBoxLayout(scope_box)
        from PyQt6.QtWidgets import QRadioButton, QButtonGroup
        self.rb_today  = QRadioButton("วันนี้เท่านั้น (default ไม่เปลี่ยน)")
        self.rb_save   = QRadioButton("บันทึกเป็น default ทุกวัน")
        self.rb_today.setChecked(True)
        self._scope_group = QButtonGroup(self)
        self._scope_group.addButton(self.rb_today)
        self._scope_group.addButton(self.rb_save)
        scope_v.addWidget(self.rb_today)
        scope_v.addWidget(self.rb_save)
        layout.addWidget(scope_box)

        # Buttons
        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        btns.button(QDialogButtonBox.StandardButton.Save).setText("✅ บันทึก")
        btns.button(QDialogButtonBox.StandardButton.Cancel).setText("ยกเลิก")
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _set_preset(self, value: float):
        if value <= 0:
            self.ed_limit.setText("ไม่จำกัด")
        else:
            self.ed_limit.setText(f"{value:,.0f}")

    def get_limit(self) -> float:
        txt = self.ed_limit.text().strip().lower()
        if "ไม่จำกัด" in txt or txt in ("0", "∞", ""):
            return 0
        try:
            return max(0.0, float(txt.replace(",", "").replace("บาท", "").strip()))
        except Exception:
            return 0

    def save_as_default(self) -> bool:
        return self.rb_save.isChecked()


class AddExpenseDialog(QDialog):
    """หน้าต่างเพิ่มรายการค่าใช้จ่ายเอง (ไม่ส่งเข้า FlowAccount)"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("➕ เพิ่มรายการค่าใช้จ่าย")
        self.setMinimumWidth(440)
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        info = QLabel("📌 รายการที่เพิ่มจะถูกเก็บไว้ในโปรแกรม\n(ยังไม่ส่งเข้า FlowAccount)")
        info.setStyleSheet(f"color:{C_MUTED}; font-size:11px; padding:4px 0;")
        info.setWordWrap(True)
        layout.addWidget(info)

        form = QFormLayout()
        form.setVerticalSpacing(10)

        self.ed_vendor = QLineEdit()
        self.ed_vendor.setPlaceholderText("เช่น  บริษัท ABC จำกัด")
        form.addRow("ชื่อผู้รับเงิน (Vendor) *", self.ed_vendor)

        self.ed_amount = QLineEdit()
        self.ed_amount.setPlaceholderText("เช่น  5000.00")
        form.addRow("ยอดเงิน (บาท) *", self.ed_amount)

        self.ed_due = QDateEdit(QDate.currentDate())
        self.ed_due.setCalendarPopup(True)
        self.ed_due.setDisplayFormat("dd/MM/yyyy")
        form.addRow("วันครบกำหนด *", self.ed_due)

        self.ed_doc = QLineEdit()
        self.ed_doc.setPlaceholderText("เช่น  M-2026-001  (ถ้าไม่ใส่ระบบจะสร้างให้)")
        form.addRow("เลขที่เอกสาร", self.ed_doc)

        self.ed_brand = QComboBox()
        self.ed_brand.setEditable(True)   # ให้พิมพ์เพิ่มเองได้ถ้าไม่มีใน list
        self.ed_brand.addItem("—")
        brand_list = load_config().get("brand_list") or []
        for b in brand_list:
            self.ed_brand.addItem(b)
        form.addRow("แบรนด์", self.ed_brand)

        self.ed_remarks = QTextEdit()
        self.ed_remarks.setPlaceholderText("หมายเหตุเพิ่มเติม...")
        self.ed_remarks.setFixedHeight(60)
        form.addRow("หมายเหตุ", self.ed_remarks)

        layout.addLayout(form)

        btns = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        btns.button(QDialogButtonBox.StandardButton.Save).setText("✅ บันทึก")
        btns.button(QDialogButtonBox.StandardButton.Cancel).setText("ยกเลิก")
        btns.accepted.connect(self._validate_and_accept)
        btns.rejected.connect(self.reject)
        layout.addWidget(btns)

    def _validate_and_accept(self):
        vendor = self.ed_vendor.text().strip()
        amount = self.ed_amount.text().strip()
        if not vendor:
            QMessageBox.warning(self, "กรอกข้อมูลไม่ครบ", "กรุณากรอกชื่อผู้รับเงิน")
            self.ed_vendor.setFocus(); return
        try:
            amt = float(amount)
            if amt <= 0: raise ValueError()
        except Exception:
            QMessageBox.warning(self, "ยอดเงินไม่ถูกต้อง", "กรุณากรอกยอดเงินเป็นตัวเลข (มากกว่า 0)")
            self.ed_amount.setFocus(); return
        self.accept()

    def get_data(self) -> dict:
        d = self.ed_due.date()
        due = f"{d.year()}-{d.month():02d}-{d.day():02d}"
        doc = self.ed_doc.text().strip()
        if not doc:
            doc = f"M-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
        brand = self.ed_brand.currentText().strip()
        if brand == "—":
            brand = ""
        return {
            "documentSerial": doc,
            "contactName":    self.ed_vendor.text().strip(),
            "grandTotal":     float(self.ed_amount.text().strip()),
            "dueDate":        due,
            "publishedOn":    date.today().isoformat(),
            "brandCode":      brand,
            "remarks":        self.ed_remarks.toPlainText().strip(),
            "statusString":   "approved",  # ถือว่าอนุมัติแล้ว
        }


class MarkPaidDialog(QDialog):
    def __init__(self, expenses: list, parent=None):
        super().__init__(parent)
        self.expenses = expenses
        self.setWindowTitle(f"✅ Mark จ่ายแล้ว — {len(expenses)} รายการ")
        self.setMinimumWidth(420)
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)

        total = sum(_amount(e) for e in self.expenses)
        lbl = QLabel(f"จะ Mark {len(self.expenses)} รายการ  รวม {fmt_amount(total)} บาท ว่าจ่ายแล้ว")
        lbl.setWordWrap(True)
        layout.addWidget(lbl)

        form = QFormLayout()
        self.date_edit = QDateEdit(QDate.currentDate())
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDisplayFormat("dd/MM/yyyy")
        form.addRow("วันที่จ่าย:", self.date_edit)

        self.method_cb = QComboBox()
        self.method_cb.addItems(["โอนเงิน", "เช็ค", "เงินสด", "บัตรเครดิต"])
        form.addRow("วิธีชำระ:", self.method_cb)
        layout.addLayout(form)

        self.progress = QProgressBar()
        self.progress.setVisible(False)
        layout.addWidget(self.progress)

        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.button(QDialogButtonBox.StandardButton.Ok).setText("ยืนยัน")
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        self.btn_box = btns
        layout.addWidget(btns)

    def get_values(self):
        d = self.date_edit.date()
        return (f"{d.year()}-{d.month():02d}-{d.day():02d}",
                self.method_cb.currentText())


class LineMessageDialog(QDialog):
    def __init__(self, message: str, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📱 ข้อความสรุปสำหรับ Line")
        self.setMinimumSize(480, 380)
        layout = QVBoxLayout(self)
        self.text = QTextEdit(message)
        self.text.setReadOnly(True)
        self.text.setFont(QFont("Consolas", 9))
        layout.addWidget(self.text)
        btn_copy  = QPushButton("📋 Copy ข้อความ")
        btn_close = QPushButton("ปิด")
        btn_copy.clicked.connect(self._copy)
        btn_close.clicked.connect(self.accept)
        row = QHBoxLayout()
        row.addWidget(btn_copy)
        row.addWidget(btn_close)
        layout.addLayout(row)

    def _copy(self):
        QApplication.clipboard().setText(self.text.toPlainText())
        QMessageBox.information(self, "✅", "Copy ข้อความสำเร็จแล้ว!")


# ──────────────────── Queue Tab ────────────────────

class CompanyManagerDialog(QDialog):
    """จัดการบริษัท (multi-company) — เพิ่ม/แก้/ลบ + ทดสอบ key"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("จัดการบริษัท")
        self.resize(640, 420)
        cfg = load_config()
        self._companies = [dict(c) for c in (cfg.get("companies") or [])]

        lay = QVBoxLayout(self)
        lay.addWidget(QLabel("แต่ละบริษัท = 1 API key ของ FlowAccount (1 สมุดบัญชี)"))

        self.listw = QTableWidget(0, 3)
        self.listw.setHorizontalHeaderLabels(["ชื่อบริษัท", "รหัสลูกค้า", "Client ID"])
        self.listw.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.listw.verticalHeader().setVisible(False)
        self.listw.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.listw.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        lay.addWidget(self.listw)

        btns = QHBoxLayout()
        for txt, fn in [("➕ เพิ่ม", self._add), ("✏️ แก้ไข", self._edit),
                        ("🧪 ทดสอบ", self._test), ("🗑️ ลบ", self._remove)]:
            b = QPushButton(txt)
            b.setStyleSheet("QPushButton{padding:6px 12px;border:1px solid #cbd5e1;border-radius:4px;background:white;}")
            b.clicked.connect(fn)
            btns.addWidget(b)
        btns.addStretch()
        lay.addLayout(btns)

        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Save |
                              QDialogButtonBox.StandardButton.Cancel)
        bb.accepted.connect(self._save)
        bb.rejected.connect(self.reject)
        lay.addWidget(bb)
        self._reload()

    def _reload(self):
        self.listw.setRowCount(0)
        for c in self._companies:
            r = self.listw.rowCount()
            self.listw.insertRow(r)
            self.listw.setItem(r, 0, QTableWidgetItem(c.get("label", "")))
            self.listw.setItem(r, 1, QTableWidgetItem(c.get("support_code", "")))
            self.listw.setItem(r, 2, QTableWidgetItem(c.get("client_id", "")))

    def _cur(self):
        r = self.listw.currentRow()
        return r if 0 <= r < len(self._companies) else -1

    def _add(self):
        c = CompanyEditDialog({}, self)
        if c.exec() == QDialog.DialogCode.Accepted:
            self._companies.append(c.get_data())
            self._reload()

    def _edit(self):
        i = self._cur()
        if i < 0:
            return
        c = CompanyEditDialog(self._companies[i], self)
        if c.exec() == QDialog.DialogCode.Accepted:
            self._companies[i] = c.get_data()
            self._reload()

    def _remove(self):
        i = self._cur()
        if i < 0:
            return
        if QMessageBox.question(self, "ลบบริษัท",
                f"ลบ '{self._companies[i].get('label')}' ?") == QMessageBox.StandardButton.Yes:
            self._companies.pop(i)
            self._reload()

    def _test(self):
        i = self._cur()
        if i < 0:
            return
        c = self._companies[i]
        self.setEnabled(False)
        QApplication.processEvents()
        try:
            from flowaccount import get_all_expenses
            exps = get_all_expenses(c.get("client_id", ""), c.get("client_secret", ""))
            comp = exps[0].get("company", {}) if exps else {}
            name = comp.get("companyName", "").strip() or "(ดึงได้แต่ไม่มีชื่อบริษัท)"
            web  = comp.get("companyWebsite", "")
            QMessageBox.information(self, "✅ ใช้ได้",
                f"Key นี้เชื่อมต่อได้\nบริษัท: {name}\nเว็บ: {web}\nดึงได้ {len(exps)} รายการ")
        except Exception as e:
            QMessageBox.critical(self, "❌ ใช้ไม่ได้", f"เชื่อมต่อไม่สำเร็จ:\n{e}")
        finally:
            self.setEnabled(True)

    def _save(self):
        cfg = load_config()
        cfg["companies"] = self._companies
        if cfg.get("active_company", 0) >= len(self._companies):
            cfg["active_company"] = 0
        from config import apply_active_company
        apply_active_company(cfg)
        save_config(cfg)
        self.accept()


class CompanyEditDialog(QDialog):
    """ฟอร์มกรอกข้อมูลบริษัท 1 ตัว"""
    COLORS = [("เหลือง", "#eab308"), ("ดำ/เทาเข้ม", "#334155"),
              ("เขียว", "#16a34a"), ("ฟ้า", "#2563eb"),
              ("ม่วง", "#7c3aed"), ("ส้ม", "#ea580c")]

    def __init__(self, data: dict, parent=None):
        super().__init__(parent)
        self.setWindowTitle("ข้อมูลบริษัท")
        self.resize(460, 0)
        form = QFormLayout(self)
        self.ed_label = QLineEdit(data.get("label", ""))
        self.ed_code  = QLineEdit(data.get("support_code", ""))
        self.ed_id    = QLineEdit(data.get("client_id", ""))
        self.ed_sec   = QLineEdit(data.get("client_secret", ""))
        self.cmb_color = QComboBox()
        for nm, hx in self.COLORS:
            self.cmb_color.addItem(nm, hx)
        cur = data.get("color", "#eab308")
        for idx in range(self.cmb_color.count()):
            if self.cmb_color.itemData(idx) == cur:
                self.cmb_color.setCurrentIndex(idx); break
        form.addRow("ชื่อบริษัท/แบรนด์:", self.ed_label)
        form.addRow("รหัสลูกค้า (เช่น N609850):", self.ed_code)
        form.addRow("Client ID:", self.ed_id)
        form.addRow("Client Secret:", self.ed_sec)
        form.addRow("สีประจำบริษัท:", self.cmb_color)
        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok |
                              QDialogButtonBox.StandardButton.Cancel)
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)
        form.addRow(bb)

    def get_data(self) -> dict:
        return {
            "label":         self.ed_label.text().strip(),
            "support_code":  self.ed_code.text().strip(),
            "client_id":     self.ed_id.text().strip(),
            "client_secret": self.ed_sec.text().strip(),
            "color":         self.cmb_color.currentData(),
        }


class ExpenseDetailDialog(QDialog):
    """รายละเอียดของรายการ (ข้อ 3) — line items + วันที่จ่าย + ลิงก์ดูเพิ่มใน FlowAccount"""
    def __init__(self, exp, assignments=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"รายละเอียด — {_doc_no(exp)}")
        self.resize(620, 480)
        lay = QVBoxLayout(self)

        # หัว
        head = QLabel(
            f"<b style='font-size:15px'>{_vendor(exp) or '(ไม่ระบุ)'}</b><br>"
            f"<span style='color:#64748b'>เลขที่ {_doc_no(exp)} • แบรนด์ {_brand_name(exp, assignments) or '—'}</span>")
        head.setTextFormat(Qt.TextFormat.RichText)
        lay.addWidget(head)

        # ข้อมูลสรุป
        pay = _payment_date(exp)
        info = QLabel(
            f"💰 ยอดรวม: <b>{fmt_amount(_amount(exp))} บาท</b>　|　"
            f"📅 ครบกำหนด: {fmt_date(_due(exp)) if _due(exp) else '—'}　|　"
            f"💸 วันที่จ่าย: {fmt_date(pay) if pay else '— (ยังไม่จ่าย)'}")
        info.setTextFormat(Qt.TextFormat.RichText)
        info.setStyleSheet("padding:6px 0;font-size:13px;")
        lay.addWidget(info)

        self._eid = _exp_id(exp)

        # ลิงก์แก้ไข (หน้าแก้ไขใน FlowAccount — ต้อง login เว็บ)
        link = _share_link(exp)
        if link:
            lbl_link = QLabel(f"✏️ <a href='{link}'>แก้ไขเอกสารใน FlowAccount</a>")
            lbl_link.setTextFormat(Qt.TextFormat.RichText)
            lbl_link.setOpenExternalLinks(True)
            lbl_link.setStyleSheet("font-size:13px;padding:2px 0;")
            lay.addWidget(lbl_link)

        # ลิงก์ "เปิดดูเอกสาร (แชร์)" — วางเองจากปุ่มแชร์ใน FlowAccount
        self.lbl_share = QLabel()
        self.lbl_share.setTextFormat(Qt.TextFormat.RichText)
        self.lbl_share.setOpenExternalLinks(True)
        self.lbl_share.setStyleSheet("font-size:13px;padding:2px 0;")
        self.lbl_share.setWordWrap(True)
        lay.addWidget(self.lbl_share)

        hint_share = QLabel("ลิงก์เปิดดูเอกสาร: กดปุ่ม \"🔗 ดึงอัตโนมัติ\" ระบบจะขอลิงก์แชร์จาก FlowAccount ให้เอง "
                            "(หรือวางเองก็ได้) แล้ว 💾 บันทึก")
        hint_share.setWordWrap(True)
        hint_share.setStyleSheet("color:#64748b;font-size:11px;")
        lay.addWidget(hint_share)

        self._exp = exp
        share_row = QHBoxLayout()
        self.ed_share = QLineEdit(share_links.get(self._eid))
        self.ed_share.setPlaceholderText("วางลิงก์แชร์ หรือกดปุ่ม 🔗 ดึงอัตโนมัติ")
        share_row.addWidget(self.ed_share, 1)
        btn_auto = QPushButton("🔗 ดึงอัตโนมัติ")
        btn_auto.setToolTip("ขอลิงก์แชร์จาก FlowAccount อัตโนมัติ (ไม่ต้องก็อปเอง)")
        btn_auto.setStyleSheet(
            "QPushButton{padding:5px 12px;border:none;border-radius:4px;"
            "background:#16a34a;color:white;font-size:12px;font-weight:600;}"
            "QPushButton:hover{background:#15803d;}")
        btn_auto.clicked.connect(self._auto_fetch_share)
        share_row.addWidget(btn_auto)
        btn_share = QPushButton("💾 บันทึก")
        btn_share.setStyleSheet(
            "QPushButton{padding:5px 12px;border:none;border-radius:4px;"
            "background:#2563eb;color:white;font-size:12px;font-weight:600;}"
            "QPushButton:hover{background:#1d4ed8;}")
        btn_share.clicked.connect(self._save_share_link)
        share_row.addWidget(btn_share)
        lay.addLayout(share_row)
        self._refresh_share_label()

        # ตาราง line items
        lay.addWidget(QLabel("📋 รายละเอียดรายการ:"))
        items = _item_details(exp)
        tbl = QTableWidget(len(items), 4)
        tbl.setHorizontalHeaderLabels(["รายละเอียด", "จำนวน", "ราคา/หน่วย", "รวม"])
        tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        tbl.verticalHeader().setVisible(False)
        tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        tbl.setWordWrap(True)
        for r, it in enumerate(items):
            tbl.setItem(r, 0, QTableWidgetItem(str(it["desc"])))
            tbl.setItem(r, 1, QTableWidgetItem(f"{it['qty']} {it['unit']}".strip()))
            tbl.setItem(r, 2, QTableWidgetItem(fmt_amount(it["price"]) if it["price"] != "" else ""))
            tbl.setItem(r, 3, QTableWidgetItem(fmt_amount(it["total"]) if it["total"] != "" else ""))
            tbl.setRowHeight(r, max(40, 18 * (str(it["desc"]).count("\n") + 1)))
        tbl.resizeRowsToContents()
        lay.addWidget(tbl, 1)

        # หมายเหตุ
        remark = (exp.get("remarks") or exp.get("note") or "").strip()
        if remark:
            rl = QLabel(f"📝 หมายเหตุ: {remark}")
            rl.setWordWrap(True)
            rl.setStyleSheet("color:#475569;font-size:12px;padding-top:4px;")
            lay.addWidget(rl)

        btn = QPushButton("ปิด")
        btn.clicked.connect(self.accept)
        btn.setStyleSheet("QPushButton{padding:6px 18px;border:1px solid #cbd5e1;border-radius:4px;background:white;}")
        row = QHBoxLayout(); row.addStretch(); row.addWidget(btn)
        lay.addLayout(row)

    def _refresh_share_label(self):
        url = share_links.get(self._eid)
        if url:
            self.lbl_share.setText(f"🔗 <a href='{url}'>เปิดดูเอกสาร (ลิงก์แชร์)</a>")
        else:
            self.lbl_share.setText(
                "<span style='color:#94a3b8'>🔗 เปิดดูเอกสาร: ยังไม่มีลิงก์แชร์ "
                "— ก็อปจากปุ่มแชร์ใน FlowAccount มาวางด้านล่าง</span>")

    def _save_share_link(self):
        share_links.set(self._eid, self.ed_share.text())
        self._refresh_share_label()
        activity_log.log("บันทึกลิงก์แชร์", str(self._eid))

    def _auto_fetch_share(self):
        """ขอลิงก์แชร์จาก FlowAccount อัตโนมัติผ่าน API"""
        did = self._exp.get("documentId") or self._exp.get("recordId") or self._eid
        cid, csec = _exp_creds(self._exp)
        if not cid:
            QMessageBox.warning(self, "ไม่มีข้อมูล", "ไม่พบ API key ของบริษัทนี้")
            return
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        try:
            link = get_share_link(cid, csec, did,
                                  doctype=self._exp.get("_doctype", "expense"))
        except Exception as e:
            QApplication.restoreOverrideCursor()
            QMessageBox.warning(self, "ดึงลิงก์ไม่สำเร็จ", str(e))
            return
        QApplication.restoreOverrideCursor()
        if link:
            self.ed_share.setText(link)
            share_links.set(self._eid, link)
            self._refresh_share_label()
            activity_log.log("ดึงลิงก์แชร์อัตโนมัติ", str(self._eid))
        else:
            QMessageBox.information(self, "ไม่มีลิงก์", "FlowAccount ไม่ได้ส่งลิงก์กลับมา")


class AttachDialog(QDialog):
    """แนบไฟล์เข้าเอกสารใน FlowAccount (ค่าใช้จ่าย/ใบสั่งซื้อ/ใบรับสินค้า)"""
    _TYPE_TO_DOC = {"expense": "ค่าใช้จ่าย", "po": "ใบสั่งซื้อ", "gr": "ใบรับสินค้า"}

    def __init__(self, exp, parent=None):
        super().__init__(parent)
        self.exp = exp
        self.setWindowTitle("แนบไฟล์เอกสาร")
        self.setFixedWidth(440)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 18, 20, 16)
        lay.setSpacing(10)

        dt = exp.get("_doctype", "expense")
        head = QLabel(
            f"<b style='font-size:15px'>{_doc_no(exp)}</b> "
            f"<span style='color:#16a34a'>({self._TYPE_TO_DOC.get(dt, '')})</span><br>"
            f"<span style='color:#64748b'>{_vendor(exp) or '-'}</span>")
        head.setTextFormat(Qt.TextFormat.RichText)
        lay.addWidget(head)

        lay.addWidget(QLabel("เลือกประเภทไฟล์ที่จะแนบ:"))
        self.cmb_type = QComboBox()
        self.cmb_type.addItems(["ใบเสร็จ", "สลิปโอนเงิน", "PO เซ็นแล้ว",
                                "ใบกำกับภาษี", "ใบส่งของ", "อื่นๆ"])
        lay.addWidget(self.cmb_type)

        btn_pick = QPushButton("📎 เลือกไฟล์ + อัปโหลดแนบ")
        btn_pick.setStyleSheet("QPushButton{padding:8px;border:none;border-radius:5px;"
                               "background:#7c3aed;color:white;font-weight:600;}"
                               "QPushButton:hover{background:#6d28d9;}")
        btn_pick.clicked.connect(self._pick_upload)
        lay.addWidget(btn_pick)

        btn_view = QPushButton("🔍 ดูไฟล์ที่แนบ (เปิดเอกสารในเว็บ)")
        btn_view.setStyleSheet("QPushButton{padding:7px;border:1px solid #cbd5e1;border-radius:5px;"
                               "background:white;}QPushButton:hover{background:#f1f5f9;}")
        btn_view.clicked.connect(self._view)
        lay.addWidget(btn_view)

        note = QLabel("* เลือกประเภท → เลือกไฟล์ → อัปโหลดแนบเข้า FlowAccount จริง\n"
                      "* FlowAccount ไม่มี API ดูรายการไฟล์แนบ — ปุ่ม 'ดูไฟล์ที่แนบ' "
                      "จะเปิดเอกสารในเว็บให้ดูไฟล์ที่แนบไว้")
        note.setWordWrap(True)
        note.setStyleSheet("color:#94a3b8;font-size:11px;")
        lay.addWidget(note)

        btn_close = QPushButton("ปิด")
        btn_close.setStyleSheet("QPushButton{padding:6px 18px;border:1px solid #cbd5e1;"
                                "border-radius:4px;background:white;}")
        btn_close.clicked.connect(self.reject)
        row = QHBoxLayout(); row.addStretch(); row.addWidget(btn_close)
        lay.addLayout(row)

    def _pick_upload(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "เลือกไฟล์แนบ", "",
            "รูปภาพ/PDF (*.png *.jpg *.jpeg *.pdf);;ทุกไฟล์ (*)")
        if not path:
            return
        try:
            with open(path, "rb") as f:
                data = f.read()
        except Exception as e:
            QMessageBox.critical(self, "อ่านไฟล์ไม่ได้", str(e)); return

        import mimetypes
        ext = os.path.splitext(path)[1].lower()
        mime = mimetypes.guess_type(path)[0] or "application/octet-stream"
        label = self.cmb_type.currentText()
        fn = f"{label}_{_doc_no(self.exp)}{ext}"

        if QMessageBox.warning(self, "ยืนยันแนบไฟล์",
                f"จะอัปโหลด '{os.path.basename(path)}'\n"
                f"แนบเข้าเอกสาร {_doc_no(self.exp)} ใน FlowAccount จริง\n\nยืนยัน?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel,
                QMessageBox.StandardButton.Cancel) != QMessageBox.StandardButton.Yes:
            return

        cfg = load_config()
        cid = self.exp.get("_co_id") or cfg.get("flowaccount_api_key", "")
        csec = self.exp.get("_co_secret") or cfg.get("flowaccount_secret_key", "")
        dt = self.exp.get("_doctype", "expense")
        did = self.exp.get("recordId") or self.exp.get("documentId")
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        try:
            from flowaccount import upload_attachment
            upload_attachment(cid, csec, dt, did, data, fn, mime)
        except Exception as e:
            QApplication.restoreOverrideCursor()
            QMessageBox.critical(self, "แนบไฟล์ไม่สำเร็จ", str(e)); return
        QApplication.restoreOverrideCursor()
        activity_log.log("แนบไฟล์", f"{_doc_no(self.exp)} • {label}")
        QMessageBox.information(self, "✅ สำเร็จ",
            f"แนบไฟล์ '{fn}' เข้า FlowAccount แล้ว")
        self.accept()

    def _view(self):
        dt = self.exp.get("_doctype", "expense")
        sc = self.exp.get("_support_code") or ""
        rid = self.exp.get("recordId") or self.exp.get("documentId") or ""
        seg = {"expense": "expenses", "po": "purchase-orders",
               "gr": "purchases"}.get(dt, "expenses")
        if sc and rid:
            QDesktopServices.openUrl(QUrl(
                f"https://advance.flowaccount.com/{sc}/business/{seg}/{rid}"))
        else:
            QMessageBox.information(self, "เปิดไม่ได้", "ไม่พบลิงก์เอกสาร")


class LoadingDialog(QDialog):
    """หลอดโหลด + % ใหญ่ ตอนดึงข้อมูล (requirement ข้อ 4)"""
    def __init__(self, text="กำลังโหลด...", parent=None):
        super().__init__(parent)
        self.setWindowTitle("KCash")
        self.setModal(True)
        self.setFixedWidth(380)
        # มีปุ่มกากบาท (X) บนหัวหน้าต่างให้กดปิดได้
        self.setWindowFlags(Qt.WindowType.Dialog | Qt.WindowType.CustomizeWindowHint |
                            Qt.WindowType.WindowTitleHint |
                            Qt.WindowType.WindowCloseButtonHint)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(24, 22, 24, 22)
        lay.setSpacing(14)

        self.lbl_pct = QLabel("0%")
        self.lbl_pct.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_pct.setStyleSheet("font-size:42px;font-weight:700;color:#15803d;")
        lay.addWidget(self.lbl_pct)

        self.bar = QProgressBar()
        self.bar.setRange(0, 100)
        self.bar.setValue(0)
        self.bar.setTextVisible(False)
        self.bar.setFixedHeight(16)
        self.bar.setStyleSheet(
            "QProgressBar{border:1px solid #cbd5e1;border-radius:8px;background:#f1f5f9;}"
            "QProgressBar::chunk{background:#16a34a;border-radius:7px;}")
        lay.addWidget(self.bar)

        self.lbl_text = QLabel(text)
        self.lbl_text.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_text.setStyleSheet("font-size:13px;color:#475569;")
        lay.addWidget(self.lbl_text)

        # ปุ่มยกเลิก/ปิด — กดเพื่อปิด popup ได้ทุกหน้าต่างโหลด
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self.btn_cancel = QPushButton("✕ ยกเลิก")
        self.btn_cancel.setStyleSheet(
            "QPushButton{padding:5px 16px;border:1px solid #cbd5e1;border-radius:5px;"
            "background:white;color:#64748b;font-size:12px;}"
            "QPushButton:hover{background:#fee2e2;color:#dc2626;border-color:#dc2626;}")
        self.btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(self.btn_cancel)
        btn_row.addStretch()
        lay.addLayout(btn_row)
        self._total = 0

    def set_total(self, total):
        self._total = max(1, total)

    def set_indeterminate(self, text=""):
        self.bar.setRange(0, 0)   # หมุนวน
        self.lbl_pct.setText("…")
        if text:
            self.lbl_text.setText(text)

    def set_progress(self, done, total, text=""):
        total = max(1, total)
        pct = int(done / total * 100)
        self.bar.setRange(0, 100)
        self.bar.setValue(pct)
        self.lbl_pct.setText(f"{pct}%")
        if text:
            self.lbl_text.setText(text)
        QApplication.processEvents()


class NoScrollComboBox(QComboBox):
    """ComboBox ที่ไม่รับล้อเมาส์ (กันเลื่อนตารางแล้วเผลอเปลี่ยนค่า)"""
    def wheelEvent(self, e):
        e.ignore()


class RejectedListDialog(QDialog):
    """หน้ารายการไม่อนุมัติ — ติ๊กแล้วอนุมัติกลับได้ (ใช้ร่วมกันหลายหน้า)
    หลัง exec อ่าน .approved_ids เพื่อให้ผู้เรียกรีเฟรช/นำกลับเข้าคิว"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.approved_ids = []
        self.setWindowTitle("🚫 รายการไม่อนุมัติ")
        self.resize(660, 480)
        dl = QVBoxLayout(self)
        dl.addWidget(QLabel("ติ๊ก ☑️ หน้ารายการ แล้วกด 'อนุมัติกลับ'"))

        data = rejected.load()
        self._t = QTableWidget(0, 5)
        self._t.setHorizontalHeaderLabels(["✓", "เลขเอกสาร", "ผู้รับเงิน", "ยอด (บาท)", "ไม่อนุมัติเมื่อ"])
        self._t.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self._t.setColumnWidth(0, 34); self._t.setColumnWidth(1, 120)
        self._t.setColumnWidth(3, 110); self._t.setColumnWidth(4, 130)
        self._t.verticalHeader().setVisible(False)
        self._t.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._ids = []
        self._checks = []
        for eid, info in data.items():
            r = self._t.rowCount(); self._t.insertRow(r); self._ids.append(eid)
            chk = QCheckBox(); self._checks.append(chk)
            cw = QWidget(); cl = QHBoxLayout(cw); cl.addWidget(chk)
            cl.setAlignment(Qt.AlignmentFlag.AlignCenter); cl.setContentsMargins(0, 0, 0, 0)
            self._t.setCellWidget(r, 0, cw)
            vals = [info.get("doc", ""), info.get("vendor", ""),
                    fmt_amount(info.get("amount", 0)),
                    str(info.get("at", "")).replace("T", " ")[:16]]
            for c, v in enumerate(vals, start=1):
                cell = QTableWidgetItem(v)
                if c == 3:
                    cell.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self._t.setItem(r, c, cell)

        topbar = QHBoxLayout()
        btn_all = QPushButton("☑️ ติ๊กทั้งหมด"); btn_none = QPushButton("⬜ ปลดติ๊ก")
        for b, c in ((btn_all, "#16a34a"), (btn_none, "#64748b")):
            b.setStyleSheet(
                "QPushButton{padding:4px 10px;border:1px solid %s;border-radius:4px;"
                "background:white;color:%s;font-size:12px;font-weight:600;}"
                "QPushButton:hover{background:#f1f5f9;}" % (c, c))
        btn_all.clicked.connect(lambda: [c.setChecked(True) for c in self._checks])
        btn_none.clicked.connect(lambda: [c.setChecked(False) for c in self._checks])
        topbar.addWidget(btn_all); topbar.addWidget(btn_none); topbar.addStretch()
        dl.addLayout(topbar)
        dl.addWidget(self._t, 1)
        if not data:
            empty = QLabel("— ยังไม่มีรายการที่ไม่อนุมัติ —")
            empty.setStyleSheet("color:#94a3b8;padding:8px;")
            empty.setAlignment(Qt.AlignmentFlag.AlignCenter)
            dl.addWidget(empty)

        bar = QHBoxLayout(); bar.addStretch()
        btn_app = QPushButton("✅ อนุมัติกลับ (ที่ติ๊ก)")
        btn_app.setStyleSheet(
            "QPushButton{padding:6px 14px;border:none;border-radius:4px;"
            "background:#16a34a;color:white;font-weight:700;}QPushButton:hover{background:#15803d;}")
        btn_close = QPushButton("ปิด")
        btn_close.setStyleSheet(
            "QPushButton{padding:6px 16px;border:1px solid #cbd5e1;border-radius:4px;background:white;}")
        btn_app.clicked.connect(self._approve)
        btn_close.clicked.connect(self.reject)
        bar.addWidget(btn_app); bar.addWidget(btn_close)
        dl.addLayout(bar)

    def _approve(self):
        sel = [i for i, c in enumerate(self._checks) if c.isChecked()]
        if not sel:
            QMessageBox.information(self, "ยังไม่ได้เลือก", "ติ๊กรายการก่อนครับ")
            return
        for i in sel:
            eid = self._ids[i]
            rejected.approve(eid)
            self.approved_ids.append(eid)
        QMessageBox.information(self, "✅ สำเร็จ", f"อนุมัติกลับ {len(sel)} รายการแล้ว")
        self.accept()


# ──────────────────── Sensitive Data (ซ่อนยอดเงิน) ────────────────────
_SENSITIVE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "kcash_sensitive.json")
_SENSITIVE_MASK = "●●●●●"

def _sensitive_load() -> dict:
    try:
        with open(_SENSITIVE_FILE, encoding="utf-8-sig") as f:
            return json.load(f)
    except Exception:
        return {"password_hash": "", "names": []}

def _sensitive_save(data: dict):
    with open(_SENSITIVE_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def _sensitive_hash(pw: str) -> str:
    return hashlib.sha256(pw.encode("utf-8")).hexdigest()

def _sensitive_norm(s: str) -> str:
    """normalize ชื่อ — รวมช่องว่าง/บรรทัดใหม่ให้เหลือช่องว่างเดียว + ตัดคำนำหน้าทั่วไป
    เพื่อให้จับคู่ได้แม้พิมพ์มี/ไม่มีคำนำหน้า หรือมีช่องว่างต่างกัน"""
    s = re.sub(r"\s+", " ", (s or "")).strip().lower()
    # ตัดคำนำหน้าชื่อไทย/ย่อ (น.ส./นาย/นาง/บจก./บริษัท ฯลฯ) ออกหัวข้อความ
    s = re.sub(r"^(น\.ส\.|นางสาว|นาย|นาง|ด\.ช\.|ด\.ญ\.|บจก\.|บมจ\.|หจก\.|บริษัท|ห้างหุ้นส่วน)\s*", "", s)
    return s.strip()

def _sensitive_names() -> set:
    """คืนเซตชื่อ (normalize แล้ว) ที่ต้องซ่อนยอดเงิน"""
    data = _sensitive_load()
    return {_sensitive_norm(n) for n in (data.get("names") or []) if _sensitive_norm(n)}

def _sensitive_check_pw(pw: str) -> bool:
    data = _sensitive_load()
    h = data.get("password_hash") or ""
    return bool(h) and _sensitive_hash(pw) == h

def _sensitive_is_vendor(vendor_name: str, names_set: set) -> bool:
    """จับคู่แบบยืดหยุ่น — ตรงกัน หรือชื่อหนึ่งเป็นส่วนหนึ่งของอีกชื่อ
    (เผื่อสลิปมีคำนำหน้า/วงเล็บ/ชื่อเล่นต่างจากที่กรอกไว้)"""
    if not names_set:
        return False
    v = _sensitive_norm(vendor_name)
    if not v:
        return False
    for n in names_set:
        if not n:
            continue
        if n == v or n in v or v in n:
            return True
    return False


class QueueTab(QWidget):
    status_message = pyqtSignal(str)
    company_changed = pyqtSignal()   # แจ้งเมื่อสลับบริษัท

    COLS = ["", "เลขที่เอกสาร", "ผู้รับเงิน / Vendor", "แบรนด์", "ครบกำหนด", "วันที่จ่าย", "ยอดเงิน (บาท)", "สถานะ", "Statement Match", "หมายเหตุ (พิมพ์ได้)", "ลิงก์แชร์ / ลิงก์แก้ไข", "แนบไฟล์เข้า FlowAccount", "🗑️ ลบ"]
    # ตัวเลือกสถานะที่กดเปลี่ยนได้จากเซลล์ (✅ จ่ายแล้ว = บันทึกเข้า FlowAccount จริง)
    STATUS_OPTS = ["🟡 รอจ่าย", "🔵 อนุมัติแล้ว", "🟣 จ่ายแล้วรออัพเดต",
                   "🚫 ไม่อนุมัติ", "✅ จ่ายแล้ว"]

    def __init__(self):
        super().__init__()
        self._expenses: list = []
        self._selected: set  = set()
        self._matched: dict  = {}
        self._filtered: list = []
        self._page     = 1
        self._per_page = 25
        self._sort_dir = -1   # ค่าเริ่มต้น = ใหม่→เก่า (ล่าสุดก่อน); 1 = เก่า→ใหม่
        self._pin_selected = False   # ดันรายการที่เลือกขึ้นบนสุด
        try:
            self._combined = bool(load_config().get("queue_all_companies", False))
        except Exception:
            self._combined = False
        # วงเงินจ่ายวันนี้ — โหลดจาก config (0 = ไม่จำกัด)
        try:
            self._daily_limit = float(load_config().get("daily_payment_limit", 150000) or 0)
        except Exception:
            self._daily_limit = 150000.0
        # Smart Auto-Refresh — ติดตามครั้งล่าสุดที่ refresh cache
        self._last_smart_refresh = None
        self._worker = None
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(6)

        # ── แถบบริษัท (multi-company) ──
        self._build_company_bar(layout)

        # ── Toolbar ──
        toolbar = QHBoxLayout()
        self.lbl_fetched = QLabel("ยังไม่ได้ดึงข้อมูล")
        self.lbl_fetched.setStyleSheet(f"color:{C_MUTED}; font-size:11px;")

        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("พิมพ์ชื่อ/เลขเอกสาร (หลายงานเว้นวรรค/คอมมา) แล้วกดค้นหา")
        self.search_box.setClearButtonEnabled(True)
        self.search_box.setFixedWidth(280)
        # พิมพ์แล้วกด Enter หรือปุ่มค้นหา (ไม่กรองทันทีทุกตัวอักษร)
        self.search_box.returnPressed.connect(self._apply_filter)
        # ล้างช่อง (กดกากบาท) → กลับมาแสดงทั้งหมด
        self.search_box.textChanged.connect(
            lambda t: self._apply_filter() if not t.strip() else None)
        self.btn_search = QPushButton("🔍 ค้นหา")
        self.btn_search.clicked.connect(self._apply_filter)
        self.btn_search.setStyleSheet(
            "QPushButton{padding:6px 14px;border:none;border-radius:5px;background:#2563eb;"
            "color:white;font-size:12px;font-weight:600;}QPushButton:hover{background:#1d4ed8;}")

        # ── ปุ่มเรียงวัน (ปุ่มเดียว กดสลับ) ──
        self.btn_sort = QPushButton("📅⬇ ใหม่→เก่า")
        self.btn_sort.setToolTip("คลิกเพื่อสลับการเรียงตามวันที่เอกสาร (ใหม่→เก่า / เก่า→ใหม่)")
        self.btn_sort.clicked.connect(self._toggle_sort)

        # ── ปุ่มดันรายการที่เลือกขึ้นบน ──
        self.btn_pin = QPushButton("📌 ดันที่เลือกขึ้นบน")
        self.btn_pin.setCheckable(True)
        self.btn_pin.setToolTip("แสดงรายการที่เลือกไว้ไว้บนสุดของตาราง")
        self.btn_pin.toggled.connect(self._toggle_pin)

        self.btn_refresh = QPushButton("🔄 รีเฟรช")
        self.btn_add     = QPushButton("➕ เพิ่มรายการ")
        self.btn_export_link = QPushButton("🔗 Export Link")
        self.btn_csv     = QPushButton("⬇️ Export PDF / Excel")
        self.btn_paid    = QPushButton("✅ Mark จ่ายแล้ว")
        self.btn_sharelinks = QPushButton("🔗 ดึงลิงก์แชร์")
        self.btn_monday  = QPushButton("☁️ Sync → monday.com")
        self.btn_refresh_cache = QPushButton("🔁 Refresh Cache")

        self.btn_refresh.clicked.connect(self.fetch_expenses)
        self.btn_add.clicked.connect(self._open_add_dialog)
        self.btn_export_link.clicked.connect(self._export_link)
        self.btn_csv.clicked.connect(self._download_csv)
        self.btn_paid.clicked.connect(self._mark_paid)
        self.btn_sharelinks.clicked.connect(self._fetch_share_links)
        self.btn_monday.clicked.connect(self._sync_monday)
        self.btn_refresh_cache.clicked.connect(self._refresh_monday_cache)

        for btn in (self.btn_export_link, self.btn_csv, self.btn_paid):
            btn.setEnabled(False)

        self.btn_sharelinks.setToolTip("ดึงลิงก์แชร์อัตโนมัติให้รายการที่ติ๊ก (ถ้าไม่ติ๊ก = ทั้งหน้า)")

        self._style_btn(self.btn_refresh, C_MUTED)
        self._style_btn(self.btn_add,     "#0891b2")  # cyan
        self._style_btn(self.btn_export_link, "#2563eb")  # blue
        self._style_btn(self.btn_csv,     C_PRIMARY)
        self._style_btn(self.btn_paid,    C_SUCCESS)
        self._style_btn(self.btn_sharelinks, "#0d9488")  # teal
        self._style_btn(self.btn_monday,  C_PURPLE)
        self._style_btn(self.btn_refresh_cache, "#9333ea")  # purple light

        # ปุ่มเรียงวัน — สีเขียว ปรับ text ตามสถานะ
        self.btn_sort.setStyleSheet(
            "QPushButton{padding:6px 12px;border:1.5px solid %s;border-radius:5px;"
            "background:%s;color:white;font-size:12px;font-weight:600;}"
            "QPushButton:hover{background:#15803d;}"
            % (C_SUCCESS, C_SUCCESS)
        )

        # ปุ่มดันที่เลือก — เปลี่ยนสีเมื่อเปิดใช้
        self.btn_pin.setStyleSheet(
            "QPushButton{padding:6px 12px;border:1.5px solid #cbd5e1;border-radius:5px;"
            "background:white;color:#475569;font-size:12px;font-weight:600;}"
            "QPushButton:hover{background:#f1f5f9;}"
            "QPushButton:checked{background:#f59e0b;color:white;border-color:#f59e0b;}")

        # ── ตัวกรองวันที่ (ข้อ 6) — ปรับวันที่แล้วกรองทันที ไม่ต้องติ๊ก ──
        today_q = QDate.currentDate()
        month_start = QDate(today_q.year(), today_q.month(), 1)   # ต้นเดือนปัจจุบัน
        self.date_from = QDateEdit(month_start)
        self.date_to   = QDateEdit(today_q)                       # ถึงวันนี้
        # สไตล์ปฏิทินป๊อปอัป — แถบเดือน/ปี เป็นพื้นเขียว ตัวอักษรขาว (เห็นชัด)
        cal_style = (
            "QCalendarWidget QWidget#qt_calendar_navigationbar{background:#16a34a;}"
            "QCalendarWidget QToolButton{color:white;font-size:13px;font-weight:bold;"
            "background:transparent;border:none;padding:4px 10px;}"
            "QCalendarWidget QToolButton:hover{background:#15803d;border-radius:4px;}"
            "QCalendarWidget QToolButton::menu-indicator{image:none;}"
            "QCalendarWidget QSpinBox{color:#0f172a;background:white;}"
            "QCalendarWidget QMenu{color:#0f172a;background:white;}"
            "QCalendarWidget QAbstractItemView{color:#0f172a;background:white;"
            "selection-background-color:#16a34a;selection-color:white;}"
        )
        # เริ่มต้น: ยังไม่กรองวันที่ — โชว์ข้อมูลทั้งหมด (ล่าสุด→เก่าสุด)
        # จะกรองก็ต่อเมื่อผู้ใช้เลือกวันที่/กด "เดือนนี้" เอง
        self._date_active = False
        for de in (self.date_from, self.date_to):
            de.setCalendarPopup(True)
            de.setLocale(QLocale(QLocale.Language.English))   # เลขอารบิก + ค.ศ. (ไม่ใช่เลขไทย)
            de.setDisplayFormat("dd/MM/yyyy")
            de.setFixedWidth(110)
            de.calendarWidget().setStyleSheet(cal_style)
            de.calendarWidget().setLocale(QLocale(QLocale.Language.English))  # ปี ค.ศ. ในปฏิทินด้วย
            de.dateChanged.connect(self._on_date_changed)
        self.btn_today = QPushButton("📅 เดือนนี้")
        self.btn_today.setToolTip("กรองเฉพาะช่วงเดือนปัจจุบัน → วันนี้")
        self.btn_all_dates = QPushButton("🗓️ ทุกวันที่")
        self.btn_all_dates.setToolTip("ยกเลิกกรองวันที่ — โชว์ข้อมูลทั้งหมด (ล่าสุด→เก่าสุด)")
        self.btn_all_dates.setStyleSheet(
            "QPushButton{padding:4px 10px;border:1px solid #cbd5e1;border-radius:4px;"
            "background:white;color:#475569;font-size:12px;}QPushButton:hover{background:#f1f5f9;}")
        self.btn_all_dates.clicked.connect(self._clear_date_filter)
        self.btn_today.setStyleSheet(
            "QPushButton{padding:4px 10px;border:1px solid #16a34a;border-radius:4px;"
            "background:white;color:#15803d;font-size:12px;}QPushButton:hover{background:#e7f5ed;}")
        self.btn_today.clicked.connect(self._reset_dates_today)

        # ── ตัวเลือกประเภทเอกสาร (ค่าใช้จ่าย/ใบรับสินค้า — ยกเลิกใบสั่งซื้อตามมติประชุม) ──
        self.doc_filter = QComboBox()
        self.doc_filter.addItems(["💰 ค่าใช้จ่าย", "📦 ใบรับสินค้า", "📋 ทั้งหมด"])
        self.doc_filter.setFixedWidth(150)
        self.doc_filter.setCurrentIndex(2)   # default = ทั้งหมด
        self.doc_filter.currentIndexChanged.connect(self._on_doctype_changed)
        self._po_items: list = []
        self._gr_items: list = []
        self._po_loaded = False
        self._gr_loaded = False

        # ── แถวบนสุด: ประเภท + ค้นหา + ตัวกรองวันที่ (ซ้าย) + Sync monday/Refresh Cache (ขวา) ──
        top_row = QHBoxLayout()
        top_row.addWidget(QLabel("ประเภท:"))
        top_row.addWidget(self.doc_filter)
        top_row.addSpacing(10)
        top_row.addWidget(self.search_box)
        top_row.addWidget(self.btn_search)
        top_row.addSpacing(12)
        top_row.addWidget(QLabel("📅 วันที่"))
        top_row.addWidget(self.date_from)
        top_row.addWidget(QLabel("ถึง"))
        top_row.addWidget(self.date_to)
        top_row.addWidget(self.btn_today)
        top_row.addWidget(self.btn_all_dates)
        top_row.addStretch()
        top_row.addWidget(self.btn_monday)
        top_row.addWidget(self.btn_refresh_cache)
        layout.addLayout(top_row)

        # ── Toolbar หลัก ──
        toolbar.addWidget(self.btn_sort)
        toolbar.addWidget(self.btn_pin)
        toolbar.addWidget(self.btn_refresh)
        toolbar.addWidget(self.btn_add)
        toolbar.addStretch()
        toolbar.addWidget(self.lbl_fetched)
        toolbar.addWidget(self.btn_export_link)
        toolbar.addWidget(self.btn_csv)
        toolbar.addWidget(self.btn_sharelinks)
        toolbar.addWidget(self.btn_paid)
        layout.addLayout(toolbar)

        # ── Daily Limit Panel ──
        self._build_limit_panel(layout)

        # ── Select-all row ──
        sel_row = QHBoxLayout()
        self.chk_all = QCheckBox("เลือกทั้งหมด")
        self.chk_all.stateChanged.connect(self._toggle_all)
        # ปุ่มติ๊ก/ปลดติ๊ก ทุกรายการที่กรองอยู่ (ทุกหน้า ไม่ใช่แค่หน้านี้)
        self.btn_check_all = QPushButton("☑️ ติ๊กทั้งหมด")
        self.btn_uncheck_all = QPushButton("⬜ ปลดติ๊กทั้งหมด")
        for b, c in ((self.btn_check_all, "#16a34a"), (self.btn_uncheck_all, "#64748b")):
            b.setStyleSheet(
                "QPushButton{padding:4px 10px;border:1px solid %s;border-radius:5px;"
                "background:white;color:%s;font-size:12px;font-weight:600;}"
                "QPushButton:hover{background:#f1f5f9;}" % (c, c))
        self.btn_check_all.clicked.connect(self._select_all_filtered)
        self.btn_uncheck_all.clicked.connect(self._deselect_all_filtered)
        self.lbl_count = QLabel("0 รายการที่เลือก")
        self.lbl_count.setStyleSheet(f"color:{C_PRIMARY}; font-weight:600;")
        sel_row.addWidget(self.btn_check_all)
        sel_row.addWidget(self.btn_uncheck_all)
        sel_row.addWidget(self.lbl_count)
        sel_row.addSpacing(18)
        # ── ฟิลเตอร์สถานะการจ่าย ──
        sel_row.addWidget(QLabel("สถานะ:"))
        self.status_filter = QComboBox()
        self.status_filter.addItems(["ทั้งหมด", "ยังไม่จ่าย", "🟡 รอจ่าย",
                                     "🟣 จ่ายแล้วรออัพเดต", "✅ จ่ายแล้ว",
                                     "⚠️ เกินกำหนด", "📅 ใกล้ครบกำหนด (30 วัน)",
                                     "🚫 ไม่อนุมัติ"])
        self.status_filter.setFixedWidth(210)
        self.status_filter.currentIndexChanged.connect(self._apply_filter)
        sel_row.addWidget(self.status_filter)
        sel_row.addStretch()

        # ปุ่ม Mark จ่ายแล้วรออัพเดต + ยกเลิก (ทำงานกับรายการที่ติ๊กในตารางหลัก)
        self.btn_mark_pending = QPushButton("🟣 Mark จ่ายแล้ว (รออัพเดต)")
        self.btn_mark_pending.setStyleSheet(
            "QPushButton{padding:5px 12px;border:none;border-radius:5px;"
            "background:#7c3aed;color:white;font-size:12px;font-weight:600;}"
            "QPushButton:hover{background:#6d28d9;}"
            "QPushButton:disabled{background:#cbd5e1;}")
        self.btn_mark_pending.clicked.connect(self._mark_pending_selected)
        self.btn_unmark_pending = QPushButton("↩️ ยกเลิกรออัพเดต")
        self.btn_unmark_pending.setStyleSheet(
            "QPushButton{padding:5px 12px;border:1px solid #cbd5e1;border-radius:5px;"
            "background:white;color:#475569;font-size:12px;}"
            "QPushButton:hover{background:#f1f5f9;}"
            "QPushButton:disabled{color:#cbd5e1;}")
        self.btn_unmark_pending.clicked.connect(self._unmark_pending_selected)
        sel_row.addWidget(self.btn_mark_pending)
        sel_row.addWidget(self.btn_unmark_pending)
        # ── ไม่อนุมัติ / อนุมัติ (ข้อ 2) ──
        self.btn_reject = QPushButton("🚫 ไม่อนุมัติ")
        self.btn_reject.setStyleSheet(
            "QPushButton{padding:5px 12px;border:none;border-radius:5px;"
            "background:#dc2626;color:white;font-size:12px;font-weight:600;}"
            "QPushButton:hover{background:#b91c1c;}QPushButton:disabled{background:#cbd5e1;}")
        self.btn_reject.clicked.connect(self._reject_selected)
        self.btn_approve = QPushButton("📋 ดูรายการที่ไม่อนุมัติ")
        self.btn_approve.setStyleSheet(
            "QPushButton{padding:5px 12px;border:1px solid #16a34a;border-radius:5px;"
            "background:white;color:#15803d;font-size:12px;font-weight:600;}"
            "QPushButton:hover{background:#e7f5ed;}")
        self.btn_approve.clicked.connect(self._show_rejected_list)
        sel_row.addWidget(self.btn_reject)
        sel_row.addWidget(self.btn_approve)
        layout.addLayout(sel_row)

        # ── Table ──
        self.table = QTableWidget(0, len(self.COLS))
        self.table.setHorizontalHeaderLabels(self.COLS)
        _qhh = self.table.horizontalHeader()
        _qhh.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)  # ทุกคอลัมน์ลากปรับขนาดได้อิสระ
        _qhh.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)     # ช่องติ๊ก คงที่
        _qhh.setStretchLastSection(False)
        self.table.setColumnWidth(0, 34)
        self.table.setColumnWidth(1, 150)   # เลขที่เอกสาร (กว้างพอ ไม่ตัดเป็น ...)
        self.table.setColumnWidth(2, 260)   # ผู้รับเงิน/Vendor (กว้างเริ่มต้น ลากปรับได้)
        self.table.setColumnWidth(3, 110)   # แบรนด์
        self.table.setColumnWidth(4, 100)   # ครบกำหนด
        self.table.setColumnWidth(5, 100)   # วันที่จ่าย
        self.table.setColumnWidth(6, 120)   # ยอดเงิน
        self.table.setColumnWidth(7, 150)   # สถานะ
        self.table.setColumnWidth(8, 170)   # match
        self.table.setColumnWidth(9, 200)   # หมายเหตุ (พิมพ์ได้)
        self.table.setColumnWidth(10, 190)  # ลิงก์แชร์ / ลิงก์แก้ไข
        self.table.setColumnWidth(11, 70)   # แนบไฟล์
        self.table.setColumnWidth(12, 60)   # ลบ
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        # อนุญาตแก้ไขเฉพาะช่องที่ตั้ง ItemIsEditable (คอลัมน์หมายเหตุ) — ช่องอื่นแก้ไม่ได้
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked
                                   | QAbstractItemView.EditTrigger.EditKeyPressed)
        self.table.setAlternatingRowColors(False)
        # ปิด Qt sort ในตัว (มันทำ cell widget เพี้ยน) → ทำ sort เองเมื่อคลิกหัวคอลัมน์
        self.table.setSortingEnabled(False)
        hh = self.table.horizontalHeader()
        hh.setSectionsClickable(True)
        hh.setSortIndicatorShown(True)
        hh.sectionClicked.connect(self._sort_by_column)
        self.table.setStyleSheet("QTableWidget { border: 1px solid #e2e8f0; font-size: 13px; }")
        self._loading = False
        self._sort_key_col = None   # คอลัมน์ที่ใช้เรียง (None = เรียงตามวันที่เอกสารปกติ)
        self._sort_asc = True
        self.table.cellDoubleClicked.connect(self._show_detail)   # ข้อ 3: ดับเบิลคลิกดูรายละเอียด
        self.table.cellClicked.connect(self._on_cell_click)       # คลิกเลขเอกสาร = copy
        self.table.itemChanged.connect(self._on_remark_changed)   # บันทึกหมายเหตุอัตโนมัติ
        layout.addWidget(self.table)

        # ── Summary + Pagination ──
        bottom = QHBoxLayout()
        self.lbl_summary = QLabel("")
        self.lbl_summary.setStyleSheet(f"color:{C_MUTED}; font-size:11px; padding: 2px 0;")

        self.btn_prev = QPushButton("◀ ก่อนหน้า")
        self.btn_next = QPushButton("ถัดไป ▶")
        self.lbl_page = QLabel("หน้า 1/1")
        self.lbl_page.setStyleSheet(f"color:{C_PRIMARY}; font-weight:600; font-size:12px; padding:0 8px;")
        self.btn_prev.clicked.connect(self._prev_page)
        self.btn_next.clicked.connect(self._next_page)
        for b in (self.btn_prev, self.btn_next):
            b.setStyleSheet(
                "QPushButton{padding:4px 12px;border:1px solid #cbd5e1;border-radius:4px;"
                "background:white;font-size:12px;}"
                "QPushButton:hover:!disabled{background:#f1f5f9;}"
                "QPushButton:disabled{color:#cbd5e1;border-color:#e2e8f0;}")

        # ── เลือกจำนวนรายการต่อหน้า (25/50/100) ──
        self.cmb_per_page = QComboBox()
        self.cmb_per_page.addItems(["25", "50", "100"])
        self.cmb_per_page.setCurrentText(str(self._per_page))
        self.cmb_per_page.setFixedWidth(70)
        self.cmb_per_page.currentTextChanged.connect(self._change_per_page)

        bottom.addWidget(self.lbl_summary)
        bottom.addStretch()
        bottom.addWidget(QLabel("แสดง/หน้า:"))
        bottom.addWidget(self.cmb_per_page)
        bottom.addSpacing(10)
        bottom.addWidget(self.btn_prev)
        bottom.addWidget(self.lbl_page)
        bottom.addWidget(self.btn_next)
        layout.addLayout(bottom)

    def _change_per_page(self, text):
        try:
            self._per_page = int(text)
        except (TypeError, ValueError):
            return
        self._page = 1
        self._render_page()

    def _style_btn(self, btn: QPushButton, colour: str):
        btn.setStyleSheet(f"""
            QPushButton {{
                background:{colour}; color:white; border:none;
                border-radius:5px; padding:6px 14px; font-size:13px; font-weight:500;
            }}
            QPushButton:hover:!disabled {{ background: {colour}cc; }}
            QPushButton:disabled {{ background:#94a3b8; }}
        """)

    # ── แถบบริษัท (multi-company) ──

    def _build_company_bar(self, parent_layout):
        self._company_frame = QFrame()
        self._company_frame.setStyleSheet(
            "QFrame{border:none;border-radius:6px;}")
        row = QHBoxLayout(self._company_frame)
        row.setContentsMargins(12, 8, 12, 8)

        self.lbl_company_icon = QLabel("")
        self.lbl_company_icon.setFixedSize(34, 34)
        self.lbl_company_icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
        row.addWidget(self.lbl_company_icon)

        self.lbl_company_name = QLabel("")
        self.lbl_company_name.setStyleSheet(
            "font-size:14px;font-weight:600;color:#0f172a;background:transparent;")
        row.addWidget(self.lbl_company_name)

        row.addSpacing(12)
        lbl_sw = QLabel("สลับบริษัท:")
        lbl_sw.setStyleSheet("background:transparent;color:#334155;")
        row.addWidget(lbl_sw)
        self.cmb_company = QComboBox()
        self.cmb_company.setMinimumWidth(220)
        self.cmb_company.setStyleSheet(
            "QComboBox{background:white;color:#0f172a;border:1px solid #cbd5e1;"
            "border-radius:4px;padding:4px 8px;font-size:13px;}"
            "QComboBox QAbstractItemView{background:white;color:#0f172a;"
            "selection-background-color:#e7f5ed;selection-color:#0f172a;}")
        self.cmb_company.currentIndexChanged.connect(self._on_company_changed)
        row.addWidget(self.cmb_company)

        self.btn_manage_company = QPushButton("⚙️ จัดการบริษัท")
        self.btn_manage_company.setStyleSheet(
            "QPushButton{padding:4px 10px;border:1px solid #cbd5e1;border-radius:4px;"
            "background:white;font-size:12px;}QPushButton:hover{background:#f1f5f9;}")
        self.btn_manage_company.clicked.connect(self._open_company_manager)
        row.addWidget(self.btn_manage_company)

        row.addStretch()
        parent_layout.addWidget(self._company_frame)
        self._refresh_company_bar()

    def _refresh_company_bar(self):
        """โหลดรายชื่อบริษัทใส่ combo (+ ตัวเลือก 'ทุกบริษัท') + แต้มสีตามที่เลือก"""
        cfg = load_config()
        comps = cfg.get("companies") or []
        active = cfg.get("active_company", 0)

        # combo: index 0 = ทุกบริษัท, index i+1 = บริษัทที่ i
        self.cmb_company.blockSignals(True)
        self.cmb_company.clear()
        self.cmb_company.addItem("🔗 ทุกบริษัท (รวมทุกแบรนด์)")
        for c in comps:
            label = c.get("label") or "(ไม่มีชื่อ)"
            sc = c.get("support_code")
            self.cmb_company.addItem(f"{label}" + (f"  ·  {sc}" if sc else ""))
        if self._combined:
            self.cmb_company.setCurrentIndex(0)
        elif comps:
            self.cmb_company.setCurrentIndex(min(active, len(comps) - 1) + 1)
        self.cmb_company.blockSignals(False)

        try:
            from config import _debug_log
            _debug_log(f"_refresh_company_bar: combo_items={self.cmb_company.count()} "
                       f"companies={len(comps)} combined={self._combined}")
        except Exception:
            pass

        if self._combined:
            self.lbl_company_name.setText("ทุกบริษัท (รวมทุกแบรนด์)")
            self.lbl_company_name.setStyleSheet(
                "font-size:14px;font-weight:600;color:#0f172a;background:transparent;")
            self.lbl_company_icon.setText("ALL")
            self.lbl_company_icon.setStyleSheet(
                "background:#475569;color:white;border-radius:17px;"
                "font-weight:600;font-size:10px;")
            self._company_frame.setStyleSheet(
                "QFrame{background:#f1f5f9;border:1px solid #cbd5e1;"
                "border-left:5px solid #475569;border-radius:6px;}"
                "QLabel{background:transparent;}")
        elif comps and 0 <= active < len(comps):
            c = comps[active]
            color = c.get("color") or "#16a34a"
            self.lbl_company_name.setText(c.get("label") or "")
            self.lbl_company_name.setStyleSheet(
                "font-size:14px;font-weight:600;color:#0f172a;background:transparent;")
            initials = "".join(w[0] for w in (c.get("label") or "?").split()[:2]).upper()[:3]
            self.lbl_company_icon.setText(initials or "?")
            self.lbl_company_icon.setStyleSheet(
                f"background:{color};color:white;border-radius:17px;"
                "font-weight:600;font-size:12px;")
            self._company_frame.setStyleSheet(
                "QFrame{background:#f8fafc;border:1px solid #e2e8f0;"
                f"border-left:5px solid {color};border-radius:6px;}}"
                "QLabel{background:transparent;}")
        else:
            self.lbl_company_name.setText("ยังไม่ได้ตั้งค่าบริษัท — กด ⚙️ จัดการบริษัท")
            self.lbl_company_name.setStyleSheet(
                "font-size:13px;font-weight:600;color:#92400e;background:transparent;")
            self.lbl_company_icon.setText("")
            self._company_frame.setStyleSheet(
                "QFrame{background:#fef9c3;border:1px solid #fde68a;border-radius:6px;}"
                "QLabel{background:transparent;}")

    def _on_company_changed(self, idx: int):
        if idx < 0:
            return
        cfg = load_config()
        comps = cfg.get("companies") or []

        if idx == 0:
            # เลือก "ทุกบริษัท"
            if self._combined:
                return
            self._combined = True
            cfg["queue_all_companies"] = True
            save_config(cfg)
            name = "ทุกบริษัท"
        else:
            ci = idx - 1
            if not (0 <= ci < len(comps)):
                return
            if (not self._combined) and ci == cfg.get("active_company", 0):
                return
            self._combined = False
            cfg["queue_all_companies"] = False
            save_config(cfg)
            set_active_company(ci)
            name = comps[ci].get("label", "")

        self._refresh_company_bar()
        self._expenses = []
        self._selected.clear()
        self.table.setRowCount(0)
        self.status_message.emit(f"เปลี่ยนเป็น: {name} — กำลังดึงข้อมูล…")
        activity_log.log("สลับบริษัท", f"เปลี่ยนเป็น {name}")
        self.company_changed.emit()
        self.fetch_expenses()

    def _open_company_manager(self):
        dlg = CompanyManagerDialog(self)
        dlg.exec()
        self._refresh_company_bar()

    # ── Daily Limit Panel ──

    def _build_limit_panel(self, parent_layout):
        """สร้าง panel แสดงวงเงินจ่ายวันนี้ + progress bar"""
        from PyQt6.QtWidgets import QProgressBar

        frame = QFrame()
        frame.setStyleSheet(
            "QFrame{background:#f0fdf4;border:1.5px solid #bbf7d0;border-radius:8px;}"
        )
        v = QVBoxLayout(frame)
        v.setContentsMargins(10, 8, 10, 8)
        v.setSpacing(6)

        # Row 1: label + ช่องวงเงิน + ปุ่ม
        r1 = QHBoxLayout()
        lbl = QLabel("💰 <b>วงเงินจ่ายวันนี้:</b>")
        lbl.setStyleSheet("color:#15803d; font-size:13px; background:transparent;")
        r1.addWidget(lbl)

        self.ed_limit = QLineEdit(self._format_limit_text())
        self.ed_limit.setFixedWidth(130)
        self.ed_limit.setAlignment(Qt.AlignmentFlag.AlignRight)
        self.ed_limit.setStyleSheet(
            "QLineEdit{border:1.5px solid #16a34a;border-radius:4px;padding:4px 8px;"
            "background:white;font-size:13px;font-weight:600;color:#15803d;}"
            "QLineEdit:focus{border-color:#15803d;}")
        self.ed_limit.editingFinished.connect(self._on_limit_edited)
        r1.addWidget(self.ed_limit)

        r1.addWidget(QLabel("บาท"))

        self.btn_limit_adjust = QPushButton("🔧 ปรับ")
        self.btn_limit_adjust.setStyleSheet(
            "QPushButton{padding:4px 10px;border:1px solid #cbd5e1;border-radius:4px;"
            "background:white;font-size:12px;}QPushButton:hover{background:#e7f5ed;}")
        self.btn_limit_adjust.clicked.connect(self._open_adjust_limit_dialog)
        r1.addWidget(self.btn_limit_adjust)

        self.btn_auto_queue = QPushButton("🤖 จัดคิวอัตโนมัติ")
        self._style_btn(self.btn_auto_queue, "#eab308")  # accent yellow
        self.btn_auto_queue.clicked.connect(self._auto_queue)
        r1.addWidget(self.btn_auto_queue)

        self.btn_view_plan = QPushButton("📋 ดูตารางคิว")
        self.btn_view_plan.setStyleSheet(
            "QPushButton{padding:4px 10px;border:1px solid #16a34a;border-radius:4px;"
            "background:white;color:#15803d;font-size:12px;font-weight:600;}"
            "QPushButton:hover{background:#e7f5ed;}")
        self.btn_view_plan.clicked.connect(self._open_queue_plan)
        r1.addWidget(self.btn_view_plan)

        self.btn_queue_history = QPushButton("📜 ประวัติการจัดคิว")
        self.btn_queue_history.setStyleSheet(
            "QPushButton{padding:4px 10px;border:1px solid #6366f1;border-radius:4px;"
            "background:white;color:#4338ca;font-size:12px;font-weight:600;}"
            "QPushButton:hover{background:#eef2ff;}")
        self.btn_queue_history.clicked.connect(lambda: QueueHistoryDialog(self).exec())
        r1.addWidget(self.btn_queue_history)

        r1.addStretch()
        v.addLayout(r1)

        # Row 2: progress bar + ข้อความสรุป
        r2 = QHBoxLayout()
        self.progress_limit = QProgressBar()
        self.progress_limit.setRange(0, 100)
        self.progress_limit.setValue(0)
        self.progress_limit.setTextVisible(False)
        self.progress_limit.setFixedHeight(14)
        self.progress_limit.setMinimumWidth(200)
        r2.addWidget(self.progress_limit, 1)

        self.lbl_limit_info = QLabel("เลือก 0 รายการ • 0.00 บาท")
        self.lbl_limit_info.setStyleSheet(
            "color:#15803d; font-size:12px; font-weight:600; background:transparent;")
        r2.addWidget(self.lbl_limit_info)
        v.addLayout(r2)

        parent_layout.addWidget(frame)
        self._set_progress_color(0)

    def _format_limit_text(self) -> str:
        if self._daily_limit <= 0:
            return "ไม่จำกัด"
        return f"{self._daily_limit:,.0f}"

    def _set_progress_color(self, pct: float):
        if self._daily_limit <= 0:
            color = "#94a3b8"   # grey - unlimited
        elif pct < 70:
            color = "#16a34a"
        elif pct < 95:
            color = "#eab308"
        elif pct <= 100:
            color = "#f97316"
        else:
            color = "#dc2626"
        self.progress_limit.setStyleSheet(f"""
            QProgressBar{{border:1px solid #cbd5e1;border-radius:4px;
                background:#f8fafc;text-align:center;}}
            QProgressBar::chunk{{background:{color};border-radius:3px;}}
        """)

    def _on_limit_edited(self):
        """อัพเดตวงเงินทันที่ที่บัญชีพิมพ์ในช่อง"""
        txt = self.ed_limit.text().strip().lower()
        if not txt or "ไม่จำกัด" in txt or txt in ("0", "∞", "ไม่จำกัด"):
            self._daily_limit = 0
        else:
            try:
                # ตัด comma + บาท ออก
                cleaned = txt.replace(",", "").replace("บาท", "").strip()
                self._daily_limit = max(0.0, float(cleaned))
            except Exception:
                pass
        # แสดงผลใหม่ในรูปแบบที่สวย
        self.ed_limit.setText(self._format_limit_text())
        self._update_limit_panel()

    def _selected_total(self) -> float:
        # ไม่นับรายการที่ Mark จ่ายแล้วรออัพเดต (requirement ข้อ 5)
        return sum(_amount(e) for e in self._expenses
                   if _exp_id(e) in self._selected
                   and not paid_pending.is_pending(_exp_id(e)))

    def _update_limit_panel(self):
        """recompute + update UI ของ panel"""
        total = self._selected_total()
        n     = len(self._selected)

        if self._daily_limit <= 0:
            self.progress_limit.setValue(0)
            self.lbl_limit_info.setText(
                f"เลือก {n} รายการ • {fmt_amount(total)} บาท   (♾️ ไม่จำกัด)")
            self.lbl_limit_info.setStyleSheet(
                "color:#475569; font-size:12px; font-weight:600; background:transparent;")
            self._set_progress_color(0)
            return

        pct = (total / self._daily_limit) * 100 if self._daily_limit > 0 else 0
        self.progress_limit.setValue(min(int(pct), 100))
        self._set_progress_color(pct)

        remain = self._daily_limit - total
        if pct > 100:
            color = "#dc2626"
            icon  = "🔴"
            tail  = f" — เกิน {fmt_amount(-remain)} บาท!"
        elif pct >= 95:
            color = "#ea580c"; icon = "🟠"; tail = f" — เหลือ {fmt_amount(remain)} บาท"
        elif pct >= 70:
            color = "#a16207"; icon = "🟡"; tail = f" — เหลือ {fmt_amount(remain)} บาท"
        else:
            color = "#15803d"; icon = "🟢"; tail = f" — เหลือ {fmt_amount(remain)} บาท"

        self.lbl_limit_info.setText(
            f"เลือก {n} รายการ • {fmt_amount(total)} บาท   {icon} {int(pct)}%{tail}")
        self.lbl_limit_info.setStyleSheet(
            f"color:{color}; font-size:12px; font-weight:600; background:transparent;")

    # ── Data ──

    def set_matched(self, matched: dict):
        self._matched = matched
        self._render_page()

    def fetch_expenses(self):
        cfg = load_config()
        comps = cfg.get("companies") or []
        if not cfg.get("flowaccount_api_key") and not comps:
            QMessageBox.warning(self, "ยังไม่ตั้งค่า", "กรุณาตั้งค่า FlowAccount API Key ก่อนครับ")
            return

        self.btn_refresh.setEnabled(False)
        self.btn_refresh.setText("⏳ กำลังดึง...")

        # หลอดโหลด % ใหญ่ (requirement ข้อ 4)
        self._loading = LoadingDialog("กำลังดึงข้อมูล...", self)
        self._loading.show()
        QApplication.processEvents()

        if self._combined and comps:
            # โหมดรวมทุกแบรนด์ — ดึงจากทุกบริษัทมารวมกัน
            self._loading.set_total(len([c for c in comps if c.get("client_id")]))
            self._worker = FetchAllWorker(comps)
            self._worker.progress.connect(self._loading.set_progress)
        else:
            # บริษัทเดียว — ติดป้ายแบรนด์ของบริษัทที่เลือก
            active = cfg.get("active_company", 0)
            company = comps[active] if (0 <= active < len(comps)) else {}
            self._loading.set_indeterminate(f"กำลังดึง {company.get('label','FlowAccount')}...")
            self._worker = FetchWorker(cfg, company)
        self._worker.done.connect(self._on_fetch_done)
        self._worker.error.connect(self._on_fetch_error)
        self._worker.start()

    def _on_fetch_done(self, expenses: list):
        if getattr(self, "_loading", None):
            self._loading.close(); self._loading = None
        self._expenses = list(expenses)
        for e in self._expenses:
            e["_doctype"] = "expense"
        # รีเฟรช → ให้ใบรับสินค้าโหลดใหม่เมื่อเลือกดูครั้งหน้า
        self._po_loaded = self._gr_loaded = False
        self._po_items = []; self._gr_items = []
        self._load_manual_into_expenses()   # รวมรายการที่เพิ่มเองด้วย
        # ถ้าเลือกประเภทที่ต้องใช้ใบรับสินค้า (ใบรับสินค้า/ทั้งหมด) → โหลดต่อเลย
        dt = self.doc_filter.currentIndex() if hasattr(self, "doc_filter") else 0
        if dt in (1, 2):
            self._lazy_fetch_docs(False, True)
        self._apply_filter()
        ts = datetime.now().strftime("%H:%M:%S")
        self.lbl_fetched.setText(f"อัพเดตล่าสุด {ts}")
        self.btn_refresh.setEnabled(True)
        self.btn_refresh.setText("🔄 รีเฟรช")
        manual_n = sum(1 for e in self._expenses if is_manual(e))
        self.status_message.emit(
            f"ดึงข้อมูลสำเร็จ {len(expenses)} รายการ"
            + (f" + ที่เพิ่มเอง {manual_n} รายการ" if manual_n else ""))
        activity_log.log("ดึงข้อมูล", f"ดึงค่าใช้จ่าย {len(expenses)} รายการ",
                         "ทุกบริษัท" if self._combined else "")

    def _on_fetch_error(self, msg: str):
        if getattr(self, "_loading", None):
            self._loading.close(); self._loading = None
        self.btn_refresh.setEnabled(True)
        self.btn_refresh.setText("🔄 รีเฟรช")
        QMessageBox.critical(self, "เกิดข้อผิดพลาด", msg)
        self.status_message.emit("ดึงข้อมูลล้มเหลว")
        dev_log(f"[ERROR] ดึงข้อมูลล้มเหลว: {msg}", "err")

    def get_expenses(self) -> list:
        return self._expenses

    # ── รวมหลายประเภทเอกสารไว้หน้าเดียว ──
    def _current_docs(self) -> list:
        """ชุดเอกสารตามประเภทที่เลือก (0=ค่าใช้จ่าย 1=ใบรับสินค้า 2=ทั้งหมด)"""
        dt = self.doc_filter.currentIndex() if hasattr(self, "doc_filter") else 0
        if dt == 1:
            return self._gr_items
        if dt == 2:
            return list(self._expenses) + list(self._gr_items)
        return self._expenses

    def _on_doctype_changed(self):
        dt = self.doc_filter.currentIndex()
        need_gr = dt in (1, 2) and not self._gr_loaded
        if need_gr:
            self._lazy_fetch_docs(False, need_gr)
        else:
            self._apply_filter()

    def _lazy_fetch_docs(self, need_po, need_gr):
        """ดึงใบสั่งซื้อ/ใบรับสินค้า แบบ background (ไม่บล็อกจอ — คิวจ่ายโชว์ได้เลย)"""
        w = getattr(self, "_doc_worker", None)
        if w is not None and w.isRunning():
            return
        self._apply_filter()   # โชว์ที่มีอยู่ (เช่นค่าใช้จ่าย) ก่อนทันที
        cfg = load_config()
        comps = cfg.get("companies") or []
        self.status_message.emit("⏳ กำลังโหลดใบสั่งซื้อ/ใบรับสินค้า... (เบื้องหลัง)")
        dev_log("เริ่มโหลดใบสั่งซื้อ/ใบรับสินค้า (background)", "info")
        self._doc_worker = DocFetchWorker(cfg, comps, self._combined, need_po, need_gr)
        self._doc_worker.done.connect(self._on_docs_fetched)
        self._doc_worker.error.connect(self._on_fetch_error)
        self._doc_worker.start()

    def _on_docs_fetched(self, pos, grs, did_po, did_gr):
        if did_po:
            for p in pos:
                p["_doctype"] = "po"
            # ซ่อน PO ที่แปลงเป็นค่าใช้จ่าย (EXP) แล้ว — EXP โชว์สถานะจริงแทน กันโชว์ซ้ำ
            self._po_items = [p for p in pos if not _doc_converted_to_expense(p)]
            self._po_loaded = True
        if did_gr:
            for g in grs:
                g["_doctype"] = "gr"
            # ใบรับสินค้า (RI) แสดงทุกใบ — มีสถานะจ่ายของตัวเอง (ชำระเงินแล้ว/รออนุมัติ)
            self._gr_items = grs
            self._gr_loaded = True
        self._apply_filter()
        self.status_message.emit(f"ดึงใบรับสินค้าแล้ว — {len(self._gr_items)} รายการ")

    def _toggle_sort(self):
        """กดปุ่มเดียวเพื่อสลับทิศทางการเรียงวันครบกำหนด"""
        self._sort_key_col = None   # กลับมาเรียงตามวันที่เอกสาร
        self.table.horizontalHeader().setSortIndicatorShown(False)
        self._sort_dir = -1 if self._sort_dir == 1 else 1
        self.btn_sort.setText(
            "📅⬆ เก่า→ใหม่" if self._sort_dir == 1 else "📅⬇ ใหม่→เก่า"
        )
        self._apply_filter()

    def _sort_key(self, col, e):
        """คีย์สำหรับเรียงแต่ละคอลัมน์ (ยอดเงิน=ตัวเลข / วันที่=ISO / อื่นๆ=ข้อความ)"""
        if col == 1:   # เลขที่เอกสาร
            return _doc_no(e).lower()
        if col == 2:   # ผู้รับเงิน
            return (_vendor(e) or "").lower()
        if col == 3:   # แบรนด์
            return (_brand_name(e, load_brands()) or "").lower()
        if col == 4:   # ครบกำหนด
            return (_due(e) or "")[:10] or "9999"
        if col == 5:   # วันที่จ่าย
            return (_payment_date(e) or "")[:10] or "9999"
        if col == 6:   # ยอดเงิน — ตัวเลข
            return _amount(e)
        if col == 7:   # สถานะ
            return self._status_category(e)
        if col == 9:   # หมายเหตุ
            return (remarks.get(_exp_id(e)) or "").lower()
        return _doc_no(e).lower()

    def _sort_by_column(self, col):
        """คลิกหัวคอลัมน์ → เรียงตามคอลัมน์นั้น (สลับ น้อย→มาก / มาก→น้อย)"""
        if col in (0, 8, 10, 11, 12):   # ช่องติ๊ก/match/ลิงก์/แนบ/ลบ — ไม่เรียง
            return
        if self._sort_key_col == col:
            self._sort_asc = not self._sort_asc
        else:
            self._sort_key_col = col
            self._sort_asc = True
        hh = self.table.horizontalHeader()
        hh.setSortIndicatorShown(True)
        hh.setSortIndicator(col, Qt.SortOrder.AscendingOrder if self._sort_asc
                            else Qt.SortOrder.DescendingOrder)
        self._apply_filter()

    def _toggle_pin(self, checked):
        """เปิด/ปิด การดันรายการที่เลือกขึ้นบนสุด"""
        self._pin_selected = bool(checked)
        self._apply_filter()

    def _open_add_dialog(self):
        """เปิด pop-up เพิ่มรายการเอง"""
        dlg = AddExpenseDialog(self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        data = dlg.get_data()
        saved = add_manual(data)
        # รวมเข้ารายการในตาราง
        self._expenses.append(saved)
        self._apply_filter()
        self.status_message.emit(f"✅ เพิ่มรายการ '{saved['contactName']}' แล้ว")

    def _load_manual_into_expenses(self):
        """โหลดรายการที่เพิ่มเองมารวมกับ expenses"""
        existing_ids = {_exp_id(e) for e in self._expenses}
        for m in load_manual():
            if m.get("id") not in existing_ids:
                self._expenses.append(m)

    def _on_date_changed(self, *_):
        """ผู้ใช้เลือกวันที่เอง → เปิดโหมดกรองวันที่ แล้วกรองทันที"""
        self._date_active = True
        self._apply_filter()

    def _clear_date_filter(self):
        """ยกเลิกกรองวันที่ — กลับไปโชว์ข้อมูลทั้งหมด (ล่าสุด→เก่าสุด)"""
        self._date_active = False
        self._apply_filter()

    def _reset_dates_today(self):
        """กรองเฉพาะช่วงเดือนปัจจุบัน → วันนี้"""
        t = QDate.currentDate()
        self.date_from.blockSignals(True)
        self.date_to.blockSignals(True)
        self.date_from.setDate(QDate(t.year(), t.month(), 1))
        self.date_to.setDate(t)
        self.date_from.blockSignals(False)
        self.date_to.blockSignals(False)
        self._date_active = True   # กด "เดือนนี้" = เริ่มกรองวันที่
        self._apply_filter()

    def _status_category(self, e):
        """หมวดสถานะการจ่าย (ตรงกับที่โชว์ในคอลัมน์สถานะ)"""
        eid = _exp_id(e)
        st = _status(e)
        if paid_pending.is_pending(eid):
            return "pending"        # จ่ายแล้วรออัพเดต
        if "paid" in st or _payment_date(e):
            return "paid"           # จ่ายแล้ว
        if "overdue" in st:
            return "overdue"
        if "approved" in st:
            return "approved"
        return "waiting"            # รอจ่าย

    def _apply_filter(self):
        from datetime import timedelta
        today  = date.today()
        past7  = (today - timedelta(days=7)).isoformat()
        ahead  = (today + timedelta(days=30)).isoformat()
        today_s = today.isoformat()

        filtered = list(self._current_docs())

        def _docdate(e):
            return (e.get("publishedOn") or e.get("createdDate")
                    or _due(e) or "")[:10]

        # ── ตัวกรองวันที่ (ข้อ 6) — กรองเฉพาะเมื่อผู้ใช้เลือกวันที่/กด "เดือนนี้" ──
        # ค่าเริ่มต้น: ไม่กรอง → โชว์ข้อมูลทั้งหมด (เรียงล่าสุด→เก่าสุด)
        if getattr(self, "_date_active", False):
            d_from = self.date_from.date().toString("yyyy-MM-dd")
            d_to   = self.date_to.date().toString("yyyy-MM-dd")
            filtered = [e for e in filtered if d_from <= _docdate(e) <= d_to]

        # ── ฟิลเตอร์สถานะ/กำหนดจ่าย (รวมเป็นอันเดียว) ──
        sc = self.status_filter.currentIndex() if hasattr(self, "status_filter") else 0
        # รายการ "ไม่อนุมัติ" (ข้อ 2) — แสดงเฉพาะตอนเลือกฟิลเตอร์นี้ ไม่งั้นซ่อนทุกมุมมอง
        if sc == 7:
            filtered = [e for e in filtered if rejected.is_rejected(_exp_id(e))]
        else:
            filtered = [e for e in filtered if not rejected.is_rejected(_exp_id(e))]
        if sc == 1:      # ยังไม่จ่าย (ไม่รวมจ่ายแล้ว/รออัพเดต)
            filtered = [e for e in filtered
                        if self._status_category(e) in ("waiting", "overdue", "approved")]
        elif sc == 2:    # รอจ่าย
            filtered = [e for e in filtered if self._status_category(e) == "waiting"]
        elif sc == 3:    # จ่ายแล้วรออัพเดต
            filtered = [e for e in filtered if self._status_category(e) == "pending"]
        elif sc == 4:    # จ่ายแล้ว
            filtered = [e for e in filtered if self._status_category(e) == "paid"]
        elif sc == 5:    # เกินกำหนด
            filtered = [e for e in filtered
                        if (_due(e) or "")[:10] and (_due(e) or "")[:10] < today_s
                        and self._status_category(e) != "paid"]
        elif sc == 6:    # ใกล้ครบกำหนด (30 วัน)
            filtered = [e for e in filtered
                        if past7 <= (_due(e) or "")[:10] <= ahead]

        # เรียง: ถ้าคลิกหัวคอลัมน์ → เรียงตามคอลัมน์นั้น / ไม่งั้นเรียงตามวันที่เอกสาร
        col = getattr(self, "_sort_key_col", None)
        if col is not None:
            filtered.sort(key=lambda e: self._sort_key(col, e),
                          reverse=not self._sort_asc)
        elif self._sort_dir == 1:
            filtered.sort(key=lambda e: _docdate(e) or "9999-12-31")
        else:
            filtered.sort(key=lambda e: _docdate(e) or "0000-00-00", reverse=True)

        # ── ค้นหา (รองรับหลายคำพร้อมกัน — เว้นวรรค/คอมมา = หลายงาน) ──
        q = self.search_box.text().strip().lower()
        if q:
            terms = [t for t in re.split(r"[,\s]+", q) if t]
            if terms:
                filtered = [e for e in filtered
                            if any(t in _vendor(e).lower() or t in _doc_no(e).lower()
                                   or t in (remarks.get(_exp_id(e)) or "").lower()
                                   for t in terms)]

        # ── ดันรายการที่เลือกขึ้นบนสุด (คงลำดับเดิมไว้) ──
        if self._pin_selected and self._selected:
            picked  = [e for e in filtered if _exp_id(e) in self._selected]
            rest    = [e for e in filtered if _exp_id(e) not in self._selected]
            filtered = picked + rest

        self._filtered = filtered
        self._page = 1
        self._render_page()

    def _render_page(self):
        """แสดงผลเฉพาะหน้าปัจจุบัน (25 รายการ/หน้า)"""
        per   = self._per_page
        total = len(self._filtered)
        pages = max(1, (total + per - 1) // per)
        self._page = min(max(1, self._page), pages)
        start = (self._page - 1) * per
        page_items = self._filtered[start:start + per]
        self._page_rows = page_items   # ใช้ตอนดับเบิลคลิกดูรายละเอียด

        self._populate(page_items)

        total_amt = sum(_amount(e) for e in self._filtered)
        self.lbl_summary.setText(
            f"ทั้งหมด {total} รายการ  |  รวมยอด {fmt_amount(total_amt)} บาท"
            f"  |  แสดง {len(page_items)} รายการ")
        self.lbl_page.setText(f"หน้า {self._page}/{pages}")
        self.btn_prev.setEnabled(self._page > 1)
        self.btn_next.setEnabled(self._page < pages)

        # ดึงลิงก์แชร์อัตโนมัติให้รายการในหน้านี้ที่ยังไม่มี (เบื้องหลัง)
        self._auto_fetch_missing_links(page_items)

    def _auto_fetch_missing_links(self, rows):
        """ดึงลิงก์แชร์ที่ยังไม่มีของรายการในหน้าปัจจุบัน แบบเงียบๆ เบื้องหลัง"""
        w = getattr(self, "_sl_auto_worker", None)
        if w is not None and w.isRunning():
            return
        if not hasattr(self, "_sl_tried"):
            self._sl_tried = set()
        pool = [e for e in rows
                if not share_links.get(_exp_id(e))
                and _exp_id(e) not in self._sl_tried
                and not is_manual(e)
                and (e.get("documentId") or e.get("recordId"))]
        if not pool:
            return
        for e in pool:
            self._sl_tried.add(_exp_id(e))   # กันดึงซ้ำวนลูป
        self._sl_auto_worker = ShareLinkWorker(pool, self)
        self._sl_auto_worker.finished_all.connect(self._on_auto_links_done)
        self._sl_auto_worker.start()

    def _on_auto_links_done(self, ok, fail):
        dev_log(f"ดึงลิงก์แชร์อัตโนมัติ: สำเร็จ {ok}" + (f" ล้มเหลว {fail}" if fail else ""),
                "ok" if not fail else "warn")
        if ok:
            self._render_page()   # โชว์ลิงก์ที่เพิ่งดึงมา

    def _prev_page(self):
        if self._page > 1:
            self._page -= 1
            self._render_page()

    def _next_page(self):
        per   = self._per_page
        pages = max(1, (len(self._filtered) + per - 1) // per)
        if self._page < pages:
            self._page += 1
            self._render_page()

    def _populate(self, expenses: list):
        today = today_str()
        cfg   = load_config()
        brand_list = cfg.get("brand_list") or ["Hollywood", "D-MAT CNC", "DEKO", "Yellow kitchen", "DDCUT"]
        assignments = load_brands()
        self._loading = True   # กันบันทึกหมายเหตุระหว่างเติมตาราง
        self.table.setSortingEnabled(False)
        self.table.setRowCount(0)
        total = 0.0

        for exp in expenses:
            row = self.table.rowCount()
            self.table.insertRow(row)
            eid = _exp_id(exp)
            due = _due(exp)[:10] if _due(exp) else ""
            amt = _amount(exp)
            total += amt
            st  = _status(exp)

            # background colour
            if due and due < today:
                bg = C_OVERDUE
            elif due and today <= due <= (date.today().strftime("%Y-%m-") + "99")[:10]:
                # within 3 days
                diff = (datetime.strptime(due, "%Y-%m-%d").date() - date.today()).days
                bg = C_DUE_SOON if diff <= 3 else None
            else:
                bg = None

            def cell(text, align=Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft):
                item = QTableWidgetItem(text)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                item.setTextAlignment(align)
                if bg:
                    item.setBackground(QBrush(bg))
                return item

            # col 0: checkbox
            chk = QCheckBox()
            chk.setChecked(eid in self._selected)
            chk.setProperty("exp_id", eid)
            chk.stateChanged.connect(self._on_check_changed)
            chk_widget = QWidget()
            chk_layout = QHBoxLayout(chk_widget)
            chk_layout.addWidget(chk)
            chk_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            chk_layout.setContentsMargins(0, 0, 0, 0)
            self.table.setCellWidget(row, 0, chk_widget)

            dt = exp.get("_doctype", "expense")
            type_icon = {"po": "📄 ", "gr": "📦 "}.get(dt, "")
            prefix = ("✏️ " if is_manual(exp) else "") + type_icon
            self.table.setItem(row, 1, cell(prefix + _doc_no(exp)))
            self.table.setItem(row, 2, cell(_vendor(exp) or "(ไม่ระบุ)"))

            # brand → ขึ้นอัตโนมัติจากบริษัทต้นทาง (ไม่ต้องเลือกเอง)
            current_brand = _brand_name(exp, assignments)
            bcolor = exp.get("_brand_color") or "#334155"
            bitem = cell(current_brand or "—",
                         Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignHCenter)
            if current_brand:
                bitem.setForeground(QBrush(QColor(bcolor)))
                bf = bitem.font(); bf.setBold(True); bitem.setFont(bf)
            self.table.setItem(row, 3, bitem)

            self.table.setItem(row, 4, cell(fmt_date(due) if due else "-"))
            # วันที่จ่าย — จาก payments ใน FlowAccount
            pay = _payment_date(exp)
            self.table.setItem(row, 5, cell(fmt_date(pay) if pay else "—",
                               Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignHCenter))
            amt_item = cell(fmt_amount(amt),
                            Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignRight)
            self.table.setItem(row, 6, amt_item)

            # status — กดเปลี่ยนได้จากเซลล์ (dropdown). 'จ่ายแล้ว' บันทึกเข้า FlowAccount จริง
            if "paid" in st or pay:
                self.table.removeCellWidget(row, 7)
                self.table.setItem(row, 7, cell("✅ จ่ายแล้ว"))
            else:
                if rejected.is_rejected(eid):
                    cur = "🚫 ไม่อนุมัติ"
                elif paid_pending.is_pending(eid):
                    cur = "🟣 จ่ายแล้วรออัพเดต"
                elif "approved" in st:
                    cur = "🔵 อนุมัติแล้ว"
                else:
                    cur = "🟡 รอจ่าย"
                combo = NoScrollComboBox()
                combo.addItems(self.STATUS_OPTS)
                combo.setCurrentText(cur)
                combo.setStyleSheet("QComboBox{font-size:12px;padding:2px 4px;border:1px solid #e2e8f0;"
                                    "border-radius:4px;background:white;}")
                combo.currentTextChanged.connect(
                    lambda txt, e=exp: self._on_status_changed(e, txt))
                self.table.setItem(row, 7, QTableWidgetItem())  # เซลล์ว่างรองรับ widget
                self.table.setCellWidget(row, 7, combo)

            # col 8: match
            stmt = self._matched.get(eid)
            if stmt:
                match_text = f"✅ {stmt['date']}  {fmt_amount(stmt['debit'] or stmt['credit'])}"
            else:
                match_text = "—"
            self.table.setItem(row, 8, cell(match_text))

            # col 9: หมายเหตุ — พิมพ์ได้ บันทึกอัตโนมัติ
            rm_item = QTableWidgetItem(remarks.get(eid))
            rm_item.setFlags(rm_item.flags() | Qt.ItemFlag.ItemIsEditable)
            rm_item.setData(Qt.ItemDataRole.UserRole, eid)
            rm_item.setToolTip("ดับเบิลคลิกเพื่อพิมพ์หมายเหตุ")
            if bg:
                rm_item.setBackground(QBrush(bg))
            self.table.setItem(row, 9, rm_item)

            # col 10: ลิงก์แชร์ / ลิงก์แก้ไข — โชว์ทั้ง 2 ลิงก์ในช่องเดียว
            url = share_links.get(eid)
            edit = "" if is_manual(exp) else _exp_edit_link(exp)
            parts = []
            if url:
                parts.append(f"🔗 <a href='{url}'>เปิดดู</a>")
            elif not is_manual(exp):
                parts.append("<span style='color:#94a3b8'>⏳</span>")
            if edit:
                parts.append(f"✏️ <a href='{edit}' style='color:#dc2626'>แก้ไข</a>")
            if parts:
                lk = QLabel(" / ".join(parts))
                lk.setTextFormat(Qt.TextFormat.RichText)
                lk.setOpenExternalLinks(True)
                lk.setToolTip((url or "") + ("\n" + edit if edit else ""))
                lk.setStyleSheet("font-size:12px;padding-left:6px;background:transparent;")
                self.table.setCellWidget(row, 10, lk)
            else:
                self.table.removeCellWidget(row, 10)
                ph = QTableWidgetItem("—")
                ph.setFlags(ph.flags() & ~Qt.ItemFlag.ItemIsEditable)
                ph.setForeground(QBrush(QColor("#94a3b8")))
                if bg:
                    ph.setBackground(QBrush(bg))
                self.table.setItem(row, 10, ph)

            # col 11: แนบไฟล์ — กดเพื่อแนบ/ดูไฟล์แนบใน FlowAccount
            btn_a = QPushButton("📎")
            btn_a.setToolTip("แนบไฟล์ / ดูไฟล์แนบ")
            btn_a.setCursor(Qt.CursorShape.PointingHandCursor)
            btn_a.setStyleSheet("QPushButton{border:none;background:transparent;font-size:15px;}"
                                "QPushButton:hover{background:#e2e8f0;border-radius:4px;}")
            btn_a.clicked.connect(lambda _=False, ex=exp: self._open_attach(ex))
            aw = QWidget(); al = QHBoxLayout(aw)
            al.addWidget(btn_a); al.setAlignment(Qt.AlignmentFlag.AlignCenter)
            al.setContentsMargins(0, 0, 0, 0)
            self.table.setCellWidget(row, 11, aw)

            # col 12: ลบ — ลบจริงใน FlowAccount (คอลัมน์สีแดง)
            _tip = {"expense": "ลบ EXP นี้ใน FlowAccount + คืนสถานะ PO เป็นอนุมัติ",
                    "po": "ลบใบสั่งซื้อ (PO) นี้ใน FlowAccount",
                    "gr": "ลบใบรับสินค้านี้ใน FlowAccount"}.get(dt, "ลบเอกสารนี้ใน FlowAccount")
            btn_d = QPushButton("🗑️")
            btn_d.setToolTip(_tip)
            btn_d.setCursor(Qt.CursorShape.PointingHandCursor)
            btn_d.setStyleSheet(
                "QPushButton{border:none;background:transparent;font-size:15px;}"
                "QPushButton:hover{background:#fecaca;border-radius:4px;}")
            btn_d.clicked.connect(lambda _=False, ex=exp: self._delete_exp_flowaccount([ex]))
            dw = QWidget(); dl = QHBoxLayout(dw)
            dl.addWidget(btn_d); dl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            dl.setContentsMargins(0, 0, 0, 0)
            dw.setStyleSheet("background:#fee2e2;")
            self.table.setCellWidget(row, 12, dw)

        self._loading = False
        self._update_buttons()

    def _on_check_changed(self, state):
        chk = self.sender()
        eid = chk.property("exp_id")
        if state == Qt.CheckState.Checked.value:
            self._selected.add(eid)
        else:
            self._selected.discard(eid)
        self._update_buttons()

    def _auto_queue(self):
        """จัดคิวอัตโนมัติ — กระจายเป็นหลายวัน (ค่าเริ่มต้น 3 วัน) แล้วเลือกทั้งหมด"""
        from PyQt6.QtWidgets import QInputDialog
        if not self._expenses:
            QMessageBox.information(self, "ยังไม่มีข้อมูล",
                "กรุณากด 🔄 รีเฟรช เพื่อดึงข้อมูลก่อนครับ")
            return

        candidates = [e for e in self._expenses if _status(e) != "paid" and _amount(e) > 0
                      and not paid_pending.is_pending(_exp_id(e))
                      and not rejected.is_rejected(_exp_id(e))]
        if not candidates:
            QMessageBox.information(self, "ไม่มีรายการ", "ไม่มีรายการรอจ่าย")
            return

        # กดจัดคิวใหม่ → ดีดการจัดเรียงเดิมทิ้ง (ข้อ 10)
        queue_plan.clear_plan()

        # กระจายเป็นวัน ๆ (best-fit, เกินกำหนดก่อน, ข้ามเสาร์-อาทิตย์)
        all_days = [d for d in distribute_into_days(candidates, self._daily_limit) if d["items"]]
        maxd = len(all_days)

        n, ok = QInputDialog.getInt(
            self, "🤖 จัดคิวอัตโนมัติ",
            f"มีรายการรอจ่ายจัดได้ทั้งหมด {maxd} วัน\nต้องการจัดคิวกี่วัน?",
            min(3, maxd), 1, maxd)
        if not ok:
            return

        days = all_days[:n]
        picked = [e for d in days for e in d["items"]]
        total  = sum(_amount(e) for e in picked)

        # ── Preview: สรุปแต่ละวัน ──
        thai = ["จันทร์", "อังคาร", "พุธ", "พฤหัสบดี", "ศุกร์", "เสาร์", "อาทิตย์"]
        msg = [f"จัดคิว {n} วัน:"]
        for di, d in enumerate(days, 1):
            dt = d["date"]
            dtot = sum(_amount(e) for e in d["items"])
            msg.append(f"  📅 วันที่ {di} ({thai[dt.weekday()]} {fmt_date(dt.isoformat())}): "
                       f"{len(d['items'])} รายการ • {fmt_amount(dtot)} บาท")
        msg.append("─" * 40)
        msg.append(f"  รวม {len(picked)} รายการ • {fmt_amount(total)} บาท")

        box = QMessageBox(self)
        box.setWindowTitle("🤖 จัดคิวอัตโนมัติ")
        box.setText("\n".join(msg) + "\n\nเลือกรายการเหล่านี้?")
        btn_yes  = box.addButton("✅ ใช้เลย", QMessageBox.ButtonRole.AcceptRole)
        btn_plan = box.addButton("📋 ดู/แก้ตารางคิว", QMessageBox.ButtonRole.ActionRole)
        box.addButton("ยกเลิก", QMessageBox.ButtonRole.RejectRole)
        box.exec()
        clicked = box.clickedButton()
        if clicked == btn_plan:
            self._open_queue_plan()
            return
        if clicked != btn_yes:
            return

        # เลือกรายการทั้ง n วัน + ดันขึ้นบนสุด
        self._selected = {_exp_id(e) for e in picked}
        self._pin_selected = True
        self.btn_pin.setChecked(True)
        self._apply_filter()
        self.status_message.emit(
            f"🤖 จัดคิว {n} วัน: {len(picked)} รายการ • {fmt_amount(total)} บาท")
        activity_log.log("จัดคิวอัตโนมัติ",
                         f"จัดคิว {n} วัน {len(picked)} รายการ • {fmt_amount(total)} บาท")
        activity_log.log_queue(
            f"จัดคิว {n} วัน {len(picked)} รายการ • {fmt_amount(total)} บาท",
            items=[{"doc": _doc_no(e), "vendor": _vendor(e), "amount": _amount(e),
                    "brand": _brand_name(e)} for e in picked])

    def _open_queue_plan(self):
        """เปิดหน้าตารางคิวจ่ายหลายวัน — ดู/โยกรายการสลับวันได้"""
        if not self._expenses:
            QMessageBox.information(self, "ยังไม่มีข้อมูล",
                "กรุณากด 🔄 รีเฟรช เพื่อดึงข้อมูลก่อนครับ")
            return
        candidates = [e for e in self._expenses if _status(e) != "paid"
                      and _amount(e) > 0 and not paid_pending.is_pending(_exp_id(e))
                      and not rejected.is_rejected(_exp_id(e))]
        if not candidates:
            QMessageBox.information(self, "ไม่มีรายการ", "ไม่มีรายการรอจ่าย")
            return
        assignments = load_brands()
        dlg = QueuePlanDialog(candidates, self._daily_limit, assignments, self)
        result = dlg.exec()
        # ถ้ามีการ Mark จ่ายรออัพเดต → รีเฟรชสถานะหน้าหลัก
        if getattr(dlg, "_marked_pending", None):
            self._apply_filter()
        if result == QDialog.DialogCode.Accepted:
            picked = dlg.selected_day_items()
            if picked:
                self._selected = {_exp_id(e) for e in picked}
                self._pin_selected = True
                self.btn_pin.setChecked(True)
                self._apply_filter()   # re-render + ดันที่เลือกขึ้นบน
                total = sum(_amount(e) for e in picked)
                self.status_message.emit(
                    f"📋 เลือกจากตารางคิว: {len(picked)} รายการ • {fmt_amount(total)} บาท")

    def _open_adjust_limit_dialog(self):
        dlg = AdjustLimitDialog(self._daily_limit, self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        new_limit = dlg.get_limit()
        self._daily_limit = new_limit
        if dlg.save_as_default():
            cfg = load_config()
            cfg["daily_payment_limit"] = new_limit
            save_config(cfg)
            self.status_message.emit(f"💾 บันทึก default = {self._format_limit_text()} บาท")
        self.ed_limit.setText(self._format_limit_text())
        self._update_limit_panel()

    def _on_brand_changed(self, text):
        """บัญชีเปลี่ยนแบรนด์ใน dropdown → บันทึกลงไฟล์ทันที"""
        cb = self.sender()
        if cb is None:
            return
        eid = cb.property("exp_id")
        if not eid:
            return
        # ถ้าเลือก "—" จะลบ assignment ออก
        set_brand(eid, "" if text == "—" else text)

    def _toggle_all(self, state):
        checked = state == Qt.CheckState.Checked.value
        self._selected.clear()
        for row in range(self.table.rowCount()):
            w = self.table.cellWidget(row, 0)
            if w:
                chk = w.findChild(QCheckBox)
                if chk:
                    chk.blockSignals(True)
                    chk.setChecked(checked)
                    chk.blockSignals(False)
                    if checked:
                        self._selected.add(chk.property("exp_id"))
        self._update_buttons()

    def _select_all_filtered(self):
        """ติ๊กทุกรายการที่กรองอยู่ (ทุกหน้า)"""
        for e in self._filtered:
            self._selected.add(_exp_id(e))
        self._render_page()
        self._update_buttons()

    def _deselect_all_filtered(self):
        """ปลดติ๊กทั้งหมด"""
        self._selected.clear()
        self._render_page()
        self._update_buttons()

    def _update_buttons(self):
        n = len(self._selected)
        self.lbl_count.setText(f"{n} รายการที่เลือก")
        # Export ใช้ได้เมื่อมีข้อมูล (ไม่ติ๊ก = export คิวทั้งหมดหลายวัน)
        self.btn_csv.setEnabled(bool(self._expenses))
        self.btn_export_link.setEnabled(bool(self._expenses))
        self.btn_paid.setEnabled(n > 0)
        self.btn_mark_pending.setEnabled(n > 0)
        self.btn_unmark_pending.setEnabled(n > 0)
        if hasattr(self, "btn_reject"):
            self.btn_reject.setEnabled(n > 0)
            # ปุ่มดูรายการไม่อนุมัติ เปิดได้เสมอ
        self._update_limit_panel()

    def _selected_expenses(self) -> list:
        return [e for e in self._expenses if _exp_id(e) in self._selected]

    def _on_status_changed(self, exp, label):
        """เปลี่ยนสถานะจาก dropdown ในเซลล์
        - รอจ่าย/รออัพเดต/ไม่อนุมัติ = เก็บในเครื่อง
        - จ่ายแล้ว = บันทึก payment เข้า FlowAccount จริง (ผ่านหน้ายืนยันเดิม)"""
        if getattr(self, "_loading", False):
            return
        eid = _exp_id(exp)
        doc = _doc_no(exp)
        # ── popup เตือนทุกครั้งกันพลาด ── (FlowAccount-linked = เตือนแรงกว่า)
        if "จ่ายแล้ว" in label:
            if QMessageBox.warning(self, "⚠️ ยืนยันเปลี่ยนสถานะ (กระทบ FlowAccount จริง)",
                    f"จะเปลี่ยนสถานะ {doc} เป็น '✅ จ่ายแล้ว'\n\n"
                    "⚠️ การนี้จะบันทึก payment เข้า FlowAccount จริง (แก้ย้อนยาก)\n"
                    "ระบบจะถามวันที่/วิธีจ่ายต่อไป\n\nยืนยันดำเนินการ?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel,
                    QMessageBox.StandardButton.Cancel) != QMessageBox.StandardButton.Yes:
                self._apply_filter()   # คืน dropdown เดิม
                return
        elif "อนุมัติแล้ว" not in label:   # สถานะ local (รอจ่าย/รออัพเดต/ไม่อนุมัติ) — ยืนยันเบา ๆ
            if QMessageBox.question(self, "ยืนยันเปลี่ยนสถานะ",
                    f"เปลี่ยนสถานะ {doc} เป็น '{label}' ?",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel,
                    QMessageBox.StandardButton.Cancel) != QMessageBox.StandardButton.Yes:
                self._apply_filter()
                return
        if "รอจ่าย" in label:
            paid_pending.unmark(eid); rejected.approve(eid)
            self.status_message.emit(f"↩️ คืนสถานะรอจ่าย: {_doc_no(exp)}")
            self._apply_filter()
        elif "รออัพเดต" in label:
            rejected.approve(eid)
            paid_pending.mark(eid, {"amount": _amount(exp), "vendor": _vendor(exp),
                                    "doc": _doc_no(exp), "user": activity_log.current_user()})
            self.status_message.emit(f"🟣 จ่ายแล้วรออัพเดต: {_doc_no(exp)}")
            self._apply_filter()
        elif "ไม่อนุมัติ" in label:
            # ไม่อนุมัติ = ซ่อนในเครื่อง (ย้อนได้) — ถ้าจะลบจริงใช้ปุ่ม 🗑️ ลบ คอลัมน์สุดท้าย
            paid_pending.unmark(eid)
            rejected.reject(eid, {"vendor": _vendor(exp), "amount": _amount(exp),
                                  "doc": _doc_no(exp), "user": activity_log.current_user()})
            self.status_message.emit(f"🚫 ไม่อนุมัติ (ซ่อน): {_doc_no(exp)}")
            self._apply_filter()
        elif "อนุมัติแล้ว" in label:
            QMessageBox.information(self, "ตั้งใน FlowAccount เท่านั้น",
                "สถานะ 'อนุมัติแล้ว' กำหนดจากใน FlowAccount เท่านั้น\n"
                "(กดลิงก์แก้ไขใบเพื่อไปอนุมัติใน FlowAccount)")
            self._apply_filter()   # คืนค่า dropdown เดิม
        elif "จ่ายแล้ว" in label:
            # บันทึก payment เข้า FlowAccount จริง — ใช้หน้ายืนยันเดิม (ถามวันที่/วิธีจ่าย)
            self._selected = {eid}
            self._mark_paid()
            # รีเฟรชเพื่อรีเซ็ต dropdown (เผื่อกดยกเลิกในหน้ายืนยัน)
            self._apply_filter()

    def _reject_selected(self):
        """ไม่อนุมัติรายการที่ติ๊ก — ซ่อนในเครื่อง (ย้อนได้) ไม่ลบ"""
        exps = self._selected_expenses()
        if not exps:
            QMessageBox.information(self, "ยังไม่ได้เลือก", "ติ๊กเลือกรายการก่อนครับ")
            return
        if QMessageBox.question(self, "🚫 ไม่อนุมัติ",
                f"ทำเครื่องหมาย 'ไม่อนุมัติ' {len(exps)} รายการ?\n\n"
                "• ซ่อนจากคิว/คิดยอด/Export (ยังไม่ลบ)\n"
                "• กด 'ดูรายการที่ไม่อนุมัติ' เพื่ออนุมัติกลับ\n"
                "• ถ้าต้องการลบจริง ใช้ปุ่ม 🗑️ คอลัมน์สุดท้าย",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel
                ) != QMessageBox.StandardButton.Yes:
            return
        for e in exps:
            rejected.reject(_exp_id(e), {"vendor": _vendor(e), "amount": _amount(e),
                                         "doc": _doc_no(e), "user": activity_log.current_user()})
        activity_log.log("ไม่อนุมัติ (ซ่อน)", f"{len(exps)} รายการ")
        self._selected.clear()
        self._apply_filter()
        self.status_message.emit(f"🚫 ไม่อนุมัติ {len(exps)} รายการ (ซ่อน)")

    def _delete_exp_flowaccount(self, docs):
        """🗑️ ลบเอกสารจริงใน FlowAccount (EXP/PO/GR)
        EXP → ลบ + คืนสถานะ PO ที่ผูกกันเป็น 'อนุมัติ'"""
        if not docs:
            return
        _tname = {"expense": "EXP (ค่าใช้จ่าย)", "po": "PO (ใบสั่งซื้อ)", "gr": "ใบรับสินค้า"}
        lines = []
        for e in docs[:15]:
            dt = e.get("_doctype", "expense")
            extra = ""
            if dt == "expense":
                _, po_ser = find_linked_po(e)
                extra = f"   →  คืน {po_ser} เป็น 'อนุมัติ'" if po_ser else ""
            lines.append(f"• {_doc_no(e)}  [{_tname.get(dt, dt)}]{extra}")
        more = f"\n... และอีก {len(docs) - 15} รายการ" if len(docs) > 15 else ""
        if QMessageBox.warning(self, "⚠️ ลบเอกสารออกจาก FlowAccount จริง",
                f"จะ 'ลบ' เอกสาร {len(docs)} รายการ ออกจาก FlowAccount จริง\n"
                "⚠️ ลบแล้วกู้คืนยาก! (EXP จะคืนสถานะ PO ที่ผูกกันเป็น 'อนุมัติ' ให้)\n\n"
                + "\n".join(lines) + more + "\n\nยืนยันลบจริง?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel,
                QMessageBox.StandardButton.Cancel) != QMessageBox.StandardButton.Yes:
            return

        dlg = LoadingDialog("กำลังลบเอกสารใน FlowAccount...", self)
        self._reject_worker = RejectFlowAccountWorker(exp_docs, self)
        self._reject_worker.progress.connect(
            lambda d, t: dlg.set_progress(d, t, f"ลบ {d}/{t}"))
        self._reject_worker.finished_all.connect(
            lambda ok_ids, probs: (dlg.accept(), self._on_reject_done(ok_ids, probs)))
        self._reject_worker.start()
        dlg.exec()

    def _on_reject_done(self, ok_ids, problems):
        # เอา EXP ที่ลบสำเร็จออกจากตารางทันที (จาก FlowAccount หายแล้ว)
        if ok_ids:
            ok_set = set(ok_ids)
            self._expenses = [e for e in self._expenses if _exp_id(e) not in ok_set]
            for eid in ok_ids:
                self._selected.discard(eid)
        activity_log.log("ไม่อนุมัติ (ลบ EXP)",
                         f"ลบสำเร็จ {len(ok_ids)} • มีปัญหา {len(problems)}")
        self._apply_filter()
        if not problems:
            QMessageBox.information(self, "✅ เสร็จสมบูรณ์",
                f"ลบ EXP {len(ok_ids)} รายการ + คืนสถานะ PO เป็น 'อนุมัติ' แล้ว\n\n"
                "(กด 🔄 รีเฟรช เพื่อยืนยันจาก FlowAccount อีกครั้ง)")
            return
        box = QMessageBox(self)
        box.setIcon(QMessageBox.Icon.Warning)
        box.setWindowTitle("มีรายการที่มีปัญหา")
        box.setText(f"ลบสำเร็จ {len(ok_ids)} • มีปัญหา {len(problems)} รายการ")
        box.setDetailedText("\n".join(f"• {p['doc']}: {p['err']}" for p in problems))
        box.exec()

    def _show_rejected_list(self):
        """เปิดหน้ารายการไม่อนุมัติ — ติ๊กแล้วอนุมัติกลับได้"""
        dlg = RejectedListDialog(self)
        dlg.exec()
        if dlg.approved_ids:
            activity_log.log("อนุมัติกลับ", f"{len(dlg.approved_ids)} รายการ")
            self._apply_filter()
            self.status_message.emit(f"✅ อนุมัติกลับ {len(dlg.approved_ids)} รายการ")

    def _show_detail(self, row, col):
        """ดับเบิลคลิกแถว → ดูรายละเอียดรายการ (ข้อ 3)"""
        if col in (1, 9, 10):    # เลขเอกสาร(คลิก=copy)/หมายเหตุ/ลิงก์แชร์ → ไม่เปิด popup
            return
        rows = getattr(self, "_page_rows", [])
        if 0 <= row < len(rows):
            ExpenseDetailDialog(rows[row], load_brands(), self).exec()

    def _on_cell_click(self, row, col):
        """คลิกคอลัมน์เลขเอกสาร (1) → คัดลอกรหัส EXP/PO ลงคลิปบอร์ด"""
        if col != 1:
            return
        item = self.table.item(row, col)
        if not item:
            return
        code = re.sub(r"^[\s✏️📄📦]+", "", item.text()).strip()
        if not code:
            return
        QApplication.clipboard().setText(code)
        self.status_message.emit(f"📋 คัดลอกเลขเอกสาร: {code}")
        from PyQt6.QtWidgets import QToolTip
        from PyQt6.QtGui import QCursor
        QToolTip.showText(QCursor.pos(), f"📋 คัดลอกแล้ว: {code}")

    def _open_attach(self, exp):
        AttachDialog(exp, self).exec()

    def _on_remark_changed(self, item):
        """พิมพ์หมายเหตุ (คอลัมน์ 9) / วางลิงก์แชร์ (คอลัมน์ 10) → บันทึกอัตโนมัติ"""
        if getattr(self, "_loading", False):
            return
        eid = item.data(Qt.ItemDataRole.UserRole)
        if not eid:
            return
        if item.column() == 9:
            remarks.set(eid, item.text())
            self.status_message.emit(f"💾 บันทึกหมายเหตุ {eid} แล้ว")
        elif item.column() == 10:
            share_links.set(eid, item.text())
            self.status_message.emit(f"🔗 บันทึกลิงก์แชร์ {eid} แล้ว")

    def _mark_pending_selected(self):
        exps = self._selected_expenses()
        if not exps:
            QMessageBox.information(self, "ยังไม่ได้เลือก",
                "ติ๊กเลือกรายการที่จ่ายแล้วก่อนครับ")
            return
        if QMessageBox.question(self, "Mark จ่ายแล้ว (รออัพเดต)",
                f"ทำเครื่องหมาย {len(exps)} รายการว่า 'จ่ายแล้ว รออัพเดตใน FlowAccount'?\n"
                "รายการนี้จะไม่ถูกนำไปคิดยอด/จัดคิว/Export อีก",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel
                ) != QMessageBox.StandardButton.Yes:
            return
        for e in exps:
            paid_pending.mark(_exp_id(e), {
                "amount": _amount(e), "vendor": _vendor(e),
                "doc": _doc_no(e), "user": activity_log.current_user()})
        activity_log.log("Mark จ่ายแล้วรออัพเดต", f"{len(exps)} รายการ (หน้าหลัก)")
        self._selected.clear()
        self._apply_filter()
        self.status_message.emit(f"🟣 Mark จ่ายแล้วรออัพเดต {len(exps)} รายการ")

    def _unmark_pending_selected(self):
        exps = self._selected_expenses()
        targets = [e for e in exps if paid_pending.is_pending(_exp_id(e))]
        if not targets:
            QMessageBox.information(self, "ยกเลิกไม่ได้",
                "ติ๊กเลือกรายการที่เป็น '🟣 จ่ายแล้วรออัพเดต' ที่ต้องการยกเลิกก่อนครับ")
            return
        for e in targets:
            paid_pending.unmark(_exp_id(e))
        activity_log.log("ยกเลิก Mark รออัพเดต", f"{len(targets)} รายการ")
        self._apply_filter()
        self.status_message.emit(f"↩️ ยกเลิกรออัพเดต {len(targets)} รายการ")

    # ── Actions ──

    def _sync_monday(self):
        cfg = load_config()
        if not cfg.get("monday_api_token") or not cfg.get("monday_board_id"):
            QMessageBox.warning(self, "ยังไม่ตั้งค่า",
                "กรุณาตั้งค่า monday.com API Token และ Board ID ก่อนครับ\n(เมนู ไฟล์ → ตั้งค่า)")
            return
        if not self._expenses:
            QMessageBox.warning(self, "ไม่มีข้อมูล",
                "กรุณากด รีเฟรช เพื่อดึงข้อมูล Expense ก่อนครับ")
            return
        # ── ถ้าเลือกรายการไว้ → sync เฉพาะที่เลือก, ถ้าไม่เลือก → sync ทุกที่ยังไม่จ่าย ──
        if self._selected:
            to_sync = [e for e in self._expenses if _exp_id(e) in self._selected]
            detail  = f"Sync {len(to_sync)} รายการที่เลือก"
        else:
            to_sync = [e for e in self._expenses if _status(e) != "paid"]
            skipped = len(self._expenses) - len(to_sync)
            detail  = f"Sync {len(to_sync)} รายการที่ยังไม่จ่าย\n(ข้าม {skipped} รายการที่จ่ายแล้ว เพื่อประหยัด API)"
        if not to_sync:
            QMessageBox.information(self, "ไม่มีรายการ", "ไม่มีรายการที่จะ Sync")
            return

        # ── เติมชื่อแบรนด์ลงในแต่ละรายการก่อนส่ง (ไม่กระทบ FlowAccount) ──
        assignments = load_brands()
        enriched = []
        for e in to_sync:
            ec = dict(e)
            ec["_brand_name"] = _brand_name(e, assignments)
            enriched.append(ec)
        to_sync = enriched
        reply = QMessageBox.question(self, "ยืนยัน Sync",
            f"{detail}\n\nดำเนินการต่อ?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel)
        if reply != QMessageBox.StandardButton.Yes:
            return
        self.btn_monday.setEnabled(False)
        self.btn_monday.setText("⏳ กำลัง sync...")
        self.status_message.emit(f"กำลัง sync {len(to_sync)} รายการไป monday.com...")
        self._mon_worker = MondaySyncWorker(cfg, to_sync)
        self._mon_worker.done.connect(self._on_monday_done)
        self._mon_worker.error.connect(self._on_monday_error)
        self._mon_worker.start()

    def _on_monday_done(self, result):
        self.btn_monday.setEnabled(True)
        self.btn_monday.setText("☁️ Sync → monday.com")
        msg = (f"✅ Sync สำเร็จ!\n\n"
               f"สร้าง: {result.get('created', 0)} รายการ\n"
               f"อัพเดต: {result.get('updated', 0)} รายการ")
        if result.get("errors"):
            msg += f"\nข้อผิดพลาด: {len(result['errors'])} รายการ"
        QMessageBox.information(self, "monday.com Sync", msg)
        self.status_message.emit(
            f"✅ Sync เสร็จ: {result.get('created',0)} created, {result.get('updated',0)} updated")
        # Smart Auto-Refresh: ถ้ามี error ที่เป็น cache เก่า → refresh อัตโนมัติ
        self._maybe_smart_refresh(result.get("errors", []))

    def _on_monday_error(self, msg):
        self.btn_monday.setEnabled(True)
        self.btn_monday.setText("☁️ Sync → monday.com")
        QMessageBox.critical(self, "Sync ล้มเหลว", msg)
        dev_log(f"[ERROR] monday sync: {msg}", "err")
        low = (msg or "").lower()
        if DevConsole._instance and any(k in low for k in
                ("limit", "complexity", "429", "exceeded", "budget")):
            DevConsole._instance.set_service("monday", "limit")

    def _refresh_monday_cache(self):
        """ดึงข้อมูลทั้ง board มาทำ cache ใหม่ — ใช้เมื่อมีคนแก้ใน monday โดยตรง"""
        cfg = load_config()
        if not cfg.get("monday_api_token") or not cfg.get("monday_board_id"):
            QMessageBox.warning(self, "ยังไม่ตั้งค่า",
                "กรุณาตั้งค่า monday.com API Token และ Board ID ก่อนครับ")
            return
        reply = QMessageBox.question(self, "ยืนยัน Refresh Cache",
            "จะดึงข้อมูลทั้ง board ใน monday.com มาสร้าง cache ใหม่\n"
            "(กิน API quota ครั้งเดียว จากนั้น Sync จะเบาขึ้นมาก)\n\nดำเนินการต่อ?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel)
        if reply != QMessageBox.StandardButton.Yes:
            return
        self.btn_refresh_cache.setEnabled(False)
        self.btn_refresh_cache.setText("⏳ กำลังดึง...")
        self.status_message.emit("กำลังดึงข้อมูลจาก monday.com มาทำ cache...")
        self._cache_worker = MondayCacheBuildWorker(cfg)
        self._cache_worker.done.connect(self._on_cache_built)
        self._cache_worker.error.connect(self._on_cache_error)
        self._cache_worker.start()

    def _maybe_smart_refresh(self, errors):
        """ถ้า sync เจอ error ที่ดูเหมือน cache เก่า (item ถูกลบ/เปลี่ยน) → trigger refresh
        จำกัดสูงสุด 1 ครั้ง/ชม. เพื่อกันลูป + ประหยัด quota"""
        if not errors:
            return

        # คำที่บ่งบอกว่า cache เก่า
        stale_keywords = ["not found", "doesn't exist", "does not exist",
                          "invalid", "item", "404", "permission denied"]
        has_stale = False
        for e in errors:
            txt = (str(e.get("error", "")) if isinstance(e, dict) else str(e)).lower()
            if any(kw in txt for kw in stale_keywords):
                has_stale = True
                break
        if not has_stale:
            return

        # throttle: สูงสุด 1 ครั้ง/ชม.
        now = datetime.now()
        if self._last_smart_refresh:
            elapsed = (now - self._last_smart_refresh).total_seconds()
            if elapsed < 3600:
                return  # เพิ่ง refresh ไป — ข้าม

        cfg = load_config()
        if not cfg.get("monday_api_token") or not cfg.get("monday_board_id"):
            return

        self._last_smart_refresh = now
        self.status_message.emit("🤖 Smart Auto-Refresh — ตรวจพบ cache เก่า กำลังรีเฟรช...")
        self._smart_worker = MondayCacheBuildWorker(cfg)
        self._smart_worker.done.connect(self._on_smart_refresh_done)
        self._smart_worker.error.connect(self._on_smart_refresh_error)
        self._smart_worker.start()

    def _on_smart_refresh_done(self, result):
        n = result.get("count", 0)
        self.status_message.emit(
            f"✅ Smart Auto-Refresh เสร็จ — cache มี {n} รายการ (Sync ครั้งหน้าจะถูกต้อง)")

    def _on_smart_refresh_error(self, msg):
        self.status_message.emit(f"⚠️ Smart Auto-Refresh ล้มเหลว: {msg[:60]}")

    def _on_cache_built(self, result):
        self.btn_refresh_cache.setEnabled(True)
        self.btn_refresh_cache.setText("🔁 Refresh Cache")
        count = result.get("count", 0)
        QMessageBox.information(self, "✅ Refresh Cache สำเร็จ",
            f"บันทึก cache แล้ว {count} รายการ\n"
            f"Sync ครั้งถัดไปจะเบาและเร็วขึ้นมาก")
        self.status_message.emit(f"✅ Cache พร้อม — {count} รายการ")

    def _on_cache_error(self, msg):
        self.btn_refresh_cache.setEnabled(True)
        self.btn_refresh_cache.setText("🔁 Refresh Cache")
        QMessageBox.critical(self, "Refresh Cache ล้มเหลว", msg)

    def _days_saved_or_distribute(self, pool):
        """ถ้ามีการจัดเรียงที่บันทึกในตารางคิว → ใช้อันนั้น (Export ตรงกับตารางคิว)
        ไม่งั้นกระจายอัตโนมัติตามวงเงิน"""
        saved = queue_plan.load_plan()
        if saved and saved.get("days"):
            by_id = {_exp_id(e): e for e in pool}
            used, groups = set(), []
            for day_ids in saved["days"]:
                grp = [by_id[i] for i in day_ids if i in by_id and i not in used]
                for i in day_ids:
                    used.add(i)
                groups.append(grp)
            leftover = [e for e in pool if _exp_id(e) not in used]
            if leftover:
                for d in distribute_into_days(leftover, self._daily_limit):
                    groups.append(d["items"])
            groups = [g for g in groups if g]
            return assign_dates(groups)
        return distribute_into_days(pool, self._daily_limit)

    def _ensure_share_links_for_export(self, exps):
        """ก่อน export: ดึงลิงก์แชร์ที่ยังไม่มีให้ครบ (โชว์หลอดโหลด) เพื่อให้ไฟล์มีลิงก์แชร์จริง"""
        missing = [e for e in exps
                   if not share_links.get(_exp_id(e))
                   and not is_manual(e)
                   and (e.get("documentId") or e.get("recordId"))]
        if not missing:
            return
        dlg = LoadingDialog("กำลังเตรียมลิงก์แชร์สำหรับไฟล์...", self)
        worker = ShareLinkWorker(missing, self)
        worker.progress.connect(
            lambda d, t: dlg.set_progress(d, t, f"ดึงลิงก์แชร์ {d}/{t}"))
        worker.finished_all.connect(lambda ok, fail: dlg.accept())
        worker.start()
        dlg.exec()

    def _fetch_share_links(self):
        """ดึงลิงก์แชร์อัตโนมัติให้รายการที่ติ๊ก (ถ้าไม่ติ๊ก = ทั้งหน้าปัจจุบัน)"""
        sel = self._selected_expenses()
        pool = sel if sel else list(getattr(self, "_page_rows", []))
        # ข้ามรายการที่มีลิงก์อยู่แล้ว
        pool = [e for e in pool if not share_links.get(_exp_id(e))]
        if not pool:
            QMessageBox.information(self, "ลิงก์แชร์",
                "รายการที่เลือกมีลิงก์แชร์ครบแล้ว\n(ถ้าต้องการดึงใหม่ ให้ลบลิงก์เดิมในคอลัมน์ก่อน)")
            return
        if QMessageBox.question(self, "ดึงลิงก์แชร์",
                f"จะดึงลิงก์แชร์อัตโนมัติจาก FlowAccount {len(pool)} รายการ\n"
                f"(รายการที่มีลิงก์แล้วข้ามให้)\n\nเริ่มเลยไหม?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel
                ) != QMessageBox.StandardButton.Yes:
            return
        self._sl_dlg = LoadingDialog("กำลังดึงลิงก์แชร์...", self)
        self._sl_worker = ShareLinkWorker(pool, self)
        self._sl_worker.progress.connect(
            lambda d, t: self._sl_dlg.set_progress(d, t, f"ดึงลิงก์แชร์ {d}/{t}"))
        self._sl_worker.finished_all.connect(self._on_share_links_done)
        self._sl_worker.start()
        self._sl_dlg.exec()

    def _on_share_links_done(self, ok, fail):
        if getattr(self, "_sl_dlg", None):
            self._sl_dlg.accept()
        activity_log.log("ดึงลิงก์แชร์อัตโนมัติ", f"สำเร็จ {ok} ล้มเหลว {fail}")
        self._render_page()   # โหลดคอลัมน์ลิงก์ใหม่
        QMessageBox.information(self, "✅ เสร็จแล้ว",
            f"ดึงลิงก์แชร์สำเร็จ {ok} รายการ" + (f"\nไม่สำเร็จ {fail} รายการ" if fail else ""))

    def _prepare_export_days(self):
        """เตรียมข้อมูลสำหรับ Export (ใช้ร่วม PDF/Excel/Link)
        คืน (days, exps) หรือ None ถ้ายกเลิก/ไม่มีข้อมูล"""
        from PyQt6.QtWidgets import QInputDialog
        assignments = load_brands()
        # ใช้รายการที่ติ๊กถ้าติ๊กไว้ ไม่งั้นใช้ทั้งหมดที่ยังไม่จ่าย (ไม่รวมที่ Mark รออัพเดต)
        sel = self._selected_expenses()
        pool = sel if sel else [e for e in self._expenses
                                if _status(e) != "paid" and _amount(e) > 0
                                and not paid_pending.is_pending(_exp_id(e))]
        if not pool:
            QMessageBox.information(self, "ไม่มีรายการ",
                "ไม่มีรายการให้ export — กด 🔄 รีเฟรช ดึงข้อมูลก่อนครับ")
            return None

        for e in pool:
            if not e.get("_brand_name"):
                b = _brand_name(e, assignments)
                if b:
                    e["_brand_name"] = b
            rm = remarks.get(_exp_id(e))
            if rm:
                e["_custom_remark"] = rm
            su = share_links.get(_exp_id(e))
            if su:
                e["_share_url"] = su

        # ใช้การจัดเรียงที่บันทึกจากตารางคิว (ถ้ามี) เพื่อให้ Export ตรงกับตารางคิว
        all_days = [d for d in self._days_saved_or_distribute(pool) if d["items"]]
        maxd = len(all_days)
        if sel:
            days = all_days
        else:
            n, ok = QInputDialog.getInt(
                self, "Export คิวจ่ายกี่วัน?",
                f"คิวจ่ายทั้งหมดจัดได้ {maxd} วัน (จากรายการที่ยังไม่จ่ายทั้งหมด)\n"
                f"ต้องการ export กี่วัน?",
                min(3, maxd), 1, maxd)
            if not ok:
                return None
            days = all_days[:n]
        exps = [e for d in days for e in d["items"]]

        # ดึงลิงก์แชร์ที่ยังขาดให้ครบก่อน เพื่อให้ไฟล์ใช้ "ลิงก์แชร์จริง" ไม่ใช่ลิงก์แก้ไข
        self._ensure_share_links_for_export(exps)
        for e in exps:
            su = share_links.get(_exp_id(e))
            if su:
                e["_share_url"] = su
        return days, exps

    def _export_link(self):
        """สร้างหน้า HTML (เหมือน PDF) → อัปขึ้น cloud → ลิงก์สุ่มหมดอายุเอง (เปิดบนมือถือได้)"""
        prepared = self._prepare_export_days()
        if not prepared:
            return
        days, exps = prepared
        cfg = load_config()
        gh_token = cfg.get("github_token", "").strip()
        permanent = bool(gh_token)
        expiry_txt = ("ลิงก์ถาวร (ไม่หมดอายุ)" if permanent
                      else "ลิงก์สุ่ม หมดอายุอัตโนมัติใน 72 ชั่วโมง")
        # เตือนเรื่องข้อมูลการเงินเป็นลิงก์สาธารณะ (ขออนุญาตก่อนเผยแพร่ออกนอก)
        if QMessageBox.question(self, "Export Link (ลิงก์สาธารณะ)",
                f"จะสร้างลิงก์เปิดดูคิวจ่าย {len(exps)} รายการ บนอินเทอร์เน็ต\n\n"
                f"• {expiry_txt}\n"
                "• ใครก็ตามที่ได้ลิงก์นี้จะเปิดดูข้อมูลได้ (มีชื่อผู้รับ/ยอดเงิน)\n\n"
                "ยืนยันสร้างลิงก์?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel
                ) != QMessageBox.StandardButton.Yes:
            return
        try:
            html = build_bank_html_multiday(days, cfg)
        except Exception as e:
            QMessageBox.critical(self, "❌ ผิดพลาด", f"สร้างหน้าเว็บไม่สำเร็จ:\n{e}")
            return
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        self.status_message.emit("กำลังอัปโหลดลิงก์...")
        try:
            if permanent:
                url = upload_html_github(html, gh_token)
            else:
                url = upload_html_temp(html, "72h")
        except Exception as e:
            QApplication.restoreOverrideCursor()
            QMessageBox.critical(self, "❌ อัปโหลดไม่สำเร็จ",
                f"สร้างลิงก์ไม่สำเร็จ (เช็คอินเทอร์เน็ต/Token):\n{e}")
            return
        QApplication.restoreOverrideCursor()
        QApplication.clipboard().setText(url)
        QDesktopServices.openUrl(QUrl(url))
        n_days = len(days)
        activity_log.log("Export Link", f"{len(exps)} รายการ • {n_days} วัน → {url}")
        self.status_message.emit(f"🔗 Export Link: คัดลอกลิงก์แล้ว ({len(exps)} รายการ)")
        note = ("• ลิงก์ถาวร ไม่หมดอายุ" if permanent
                else "• หมดอายุอัตโนมัติใน 72 ชม.")
        dlg = QMessageBox(self)
        dlg.setWindowTitle("✅ สร้างลิงก์แล้ว")
        dlg.setTextFormat(Qt.TextFormat.RichText)
        dlg.setText(
            f"คัดลอกลิงก์ให้แล้ว + เปิดในเบราว์เซอร์<br><br>"
            f"<b>ลิงก์ (เปิดบนมือถือได้):</b><br>"
            f"<a href='{url}'>{url}</a><br><br>"
            f"<span style='color:#64748b'>{note}<br>"
            f"• ส่งให้ใครก็เปิดได้ (สแกน/วางลิงก์)</span>")
        dlg.exec()

    def _download_csv(self):
        prepared = self._prepare_export_days()
        if not prepared:
            return
        days, exps = prepared
        n_days = len(days)
        cfg = load_config()

        base, _ = QFileDialog.getSaveFileName(
            self, "บันทึกไฟล์ (จะได้ทั้ง Excel และ PDF)",
            f"KCash_Payment_{date.today().strftime('%Y%m%d')}",
            "ไฟล์ทำจ่าย (*.xlsx)")
        if not base:
            return
        # ตัดนามสกุลออกเพื่อใช้เป็นชื่อฐาน
        for ext in (".xlsx", ".pdf", ".csv"):
            if base.lower().endswith(ext):
                base = base[:-len(ext)]
        xlsx_path = base + ".xlsx"
        pdf_path  = base + ".pdf"

        try:
            build_bank_excel_multiday(days, cfg, xlsx_path)
            build_bank_pdf_multiday(days, cfg, pdf_path)
        except Exception as e:
            QMessageBox.critical(self, "❌ ผิดพลาด", f"บันทึกไฟล์ไม่สำเร็จ:\n{e}")
            return

        self.status_message.emit(f"Export แล้ว: {len(exps)} รายการ • {n_days} วัน (Excel + PDF)")
        activity_log.log("Export", f"{len(exps)} รายการ • {n_days} วัน (Excel+PDF) → {os.path.basename(xlsx_path)}")
        QMessageBox.information(self, "✅ สำเร็จ",
            f"บันทึกแล้ว {len(exps)} รายการ • จัดคิว {n_days} วัน:\n\n"
            f"📊 {xlsx_path}\n📄 {pdf_path}")

    def _mark_paid(self):
        exps = self._selected_expenses()
        if not exps:
            return

        dlg = MarkPaidDialog(exps, self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        payment_date, method = dlg.get_values()
        cfg = load_config()

        # ── แยก manual กับ FlowAccount ──
        manual_exps = [e for e in exps if is_manual(e)]
        fa_exps     = [e for e in exps if not is_manual(e)]

        # Manual: mark paid ในไฟล์ local ทันที (ไม่เรียก API)
        manual_ok = []
        for m in manual_exps:
            try:
                update_manual(m["id"], {
                    "statusString":   "paid",
                    "paymentDate":    payment_date,
                    "paymentMethod":  method,
                })
                manual_ok.append(m)
            except Exception:
                pass

        if not fa_exps:
            # ไม่มีรายการ FlowAccount → จบเลย ไม่ต้องเรียก worker
            self._on_mark_done(manual_ok, [], dlg)
            return

        dlg.progress.setVisible(True)
        dlg.progress.setRange(0, len(fa_exps))
        dlg.btn_box.setEnabled(False)

        self._mark_worker = MarkPaidWorker(cfg, fa_exps, payment_date, method)
        self._mark_worker.progress.connect(lambda i, n: dlg.progress.setValue(i))
        self._mark_worker.done.connect(
            lambda ok, errs: self._on_mark_done(ok + manual_ok, errs, dlg))
        self._mark_worker.start()

    def _on_mark_done(self, ok: list, errs: list, dlg: QDialog):
        dlg.close()
        total = sum(_amount(e) for e in ok)
        activity_log.log("Mark จ่ายแล้ว",
                         f"จ่าย {len(ok)} รายการ • {fmt_amount(total)} บาท"
                         + (f" (ไม่สำเร็จ {len(errs)})" if errs else ""))
        self._selected.clear()
        self._expenses = []
        self.fetch_expenses()

        cfg = load_config()
        line_msg = build_line_message(ok, today_str(), cfg)
        if errs:
            err_txt = "\n".join(f"• {e['id']}: {e['error']}" for e in errs)
            QMessageBox.warning(self, "มีบางรายการไม่สำเร็จ",
                                f"สำเร็จ {len(ok)} รายการ ไม่สำเร็จ {len(errs)} รายการ:\n{err_txt}")

        line_dlg = LineMessageDialog(line_msg, self)
        line_dlg.exec()


# ──────────────────── PO Workers ────────────────────

class POFetchWorker(QThread):
    done  = pyqtSignal(list)
    error = pyqtSignal(str)
    def __init__(self, cfg, company: dict = None):
        super().__init__()
        self.cfg = cfg
        self.company = company or {}
    def run(self):
        try:
            items = get_all_purchase_orders(self.cfg["flowaccount_api_key"],
                                             self.cfg["flowaccount_secret_key"])
            if self.company:
                _tag_brand(items, self.company)
            self.done.emit(items)
        except FlowAccountError as e:
            self.error.emit(str(e))
        except Exception as e:
            self.error.emit(f"เกิดข้อผิดพลาด: {e}")


class POFetchAllWorker(QThread):
    """ดึงใบสั่งซื้อจากทุกบริษัทมารวมกัน + ติดป้ายแบรนด์"""
    done     = pyqtSignal(list)
    error    = pyqtSignal(str)
    progress = pyqtSignal(str)
    def __init__(self, companies):
        super().__init__()
        self.companies = companies
    def run(self):
        all_items, fails = [], []
        for c in self.companies:
            cid, csec = c.get("client_id"), c.get("client_secret")
            if not cid:
                continue
            self.progress.emit(f"กำลังดึงใบสั่งซื้อ {c.get('label','')}...")
            try:
                items = get_all_purchase_orders(cid, csec)
                _tag_brand(items, c)
                all_items.extend(items)
            except Exception as e:
                fails.append(f"{c.get('label','')}: {e}")
        if fails and not all_items:
            self.error.emit("ดึงข้อมูลไม่สำเร็จ:\n" + "\n".join(fails))
        else:
            self.done.emit(all_items)


class POConvertWorker(QThread):
    progress = pyqtSignal(int, int)
    done     = pyqtSignal(list, list)
    def __init__(self, cfg, pos):
        super().__init__()
        self.cfg = cfg
        self.pos = pos   # list ของ po object (มี _co_id/_co_secret)
    def run(self):
        ok, errs = [], []
        for i, po in enumerate(self.pos):
            self.progress.emit(i+1, len(self.pos))
            pid  = _exp_id(po)
            cid  = po.get("_co_id")     or self.cfg["flowaccount_api_key"]
            csec = po.get("_co_secret") or self.cfg["flowaccount_secret_key"]
            try:
                convert_po_to_expense(cid, csec, pid)
                ok.append(pid)
            except FlowAccountError as e:
                errs.append({"id":pid,"error":str(e)})
        self.done.emit(ok, errs)


class MondaySyncWorker(QThread):
    done  = pyqtSignal(dict)
    error = pyqtSignal(str)
    def __init__(self, cfg, expenses):
        super().__init__()
        self.cfg      = cfg
        self.expenses = expenses
    def run(self):
        try:
            from monday import sync_expenses_to_monday
            from monday_cache import load_cache, save_cache
            cache = load_cache()
            result = sync_expenses_to_monday(
                self.cfg["monday_api_token"],
                self.cfg["monday_board_id"],
                self.expenses,
                cache=cache,
            )
            # บันทึก cache ที่อัพเดตแล้วกลับลงไฟล์
            save_cache(result.get("cache", cache))
            self.done.emit(result)
        except Exception as e:
            self.error.emit(str(e))


class MondayCacheBuildWorker(QThread):
    """Worker สำหรับสร้าง cache จาก board (กด Refresh Cache)"""
    done  = pyqtSignal(dict)   # {count, cache}
    error = pyqtSignal(str)
    def __init__(self, cfg):
        super().__init__()
        self.cfg = cfg
    def run(self):
        try:
            from monday import build_cache_from_board
            from monday_cache import save_cache
            cache = build_cache_from_board(
                self.cfg["monday_api_token"],
                self.cfg["monday_board_id"],
            )
            save_cache(cache)
            self.done.emit({"count": len(cache), "cache": cache})
        except Exception as e:
            self.error.emit(str(e))


# ──────────────────── PO Tab ────────────────────

class POTab(QWidget):
    status_message = pyqtSignal(str)

    COLS = ["", "เลขที่เอกสาร", "ผู้รับเงิน / Vendor", "แบรนด์", "วันที่", "ครบกำหนด", "ยอดเงิน (บาท)", "สถานะ", "ไฟล์แนบ"]

    def __init__(self):
        super().__init__()
        self._pos:      list = []
        self._selected: set  = set()
        self._filtered: list = []
        self._page     = 1
        self._per_page = 25
        self._worker = None
        try:
            self._combined = bool(load_config().get("queue_all_companies", False))
        except Exception:
            self._combined = False
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8,8,8,8)
        layout.setSpacing(6)

        # Toolbar
        tb = QHBoxLayout()
        self.lbl_fetched = QLabel("ยังไม่ได้ดึงข้อมูล")
        self.lbl_fetched.setStyleSheet(f"color:{C_MUTED};font-size:11px;")
        self.btn_refresh  = QPushButton("🔄 รีเฟรช")
        self.btn_convert  = QPushButton("🔄 เปลี่ยนเป็นใบทำจ่าย")
        self.btn_attach   = QPushButton("📎 อัปโหลดรูป PO")

        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("🔍 ค้นหา vendor / เลขเอกสาร")
        self.search_box.setClearButtonEnabled(True)
        self.search_box.setFixedWidth(230)
        self.search_box.textChanged.connect(self._apply_filter)

        self.btn_auto_queue = QPushButton("🤖 จัดคิวอัตโนมัติ")
        self.btn_export     = QPushButton("⬇️ Export")

        self.btn_refresh.clicked.connect(self.fetch_pos)
        self.btn_convert.clicked.connect(self._convert_selected)
        self.btn_attach.clicked.connect(self._upload_attachment)
        self.btn_auto_queue.clicked.connect(self._auto_queue)
        self.btn_export.clicked.connect(self._export)

        for btn in (self.btn_convert, self.btn_attach):
            btn.setEnabled(False)

        self._style_btn(self.btn_refresh, C_MUTED)
        self._style_btn(self.btn_auto_queue, "#eab308")
        self._style_btn(self.btn_export, C_PRIMARY)
        self._style_btn(self.btn_convert, C_WARNING)
        self._style_btn(self.btn_attach,  C_SUCCESS)

        tb.addWidget(QLabel("ใบสั่งซื้อ"))
        tb.addWidget(self.search_box)
        tb.addWidget(self.btn_refresh)
        tb.addWidget(self.btn_auto_queue)
        tb.addStretch()
        tb.addWidget(self.lbl_fetched)
        tb.addWidget(self.btn_export)
        tb.addWidget(self.btn_convert)
        tb.addWidget(self.btn_attach)
        layout.addLayout(tb)

        # Select-all
        sel_row = QHBoxLayout()
        self.chk_all   = QCheckBox("เลือกทั้งหมด")
        self.lbl_count = QLabel("0 รายการที่เลือก")
        self.lbl_count.setStyleSheet(f"color:{C_PRIMARY};font-weight:600;")
        self.chk_all.stateChanged.connect(self._toggle_all)
        sel_row.addWidget(self.chk_all)
        sel_row.addWidget(self.lbl_count)
        sel_row.addStretch()
        layout.addLayout(sel_row)

        # Table
        self.table = QTableWidget(0, len(self.COLS))
        self.table.setHorizontalHeaderLabels(self.COLS)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.table.setColumnWidth(0, 34)
        self.table.setColumnWidth(1, 130)
        self.table.setColumnWidth(3, 110)   # แบรนด์
        self.table.setColumnWidth(4, 100)
        self.table.setColumnWidth(5, 110)
        self.table.setColumnWidth(6, 130)
        self.table.setColumnWidth(7, 100)
        self.table.setColumnWidth(8, 120)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSortingEnabled(True)
        self.table.setStyleSheet("QTableWidget{border:1px solid #e2e8f0;font-size:13px;}")
        layout.addWidget(self.table)

        # ── Summary + Pagination ──
        bottom = QHBoxLayout()
        self.lbl_summary = QLabel("")
        self.lbl_summary.setStyleSheet(f"color:{C_MUTED};font-size:11px;padding:2px 0;")

        self.btn_prev = QPushButton("◀ ก่อนหน้า")
        self.btn_next = QPushButton("ถัดไป ▶")
        self.lbl_page = QLabel("หน้า 1/1")
        self.lbl_page.setStyleSheet(f"color:{C_PRIMARY};font-weight:600;font-size:12px;padding:0 8px;")
        self.btn_prev.clicked.connect(self._prev_page)
        self.btn_next.clicked.connect(self._next_page)
        for b in (self.btn_prev, self.btn_next):
            b.setStyleSheet(
                "QPushButton{padding:4px 12px;border:1px solid #cbd5e1;border-radius:4px;"
                "background:white;font-size:12px;}"
                "QPushButton:hover:!disabled{background:#f1f5f9;}"
                "QPushButton:disabled{color:#cbd5e1;border-color:#e2e8f0;}")

        bottom.addWidget(self.lbl_summary)
        bottom.addStretch()
        bottom.addWidget(self.btn_prev)
        bottom.addWidget(self.lbl_page)
        bottom.addWidget(self.btn_next)
        layout.addLayout(bottom)

    def _style_btn(self, btn, colour):
        btn.setStyleSheet(f"""
            QPushButton{{background:{colour};color:white;border:none;
              border-radius:5px;padding:6px 14px;font-size:13px;font-weight:500;}}
            QPushButton:hover:!disabled{{background:{colour}cc;}}
            QPushButton:disabled{{background:#94a3b8;}}
        """)

    def fetch_pos(self):
        cfg = load_config()
        comps = cfg.get("companies") or []
        self._combined = bool(cfg.get("queue_all_companies", False))
        if not cfg.get("flowaccount_api_key") and not comps:
            QMessageBox.warning(self,"ยังไม่ตั้งค่า","กรุณาตั้งค่า FlowAccount API Key ก่อนครับ")
            return
        self.btn_refresh.setEnabled(False)
        self.btn_refresh.setText("⏳ กำลังดึง...")
        if self._combined and comps:
            self.status_message.emit("กำลังดึงใบสั่งซื้อจากทุกบริษัท...")
            self._worker = POFetchAllWorker(comps)
            self._worker.progress.connect(self.status_message.emit)
        else:
            active = cfg.get("active_company", 0)
            company = comps[active] if (0 <= active < len(comps)) else {}
            self.status_message.emit("กำลังดึงข้อมูลใบสั่งซื้อ...")
            self._worker = POFetchWorker(cfg, company)
        self._worker.done.connect(self._on_fetch_done)
        self._worker.error.connect(self._on_fetch_error)
        self._worker.start()

    def _on_fetch_done(self, pos):
        self._pos = pos
        self._apply_filter()
        self.lbl_fetched.setText(f"อัพเดตล่าสุด {datetime.now().strftime('%H:%M:%S')}")
        self.btn_refresh.setEnabled(True)
        self.btn_refresh.setText("🔄 รีเฟรช")
        self.status_message.emit(f"ดึง PO สำเร็จ {len(pos)} รายการ")

    def _apply_filter(self):
        """กรองตามคำค้นหา แล้วแสดงหน้าแรก"""
        q = self.search_box.text().strip().lower()
        if q:
            self._filtered = [p for p in self._pos if
                              q in _vendor(p).lower() or
                              q in _doc_no(p).lower()]
        else:
            self._filtered = list(self._pos)
        self._page = 1
        self._render_page()

    def _render_page(self):
        """แสดงผลเฉพาะหน้าปัจจุบัน (25 รายการ/หน้า)"""
        per   = self._per_page
        total = len(self._filtered)
        pages = max(1, (total + per - 1) // per)
        self._page = min(max(1, self._page), pages)
        start = (self._page - 1) * per
        page_items = self._filtered[start:start + per]

        self._populate(page_items)

        total_amt = sum(_amount(p) for p in self._filtered)
        self.lbl_summary.setText(
            f"ทั้งหมด {total} รายการ  |  รวม {fmt_amount(total_amt)} บาท"
            f"  |  แสดง {len(page_items)} รายการ")
        self.lbl_page.setText(f"หน้า {self._page}/{pages}")
        self.btn_prev.setEnabled(self._page > 1)
        self.btn_next.setEnabled(self._page < pages)

    def _prev_page(self):
        if self._page > 1:
            self._page -= 1
            self._render_page()

    def _next_page(self):
        per   = self._per_page
        pages = max(1, (len(self._filtered) + per - 1) // per)
        if self._page < pages:
            self._page += 1
            self._render_page()

    def _on_fetch_error(self, msg):
        self.btn_refresh.setEnabled(True)
        self.btn_refresh.setText("🔄 รีเฟรช")
        QMessageBox.critical(self,"เกิดข้อผิดพลาด",msg)

    def _populate(self, pos):
        today = today_str()
        self.table.setSortingEnabled(False)
        self.table.setRowCount(0)
        total = 0.0
        for po in pos:
            row = self.table.rowCount()
            self.table.insertRow(row)
            pid     = _exp_id(po)
            due     = _due(po)[:10] if _due(po) else ""
            doc_d   = (po.get("publishedOn") or "")[:10]
            amt     = _amount(po)
            total  += amt
            st      = _status(po)
            bg = C_OVERDUE if due and due < today else None

            def cell(text, align=Qt.AlignmentFlag.AlignVCenter|Qt.AlignmentFlag.AlignLeft):
                item = QTableWidgetItem(text)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                item.setTextAlignment(align)
                if bg: item.setBackground(QBrush(bg))
                return item

            chk = QCheckBox()
            chk.setChecked(pid in self._selected)
            chk.setProperty("po_id", pid)
            chk.stateChanged.connect(self._on_check)
            cw = QWidget(); cl = QHBoxLayout(cw)
            cl.addWidget(chk); cl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            cl.setContentsMargins(0,0,0,0)
            self.table.setCellWidget(row,0,cw)

            self.table.setItem(row,1,cell(_doc_no(po)))
            self.table.setItem(row,2,cell(_vendor(po) or "(ไม่ระบุ)"))
            # แบรนด์ — ขึ้นอัตโนมัติจากบริษัทต้นทาง
            brand = _brand_name(po)
            bitem = cell(brand or "—", Qt.AlignmentFlag.AlignVCenter|Qt.AlignmentFlag.AlignHCenter)
            if brand:
                bitem.setForeground(QBrush(QColor(po.get("_brand_color") or "#334155")))
                bf = bitem.font(); bf.setBold(True); bitem.setFont(bf)
            self.table.setItem(row,3,bitem)
            self.table.setItem(row,4,cell(fmt_date(doc_d) if doc_d else "-"))
            self.table.setItem(row,5,cell(fmt_date(due) if due else "-"))
            self.table.setItem(row,6,cell(fmt_amount(amt),
                Qt.AlignmentFlag.AlignVCenter|Qt.AlignmentFlag.AlignRight))
            st_map = {"paid":"✅ จ่ายแล้ว","approved":"🔵 อนุมัติ","awaiting_approval":"🟡 รออนุมัติ"}
            self.table.setItem(row,7,cell(st_map.get(st,"🟡 "+st if st else "🟡 ร่าง")))
            # attach button
            btn_a = QPushButton("📎 แนบ")
            btn_a.setStyleSheet("font-size:11px;padding:2px 8px;border-radius:3px;background:#e2e8f0;border:none;cursor:pointer;")
            btn_a.clicked.connect(lambda checked, p=pid: self._upload_one(p))
            self.table.setCellWidget(row,8,btn_a)

        self.table.setSortingEnabled(True)
        self._update_buttons()

    def _on_check(self, state):
        chk = self.sender()
        pid = chk.property("po_id")
        if state == Qt.CheckState.Checked.value: self._selected.add(pid)
        else: self._selected.discard(pid)
        self._update_buttons()

    def _toggle_all(self, state):
        checked = state == Qt.CheckState.Checked.value
        self._selected.clear()
        for row in range(self.table.rowCount()):
            w = self.table.cellWidget(row,0)
            if w:
                chk = w.findChild(QCheckBox)
                if chk:
                    chk.blockSignals(True); chk.setChecked(checked); chk.blockSignals(False)
                    if checked: self._selected.add(chk.property("po_id"))
        self._update_buttons()

    def _update_buttons(self):
        n = len(self._selected)
        self.lbl_count.setText(f"{n} รายการที่เลือก")
        self.btn_convert.setEnabled(n > 0)
        self.btn_attach.setEnabled(n > 0)
        self.btn_export.setEnabled(bool(self._pos))

    def _selected_pos(self):
        return [p for p in self._pos if _exp_id(p) in self._selected]

    def _convert_selected(self):
        pos = self._selected_pos()
        if not pos: return
        reply = QMessageBox.question(self,"ยืนยัน",
            f"จะเปลี่ยนสถานะ {len(pos)} รายการ\nจาก ใบสั่งซื้อ → ใบทำจ่าย\n\n⚠️ ไม่สามารถย้อนกลับได้",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel)
        if reply != QMessageBox.StandardButton.Yes: return
        cfg = load_config()
        self.btn_convert.setEnabled(False)
        self.btn_convert.setText("⏳ กำลังดำเนินการ...")
        self._conv_worker = POConvertWorker(cfg, pos)
        self._conv_worker.done.connect(self._on_convert_done)
        self._conv_worker.start()

    def _on_convert_done(self, ok, errs):
        self.btn_convert.setEnabled(True)
        self.btn_convert.setText("🔄 เปลี่ยนเป็นใบทำจ่าย")
        if errs:
            err_txt = "\n".join(f"• {e['id']}: {e['error']}" for e in errs[:5])
            QMessageBox.warning(self,"มีข้อผิดพลาด",f"สำเร็จ {len(ok)} รายการ\nข้อผิดพลาด:\n{err_txt}")
        else:
            QMessageBox.information(self,"✅ สำเร็จ",f"เปลี่ยนสถานะ {len(ok)} รายการเรียบร้อยแล้ว")
        self._selected.clear()
        self.fetch_pos()

    def _upload_one(self, po_id):
        self._do_upload(po_id)

    def _upload_attachment(self):
        pos = self._selected_pos()
        if not pos: return
        self._do_upload(_exp_id(pos[0]))

    def _do_upload(self, po_id):
        path, _ = QFileDialog.getOpenFileName(self,"เลือกไฟล์รูป PO ที่เซ็นแล้ว","",
            "Images (*.png *.jpg *.jpeg);;PDF (*.pdf);;All Files (*)")
        if not path: return
        cfg = load_config()
        po = next((p for p in self._pos if _exp_id(p) == po_id), {})
        cid  = po.get("_co_id")     or cfg["flowaccount_api_key"]
        csec = po.get("_co_secret") or cfg["flowaccount_secret_key"]
        try:
            with open(path,"rb") as f: data = f.read()
            import os as _os
            upload_po_attachment(cid, csec, po_id, data, _os.path.basename(path))
            QMessageBox.information(self,"✅ สำเร็จ","อัปโหลดไฟล์เรียบร้อยแล้ว")
        except FlowAccountError as e:
            QMessageBox.critical(self,"อัปโหลดล้มเหลว",str(e))

    def _daily_limit(self):
        try:
            return float(load_config().get("daily_payment_limit", 150000) or 0)
        except Exception:
            return 150000.0

    def _auto_queue(self):
        """จัดคิวใบสั่งซื้อหลายวัน (ค่าเริ่มต้น 3 วัน) แล้วเลือกทั้งหมด"""
        from PyQt6.QtWidgets import QInputDialog
        candidates = [p for p in self._pos if _status(p) != "paid" and _amount(p) > 0]
        if not candidates:
            QMessageBox.information(self, "ไม่มีรายการ",
                "ไม่มีใบสั่งซื้อให้จัดคิว — กด 🔄 รีเฟรช ก่อนครับ")
            return
        all_days = [d for d in distribute_into_days(candidates, self._daily_limit()) if d["items"]]
        maxd = len(all_days)
        n, ok = QInputDialog.getInt(self, "🤖 จัดคิวใบสั่งซื้อ",
            f"จัดได้ทั้งหมด {maxd} วัน\nต้องการจัดคิวกี่วัน?", min(3, maxd), 1, maxd)
        if not ok:
            return
        days = all_days[:n]
        picked = [p for d in days for p in d["items"]]
        thai = ["จันทร์","อังคาร","พุธ","พฤหัสบดี","ศุกร์","เสาร์","อาทิตย์"]
        msg = [f"จัดคิว {n} วัน:"]
        for di, d in enumerate(days, 1):
            dt = d["date"]; dtot = sum(_amount(e) for e in d["items"])
            msg.append(f"  📅 วันที่ {di} ({thai[dt.weekday()]} {fmt_date(dt.isoformat())}): "
                       f"{len(d['items'])} รายการ • {fmt_amount(dtot)} บาท")
        msg.append(f"\nรวม {len(picked)} รายการ • {fmt_amount(sum(_amount(e) for e in picked))} บาท")
        if QMessageBox.question(self, "🤖 จัดคิวใบสั่งซื้อ", "\n".join(msg) + "\n\nเลือกรายการเหล่านี้?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel) != QMessageBox.StandardButton.Yes:
            return
        self._selected = {_exp_id(p) for p in picked}
        self._render_page()
        self._update_buttons()
        self.status_message.emit(f"🤖 จัดคิว {n} วัน: {len(picked)} ใบสั่งซื้อ")

    def _export(self):
        from PyQt6.QtWidgets import QInputDialog
        sel = self._selected_pos()
        pool = sel if sel else [p for p in self._pos if _status(p) != "paid" and _amount(p) > 0]
        if not pool:
            QMessageBox.information(self, "ไม่มีรายการ", "ไม่มีใบสั่งซื้อให้ export")
            return
        cfg = load_config()
        for e in pool:
            if not e.get("_brand_name"):
                b = _brand_name(e)
                if b: e["_brand_name"] = b
        all_days = [d for d in distribute_into_days(pool, self._daily_limit()) if d["items"]]
        if sel:
            days = all_days
        else:
            maxd = len(all_days)
            n, ok = QInputDialog.getInt(self, "Export ใบสั่งซื้อกี่วัน?",
                f"จัดได้ทั้งหมด {maxd} วัน\nต้องการ export กี่วัน?", min(3, maxd), 1, maxd)
            if not ok: return
            days = all_days[:n]
        exps = [e for d in days for e in d["items"]]
        base, _ = QFileDialog.getSaveFileName(self, "บันทึกไฟล์ (Excel + PDF)",
            f"KCash_PO_{date.today().strftime('%Y%m%d')}", "ไฟล์ใบสั่งซื้อ (*.xlsx)")
        if not base: return
        for ext in (".xlsx", ".pdf", ".csv"):
            if base.lower().endswith(ext): base = base[:-len(ext)]
        try:
            build_bank_excel_multiday(days, cfg, base + ".xlsx")
            build_bank_pdf_multiday(days, cfg, base + ".pdf")
        except Exception as e:
            QMessageBox.critical(self, "❌ ผิดพลาด", f"บันทึกไฟล์ไม่สำเร็จ:\n{e}")
            return
        QMessageBox.information(self, "✅ สำเร็จ",
            f"บันทึกแล้ว {len(exps)} ใบสั่งซื้อ • {len(days)} วัน:\n\n"
            f"📊 {base}.xlsx\n📄 {base}.pdf")

    def get_pos(self): return self._pos

    def clear_pos(self):
        """ล้างรายการ PO (ใช้ตอนสลับบริษัท)"""
        self._pos = []
        if hasattr(self, "table"):
            self.table.setRowCount(0)


# ──────────────────── Statement Tab ────────────────────

class StatementTab(QWidget):
    matched_updated = pyqtSignal(dict)

    STMT_COLS = ["วันที่", "รายการ", "เดบิต (จ่าย)", "เครดิต (รับ)"]
    MATCH_COLS = ["เลขที่เอกสาร", "Vendor", "ยอด Expense", "Statement ที่จับคู่"]

    def __init__(self):
        super().__init__()
        self._statements: list = []
        self._expenses: list   = []
        self._matched: dict    = {}
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(8)

        # ── หัวข้อ + คำอธิบาย ──
        header = QLabel("🏦 จับคู่ Statement ธนาคาร กับ ค่าใช้จ่าย")
        header.setStyleSheet("font-size:16px;font-weight:700;color:#15803d;")
        layout.addWidget(header)
        sub = QLabel("นำเข้าไฟล์ CSV รายการเดินบัญชีจากธนาคาร → ระบบจับคู่กับค่าใช้จ่ายอัตโนมัติ "
                     "(ยอด+วันที่) เพื่อยืนยันว่าจ่ายจริงแล้ว")
        sub.setWordWrap(True)
        sub.setStyleSheet("color:#64748b;font-size:12px;")
        layout.addWidget(sub)

        # ── Toolbar (การ์ดสีเทาอ่อน) ──
        tbar = QWidget()
        tbar.setStyleSheet("background:#f8fafc;border:1px solid #e2e8f0;border-radius:6px;")
        tb = QHBoxLayout(tbar)
        tb.setContentsMargins(10, 8, 10, 8)
        btn_import = QPushButton("📂 Import CSV Statement")
        btn_clear  = QPushButton("🗑️ ล้าง Statement")
        btn_auto   = QPushButton("🔗 Auto-Match อีกครั้ง")
        btn_import.clicked.connect(self._import_stmt)
        btn_clear.clicked.connect(self._clear)
        btn_auto.clicked.connect(self._auto_match)
        btn_import.setStyleSheet(
            "QPushButton{padding:7px 14px;border:none;border-radius:5px;background:#2563eb;"
            "color:white;font-size:12px;font-weight:600;}QPushButton:hover{background:#1d4ed8;}")
        for btn in (btn_clear, btn_auto):
            btn.setStyleSheet(
                "QPushButton{padding:7px 14px;border:1px solid #cbd5e1;border-radius:5px;"
                "background:white;color:#475569;font-size:12px;font-weight:600;}"
                "QPushButton:hover{background:#f1f5f9;}")
        self.lbl_stats = QLabel("ยังไม่มี statement")
        self.lbl_stats.setStyleSheet("color:#64748b;font-size:12px;font-weight:600;")
        tb.addWidget(btn_import)
        tb.addWidget(btn_clear)
        tb.addWidget(btn_auto)
        tb.addStretch()
        tb.addWidget(self.lbl_stats)
        layout.addWidget(tbar)

        # ── Splitter: ซ้าย=ตารางจับคู่ / ขวา=รายการ statement ──
        splitter = QSplitter(Qt.Orientation.Horizontal)
        tbl_css = ("QTableWidget{border:1px solid #e2e8f0;border-radius:4px;font-size:12px;"
                   "gridline-color:#eef2f7;}"
                   "QHeaderView::section{background:#dcfce7;color:#15803d;font-weight:700;"
                   "padding:5px;border:none;border-right:1px solid #cbd5e1;}")

        def _panel(title, table):
            w = QWidget(); vl = QVBoxLayout(w)
            vl.setContentsMargins(0, 0, 0, 0); vl.setSpacing(4)
            lab = QLabel(title)
            lab.setStyleSheet("font-size:13px;font-weight:600;color:#334155;")
            vl.addWidget(lab)
            vl.addWidget(table, 1)
            return w

        self.match_table = QTableWidget(0, len(self.MATCH_COLS))
        self.match_table.setHorizontalHeaderLabels(self.MATCH_COLS)
        self.match_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.match_table.horizontalHeader().setStretchLastSection(False)
        self.match_table.setColumnWidth(1, 240)
        self.match_table.verticalHeader().setVisible(False)
        self.match_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.match_table.setAlternatingRowColors(True)
        self.match_table.setStyleSheet(tbl_css)
        splitter.addWidget(_panel("📋 ค่าใช้จ่าย ↔ Statement (จับคู่แล้ว)", self.match_table))

        self.stmt_table = QTableWidget(0, len(self.STMT_COLS))
        self.stmt_table.setHorizontalHeaderLabels(self.STMT_COLS)
        self.stmt_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.stmt_table.horizontalHeader().setStretchLastSection(False)
        self.stmt_table.setColumnWidth(1, 240)
        self.stmt_table.verticalHeader().setVisible(False)
        self.stmt_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.stmt_table.setAlternatingRowColors(True)
        self.stmt_table.setStyleSheet(tbl_css)
        splitter.addWidget(_panel("📑 รายการเดินบัญชีใน Statement", self.stmt_table))

        splitter.setSizes([620, 420])
        layout.addWidget(splitter, 1)   # ขยายเต็มพื้นที่ที่เหลือ (แก้ช่องว่าง)

    def set_expenses(self, expenses: list):
        self._expenses = expenses
        if self._statements:
            self._auto_match()
        self._refresh_match_table()

    def _import_stmt(self):
        path, _ = QFileDialog.getOpenFileName(self, "เลือกไฟล์ CSV Statement", "", "CSV Files (*.csv);;All Files (*)")
        if not path:
            return
        try:
            with open(path, encoding="utf-8-sig", errors="replace") as f:
                content = f.read()
            self._statements = parse_statement_csv(content)
        except Exception as exc:
            QMessageBox.critical(self, "ไม่สามารถอ่านไฟล์ได้", str(exc))
            return
        self._auto_match()
        self._refresh_stmt_table()
        self._refresh_match_table()
        self.lbl_stats.setText(f"{len(self._statements)} รายการ | จับคู่ {len(self._matched)} รายการ")

    def _clear(self):
        self._statements = []
        self._matched    = {}
        self._refresh_stmt_table()
        self._refresh_match_table()
        self.lbl_stats.setText("ยังไม่มี statement")
        self.matched_updated.emit({})

    def _auto_match(self):
        if self._statements and self._expenses:
            self._matched = match_statements(self._statements, self._expenses)
            self.matched_updated.emit(self._matched)
        self._refresh_match_table()
        self.lbl_stats.setText(f"{len(self._statements)} รายการ | จับคู่ {len(self._matched)} รายการ")

    def _refresh_stmt_table(self):
        t = self.stmt_table
        t.setRowCount(0)
        used = set(id(s) for s in self._matched.values())
        for stmt in self._statements:
            row = t.rowCount()
            t.insertRow(row)
            bg = C_MATCHED if id(stmt) in used else None
            for col, txt in enumerate([
                stmt.get("date", ""),
                stmt.get("description", ""),
                fmt_amount(stmt["debit"])  if stmt["debit"]  else "—",
                fmt_amount(stmt["credit"]) if stmt["credit"] else "—",
            ]):
                item = QTableWidgetItem(txt)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if bg:
                    item.setBackground(QBrush(bg))
                t.setItem(row, col, item)

    def _refresh_match_table(self):
        t = self.match_table
        t.setRowCount(0)
        for exp in self._expenses:
            eid = _exp_id(exp)
            row = t.rowCount()
            t.insertRow(row)
            stmt = self._matched.get(eid)
            match_txt = (f"✅ {stmt['date']}  {fmt_amount(stmt['debit'] or stmt['credit'])}"
                         if stmt else "—")
            for col, txt in enumerate([
                _doc_no(exp),
                _vendor(exp) or "(ไม่ระบุ)",
                fmt_amount(_amount(exp)),
                match_txt,
            ]):
                item = QTableWidgetItem(txt)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if stmt:
                    item.setBackground(QBrush(C_MATCHED))
                t.setItem(row, col, item)


# ──────────────────── Slip Matching Tab (ตัดบิล) ────────────────────

class SensitiveEyeWidget(QWidget):
    """ปุ่ม 👁 toggle ซ่อน/แสดงยอดเงิน — ใช้ใน SlipMatchTab"""
    def __init__(self, amount: float, is_revealed=False,
                 on_unlock=None, on_relock=None, parent=None):
        super().__init__(parent)
        self._amount = amount
        self._revealed = is_revealed
        self._on_unlock = on_unlock
        self._on_relock = on_relock

        lay = QHBoxLayout(self)
        lay.setContentsMargins(4, 0, 4, 0)
        lay.setSpacing(4)

        self.lbl = QLabel()
        self.lbl.setAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        self.lbl.setStyleSheet("font-size:13px;background:transparent;")
        lay.addWidget(self.lbl, 1)

        self.btn = QPushButton()
        self.btn.setFixedSize(26, 26)
        self.btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn.clicked.connect(self._toggle)
        lay.addWidget(self.btn)

        self._refresh()

    def _refresh(self):
        if self._revealed:
            self.lbl.setText(fmt_amount(self._amount))
            self.lbl.setStyleSheet("font-size:13px;color:#0f172a;background:transparent;")
            self.btn.setText("👁")
            self.btn.setToolTip("คลิกเพื่อซ่อนยอดเงิน")
            self.btn.setStyleSheet(
                "QPushButton{border:none;border-radius:13px;"
                "background:#dcfce7;color:#15803d;font-size:13px;}"
                "QPushButton:hover{background:#bbf7d0;}")
        else:
            self.lbl.setText("●●●●●")
            self.lbl.setStyleSheet("font-size:13px;color:#94a3b8;background:transparent;")
            self.btn.setText("🔒")
            self.btn.setToolTip("คลิกเพื่อดูยอดเงิน (ต้องใส่รหัสผ่าน)")
            self.btn.setStyleSheet(
                "QPushButton{border:none;border-radius:13px;"
                "background:#dbeafe;color:#2563eb;font-size:13px;}"
                "QPushButton:hover{background:#bfdbfe;}")

    def _toggle(self):
        if self._revealed:
            self._revealed = False
            self._refresh()
            if self._on_relock:
                self._on_relock()
            return
        # ซ่อนอยู่ → ขอรหัส
        data = _sensitive_load()
        if not data.get("password_hash"):
            QMessageBox.information(self, "ข้อมูลลับ",
                "ยังไม่ได้ตั้งรหัสผ่าน\nAdmin ตั้งได้ที่เมนู 'จัดการผู้ใช้ → จัดการข้อมูลลับ'")
            return
        pw, ok = QInputDialog.getText(
            self, "🔒 ยืนยันตัวตน", "ใส่รหัสผ่านเพื่อดูยอดเงิน:",
            QLineEdit.EchoMode.Password)
        if not ok:
            return
        if not pw or not _sensitive_check_pw(pw):
            QMessageBox.warning(self, "รหัสผ่านไม่ถูกต้อง",
                "รหัสผ่านไม่ถูกต้อง กรุณาลองใหม่")
            return
        self._revealed = True
        self._refresh()
        if self._on_unlock:
            self._on_unlock()


class SlipPDFWorker(QThread):
    """อ่านไฟล์สลิป PDF แบบ background — ไม่ให้ UI ค้างระหว่างโหลด"""
    done  = pyqtSignal(list)
    error = pyqtSignal(str)
    def __init__(self, path, parent=None):
        super().__init__(parent)
        self.path = path
    def run(self):
        try:
            records = parse_payment_report(self.path)
            self.done.emit(records)
        except Exception as exc:
            self.error.emit(str(exc))


class SlipImageWorker(QThread):
    """อ่านสลิปจาก 'รูปภาพ' หลายไฟล์ด้วย OCR (Windows OCR) + PDF (ถ้ามี) — background
    บัญชีเขียนแท็ก 'ai#EXP<เลข>' ในรูป → อ่านเลข EXP ไปจับคู่แบบเลขตรง"""
    done  = pyqtSignal(list)
    error = pyqtSignal(str)
    def __init__(self, img_paths, pdf_path=None, parent=None):
        super().__init__(parent)
        self.img_paths = img_paths
        self.pdf_path = pdf_path
    def run(self):
        try:
            import slip_ocr
            recs = []
            if self.pdf_path:                       # เผื่อเลือก PDF ปนกับรูป
                recs.extend(parse_payment_report(self.pdf_path))
            for p in self.img_paths:
                recs.append(slip_ocr.parse_slip_image(p))
            self.done.emit(recs)
        except Exception as exc:
            self.error.emit(str(exc))


class SlipAttachWorker(QThread):
    """ตัดบิลเป็นรูป + แนบเข้า FlowAccount + เปลี่ยนสถานะเป็นจ่ายแล้ว (background) — Phase 2"""
    progress = pyqtSignal(int, int)
    # ok_count, fail_count, problems(list of dict: doc, attach, status)
    finished_all = pyqtSignal(int, int, list)

    def __init__(self, pdf_path, matched_results, parent=None, save_dir=None):
        super().__init__(parent)
        self.pdf_path = pdf_path
        self.items = matched_results
        self.save_dir = save_dir   # โฟลเดอร์บันทึกรูปที่ตัด (None = ไม่บันทึก)

    def run(self):
        from bank_slip import crop_all_rows
        from flowaccount import (upload_expense_attachment, mark_expense_paid,
                                 update_expense_dates)
        recs = [r["slip"] for r in self.items]
        try:
            pngs = crop_all_rows(self.pdf_path, recs)
        except Exception as e:
            self.finished_all.emit(0, len(self.items),
                [{"doc": "(ทั้งหมด)", "attach": f"ตัดรูปไม่สำเร็จ: {e}", "status": ""}])
            return
        cfg = load_config()
        ok = fail = 0
        problems = []
        for i, (res, png) in enumerate(zip(self.items, pngs), 1):
            e = res["expense"]
            s = res["slip"]
            doc = _slip_exp_serial(e)
            cid = e.get("_co_id") or cfg.get("flowaccount_api_key", "")
            csec = e.get("_co_secret") or cfg.get("flowaccount_secret_key", "")
            eid = e.get("recordId") or e.get("documentId")
            attach_err = status_err = ""

            # 1) แนบสลิป — ชื่อไฟล์ = เลขเอกสาร_ชื่อคน
            try:
                if not png:
                    raise RuntimeError("ตัดรูปไม่ได้")
                import re as _re
                nm = _re.sub(r"[\\/:*?\"<>|]", "", (s.get("recv_name") or "")).strip()[:30]
                fn = (f"{doc}_{nm}.png".strip("_")) if doc else f"slip_{nm}.png"
                # บันทึกรูปที่ตัดลงโฟลเดอร์ที่เลือก (ถ้ามี) ก่อนอัปโหลด
                if self.save_dir:
                    try:
                        with open(os.path.join(self.save_dir, fn), "wb") as _f:
                            _f.write(png)
                    except Exception:
                        pass
                upload_expense_attachment(cid, csec, eid, png, fn)
            except Exception as ex:
                attach_err = str(ex)[:120]

            # 2) แก้วันที่เอกสาร + ครบกำหนด = วันที่แนบ (วันนี้) + เครดิต 0
            #    (ทำก่อน mark จ่าย — ให้วันเอกสารตรงกับวันแนบตามที่ลูกค้าต้องการ)
            attach_date = today_str()
            datefix_err = ""
            try:
                update_expense_dates(cid, csec, eid, attach_date)
                e["publishedOn"] = attach_date
                e["dueDate"] = attach_date
            except Exception as ex:
                datefix_err = str(ex)[:150]

            # 3) เปลี่ยนสถานะเป็นจ่ายแล้วใน FlowAccount (วันจ่าย = วันที่โอนจริงจากสลิป)
            pay_date = s.get("pay_date") or attach_date
            try:
                mark_expense_paid(cid, csec, eid, {
                    "paymentDate": pay_date,
                    "amount": _amount(e) or s.get("amount") or 0,
                    "paymentMethod": 5,            # โอนเงิน
                    "payAccount": s.get("pay_acct"),  # บัญชีหักเงินจากสลิป → จับคู่ bankAccountId
                })
                # อัปเดตสถานะในโปรแกรม (รายการเดียวกัน ref ร่วมกับหน้าคิว)
                e["statusString"] = "paid"
                e["payments"] = [{"paymentDate": pay_date}]
            except Exception as ex:
                status_err = str(ex)[:120]

            if attach_err or status_err or datefix_err:
                fail += 1
                problems.append({"doc": doc, "attach": attach_err,
                                 "status": status_err or datefix_err})
            else:
                ok += 1
            self.progress.emit(i, len(self.items))
        self.finished_all.emit(ok, fail, problems)


class MatchExpenseFetchWorker(QThread):
    """ดึงค่าใช้จ่ายให้ครอบคลุมช่วงวันที่ของสลิป (สำหรับจับคู่ให้แม่น) — background"""
    done = pyqtSignal(list)
    error = pyqtSignal(str)

    def __init__(self, since_iso, parent=None):
        super().__init__(parent)
        self.since = since_iso

    def run(self):
        from concurrent.futures import ThreadPoolExecutor
        cfg = load_config()
        comps = cfg.get("companies") or []
        targets = [(c.get("client_id"), c.get("client_secret"), c)
                   for c in comps if c.get("client_id")]
        if not targets:
            targets = [(cfg.get("flowaccount_api_key"),
                        cfg.get("flowaccount_secret_key"), {})]

        def _one(t):
            cid, csec, co = t
            # ดึงเยอะ (ใหม่→เก่า) ให้ครอบคลุมช่วงสลิป — ไม่ใช้วันที่หยุด (วันที่ย้อนหลังได้)
            items = get_all_expenses(cid, csec, max_records=10000)
            if co:
                _tag_brand(items, co)
            return items
        try:
            out = []
            with ThreadPoolExecutor(max_workers=min(4, max(1, len(targets)))) as ex:
                for items in ex.map(_one, targets):
                    out.extend(items)
            self.done.emit(out)
        except Exception as e:
            self.error.emit(str(e))


class SlipManualMatchDialog(QDialog):
    """จับคู่สลิปกับค่าใช้จ่ายเอง — ค้นหา + เลือก expense"""
    def __init__(self, slip, expenses, parent=None):
        super().__init__(parent)
        self.setWindowTitle("จับคู่ค่าใช้จ่ายเอง")
        self.resize(620, 480)
        self.expenses = expenses
        self.picked = None
        amt = slip.get("amount") or 0

        lay = QVBoxLayout(self)
        lay.addWidget(QLabel(
            f"<b>สลิป:</b> {slip.get('recv_name','')} • "
            f"<b>{fmt_amount(amt)} บาท</b> • วันที่ {fmt_date(slip.get('pay_date','')) or '-'}"))

        self.search = QLineEdit()
        self.search.setPlaceholderText("🔍 ค้นหา ชื่อผู้รับ / เลขเอกสาร / ยอดเงิน")
        self.search.textChanged.connect(self._filter)
        lay.addWidget(self.search)

        self.listw = QListWidget()
        self.listw.itemDoubleClicked.connect(lambda *_: self._ok())
        lay.addWidget(self.listw, 1)

        # เรียงให้ "ยอดตรง" ขึ้นก่อน
        def sort_key(e):
            return (abs(_amount(e) - amt) > 0.01, abs(_amount(e) - amt))
        self._sorted = sorted(expenses, key=sort_key)
        self._fill(self._sorted)

        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        bb.accepted.connect(self._ok); bb.rejected.connect(self.reject)
        lay.addWidget(bb)

    def _fill(self, items):
        self.listw.clear()
        self._cur = items
        for e in items[:400]:
            txt = f"{_doc_no(e)}  |  {_vendor(e) or '-'}  |  {fmt_amount(_amount(e))} บาท"
            it = QListWidgetItem(txt)
            self.listw.addItem(it)

    def _filter(self, q):
        q = (q or "").strip().lower()
        if not q:
            self._fill(self._sorted); return
        out = [e for e in self._sorted if
               q in (_vendor(e) or "").lower() or
               q in (_doc_no(e) or "").lower() or
               q in f"{_amount(e):.2f}"]
        self._fill(out)

    def _ok(self):
        i = self.listw.currentRow()
        if 0 <= i < len(self._cur[:400]):
            self.picked = self._cur[i]
            self.accept()


class SlipMatchTab(QWidget):
    """
    ตัดบิล (พรีวิว) — อัปโหลดสลิป PDF จากธนาคาร → อ่านทีละแถว →
    จับคู่กับค่าใช้จ่ายในระบบด้วย ชื่อ+ยอด → แสดง matched / unmatched
    *** อ่านอย่างเดียว ยังไม่ตัดรูป ไม่แนบเข้า FlowAccount ไม่ mark จ่าย ***
    """
    status_message = pyqtSignal(str)

    COLS = ["✓", "#", "วันที่จ่าย", "ชื่อผู้รับ (สลิป)", "ยอด (บาท)",
            "จับคู่กับ EXP", "ชื่อในระบบ", "แบรนด์", "คะแนน", "สถานะ", "ดูใน FlowAccount", "รูป"]
    # หน้า "ไม่เจอคู่" — รวม "ไม่เจอ" + "น่าจะใช่" (ยังไม่ชัวร์ ต้องตรวจ)
    # โชว์รายละเอียดในสลิป + EXP ที่ระบบเดา เพื่อช่วยยืนยัน/ลงเอกสาร
    # คอลัมน์แรก ✓ = ติ๊กเลือกเพื่อนำกลับเข้าหน้า "จับคู่สำเร็จเตรียมดำเนินการต่อ"
    COLS_UN = ["✓", "#", "วันที่จ่าย", "ชื่อผู้รับ (สลิป)", "ยอด (บาท)",
               "รายละเอียดในสลิป (เลขบิล/บันทึก)", "EXP ที่ระบบเดา",
               "ชื่อในระบบ", "สถานะ", "ดูใน FlowAccount", "รูป"]

    def __init__(self):
        super().__init__()
        self._expenses: list = []
        self._records:  list = []
        self._results:  list = []
        self._pdf_path = ""
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(6)

        tb = QHBoxLayout()
        self.btn_open = QPushButton("📂 เลือกไฟล์สลิป PDF/รูปภาพ")
        self.btn_match = QPushButton("🔗 จับคู่")
        self.btn_savepng = QPushButton("✂️ ตัดบิลเป็นรูป (ลงโฟลเดอร์)")
        self.btn_attach = QPushButton("📎 ตัดบิล + แนบเข้า FlowAccount")
        self.btn_export = QPushButton("⬇️ Export Excel/PDF")
        self.btn_open.clicked.connect(self._open_pdf)
        self.btn_match.clicked.connect(self._do_match)
        self.btn_savepng.clicked.connect(self._cut_to_folder)
        self.btn_attach.clicked.connect(self._cut_and_attach)
        self.btn_export.clicked.connect(self._export_matched)
        self.btn_match.setEnabled(False)
        self.btn_savepng.setEnabled(False)
        self.btn_attach.setEnabled(False)
        self.btn_export.setEnabled(False)
        self._style(self.btn_open, "#16a34a")
        self._style(self.btn_match, C_PRIMARY)
        self._style(self.btn_savepng, "#0d9488")
        self._style(self.btn_attach, "#7c3aed")
        self._style(self.btn_export, "#0891b2")
        self.lbl_file = QLabel("ยังไม่ได้เลือกไฟล์")
        self.lbl_file.setStyleSheet(f"color:{C_MUTED}; font-size:11px;")
        tb.addWidget(self.btn_open)
        tb.addWidget(self.btn_match)
        tb.addWidget(self.btn_savepng)
        tb.addWidget(self.btn_attach)
        tb.addWidget(self.btn_export)
        tb.addWidget(self.lbl_file)
        tb.addStretch()
        # เลือกทั้งหมด (หน้าหลักโชว์เฉพาะเจอคู่ — ไม่เจอ/น่าจะใช่ อยู่แท็บ 'ไม่เจอคู่')
        self.chk_all_slip = QCheckBox("เลือกทั้งหมด")
        self.chk_all_slip.stateChanged.connect(self._toggle_all_slips)
        tb.addWidget(self.chk_all_slip)
        layout.addLayout(tb)
        self._checked_slips = set()   # เก็บ id() ของ result ที่ติ๊ก

        note = QLabel("ตัดบิล: เลือกไฟล์ → จับคู่ → ติ๊กเลือกรายการที่ต้องการ (ไม่ติ๊ก = ทำทั้งหมด) "
                      "→ 'ตัดบิล+แนบ' จะตัดสลิปเป็นรูปแล้วแนบเข้ารายการที่จับคู่ได้ + เปลี่ยนสถานะจ่ายแล้ว")
        note.setStyleSheet("color:#92400e; background:#fef9c3; border-radius:4px;"
                           "padding:5px 10px; font-size:11px;")
        note.setWordWrap(True)
        layout.addWidget(note)

        # สรุป
        self.lbl_summary = QLabel("")
        self.lbl_summary.setStyleSheet("font-size:12px; font-weight:600; padding:2px;")
        layout.addWidget(self.lbl_summary)

        # แท็บย่อย: ทั้งหมด / จับคู่สำเร็จเตรียมดำเนินการต่อ / ไม่เจอคู่
        self.subtabs = QTabWidget()

        # ── หน้า 0: ทั้งหมด (รวมทุกรายการ เจอคู่ + ไม่เจอคู่) ──
        self.table_all = self._make_match_table()
        self.table_all.cellDoubleClicked.connect(self._manual_match_all)
        self.subtabs.addTab(self.table_all, "📋 ทั้งหมด")
        self._shown_all = []

        # ── หน้า 1: จับคู่สำเร็จ เตรียมดำเนินการต่อ (เฉพาะที่เจอคู่ สีเขียว) ──
        page_ready = QWidget()
        lr = QVBoxLayout(page_ready)
        lr.setContentsMargins(0, 0, 0, 0)
        lr.setSpacing(4)
        bar_r = QHBoxLayout()
        hint_r = QLabel("รายการที่จับคู่สำเร็จ พร้อมตัดบิล/แนบ FlowAccount — "
                        "ติ๊กเลือกแล้วกด 'ส่งกลับ' เพื่อย้ายไปหน้า 'ไม่เจอคู่' "
                        "(รายการสีแดง = นำกลับมาเอง ควรตรวจก่อน)")
        hint_r.setStyleSheet("color:#166534;background:#dcfce7;border-radius:4px;"
                             "padding:5px 10px;font-size:11px;")
        hint_r.setWordWrap(True)
        bar_r.addWidget(hint_r, 1)
        self.btn_demote = QPushButton("➡️ ส่งกลับไป 'ไม่เจอคู่'")
        self._style(self.btn_demote, "#dc2626")
        self.btn_demote.clicked.connect(self._demote_to_unmatched)
        bar_r.addWidget(self.btn_demote)
        lr.addLayout(bar_r)
        self.table = self._make_match_table()
        self.table.cellDoubleClicked.connect(self._manual_match)
        lr.addWidget(self.table)
        self.subtabs.addTab(page_ready, "✅ จับคู่สำเร็จเตรียมดำเนินการต่อ")

        # ── หน้า 3: ไม่เจอคู่ (ต้องไปลงเอกสาร/จับคู่เอง) ──
        page_un = QWidget()
        lu = QVBoxLayout(page_un)
        lu.setContentsMargins(0, 0, 0, 0)
        lu.setSpacing(4)
        bar_un = QHBoxLayout()
        hint_un = QLabel("รายการที่ระบบจับคู่กับเอกสารใน FlowAccount ไม่ได้ — "
                         "ดับเบิลคลิกเพื่อจับคู่เอง หรือไปลงเอกสารใน FlowAccount "
                         "ก่อนแล้วกด 'จับคู่' ใหม่")
        hint_un.setStyleSheet("color:#991b1b;background:#fee2e2;border-radius:4px;"
                              "padding:5px 10px;font-size:11px;")
        hint_un.setWordWrap(True)
        bar_un.addWidget(hint_un, 1)
        # ปุ่มนำกลับ: ติ๊กเลือกรายการ → ย้ายกลับเข้าหน้า "จับคู่สำเร็จเตรียมดำเนินการต่อ"
        self.btn_promote = QPushButton("↩️ นำกลับไปไว้ใน 'ผลจับคู่สำเร็จเตรียมดำเนินการ'")
        self._style(self.btn_promote, "#16a34a")
        self.btn_promote.clicked.connect(self._promote_unmatched)
        bar_un.addWidget(self.btn_promote)
        self.btn_step2 = QPushButton("🔁 จับคู่รอบ 2 (ตามยอดเงิน)")
        self._style(self.btn_step2, "#0d9488")
        self.btn_step2.clicked.connect(self._match_step2_by_amount)
        bar_un.addWidget(self.btn_step2)
        self.btn_un_export = QPushButton("⬇️ บันทึกรายการไม่เจอคู่ (Excel)")
        self._style(self.btn_un_export, "#dc2626")
        self.btn_un_export.clicked.connect(self._export_unmatched)
        bar_un.addWidget(self.btn_un_export)
        lu.addLayout(bar_un)

        self.table_un = QTableWidget(0, len(self.COLS_UN))
        self.table_un.setHorizontalHeaderLabels(self.COLS_UN)
        _uhh = self.table_un.horizontalHeader()
        _uhh.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)  # ลากปรับขนาดได้อิสระ
        _uhh.setStretchLastSection(False)
        self.table_un.setColumnWidth(0, 34)    # ✓
        self.table_un.setColumnWidth(1, 40)    # #
        self.table_un.setColumnWidth(2, 90)    # วันที่จ่าย
        self.table_un.setColumnWidth(3, 220)   # ชื่อผู้รับ (สลิป)
        self.table_un.setColumnWidth(4, 100)   # ยอด
        self.table_un.setColumnWidth(5, 220)   # รายละเอียด
        self.table_un.setColumnWidth(6, 120)   # EXP ที่ระบบเดา
        self.table_un.setColumnWidth(7, 200)   # ชื่อในระบบ
        self.table_un.setColumnWidth(8, 140)   # สถานะ
        self.table_un.setColumnWidth(9, 130)   # ดูใน FlowAccount
        self.table_un.setColumnWidth(10, 80)   # รูป
        self.table_un.verticalHeader().setVisible(False)
        self.table_un.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table_un.setStyleSheet("QTableWidget{border:1px solid #fecaca;font-size:12px;}")
        self.table_un.cellDoubleClicked.connect(self._manual_match_un)
        lu.addWidget(self.table_un)
        self.subtabs.addTab(page_un, "🔴 ไม่เจอคู่")

        layout.addWidget(self.subtabs)
        self._shown = []
        self._shown_un = []
        self._checked_un = set()        # id(res) ของรายการไม่เจอคู่ที่ติ๊ก
        self._revealed_slip_ids = set() # id(res) ที่ unlock ดูยอดแล้ว (session)

    def _make_match_table(self):
        """สร้างตารางผลจับคู่ (คอลัมน์ COLS) — ใช้ทั้งหน้า 'เตรียมดำเนินการ' และ 'ทั้งหมด'"""
        t = QTableWidget(0, len(self.COLS))
        t.setHorizontalHeaderLabels(self.COLS)
        _thh = t.horizontalHeader()
        _thh.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)  # ลากปรับขนาดได้อิสระ
        _thh.setStretchLastSection(False)
        t.setColumnWidth(0, 34)    # ✓
        t.setColumnWidth(1, 40)    # #
        t.setColumnWidth(2, 90)    # วันที่จ่าย
        t.setColumnWidth(3, 220)   # ชื่อผู้รับ (สลิป)
        t.setColumnWidth(4, 100)   # ยอด
        t.setColumnWidth(5, 120)   # จับคู่กับ EXP
        t.setColumnWidth(6, 200)   # ชื่อในระบบ
        t.setColumnWidth(7, 110)   # แบรนด์
        t.setColumnWidth(8, 60)    # คะแนน
        t.setColumnWidth(9, 130)   # สถานะ
        t.setColumnWidth(10, 130)  # ดูใน FlowAccount
        t.setColumnWidth(11, 80)   # รูป
        t.verticalHeader().setVisible(False)
        t.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        t.setStyleSheet("QTableWidget{border:1px solid #e2e8f0;font-size:12px;}")
        return t

    def _style(self, btn, color):
        btn.setStyleSheet(
            "QPushButton{padding:6px 14px;border:none;border-radius:5px;"
            f"background:{color};color:white;font-size:12px;font-weight:600;}}"
            "QPushButton:disabled{background:#cbd5e1;}")

    def set_expenses(self, expenses: list):
        self._expenses = expenses or []

    def _open_pdf(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self, "เลือกไฟล์สลิป — PDF หรือรูปภาพ (เลือกได้หลายไฟล์)", "",
            "สลิปทั้งหมด (*.pdf *.png *.jpg *.jpeg *.bmp);;"
            "PDF (*.pdf);;รูปภาพ (*.png *.jpg *.jpeg *.bmp)")
        if not paths:
            return
        _IMG = (".png", ".jpg", ".jpeg", ".bmp")
        imgs = [p for p in paths if p.lower().endswith(_IMG)]
        pdfs = [p for p in paths if p.lower().endswith(".pdf")]
        if not imgs and not pdfs:
            QMessageBox.information(self, "ไฟล์ไม่รองรับ", "รองรับเฉพาะ PDF หรือรูปภาพ")
            return
        self._clear_slip_images()
        self._pdf_path = pdfs[0] if pdfs else None   # ใช้ตัดรูปแถว PDF (รูปภาพใช้ไฟล์เอง)
        # ป้ายชื่อไฟล์
        if imgs and pdfs:
            label = f"{len(pdfs)} PDF + {len(imgs)} รูป"
        elif imgs:
            label = (os.path.basename(imgs[0]) if len(imgs) == 1 else f"{len(imgs)} รูปภาพ")
        else:
            label = (os.path.basename(pdfs[0]) if len(pdfs) == 1 else f"{len(pdfs)} ไฟล์ PDF")
        self.lbl_file.setText("กำลังอ่านสลิป…")
        use_ocr = bool(imgs)
        msg = ("กำลังอ่านสลิปจากรูป (OCR อาจใช้เวลาเล็กน้อย)..."
               if use_ocr else "กำลังอ่านไฟล์สลิป PDF...")
        dlg = LoadingDialog(msg, self)
        dlg.set_indeterminate(msg)
        if use_ocr:
            self._slip_worker = SlipImageWorker(imgs, pdfs[0] if pdfs else None, self)
        else:
            self._slip_worker = SlipPDFWorker(pdfs[0], self)
        self._slip_worker.done.connect(
            lambda recs: (dlg.accept(), self._on_pdf_loaded(label, recs)))
        self._slip_worker.error.connect(
            lambda m: (dlg.reject(), self._on_pdf_error(m)))
        self._slip_worker.start()
        dlg.exec()

    def _on_pdf_loaded(self, label, records):
        self._records = records
        total = sum((r.get("amount") or 0) for r in self._records)
        self.lbl_file.setText(
            f"{label} • {len(self._records)} รายการ • {fmt_amount(total)} บาท")
        self.btn_match.setEnabled(bool(self._records))
        self.btn_savepng.setEnabled(bool(self._records))
        self.status_message.emit(f"อ่านสลิปได้ {len(self._records)} รายการ")
        self._show_records_only()

    def _on_pdf_error(self, msg):
        QMessageBox.critical(self, "อ่านสลิปไม่สำเร็จ", msg)
        self.lbl_file.setText("อ่านไฟล์ไม่สำเร็จ")

    def _show_records_only(self):
        """แสดงรายการสลิปก่อนจับคู่ (คอลัมน์จับคู่ว่างไว้)"""
        self._checked_slips = set()
        self._checked_un = set()
        self._revealed_slip_ids = set()
        self._results = [{"slip": r, "expense": None, "score": None, "status": ""}
                         for r in self._records]
        # ก่อนจับคู่: โชว์ทุกแถวในหน้า "ทั้งหมด" (หน้าเตรียมดำเนินการยังว่าง)
        self.table.setRowCount(0)
        self._shown = []
        self.table_all.setRowCount(0)
        self._shown_all = []
        for i, res in enumerate(self._results, 1):
            self._shown_all.append(res)
            self._add_row(i, res, table=self.table_all)
        # ล้างหน้า "ไม่เจอคู่" จากไฟล์ก่อนหน้า
        if hasattr(self, "table_un"):
            self.table_un.setRowCount(0)
            self._shown_un = []
            self.subtabs.setTabText(2, "🔴 ไม่เจอคู่")
            self.btn_un_export.setEnabled(False)

    def _do_match(self):
        if not self._records:
            return
        # ดึงเอกสารให้ครอบคลุมช่วงวันที่ของสลิป (ไม่จำกัดแค่ N ล่าสุด → จับคู่แม่นขึ้น)
        from datetime import date as _date, timedelta
        dts = [(r.get("pay_date") or "")[:10] for r in self._records if r.get("pay_date")]
        oldest = min(dts) if dts else _date.today().isoformat()
        try:
            since = (_date.fromisoformat(oldest) - timedelta(days=45)).isoformat()
        except Exception:
            since = "2000-01-01"
        dlg = LoadingDialog("กำลังดึงเอกสารให้ครบช่วงสลิป...", self)
        self._mfw = MatchExpenseFetchWorker(since, self)
        self._mfw.done.connect(lambda exps: (dlg.accept(), self._run_match(exps)))
        self._mfw.error.connect(lambda m: (dlg.accept(), self._run_match(None)))
        self._mfw.start()
        dlg.exec()

    def _run_match(self, expenses):
        exps = expenses if expenses else self._expenses
        if not exps:
            QMessageBox.information(self, "ยังไม่มีข้อมูลค่าใช้จ่าย",
                "ไปที่แท็บ 'คิวจ่ายเงิน' กด 🔄 รีเฟรช เพื่อดึงค่าใช้จ่ายก่อน")
            return
        self._match_expenses = exps   # ชุดเอกสารที่ใช้จับคู่ (ครบช่วงสลิป)
        self.status_message.emit(f"กำลังจับคู่กับ {len(exps)} เอกสาร…")
        QApplication.processEvents()
        self._results = match_slip_to_expenses(self._records, exps)
        self._checked_slips = set()
        self._checked_un = set()
        if hasattr(self, "chk_all_slip"):
            self.chk_all_slip.blockSignals(True)
            self.chk_all_slip.setChecked(False)
            self.chk_all_slip.blockSignals(False)
        # ตัดรูปสลิปทุกแถวไว้ใน RAM เพื่อโชว์ในคอลัมน์ "รูป" (ไม่เซฟลงดิสก์)
        self._crop_slip_images()
        self._render_results()

    def _crop_slip_images(self):
        """ตัดรูปสลิปทุกแถว เก็บใน RAM (slip['_png']) — กดดูในโปรแกรมได้
        (รายการจากรูปภาพใช้ไฟล์รูปเองผ่าน _img_path — ไม่ต้องมี PDF)"""
        if not self._records:
            return
        if not self._pdf_path and not any(r.get("_img_path") for r in self._records):
            return
        self.status_message.emit("กำลังตัดรูปสลิปเพื่อแสดง…")
        QApplication.processEvents()
        try:
            from bank_slip import crop_all_rows
            pngs = crop_all_rows(self._pdf_path, self._records)
            for rec, png in zip(self._records, pngs):
                rec["_png"] = png
        except Exception:
            pass   # ไม่มี Pillow/ตัดไม่ได้ → ข้ามรูปไป (ไม่ทำให้จับคู่พัง)

    def _clear_slip_images(self):
        """ล้างรูปสลิปออกจาก RAM (กันเปลืองหน่วยความจำ)"""
        for rec in (self._records or []):
            rec.pop("_png", None)

    def _ask_cut_scope(self, title):
        """ถามว่าจะทำกับ 'เฉพาะที่เจอคู่' หรือ 'ทั้งหมด' → คืน 'matched'/'all'/None"""
        box = QMessageBox(self)
        box.setIcon(QMessageBox.Icon.Question)
        box.setWindowTitle(title)
        box.setText("ต้องการทำกับรายการแบบไหน?")
        box.setInformativeText(
            "• ดำเนินการเฉพาะที่เจอคู่ = ทุกรายการในหน้า 'จับคู่สำเร็จเตรียมดำเนินการต่อ' (ชัวร์)\n"
            "• ดำเนินการทั้งหมด = รวมที่ไม่เจอคู่/น่าจะใช่ด้วย")
        b_match = box.addButton("✅ ดำเนินการเฉพาะที่เจอคู่", QMessageBox.ButtonRole.AcceptRole)
        b_all = box.addButton("📋 ดำเนินการทั้งหมด", QMessageBox.ButtonRole.AcceptRole)
        box.addButton("ยกเลิก", QMessageBox.ButtonRole.RejectRole)
        b_match.setStyleSheet(
            "QPushButton{background:#dc2626;color:white;border:none;border-radius:5px;"
            "padding:6px 16px;font-weight:600;}QPushButton:hover{background:#b91c1c;}")
        b_all.setStyleSheet(
            "QPushButton{background:#2563eb;color:white;border:none;border-radius:5px;"
            "padding:6px 16px;font-weight:600;}QPushButton:hover{background:#1d4ed8;}")
        box.exec()
        clk = box.clickedButton()
        if clk is b_match:
            return "matched"
        if clk is b_all:
            return "all"
        return None

    def _cut_to_folder(self):
        """ตัดบิลเป็นรูป PNG บันทึกลงโฟลเดอร์ (พรีวิว ไม่แตะ FlowAccount)
        ชื่อไฟล์ = เลขเอกสาร (EXP) + ชื่อคน"""
        # ถ้ายังไม่ได้จับคู่ (ไม่มี result ไหนมี expense) → จับคู่ให้อัตโนมัติก่อน
        if self._records and self._expenses and not any(
                r.get("expense") for r in self._results):
            self.status_message.emit("กำลังจับคู่อัตโนมัติก่อนตัดบิล…")
            QApplication.processEvents()
            self._results = match_slip_to_expenses(self._records, self._expenses)
            self._checked_slips = set()
            self._render_results()
        sel = self._selected_results()
        if sel is not None:
            base = sel                       # ติ๊กไว้ → ใช้ที่ติ๊ก
        else:
            scope = self._ask_cut_scope("ตัดบิลเป็นรูป")
            if scope is None:
                return
            base = ([r for r in self._results if r["status"] == "matched"]
                    if scope == "matched" else list(self._results))
        if not base:
            QMessageBox.information(self, "ไม่มีรายการ", "ไม่มีรายการให้ตัด")
            return
        pairs = [(r["slip"], r.get("expense")) for r in base]
        folder = QFileDialog.getExistingDirectory(self, "เลือกโฟลเดอร์บันทึกรูปสลิป")
        if not folder:
            return
        from bank_slip import crop_all_rows
        dlg = LoadingDialog("กำลังตัดบิลเป็นรูป...", self)
        dlg.show(); QApplication.processEvents()
        try:
            pngs = crop_all_rows(self._pdf_path, [p[0] for p in pairs])
        except Exception as e:
            dlg.accept()
            QMessageBox.critical(self, "❌ ผิดพลาด", f"ตัดบิลไม่สำเร็จ:\n{e}")
            return
        import re as _re
        saved = 0
        for i, ((s, e), png) in enumerate(zip(pairs, pngs), 1):
            if not png:
                continue
            doc = _slip_exp_serial(e) if e else ""
            nm = _re.sub(r"[\\/:*?\"<>|]", "", (s.get("recv_name") or "")).strip()[:30]
            # ชื่อไฟล์: เลขเอกสาร_ชื่อคน  (ถ้าไม่เจอคู่ใช้ ไม่พบเอกสาร_ชื่อคน)
            head = doc if doc else "ไม่พบเอกสาร"
            fn = f"{head}_{nm}.png".strip("_")
            try:
                with open(os.path.join(folder, fn), "wb") as f:
                    f.write(png)
                saved += 1
            except Exception:
                pass
            dlg.set_progress(i, len(pngs), f"ตัดบิล {i}/{len(pngs)}")
            QApplication.processEvents()
        dlg.accept()
        activity_log.log("ตัดบิลเป็นรูป", f"{saved} รูป → {folder}")
        QMessageBox.information(self, "✅ เสร็จแล้ว",
            f"บันทึกรูปสลิป {saved} รูปแล้ว\n{folder}")
        QDesktopServices.openUrl(QUrl.fromLocalFile(folder))

    def _cut_and_attach(self):
        """ตัดบิล + แนบเข้า FlowAccount (รายการที่มีคู่ค่าใช้จ่าย และที่ติ๊กเลือก)"""
        sel = self._selected_results()
        if sel is not None:
            # ติ๊กไว้ → ทำทุกตัวที่ติ๊กและมีคู่ (รวม 'น่าจะใช่' + จับคู่เอง)
            matched = [r for r in sel if r.get("expense")]
        else:
            scope = self._ask_cut_scope("ตัดบิล + แนบเข้า FlowAccount")
            if scope is None:
                return
            if scope == "matched":
                matched = [r for r in self._results if r["status"] == "matched"]
            else:
                # ทั้งหมดที่มีคู่ค่าใช้จ่าย (รวม 'น่าจะใช่' ที่ระบบเดา)
                matched = [r for r in self._results if r.get("expense")]
        if not matched:
            QMessageBox.information(self, "ไม่มีรายการ",
                "ไม่มีรายการที่มีคู่ค่าใช้จ่ายให้แนบ\n"
                "(เลือก 'ทั้งหมด' เพื่อรวมที่ระบบเดา หรือจับคู่เองก่อน)")
            return
        if QMessageBox.warning(self, "⚠️ ยืนยันแนบสลิป + เปลี่ยนสถานะจ่ายแล้ว",
                f"จะดำเนินการกับรายการที่จับคู่ได้ {len(matched)} รายการ:\n\n"
                "1) ตัดสลิปแนบเข้ารายการจริงใน FlowAccount\n"
                "2) เปลี่ยนสถานะเป็น 'จ่ายแล้ว' ทั้งในโปรแกรมและ FlowAccount\n\n"
                "⚠️ เปลี่ยนข้อมูลใน FlowAccount จริง — ยืนยันดำเนินการ?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel,
                QMessageBox.StandardButton.Cancel) != QMessageBox.StandardButton.Yes:
            return
        # ถามว่าจะบันทึกรูปที่ตัดไว้ด้วยไหม + เลือกโฟลเดอร์
        save_dir = None
        if QMessageBox.question(self, "บันทึกรูปสลิปที่ตัด?",
                "ต้องการบันทึกรูปสลิปที่ตัด ลงโฟลเดอร์ด้วยไหม?\n"
                "(นอกจากแนบเข้า FlowAccount แล้ว)",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
            save_dir = QFileDialog.getExistingDirectory(self, "เลือกโฟลเดอร์บันทึกรูปสลิป")
            if not save_dir:
                save_dir = None
        dlg = LoadingDialog("กำลังตัดบิล + แนบ + เปลี่ยนสถานะ...", self)
        self._attach_worker = SlipAttachWorker(self._pdf_path, matched, self, save_dir=save_dir)
        self._attach_worker.progress.connect(
            lambda d, t: dlg.set_progress(d, t, f"แนบสลิป {d}/{t}"))
        self._attach_worker.finished_all.connect(
            lambda ok, fail, errs: (dlg.accept(), self._on_attach_done(ok, fail, errs)))
        self._attach_worker.start()
        dlg.exec()

    def _on_attach_done(self, ok, fail, problems):
        activity_log.log("ตัดบิล+แนบ+จ่ายแล้ว", f"สำเร็จ {ok} มีปัญหา {fail}")
        # แนบเข้า FlowAccount แล้ว → เคลียร์รูปออกจาก RAM (ไม่เปลืองหน่วยความจำ)
        self._clear_slip_images()
        # รีเฟรชหน้าจับคู่ + แจ้งหน้าคิวให้รู้ว่าสถานะเปลี่ยน
        self._render_results()
        if not problems:
            QMessageBox.information(self, "✅ เสร็จสมบูรณ์",
                f"แนบสลิป + เปลี่ยนสถานะเป็นจ่ายแล้ว สำเร็จครบ {ok} รายการ\n\n"
                "(ไปหน้าคิวจ่ายเงิน กด 🔄 รีเฟรช เพื่อดูสถานะล่าสุดจาก FlowAccount)")
            return
        # มีรายการทำไม่ได้ → โชว์รายละเอียดเป็นรายตัว
        lines = []
        for p in problems:
            parts = []
            if p.get("attach"):
                parts.append(f"แนบไฟล์ไม่ได้: {p['attach']}")
            if p.get("status"):
                parts.append(f"เปลี่ยนสถานะไม่ได้: {p['status']}")
            lines.append(f"• {p['doc']} — " + " | ".join(parts))
        dlg = QMessageBox(self)
        dlg.setIcon(QMessageBox.Icon.Warning)
        dlg.setWindowTitle("มีรายการที่ทำไม่สำเร็จ")
        dlg.setText(f"สำเร็จ {ok} รายการ • มีปัญหา {fail} รายการ\n\nรายการที่ทำไม่ได้:")
        dlg.setDetailedText("\n".join(lines))
        # โชว์ส่วนต้น ๆ ในกล่องเลย (ไม่ต้องกดดูรายละเอียด)
        preview = "\n".join(lines[:8]) + ("\n..." if len(lines) > 8 else "")
        dlg.setInformativeText(preview)
        dlg.exec()

    def _render_results(self):
        self.table.setRowCount(0)
        self._shown = []   # ลำดับ result ตามที่แสดง (ใช้ตอนดับเบิลคลิกจับคู่เอง)
        # นับสถานะจากผลทั้งหมด (ไม่ขึ้นกับการเรียง/กรอง)
        n_match = sum(1 for r in self._results if r["status"] == "matched")
        n_maybe = sum(1 for r in self._results if r["status"] == "amount_only")
        n_un    = sum(1 for r in self._results if r["status"] == "unmatched")

        # หน้าหลัก = เฉพาะ "เจอคู่" (เขียว) เท่านั้น
        # ไม่เจอ/น่าจะใช่ → ไปอยู่แท็บ "🔴 ไม่เจอคู่" อย่างเดียว
        matched = [r for r in self._results if r["status"] == "matched"]
        for res in matched:
            self._shown.append(res)
            self._add_row(len(self._shown), res)

        self.lbl_summary.setText(
            f"✅ เจอคู่ {n_match} (แสดงในหน้านี้)   •   🔴 ต้องตรวจ/ไม่เจอคู่ {n_maybe + n_un} "
            f"(น่าจะใช่ {n_maybe} + ไม่เจอ {n_un})   (ทั้งหมด {len(self._results)})  "
            f"— ไปแท็บ '🔴 ไม่เจอคู่' เพื่อตรวจ/จับคู่เอง")
        self.status_message.emit(
            f"จับคู่เสร็จ: เจอ {n_match}, ต้องตรวจ/ไม่เจอ {n_maybe + n_un}")
        self.btn_savepng.setEnabled(bool(self._results))
        self.btn_attach.setEnabled((n_match + n_maybe) > 0)
        self.btn_export.setEnabled(bool(self._results))
        if hasattr(self, "btn_demote"):
            self.btn_demote.setEnabled(n_match > 0)
        self._render_unmatched(n_maybe + n_un)
        self._render_all()

    def _render_all(self):
        """เติมหน้า 'ทั้งหมด' — รวมทุกรายการ (เจอคู่ + น่าจะใช่ + ไม่เจอ) ตามลำดับสลิป"""
        self.table_all.setRowCount(0)
        self._shown_all = []
        for res in self._results:
            self._shown_all.append(res)
            self._add_row(len(self._shown_all), res, table=self.table_all)
        self.subtabs.setTabText(0, f"📋 ทั้งหมด ({len(self._results)})")

    def _render_unmatched(self, n_un=None):
        """เติมหน้า 'ไม่เจอคู่' — รวมทั้ง 'ไม่เจอ'(แดง) และ 'น่าจะใช่'(เหลือง)
        เพราะเหลืองยังไม่ชัวร์ ต้องให้บัญชีตรวจ/ยืนยัน + อัปเดตชื่อแท็บ"""
        # เรียงให้ 'ไม่เจอ' จริงขึ้นก่อน แล้วตามด้วย 'น่าจะใช่'
        pend = [r for r in self._results if r["status"] == "unmatched"] + \
               [r for r in self._results if r["status"] == "amount_only"]
        if n_un is None:
            n_un = len(pend)
        self.table_un.setRowCount(0)
        self._shown_un = []
        for res in pend:
            s = res["slip"]
            e = res.get("expense")
            self._shown_un.append(res)
            row = self.table_un.rowCount()
            self.table_un.insertRow(row)
            if res["status"] == "amount_only":
                bg = C_DUE_SOON
                st_txt = "🔴 ไม่เจอคู่ (มีตัวเดา)"
            else:
                bg = C_OVERDUE
                st_txt = "🔴 ไม่เจอคู่"
            # col 0: checkbox เลือก (สำหรับ "นำกลับ" เข้าหน้าเตรียมดำเนินการ)
            chk = QCheckBox()
            chk.setChecked(id(res) in self._checked_un)
            chk.stateChanged.connect(lambda st, r=res: self._on_un_check(r, st))
            cw = QWidget(); cl = QHBoxLayout(cw)
            cl.addWidget(chk); cl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            cl.setContentsMargins(0, 0, 0, 0)
            self.table_un.setCellWidget(row, 0, cw)
            vals = [
                str(len(self._shown_un)),
                fmt_date(s["pay_date"]) if s.get("pay_date") else "-",
                s.get("recv_name", ""),
                fmt_amount(s["amount"]),
                (s.get("detail") or "").strip() or "—",
                _slip_exp_serial(e) if e else "—",
                (e.get("contactName") or "") if e else "",
                st_txt,
            ]
            _sns_un = _sensitive_names()
            _is_sens_un = _sensitive_is_vendor(s.get("recv_name", ""), _sns_un)
            for col, v in enumerate(vals, start=1):   # ขยับไป col 1 (col 0 = ✓)
                if col == 4 and _is_sens_un:
                    continue   # จัด cell widget แยกด้านล่าง
                item = QTableWidgetItem(v)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if col == 4:   # ยอด ชิดขวา
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                item.setBackground(QBrush(bg))
                self.table_un.setItem(row, col, item)
            if _is_sens_un:
                is_rev_un = id(res) in self._revealed_slip_ids
                eye_un = SensitiveEyeWidget(
                    amount=s["amount"],
                    is_revealed=is_rev_un,
                    on_unlock=lambda r=res: self._revealed_slip_ids.add(id(r)),
                    on_relock=lambda r=res: self._revealed_slip_ids.discard(id(r)),
                )
                eye_un.setAutoFillBackground(True)
                p_un = eye_un.palette(); p_un.setColor(eye_un.backgroundRole(), bg); eye_un.setPalette(p_un)
                self.table_un.setCellWidget(row, 4, eye_un)
            # ปุ่มดูใบ (เฉพาะที่ระบบเดาได้) — col 9
            if e:
                btn = QPushButton("🔗 ดูใบ")
                btn.setCursor(Qt.CursorShape.PointingHandCursor)
                btn.setStyleSheet(
                    "QPushButton{padding:2px 6px;border:1px solid #2563eb;border-radius:4px;"
                    "background:white;color:#2563eb;font-size:11px;}"
                    "QPushButton:hover{background:#dbeafe;}")
                btn.clicked.connect(lambda _=False, ex=e: self._open_doc(ex))
                bw = QWidget(); bl = QHBoxLayout(bw)
                bl.addWidget(btn); bl.setAlignment(Qt.AlignmentFlag.AlignCenter)
                bl.setContentsMargins(0, 0, 0, 0)
                self.table_un.setCellWidget(row, 9, bw)
            # col 10: รูปสลิป
            self._add_image_cell(self.table_un, row, 10, s)
        # อัปเดตชื่อแท็บ (แท็บไม่เจอคู่ = index 2)
        if hasattr(self, "subtabs"):
            self.subtabs.setTabText(2, f"🔴 ไม่เจอคู่ ({n_un})")
        self.btn_un_export.setEnabled(n_un > 0)
        self.btn_promote.setEnabled(n_un > 0)

    def _export_unmatched(self):
        """บันทึกรายการ 'ไม่เจอคู่' (รวมน่าจะใช่) เป็น Excel ให้บัญชีจัดการต่อ"""
        rows = [r for r in self._results if r["status"] == "unmatched"] + \
               [r for r in self._results if r["status"] == "amount_only"]
        if not rows:
            QMessageBox.information(self, "ไม่มีรายการ", "ไม่มีรายการที่ไม่เจอคู่")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "บันทึกรายการไม่เจอคู่", "รายการไม่เจอคู่.xlsx",
            "Excel (*.xlsx)")
        if not path:
            return
        try:
            from openpyxl import Workbook
            wb = Workbook(); ws = wb.active; ws.title = "ไม่เจอคู่"
            ws.append(["#", "วันที่จ่าย", "ชื่อผู้รับ (สลิป)", "ยอด (บาท)",
                       "เลขบัญชีผู้รับ", "รายละเอียดในสลิป",
                       "EXP ที่ระบบเดา", "ชื่อในระบบ", "สถานะ"])
            for i, r in enumerate(rows, 1):
                s = r["slip"]; e = r.get("expense")
                ws.append([i,
                           fmt_date(s["pay_date"]) if s.get("pay_date") else "",
                           s.get("recv_name", ""),
                           s.get("amount", 0),
                           s.get("recv_acct", ""),
                           (s.get("detail") or "").strip(),
                           _slip_exp_serial(e) if e else "",
                           (e.get("contactName") or "") if e else "",
                           "น่าจะใช่ (ต้องตรวจ)" if r["status"] == "amount_only"
                           else "ไม่เจอคู่"])
            wb.save(path)
        except Exception as e:
            QMessageBox.critical(self, "บันทึกไม่สำเร็จ", str(e))
            return
        activity_log.log("Export ไม่เจอคู่", f"{len(rows)} รายการ → {path}")
        QMessageBox.information(self, "✅ บันทึกแล้ว",
            f"บันทึกรายการไม่เจอคู่ {len(rows)} รายการแล้ว")
        QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.dirname(path)))

    def _match_step2_by_amount(self):
        """จับคู่รอบ 2: สำหรับรายการไม่เจอคู่ → ค้นเอกสารที่ 'ยอดตรง' แล้วให้ยืนยันทีละรายการ"""
        pending = [r for r in self._results if r["status"] in ("unmatched", "amount_only")]
        if not pending:
            QMessageBox.information(self, "ไม่มีรายการ", "ไม่มีรายการที่ต้องจับคู่")
            return
        pool = getattr(self, "_match_expenses", None) or self._expenses
        if not pool:
            QMessageBox.information(self, "ยังไม่มีข้อมูล",
                "ไปแท็บคิวจ่ายเงิน กดรีเฟรชเพื่อดึงค่าใช้จ่ายก่อน")
            return
        used = {_exp_id(r["expense"]) for r in self._results
                if r["status"] == "matched" and r.get("expense")}
        matched_n = 0
        for r in pending:
            s = r["slip"]; amt = s.get("amount", 0)
            # ค้นเอกสารยอดตรง (ยังไม่ถูกใช้)
            cands = [e for e in pool
                     if abs(_amount(e) - amt) <= 0.01 and _exp_id(e) not in used]
            if not cands:
                continue
            e = cands[0]
            lines = "\n".join(
                f"   • {_slip_exp_serial(c)} | {_vendor(c)} | {fmt_amount(_amount(c))}"
                for c in cands[:5])
            box = QMessageBox(self)
            box.setIcon(QMessageBox.Icon.Question)
            box.setWindowTitle("จับคู่รอบ 2 — เจอยอดตรง")
            box.setText(f"สลิป: {s.get('recv_name','')}\nยอด: {fmt_amount(amt)}\n\n"
                        f"เจอเอกสารยอดตรง {len(cands)} ใบ:\n{lines}\n\n"
                        f"จับคู่กับใบแรก ({_slip_exp_serial(e)}) ?")
            b_ok = box.addButton("✅ ตกลง (จับคู่)", QMessageBox.ButtonRole.AcceptRole)
            b_skip = box.addButton("ข้าม", QMessageBox.ButtonRole.RejectRole)
            b_stop = box.addButton("❌ ยกเลิกทำรายการ", QMessageBox.ButtonRole.DestructiveRole)
            box.exec()
            clk = box.clickedButton()
            if clk is b_stop:
                break
            if clk is b_ok:
                r["expense"] = e; r["status"] = "matched"; r["score"] = 1.0
                r["via"] = "amount_detail"; r["match_note"] = "จับคู่รอบ 2 (ยอดตรง)"
                used.add(_exp_id(e))
                self._checked_slips.add(id(r))
                matched_n += 1
        self._render_results()
        QMessageBox.information(self, "เสร็จแล้ว",
            f"จับคู่รอบ 2 สำเร็จ {matched_n} รายการ (ย้ายไปหน้า 'จับคู่สำเร็จเตรียมดำเนินการต่อ')")

    def _export_matched(self):
        """Export ผลจับคู่ทั้งหมด เป็น Excel + PDF"""
        if not self._results:
            QMessageBox.information(self, "ไม่มีข้อมูล", "ยังไม่ได้จับคู่")
            return
        base, _ = QFileDialog.getSaveFileName(
            self, "บันทึกผลจับคู่ (Excel + PDF)",
            f"ผลจับคู่สลิป_{date.today().strftime('%Y%m%d')}", "Excel (*.xlsx)")
        if not base:
            return
        for ext in (".xlsx", ".pdf"):
            if base.lower().endswith(ext):
                base = base[:-len(ext)]
        brands = load_brands()
        headers = ["#", "วันที่จ่าย", "ชื่อผู้รับ (สลิป)", "ยอด (บาท)", "จับคู่กับ EXP",
                   "ชื่อในระบบ", "แบรนด์", "สถานะ"]

        def _brand_of(e):
            return (e.get("_brand_name") or _brand_name(e, brands) or "") if e else ""

        def _stat(r):
            return {"matched": "เจอคู่", "amount_only": "ไม่เจอคู่ (มีตัวเดา)"}.get(
                r["status"], "ไม่เจอคู่")
        rows = []
        for i, r in enumerate(self._results, 1):
            s = r["slip"]; e = r.get("expense")
            rows.append([i, fmt_date(s["pay_date"]) if s.get("pay_date") else "",
                         s.get("recv_name", ""), s.get("amount", 0),
                         _slip_exp_serial(e) if e else "",
                         (e.get("contactName") or "") if e else "",
                         _brand_of(e), _stat(r)])
        # Excel
        try:
            from openpyxl import Workbook
            wb = Workbook(); ws = wb.active; ws.title = "ผลจับคู่สลิป"
            ws.append(headers)
            for rw in rows:
                ws.append(rw)
            wb.save(base + ".xlsx")
        except Exception as e:
            QMessageBox.critical(self, "บันทึก Excel ไม่สำเร็จ", str(e)); return
        # PDF (ตาราง)
        try:
            from utils import build_simple_table_pdf
            build_simple_table_pdf(base + ".pdf", "ผลจับคู่สลิปกับ FlowAccount",
                                   headers, rows)
        except Exception:
            pass   # ไม่มี builder/ทำ PDF ไม่ได้ → ได้ Excel อย่างเดียวก็พอ
        activity_log.log("Export ผลจับคู่สลิป", f"{len(rows)} รายการ")
        QMessageBox.information(self, "✅ บันทึกแล้ว",
            f"บันทึกผลจับคู่ {len(rows)} รายการแล้ว\n📊 {base}.xlsx\n📄 {base}.pdf")
        QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.dirname(base)))

    def _add_row(self, idx, res, table=None):
        table = table if table is not None else self.table
        slip, expense = res["slip"], res["expense"]
        score, status = res["score"], res["status"]
        bg = {"matched": C_MATCHED, "amount_only": C_DUE_SOON}.get(status, C_OVERDUE)

        exp_serial = _slip_exp_serial(expense) if expense else ""
        exp_name   = (expense.get("contactName") or "") if expense else ""
        # 'น่าจะใช่'(ยอดตรงแต่ไม่ชัวร์) ถือเป็น 'ไม่เจอคู่' (ตามที่พี่เบิร์ดสั่ง)
        status_txt = {"matched": "✅ เจอคู่",
                      "amount_only": "🔴 ไม่เจอคู่ (มีตัวเดา)",
                      "unmatched": "🔴 ไม่เจอคู่"}.get(status, "")
        if res.get("via") == "account":
            status_txt = "✅ เจอคู่ (จำเลขบัญชี)"
        elif res.get("via") == "reference":
            status_txt = "✅ เจอคู่ (เลขอ้างอิง)"
        elif res.get("via") == "detail":
            status_txt = "✅ เจอคู่ (รายละเอียด)"
        elif res.get("via") == "amount_detail":
            status_txt = "✅ เจอคู่ (ยอดย่อย/หัก)"
        # รายการที่ผู้ใช้ 'นำกลับ' จากหน้าไม่เจอคู่ → ทำเป็นสีแดง + ป้ายเตือนให้ตรวจ
        if res.get("promoted"):
            bg = C_OVERDUE
            status_txt = "↩️ นำกลับเข้าคิว (ตรวจ)"
        brand = ""
        if expense:
            brand = expense.get("_brand_name") or _brand_name(expense, load_brands()) or ""
        vals = [
            str(idx),
            fmt_date(slip["pay_date"]) if slip.get("pay_date") else "-",
            slip["recv_name"],
            fmt_amount(slip["amount"]),
            exp_serial,
            exp_name,
            brand,
            (f"{score:.0%}" if score not in (None, "") else ""),
            status_txt,
        ]
        row = table.rowCount()
        table.insertRow(row)

        # col 0: checkbox เลือกรายการ
        chk = QCheckBox()
        chk.setChecked(id(res) in self._checked_slips)
        chk.stateChanged.connect(lambda st, r=res: self._on_slip_check(r, st))
        cw = QWidget(); cl = QHBoxLayout(cw)
        cl.addWidget(chk); cl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        cl.setContentsMargins(0, 0, 0, 0)
        table.setCellWidget(row, 0, cw)

        _sns = _sensitive_names()
        _is_sensitive = _sensitive_is_vendor(slip.get("recv_name", ""), _sns)
        for col, v in enumerate(vals, start=1):   # ขยับไป col 1 เป็นต้นไป
            if col == 4 and _is_sensitive:
                continue   # จัด cell widget แยกด้านล่าง
            item = QTableWidgetItem(v)
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            if col in (4, 8):   # ยอด(4) / คะแนน(8) ชิดขวา
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            if bg:
                item.setBackground(QBrush(bg))
            # บอกเหตุผลการจับ (หัก ณ ที่จ่าย/ยอดย่อย) เมื่อชี้ที่ช่องสถานะ (col 9)
            if col == 9 and res.get("match_note"):
                item.setToolTip(res["match_note"])
            table.setItem(row, col, item)
        if _is_sensitive:
            is_rev = id(res) in self._revealed_slip_ids
            eye = SensitiveEyeWidget(
                amount=slip["amount"],
                is_revealed=is_rev,
                on_unlock=lambda r=res: self._revealed_slip_ids.add(id(r)),
                on_relock=lambda r=res: self._revealed_slip_ids.discard(id(r)),
            )
            if bg:
                eye.setAutoFillBackground(True)
                p = eye.palette(); p.setColor(eye.backgroundRole(), bg); eye.setPalette(p)
            table.setCellWidget(row, 4, eye)

        # col 10: ปุ่ม "ดูใบ" — เปิดลิงก์แชร์เอกสารใน FlowAccount เพื่อเช็คว่าถูกไหม
        if expense:
            btn = QPushButton("🔗 ดูใบ")
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setStyleSheet(
                "QPushButton{padding:2px 6px;border:1px solid #2563eb;border-radius:4px;"
                "background:white;color:#2563eb;font-size:11px;}"
                "QPushButton:hover{background:#dbeafe;}")
            btn.clicked.connect(lambda _=False, e=expense: self._open_doc(e))
            bw = QWidget(); bl = QHBoxLayout(bw)
            bl.addWidget(btn); bl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            bl.setContentsMargins(0, 0, 0, 0)
            table.setCellWidget(row, 10, bw)

        # col 11: รูปสลิปที่ตัดแล้ว (เก็บใน RAM) — กดดูในโปรแกรม
        self._add_image_cell(table, row, 11, slip)

    def _add_image_cell(self, table, row, col, slip):
        """ใส่ปุ่มดูรูปสลิป (ถ้ามีรูปที่ตัดไว้ใน RAM)"""
        png = slip.get("_png") if isinstance(slip, dict) else None
        if not png:
            return
        btn = QPushButton("🖼 ดูรูป")
        btn.setCursor(Qt.CursorShape.PointingHandCursor)
        btn.setStyleSheet(
            "QPushButton{padding:2px 6px;border:1px solid #0d9488;border-radius:4px;"
            "background:white;color:#0d9488;font-size:11px;}"
            "QPushButton:hover{background:#ccfbf1;}")
        btn.clicked.connect(lambda _=False, p=png, s=slip: self._show_slip_image(p, s))
        w = QWidget(); l = QHBoxLayout(w)
        l.addWidget(btn); l.setAlignment(Qt.AlignmentFlag.AlignCenter)
        l.setContentsMargins(0, 0, 0, 0)
        table.setCellWidget(row, col, w)

    def _show_slip_image(self, png, slip=None):
        """แสดงรูปสลิปในโปรแกรม"""
        if not png:
            return
        from PyQt6.QtGui import QPixmap
        from PyQt6.QtWidgets import QScrollArea
        dlg = QDialog(self)
        dlg.setWindowTitle(f"รูปสลิป — {slip.get('recv_name','') if slip else ''}")
        dlg.resize(780, 620)
        v = QVBoxLayout(dlg)
        pix = QPixmap()
        pix.loadFromData(png)
        lbl = QLabel()
        lbl.setPixmap(pix.scaledToWidth(740, Qt.TransformationMode.SmoothTransformation))
        scroll = QScrollArea(); scroll.setWidget(lbl); scroll.setWidgetResizable(True)
        v.addWidget(scroll, 1)
        btn = QPushButton("ปิด"); btn.clicked.connect(dlg.accept)
        v.addWidget(btn)
        dlg.exec()

    def _open_doc(self, expense):
        """เปิดเอกสารใน FlowAccount (ลิงก์แชร์) เพื่อตรวจว่าจับคู่ถูกไหม"""
        if not expense:
            return
        eid = _exp_id(expense)
        url = share_links.get(eid)
        if not url:
            self.status_message.emit("กำลังดึงลิงก์แชร์เอกสาร…")
            QApplication.processEvents()
            try:
                did = expense.get("documentId") or expense.get("recordId")
                cid, csec = _exp_creds(expense)
                url = get_share_link(cid, csec, did,
                                     doctype=expense.get("_doctype", "expense"))
                if url:
                    share_links.set(eid, url)
            except Exception as ex:
                QMessageBox.warning(self, "เปิดใบไม่ได้",
                                    f"ดึงลิงก์แชร์ไม่สำเร็จ:\n{ex}")
                return
        if url:
            QDesktopServices.openUrl(QUrl(url))
        else:
            QMessageBox.information(self, "ไม่มีลิงก์แชร์",
                "ดึงลิงก์แชร์ของเอกสารนี้ไม่สำเร็จ")

    def _on_slip_check(self, res, state):
        if state == Qt.CheckState.Checked.value:
            self._checked_slips.add(id(res))
        else:
            self._checked_slips.discard(id(res))

    def _on_un_check(self, res, state):
        """ติ๊ก/เลิกติ๊ก รายการในหน้า 'ไม่เจอคู่' (สำหรับนำกลับ)"""
        if state == Qt.CheckState.Checked.value:
            self._checked_un.add(id(res))
        else:
            self._checked_un.discard(id(res))

    def _promote_unmatched(self):
        """นำรายการที่ติ๊กในหน้า 'ไม่เจอคู่' กลับเข้าหน้า 'จับคู่สำเร็จเตรียมดำเนินการต่อ'"""
        sel = [r for r in self._results
               if r["status"] in ("unmatched", "amount_only")
               and id(r) in self._checked_un]
        if not sel:
            QMessageBox.information(self, "ยังไม่ได้เลือก",
                "ติ๊กเลือกรายการที่ต้องการนำกลับเข้าหน้า "
                "'จับคู่สำเร็จเตรียมดำเนินการต่อ' ก่อน")
            return
        no_exp = [r for r in sel if not r.get("expense")]
        if no_exp and QMessageBox.question(self, "บางรายการยังไม่มีคู่เอกสาร",
                f"มี {len(no_exp)} รายการที่ยังไม่ได้จับคู่กับเอกสาร (EXP)\n"
                "ถ้านำกลับ จะยังตัดบิลแนบเข้า FlowAccount ไม่ได้ "
                "จนกว่าจะดับเบิลคลิกจับคู่เอกสารเอง\n\nยืนยันนำกลับทั้งหมด?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel,
                QMessageBox.StandardButton.Cancel) != QMessageBox.StandardButton.Yes:
            return
        for r in sel:
            r["_orig_status"] = r["status"]   # จำสถานะเดิมไว้ เผื่อกด 'ส่งกลับ'
            r["status"] = "matched"
            r["promoted"] = True              # ทำเป็นแถวสีแดง (นำกลับมาเอง)
            r.pop("via", None)
            r["match_note"] = "นำกลับเข้าหน้าเตรียมดำเนินการเอง"
            self._checked_slips.add(id(r))
            self._checked_un.discard(id(r))
        self._render_results()
        self.status_message.emit(
            f"นำกลับ {len(sel)} รายการเข้าหน้า 'จับคู่สำเร็จเตรียมดำเนินการต่อ'")
        self.subtabs.setCurrentIndex(1)

    def _demote_to_unmatched(self):
        """ส่งรายการที่ติ๊กในหน้า 'จับคู่สำเร็จเตรียมดำเนินการต่อ' กลับไปหน้า 'ไม่เจอคู่'"""
        sel = [r for r in self._results
               if r["status"] == "matched" and id(r) in self._checked_slips]
        if not sel:
            QMessageBox.information(self, "ยังไม่ได้เลือก",
                "ติ๊กเลือกรายการในหน้านี้ที่ต้องการส่งกลับไป 'ไม่เจอคู่' ก่อน")
            return
        for r in sel:
            # คืนสถานะเดิมถ้าเคยถูกนำกลับ ไม่งั้นเดาจากว่ามีคู่เอกสารไหม
            orig = r.pop("_orig_status", None)
            if orig in ("unmatched", "amount_only"):
                r["status"] = orig
            else:
                r["status"] = "amount_only" if r.get("expense") else "unmatched"
            r.pop("promoted", None)
            r.pop("match_note", None)
            self._checked_slips.discard(id(r))
        self._render_results()
        self.status_message.emit(f"ส่งกลับ {len(sel)} รายการไปหน้า 'ไม่เจอคู่'")
        self.subtabs.setCurrentIndex(2)

    def _toggle_all_slips(self, state):
        checked = state == Qt.CheckState.Checked.value
        # ติ๊กเฉพาะที่แสดงในหน้าหลัก (เจอคู่) เท่านั้น
        self._checked_slips = ({id(r) for r in self._results if r["status"] == "matched"}
                               if checked else set())
        self._render_results()

    def _selected_results(self):
        """รายการที่ติ๊กไว้ (ถ้าไม่ติ๊ก = None = ทำทั้งหมด)"""
        if not self._checked_slips:
            return None
        return [r for r in self._results if id(r) in self._checked_slips]

    def _manual_match(self, row, col):
        """ดับเบิลคลิกแถวหน้า 'ผลทั้งหมด' → จับคู่ค่าใช้จ่ายเอง"""
        if 0 <= row < len(self._shown):
            self._do_manual_match(self._shown[row])

    def _manual_match_un(self, row, col):
        """ดับเบิลคลิกแถวหน้า 'ไม่เจอคู่' → จับคู่ค่าใช้จ่ายเอง"""
        if 0 <= row < len(self._shown_un):
            self._do_manual_match(self._shown_un[row])

    def _manual_match_all(self, row, col):
        """ดับเบิลคลิกแถวหน้า 'ทั้งหมด' → จับคู่ค่าใช้จ่ายเอง"""
        if 0 <= row < len(getattr(self, "_shown_all", [])):
            self._do_manual_match(self._shown_all[row])

    def _do_manual_match(self, res):
        pool = getattr(self, "_match_expenses", None) or self._expenses
        if not pool:
            QMessageBox.information(self, "ยังไม่มีข้อมูล",
                "ไปแท็บคิวจ่ายเงิน กดรีเฟรชเพื่อดึงค่าใช้จ่ายก่อน")
            return
        dlg = SlipManualMatchDialog(res["slip"], pool, self)
        if dlg.exec() == QDialog.DialogCode.Accepted and dlg.picked is not None:
            res["expense"] = dlg.picked
            res["status"] = "matched"
            res["score"] = 1.0   # จับคู่เอง
            res.pop("via", None)
            # สอนระบบจำ: เลขบัญชีนี้ = ผู้จำหน่ายคนนี้ (บัญชียืนยันเอง = แม่นสุด)
            try:
                import account_memory as _am
                _am.learn(res["slip"].get("recv_acct"),
                          dlg.picked.get("contactName") or "")
                _am.save()
            except Exception:
                pass
            # ติ๊กให้อัตโนมัติ จะได้ตัด/แนบได้เลย
            self._checked_slips.add(id(res))
            self._render_results()
            self.status_message.emit(
                f"จับคู่เอง: {res['slip'].get('recv_name','')} → {_slip_exp_serial(dlg.picked)}")


# ──────────────────── Goods Receipt Tab (ใบรับสินค้า) ────────────────────

class GoodsReceiptFetchWorker(QThread):
    done  = pyqtSignal(list)
    error = pyqtSignal(str)
    def __init__(self, cfg, company: dict = None):
        super().__init__()
        self.cfg = cfg
        self.company = company or {}
    def run(self):
        try:
            items = get_all_purchases(self.cfg["flowaccount_api_key"],
                                      self.cfg["flowaccount_secret_key"])
            if self.company:
                _tag_brand(items, self.company)
            self.done.emit(items)
        except FlowAccountError as e:
            self.error.emit(str(e))
        except Exception as e:
            self.error.emit(f"เกิดข้อผิดพลาด: {e}")


class GoodsReceiptAllWorker(QThread):
    """ดึงใบรับสินค้าจากทุกบริษัทมารวมกัน + ติดป้ายแบรนด์"""
    done     = pyqtSignal(list)
    error    = pyqtSignal(str)
    progress = pyqtSignal(str)
    def __init__(self, companies):
        super().__init__()
        self.companies = companies
    def run(self):
        all_items, fails = [], []
        for c in self.companies:
            cid, csec = c.get("client_id"), c.get("client_secret")
            if not cid:
                continue
            self.progress.emit(f"กำลังดึงใบรับสินค้า {c.get('label','')}...")
            try:
                items = get_all_purchases(cid, csec)
                _tag_brand(items, c)
                all_items.extend(items)
            except Exception as e:
                fails.append(f"{c.get('label','')}: {e}")
        if fails and not all_items:
            self.error.emit("ดึงข้อมูลไม่สำเร็จ:\n" + "\n".join(fails))
        else:
            self.done.emit(all_items)


class GoodsReceiptTab(QWidget):
    """ใบรับสินค้า (Goods Receipt / RI) — ดึงจาก FlowAccount /purchases ตามบริษัทที่เลือก"""
    status_message = pyqtSignal(str)

    COLS = ["เลขที่เอกสาร", "ผู้จำหน่าย / Vendor", "แบรนด์", "วันที่", "ครบกำหนด", "ยอดรวม (บาท)", "สถานะ"]

    STATUS_TH = {
        "awaiting": "รออนุมัติ", "awaiting_approval": "รออนุมัติ", "pending": "รออนุมัติ",
        "approved": "อนุมัติแล้ว", "paid": "ชำระเงินแล้ว", "partial_paid": "จ่ายบางส่วน",
        "rejected": "ไม่อนุมัติ", "not_approved": "ไม่อนุมัติ", "void": "ยกเลิก",
        "overdue": "เกินกำหนด", "draft": "ฉบับร่าง", "open": "เปิดอยู่",
    }

    def __init__(self):
        super().__init__()
        self._items: list   = []
        self._filtered: list = []
        self._page     = 1
        self._per_page = 25
        self._worker = None
        self._build_ui()

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(6)

        tb = QHBoxLayout()
        self.btn_refresh = QPushButton("🔄 รีเฟรช")
        self.btn_refresh.setStyleSheet(
            "QPushButton{padding:6px 14px;border:none;border-radius:5px;"
            f"background:{C_PRIMARY};color:white;font-size:12px;font-weight:600;}}")
        self.btn_refresh.clicked.connect(self.fetch_items)
        tb.addWidget(self.btn_refresh)

        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("🔍 ค้นหา เลขที่เอกสาร / ผู้จำหน่าย")
        self.search_box.setClearButtonEnabled(True)
        self.search_box.setFixedWidth(260)
        self.search_box.textChanged.connect(self._apply_filter)
        tb.addWidget(self.search_box)

        tb.addStretch()
        self.lbl_fetched = QLabel("ยังไม่ได้ดึงข้อมูล")
        self.lbl_fetched.setStyleSheet(f"color:{C_MUTED};font-size:11px;")
        tb.addWidget(self.lbl_fetched)
        layout.addLayout(tb)

        self.table = QTableWidget(0, len(self.COLS))
        self.table.setHorizontalHeaderLabels(self.COLS)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.table.setColumnWidth(0, 130)
        self.table.setColumnWidth(2, 110)   # แบรนด์
        self.table.setColumnWidth(3, 110)
        self.table.setColumnWidth(4, 110)
        self.table.setColumnWidth(5, 130)
        self.table.setColumnWidth(6, 130)
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSortingEnabled(True)
        self.table.setStyleSheet("QTableWidget{border:1px solid #e2e8f0;font-size:12px;}")
        layout.addWidget(self.table)

        # pagination
        pg = QHBoxLayout()
        self.lbl_summary = QLabel("")
        self.lbl_summary.setStyleSheet(f"color:{C_MUTED};font-size:11px;")
        pg.addWidget(self.lbl_summary)
        pg.addStretch()
        self.btn_prev = QPushButton("◀ ก่อนหน้า")
        self.btn_next = QPushButton("ถัดไป ▶")
        self.lbl_page = QLabel("หน้า 1/1")
        self.lbl_page.setStyleSheet(f"color:{C_PRIMARY};font-weight:600;font-size:12px;padding:0 8px;")
        for b in (self.btn_prev, self.btn_next):
            b.setStyleSheet("QPushButton{padding:4px 10px;border:1px solid #cbd5e1;border-radius:4px;background:white;}")
        self.btn_prev.clicked.connect(self._prev_page)
        self.btn_next.clicked.connect(self._next_page)
        pg.addWidget(self.btn_prev)
        pg.addWidget(self.lbl_page)
        pg.addWidget(self.btn_next)
        layout.addLayout(pg)

    def get_items(self): return self._items

    def clear_items(self):
        self._items = []
        self.table.setRowCount(0)
        self.lbl_fetched.setText("ยังไม่ได้ดึงข้อมูล")

    def fetch_items(self):
        cfg = load_config()
        comps = cfg.get("companies") or []
        combined = bool(cfg.get("queue_all_companies", False))
        if not cfg.get("flowaccount_api_key") and not comps:
            QMessageBox.warning(self, "ยังไม่ตั้งค่า", "กรุณาตั้งค่า FlowAccount API ก่อน")
            return
        self.btn_refresh.setEnabled(False)
        self.btn_refresh.setText("⏳ กำลังดึง...")
        if combined and comps:
            self.status_message.emit("กำลังดึงใบรับสินค้าจากทุกบริษัท...")
            self._worker = GoodsReceiptAllWorker(comps)
            self._worker.progress.connect(self.status_message.emit)
        else:
            active = cfg.get("active_company", 0)
            company = comps[active] if (0 <= active < len(comps)) else {}
            self.status_message.emit("กำลังดึงใบรับสินค้า...")
            self._worker = GoodsReceiptFetchWorker(cfg, company)
        self._worker.done.connect(self._on_done)
        self._worker.error.connect(self._on_error)
        self._worker.start()

    def _on_done(self, items):
        self._items = items
        self.btn_refresh.setEnabled(True)
        self.btn_refresh.setText("🔄 รีเฟรช")
        self.lbl_fetched.setText(f"ดึงได้ {len(items)} ใบ")
        self.status_message.emit(f"ดึงใบรับสินค้าได้ {len(items)} ใบ")
        self._apply_filter()

    def _on_error(self, msg):
        self.btn_refresh.setEnabled(True)
        self.btn_refresh.setText("🔄 รีเฟรช")
        self.lbl_fetched.setText("ดึงข้อมูลไม่สำเร็จ")
        QMessageBox.critical(self, "ผิดพลาด", msg)

    def _apply_filter(self):
        q = self.search_box.text().strip().lower()
        if q:
            self._filtered = [e for e in self._items if
                              q in _doc_no(e).lower() or q in _vendor(e).lower()]
        else:
            self._filtered = list(self._items)
        self._page = 1
        self._render_page()

    def _render_page(self):
        per = self._per_page
        total = len(self._filtered)
        pages = max(1, (total + per - 1) // per)
        self._page = min(max(1, self._page), pages)
        start = (self._page - 1) * per
        page_items = self._filtered[start:start + per]

        self.table.setRowCount(0)
        for e in page_items:
            row = self.table.rowCount()
            self.table.insertRow(row)
            vals = [
                _doc_no(e),
                _vendor(e),
                _brand_name(e) or "—",
                fmt_date((e.get("publishedOn") or "")[:10]),
                fmt_date((_due(e) or "")[:10]) if _due(e) else "-",
                fmt_amount(_amount(e)),
                self._status_th(e),
            ]
            for col, v in enumerate(vals):
                item = QTableWidgetItem(v)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if col == 2 and _brand_name(e):   # แบรนด์ — สีตามบริษัท
                    item.setForeground(QBrush(QColor(e.get("_brand_color") or "#334155")))
                    bf = item.font(); bf.setBold(True); item.setFont(bf)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if col == 5:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.table.setItem(row, col, item)

        total_amt = sum(_amount(e) for e in self._filtered)
        self.lbl_summary.setText(f"ทั้งหมด {total} ใบ  |  รวมยอด {fmt_amount(total_amt)} บาท")
        self.lbl_page.setText(f"หน้า {self._page}/{pages}")
        self.btn_prev.setEnabled(self._page > 1)
        self.btn_next.setEnabled(self._page < pages)

    def _status_th(self, e):
        s = (e.get("statusString") or "").strip()
        return self.STATUS_TH.get(s.lower(), s or "-")

    def _prev_page(self):
        if self._page > 1:
            self._page -= 1
            self._render_page()

    def _next_page(self):
        per = self._per_page
        pages = max(1, (len(self._filtered) + per - 1) // per)
        if self._page < pages:
            self._page += 1
            self._render_page()


# ──────────────────── Log Tab (Activity / Queue) ────────────────────

class LogTab(QWidget):
    """หน้า Log — สลับดู Activity Log / Queue Log + ค้นหา + Export
    เตรียมรากฐานสำหรับระบบ login admin/user ในอนาคต"""
    status_message = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self._mode = "activity"   # activity | queue
        self._build_ui()
        self.reload()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(8, 8, 8, 8)
        lay.setSpacing(6)

        tb = QHBoxLayout()
        self.btn_activity = QPushButton("👤 Activity Log")
        self.btn_queue    = QPushButton("📋 Log การจัดคิว")
        self.btn_activity.setCheckable(True); self.btn_queue.setCheckable(True)
        self.btn_activity.setChecked(True)
        self.btn_activity.clicked.connect(lambda: self._switch("activity"))
        self.btn_queue.clicked.connect(lambda: self._switch("queue"))
        for b in (self.btn_activity, self.btn_queue):
            b.setStyleSheet(
                "QPushButton{padding:6px 14px;border:1px solid #cbd5e1;border-radius:5px;"
                "background:white;font-size:12px;font-weight:600;}"
                "QPushButton:checked{background:#16a34a;color:white;border-color:#16a34a;}")
        tb.addWidget(self.btn_activity)
        tb.addWidget(self.btn_queue)
        tb.addSpacing(16)

        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("🔍 ค้นหา วันที่ / ผู้ใช้ / รายการ")
        self.search_box.setClearButtonEnabled(True)
        self.search_box.setFixedWidth(280)
        self.search_box.textChanged.connect(self._render)
        tb.addWidget(self.search_box)
        tb.addStretch()
        self.btn_refresh = QPushButton("🔄 รีเฟรช")
        self.btn_export  = QPushButton("⬇️ Export CSV")
        self.btn_pdf     = QPushButton("🖨️ บันทึก PDF")
        for b in (self.btn_refresh, self.btn_export):
            b.setStyleSheet("QPushButton{padding:6px 12px;border:1px solid #cbd5e1;border-radius:5px;background:white;font-size:12px;}")
        self.btn_pdf.setStyleSheet(
            "QPushButton{padding:6px 12px;border:none;border-radius:5px;"
            "background:#dc2626;color:white;font-size:12px;font-weight:600;}")
        self.btn_pdf.setToolTip("บันทึก Log เป็นไฟล์ PDF (แก้ไขไม่ได้) ไว้ดูย้อนหลัง")
        self.btn_refresh.clicked.connect(self.reload)
        self.btn_export.clicked.connect(self._export)
        self.btn_pdf.clicked.connect(self._export_pdf)
        tb.addWidget(self.btn_refresh)
        tb.addWidget(self.btn_export)
        tb.addWidget(self.btn_pdf)
        lay.addLayout(tb)

        self.table = QTableWidget(0, 5)
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setStyleSheet("QTableWidget{border:1px solid #e2e8f0;font-size:12px;}")
        lay.addWidget(self.table)

        self.lbl_count = QLabel("")
        self.lbl_count.setStyleSheet(f"color:{C_MUTED};font-size:11px;")
        lay.addWidget(self.lbl_count)

    def _switch(self, mode):
        self._mode = mode
        self.btn_activity.setChecked(mode == "activity")
        self.btn_queue.setChecked(mode == "queue")
        self.reload()

    def reload(self):
        self._rows = (activity_log.load_activity() if self._mode == "activity"
                      else activity_log.load_queue_log())
        self._render()

    def _render(self):
        rows = activity_log.search(self._rows, self.search_box.text())
        if self._mode == "activity":
            cols = [("time", "เวลา"), ("user", "ผู้ใช้"), ("action", "การกระทำ"),
                    ("company", "บริษัท"), ("detail", "รายละเอียด")]
        else:
            cols = [("time", "เวลา"), ("user", "ผู้ใช้"), ("count", "จำนวน"),
                    ("company", "บริษัท"), ("detail", "รายละเอียด")]
        self.table.setColumnCount(len(cols))
        self.table.setHorizontalHeaderLabels([c[1] for c in cols])
        self.table.horizontalHeader().setSectionResizeMode(len(cols) - 1, QHeaderView.ResizeMode.Stretch)
        self.table.setColumnWidth(0, 150)
        self.table.setColumnWidth(1, 110)
        self.table.setColumnWidth(2, 110)
        self.table.setColumnWidth(3, 130)
        self.table.setRowCount(0)
        for r in rows:
            row = self.table.rowCount()
            self.table.insertRow(row)
            for col, (key, _) in enumerate(cols):
                v = r.get(key, "")
                if key == "time" and v:
                    v = v.replace("T", " ")
                item = QTableWidgetItem(str(v))
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                self.table.setItem(row, col, item)
        self.lbl_count.setText(f"ทั้งหมด {len(rows)} รายการ")
        self._cols = cols

    def _export(self):
        rows = activity_log.search(self._rows, self.search_box.text())
        if not rows:
            return
        name = ("activity_log" if self._mode == "activity" else "queue_log")
        path, _ = QFileDialog.getSaveFileName(self, "Export Log",
            f"KCash_{name}_{date.today().strftime('%Y%m%d')}.csv", "CSV (*.csv)")
        if not path:
            return
        activity_log.export_csv(rows, path, self._cols)
        QMessageBox.information(self, "✅ สำเร็จ", f"บันทึก Log แล้ว:\n{path}")

    def _export_pdf(self):
        rows = activity_log.search(self._rows, self.search_box.text())
        if not rows:
            QMessageBox.information(self, "ไม่มีข้อมูล", "ไม่มี Log ให้บันทึก")
            return
        title = ("Activity Log (ใครทำอะไรเมื่อไหร่)" if self._mode == "activity"
                 else "Log การจัดคิวจ่ายเงิน")
        name = ("activity_log" if self._mode == "activity" else "queue_log")
        path, _ = QFileDialog.getSaveFileName(self, "บันทึก Log เป็น PDF",
            f"KCash_{name}_{date.today().strftime('%Y%m%d')}.pdf", "PDF (*.pdf)")
        if not path:
            return
        if not path.lower().endswith(".pdf"):
            path += ".pdf"
        try:
            from utils import build_log_pdf
            build_log_pdf(rows, self._cols, title, path)
        except Exception as e:
            QMessageBox.critical(self, "❌ ผิดพลาด", f"บันทึก PDF ไม่สำเร็จ:\n{e}")
            return
        activity_log.log("บันทึก Log PDF", f"{title} • {len(rows)} รายการ")
        QMessageBox.information(self, "✅ สำเร็จ",
            f"บันทึก Log เป็น PDF แล้ว (แก้ไขไม่ได้):\n{path}")


# ──────────────────── Login + User Management (ข้อ 8) ────────────────────

class ContactDialog(QDialog):
    """ช่องทางติดต่อผู้พัฒนา (กดจาก Powered by)"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("ช่องทางติดต่อผู้พัฒนา")
        self.setFixedWidth(420)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(22, 20, 22, 18)
        lay.setSpacing(8)

        head = QLabel("ช่องทางติดต่อผู้พัฒนา")
        head.setStyleSheet("font-size:16px;font-weight:700;color:#15803d;")
        lay.addWidget(head)

        body = QLabel(
            "<div style='font-size:13px;color:#0f172a;line-height:1.7'>"
            "<b>ผู้พัฒนา :</b> คุณพชร รัชนาทสกุล (คุณมายด์)<br>"
            "<b>Line :</b> @mildpachara<br>"
            "<b>โทร :</b> 092-703-2121 (คุณมายด์)<br>"
            "<b>Email :</b> mildwork43@gmail.com<br>"
            "<b>Facebook :</b> "
            "<a href='https://www.facebook.com/MildPachara2543?locale=th_TH'>"
            "facebook.com/MildPachara2543</a>"
            "</div>")
        body.setTextFormat(Qt.TextFormat.RichText)
        body.setOpenExternalLinks(True)
        body.setWordWrap(True)
        body.setTextInteractionFlags(Qt.TextInteractionFlag.TextBrowserInteraction)
        lay.addWidget(body)

        btn = QPushButton("ปิด")
        btn.setStyleSheet("QPushButton{padding:6px 18px;border:1px solid #cbd5e1;"
                          "border-radius:4px;background:white;}")
        btn.clicked.connect(self.accept)
        row = QHBoxLayout(); row.addStretch(); row.addWidget(btn)
        lay.addLayout(row)


class LoginDialog(QDialog):
    """หน้า Login สไตล์ FB — ซ้ายแบรนด์ ขวาฟอร์ม (ไม่มีปุ่มสมัครเอง)"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("เข้าสู่ระบบ — KCash Queue System")
        self.setFixedSize(820, 460)
        self.user = None

        root = QHBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ซ้าย — แผงแบรนด์ (ไล่เฉดเขียวแบบทางการ)
        left = QFrame()
        left.setStyleSheet(
            "background:qlineargradient(x1:0,y1:0,x2:1,y2:1,"
            "stop:0 #064e3b, stop:0.45 #0f766e, stop:1 #16a34a);")
        ll = QVBoxLayout(left)
        ll.setContentsMargins(40, 36, 40, 24)
        ll.addStretch()

        # โลโก้ (ตัดกรอบสี่เหลี่ยมออก ทำมุมมนให้กลืนพื้นหลัง)
        logo_img = QLabel()
        logo_img.setStyleSheet("background:transparent;")
        logo_path = _resource_path("icon/icon.png")
        if os.path.exists(logo_path):
            pix = QPixmap(logo_path).scaled(
                150, 150, Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation)
            logo_img.setPixmap(self._rounded_pixmap(pix, 34))
        ll.addWidget(logo_img, alignment=Qt.AlignmentFlag.AlignLeft)

        logo = QLabel("KCash")
        logo.setStyleSheet("color:white;font-size:48px;font-weight:800;background:transparent;")
        sub = QLabel("ระบบจัดคิวจ่ายเงิน\nสำหรับฝ่ายบัญชี")
        sub.setStyleSheet("color:#dcfce7;font-size:18px;background:transparent;")
        ll.addSpacing(8)
        ll.addWidget(logo); ll.addWidget(sub)
        ll.addStretch()

        # เครดิตมุมล่างซ้าย — กดได้ เปิดหน้าต่างช่องทางติดต่อ
        powered = QLabel("<a href='#contact' style='color:#bbf7d0;text-decoration:none;'>"
                         "Powered by Mild Pachara Ratchanatskul</a>")
        powered.setStyleSheet("font-size:11px;background:transparent;")
        powered.setCursor(Qt.CursorShape.PointingHandCursor)
        powered.setOpenExternalLinks(False)
        powered.linkActivated.connect(lambda _: ContactDialog(self).exec())
        ll.addWidget(powered, alignment=Qt.AlignmentFlag.AlignLeft)

        left.setFixedWidth(380)
        root.addWidget(left)

        # ขวา — ฟอร์ม
        right = QFrame()
        right.setStyleSheet("background:white;")
        rl = QVBoxLayout(right)
        rl.setContentsMargins(46, 46, 46, 46)
        rl.setSpacing(14)
        title = QLabel("เข้าสู่ระบบ")
        title.setStyleSheet("font-size:24px;font-weight:700;color:#0f172a;background:transparent;")
        rl.addWidget(title)

        self.ed_user = QLineEdit()
        self.ed_user.setPlaceholderText("ชื่อผู้ใช้ (username)")
        self.ed_pw = QLineEdit()
        self.ed_pw.setPlaceholderText("รหัสผ่าน")
        self.ed_pw.setEchoMode(QLineEdit.EchoMode.Password)
        for e in (self.ed_user, self.ed_pw):
            e.setMinimumHeight(44)
            e.setStyleSheet("QLineEdit{border:1px solid #cbd5e1;border-radius:8px;"
                            "padding:8px 12px;font-size:15px;background:white;color:#0f172a;}"
                            "QLineEdit:focus{border-color:#16a34a;}")
            rl.addWidget(e)
        self.ed_pw.returnPressed.connect(self._do_login)

        # ปุ่ม 👁 ดู/ซ่อนรหัสผ่าน (ในช่องรหัสผ่าน)
        self._show_pw = False
        self.act_eye = self.ed_pw.addAction(
            self._eye_icon(False), QLineEdit.ActionPosition.TrailingPosition)
        self.act_eye.setToolTip("แสดง/ซ่อนรหัสผ่าน")
        self.act_eye.triggered.connect(self._toggle_pw)

        # checkbox จดจำ (จำเฉพาะ username) — ใส่กรอบให้เห็นชัดเสมอ
        self.chk_remember = QCheckBox("จดจำ")
        self.chk_remember.setStyleSheet(
            "QCheckBox{color:#475569;font-size:13px;background:transparent;spacing:6px;}"
            "QCheckBox::indicator{width:16px;height:16px;border:1.5px solid #94a3b8;"
            "border-radius:4px;background:white;}"
            "QCheckBox::indicator:checked{background:#16a34a;border-color:#16a34a;}")
        rl.addWidget(self.chk_remember)

        self.lbl_err = QLabel("")
        self.lbl_err.setStyleSheet("color:#dc2626;font-size:13px;background:transparent;")
        rl.addWidget(self.lbl_err)

        btn = QPushButton("เข้าสู่ระบบ")
        btn.setMinimumHeight(46)
        btn.setStyleSheet("QPushButton{background:#16a34a;color:white;border:none;"
                          "border-radius:8px;font-size:17px;font-weight:700;}"
                          "QPushButton:hover{background:#15803d;}")
        btn.clicked.connect(self._do_login)
        rl.addWidget(btn)
        rl.addStretch()
        note = QLabel("โปรแกรมชุดนี้สร้างขึ้นเพื่อใช้งานภายใน "
                      "บริษัท ฮอลลีวู้ด เอ็นเตอร์ไพร์ส จำกัด เท่านั้น\n"
                      "ห้ามนำไปเผยแพร่ หากตรวจสอบพบ มีโทษทางกฎหมาย")
        note.setWordWrap(True)
        note.setStyleSheet("color:#94a3b8;font-size:11px;background:transparent;")
        rl.addWidget(note)
        root.addWidget(right, 1)

        # เติมเฉพาะ username ที่จำไว้ (ไม่จำรหัสผ่าน)
        ru, _ = users.load_remember()
        if ru:
            self.ed_user.setText(ru)
            self.chk_remember.setChecked(True)
            self.ed_pw.setFocus()   # ไปที่ช่องรหัสผ่านให้เลย

    def _rounded_pixmap(self, src: QPixmap, radius: int = 30) -> QPixmap:
        """ทำมุมโลโก้ให้มน ตัดขอบสี่เหลี่ยมออก ให้กลืนกับพื้นหลัง"""
        if src.isNull():
            return src
        out = QPixmap(src.size())
        out.fill(Qt.GlobalColor.transparent)
        p = QPainter(out)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        p.setRenderHint(QPainter.RenderHint.SmoothPixmapTransform)
        path = QPainterPath()
        path.addRoundedRect(0, 0, src.width(), src.height(), radius, radius)
        p.setClipPath(path)
        p.drawPixmap(0, 0, src)
        p.end()
        return out

    def _eye_icon(self, shown: bool) -> QIcon:
        """วาดไอคอนรูปตาเอง (ไม่ใช้ emoji กันเรนเดอร์เพี้ยน)
        shown=False → ตาเปิด (รหัสถูกซ่อน กดเพื่อดู)
        shown=True  → ตาขีดทับ (รหัสกำลังแสดง กดเพื่อซ่อน)"""
        pix = QPixmap(24, 24)
        pix.fill(Qt.GlobalColor.transparent)
        p = QPainter(pix)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        pen = QPen(QColor("#64748b")); pen.setWidth(2)
        p.setPen(pen)
        # เปลือกตา (เส้นโค้งบน-ล่าง เป็นวงรี)
        p.drawArc(3, 6, 18, 12, 0, 180 * 16)
        p.drawArc(3, 6, 18, 12, 0, -180 * 16)
        # รูม่านตา
        p.setBrush(QColor("#64748b"))
        p.drawEllipse(9, 9, 6, 6)
        if shown:
            # ขีดทับเฉียง
            pen2 = QPen(QColor("#dc2626")); pen2.setWidth(2)
            p.setPen(pen2)
            p.drawLine(4, 20, 20, 4)
        p.end()
        return QIcon(pix)

    def _toggle_pw(self):
        self._show_pw = not self._show_pw
        self.ed_pw.setEchoMode(
            QLineEdit.EchoMode.Normal if self._show_pw else QLineEdit.EchoMode.Password)
        self.act_eye.setIcon(self._eye_icon(self._show_pw))

    def _do_login(self):
        u = users.authenticate(self.ed_user.text(), self.ed_pw.text())
        if not u:
            self.lbl_err.setText("โปรดตรวจสอบ username และ password ด้วย")
            return
        # จดจำ = จำเฉพาะ username เท่านั้น (ไม่จำรหัสผ่าน)
        if self.chk_remember.isChecked():
            users.save_remember(self.ed_user.text().strip(), "")
        else:
            users.clear_remember()
        self.user = u
        self.accept()


class UserEditDialog(QDialog):
    """ฟอร์มเพิ่ม/แก้ไขผู้ใช้"""
    def __init__(self, data=None, viewer_role="admin", parent=None):
        super().__init__(parent)
        self.setWindowTitle("ข้อมูลผู้ใช้")
        self.resize(420, 0)
        self._editing = data is not None
        data = data or {}
        form = QFormLayout(self)
        self.ed_username = QLineEdit(data.get("username", ""))
        if self._editing:
            self.ed_username.setReadOnly(True)
        self.ed_password = QLineEdit()
        self.ed_password.setPlaceholderText("(เว้นว่าง = ไม่เปลี่ยน)" if self._editing else "ตั้งรหัสผ่าน")
        self.ed_password.setEchoMode(QLineEdit.EchoMode.Password)
        self.ed_fullname = QLineEdit(data.get("fullname", ""))
        self.ed_nickname = QLineEdit(data.get("nickname", ""))
        self.cmb_role = QComboBox()
        # dev สร้าง/ตั้งได้ทุก role; admin ได้แค่ user
        roles = ["user", "admin", "dev"] if viewer_role == "dev" else ["user"]
        self.cmb_role.addItems(roles)
        if data.get("role") in roles:
            self.cmb_role.setCurrentText(data["role"])
        form.addRow("Username:", self.ed_username)
        form.addRow("รหัสผ่าน:", self.ed_password)
        form.addRow("ชื่อ-นามสกุลจริง:", self.ed_fullname)
        form.addRow("ชื่อเล่น (โชว์ใน Log):", self.ed_nickname)
        form.addRow("สิทธิ์ (role):", self.cmb_role)
        bb = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        bb.accepted.connect(self.accept); bb.rejected.connect(self.reject)
        form.addRow(bb)

    def get(self):
        return {"username": self.ed_username.text().strip(),
                "password": self.ed_password.text(),
                "fullname": self.ed_fullname.text().strip(),
                "nickname": self.ed_nickname.text().strip(),
                "role": self.cmb_role.currentText()}


class UserManagerDialog(QDialog):
    """จัดการผู้ใช้ — เพิ่ม/แก้ไข/ลบ (admin มองไม่เห็น dev)"""
    def __init__(self, viewer, parent=None):
        super().__init__(parent)
        self.viewer = viewer
        self.setWindowTitle("จัดการผู้ใช้")
        self.resize(640, 420)
        lay = QVBoxLayout(self)
        lay.addWidget(QLabel(f"เข้าใช้งานโดย: {viewer['nickname']} ({viewer['role']})"))

        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["Username", "ชื่อจริง", "ชื่อเล่น", "สิทธิ์"])
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        lay.addWidget(self.table)

        btns = QHBoxLayout()
        for txt, fn in [("➕ เพิ่ม", self._add), ("✏️ แก้ไข", self._edit), ("🗑️ ลบ", self._del)]:
            b = QPushButton(txt)
            b.setStyleSheet("QPushButton{padding:6px 12px;border:1px solid #cbd5e1;border-radius:4px;background:white;}")
            b.clicked.connect(fn); btns.addWidget(b)
        btns.addStretch()
        bclose = QPushButton("ปิด"); bclose.clicked.connect(self.accept)
        btns.addWidget(bclose)
        lay.addLayout(btns)
        self._reload()

    def _reload(self):
        rows = users.list_users(self.viewer["role"])
        self.table.setRowCount(0)
        for u in rows:
            r = self.table.rowCount(); self.table.insertRow(r)
            for c, k in enumerate(("username", "fullname", "nickname", "role")):
                self.table.setItem(r, c, QTableWidgetItem(str(u.get(k, ""))))

    def _cur_username(self):
        r = self.table.currentRow()
        if r < 0:
            return None
        return self.table.item(r, 0).text()

    def _add(self):
        dlg = UserEditDialog(None, self.viewer["role"], self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            d = dlg.get()
            try:
                users.add_user(self.viewer["role"], d["username"], d["password"],
                               d["fullname"], d["nickname"], d["role"])
                activity_log.log("เพิ่มผู้ใช้", f"{d['username']} ({d['role']})")
                self._reload()
            except Exception as e:
                QMessageBox.warning(self, "เพิ่มไม่ได้", str(e))

    def _edit(self):
        un = self._cur_username()
        if not un:
            return
        cur = next((u for u in users.list_users(self.viewer["role"])
                    if u["username"] == un), None)
        if not cur:
            return
        dlg = UserEditDialog(cur, self.viewer["role"], self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            d = dlg.get()
            try:
                users.update_user(self.viewer["role"], un,
                                  password=d["password"] or None,
                                  fullname=d["fullname"], nickname=d["nickname"],
                                  role=d["role"])
                activity_log.log("แก้ไขผู้ใช้", un)
                self._reload()
            except Exception as e:
                QMessageBox.warning(self, "แก้ไขไม่ได้", str(e))

    def _del(self):
        un = self._cur_username()
        if not un:
            return
        if QMessageBox.question(self, "ลบผู้ใช้", f"ลบผู้ใช้ '{un}'?") != QMessageBox.StandardButton.Yes:
            return
        try:
            users.delete_user(self.viewer["role"], un)
            activity_log.log("ลบผู้ใช้", un)
            self._reload()
        except Exception as e:
            QMessageBox.warning(self, "ลบไม่ได้", str(e))


# ──────────────────── Dev Console (เฉพาะ dev mode) ────────────────────

class HealthWorker(QThread):
    """เช็คสถานะบริการภายนอก (เน็ต/FlowAccount/monday/แชร์ลิงก์) — background"""
    result = pyqtSignal(dict)

    def run(self):
        import requests

        def reachable(url):
            try:
                requests.get(url, timeout=6, headers={"User-Agent": "KCash"})
                return "up"          # ได้ response ใดๆ = บริการออนไลน์
            except Exception:
                return "down"

        out = {}
        out["net"] = reachable("https://www.google.com/generate_204")
        # ถ้าเน็ตล่ม ตัวอื่นก็ล่มหมด ไม่ต้องเช็คต่อ
        if out["net"] == "down":
            out.update({"flowaccount": "down", "monday": "down", "share": "down"})
        else:
            out["flowaccount"] = reachable("https://openapi.flowaccount.com/v1")
            out["monday"] = reachable("https://api.monday.com/v2")
            out["share"] = reachable("https://cdn.statically.io")
        self.result.emit(out)


class GpuWorker(QThread):
    """อ่าน % การใช้งาน GPU แบบ background (รองรับทั้ง NVIDIA และการ์ดออนบอร์ด)"""
    result = pyqtSignal(str)

    def run(self):
        flags = 0x08000000 if os.name == "nt" else 0   # CREATE_NO_WINDOW
        # 1) NVIDIA แยก (ถ้ามี)
        try:
            out = subprocess.run(
                ["nvidia-smi", "--query-gpu=utilization.gpu",
                 "--format=csv,noheader,nounits"],
                capture_output=True, text=True, timeout=4, creationflags=flags)
            v = out.stdout.strip().splitlines()[0].strip()
            if v.isdigit():
                self.result.emit(f"{v}%"); return
        except Exception:
            pass
        # 2) การ์ดออนบอร์ด/ใดๆ บน Windows ผ่าน Performance Counter
        if os.name == "nt":
            try:
                ps = ("$ErrorActionPreference='SilentlyContinue';"
                      "$s=(Get-Counter '\\GPU Engine(*engtype_3D)\\Utilization Percentage')"
                      ".CounterSamples|Measure-Object -Property CookedValue -Sum;"
                      "[math]::Round($s.Sum)")
                out = subprocess.run(
                    ["powershell", "-NoProfile", "-Command", ps],
                    capture_output=True, text=True, timeout=8, creationflags=flags)
                v = out.stdout.strip()
                if v and v.replace(".", "").isdigit():
                    self.result.emit(f"{min(100, int(float(v)))}%"); return
            except Exception:
                pass
        self.result.emit("N/A")


class DevConsole(QWidget):
    """คอนโซลแสดงสถานะการทำงาน/ข้อผิดพลาด — เฉพาะ dev mode, พับเก็บได้"""
    _instance = None
    log_sig = pyqtSignal(str, str)   # (message, level)

    def __init__(self, parent=None):
        super().__init__(parent)
        DevConsole._instance = self
        self._collapsed = False
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # แถบหัว
        bar = QWidget()
        bar.setStyleSheet("background:#0f172a;")
        bl = QHBoxLayout(bar)
        bl.setContentsMargins(8, 3, 8, 3)
        self.btn_toggle = QPushButton("▼ Dev Console")
        self.btn_toggle.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_toggle.setStyleSheet(
            "QPushButton{color:#4ade80;background:transparent;border:none;"
            "font-weight:700;font-size:12px;}")
        self.btn_toggle.clicked.connect(self.toggle)
        title = QLabel("🖥️ Dev Console — สถานะการทำงานของโปรแกรม")
        title.setStyleSheet("color:#94a3b8;font-size:11px;background:transparent;")
        btn_clear = QPushButton("ล้าง")
        btn_clear.setStyleSheet(
            "QPushButton{color:#cbd5e1;background:#1e293b;border:none;border-radius:3px;"
            "padding:2px 10px;font-size:11px;}QPushButton:hover{background:#334155;}")
        btn_clear.clicked.connect(lambda: self.view.clear())
        bl.addWidget(self.btn_toggle)
        bl.addSpacing(8)
        bl.addWidget(title)
        bl.addStretch()
        bl.addWidget(btn_clear)
        root.addWidget(bar)

        # แถบสถานะระบบ + บริการ
        statbar = QWidget()
        statbar.setStyleSheet("background:#0b1220;border-top:1px solid #1e293b;")
        sl = QHBoxLayout(statbar)
        sl.setContentsMargins(10, 3, 10, 3)
        self.stat_sys = QLabel("CPU – · RAM – · DISK – · GPU –")
        self.stat_sys.setStyleSheet(
            "color:#94a3b8;font-family:'Consolas',monospace;font-size:11px;background:transparent;")
        self.stat_net = QLabel()
        self.stat_net.setStyleSheet("font-size:11px;background:transparent;")
        sl.addWidget(self.stat_sys)
        sl.addStretch()
        sl.addWidget(self.stat_net)
        root.addWidget(statbar)
        self._svc = {"net": "?", "flowaccount": "?", "monday": "?", "share": "?"}
        self._render_services()

        # พื้นที่ข้อความ
        self.view = QPlainTextEdit()
        self.view.setReadOnly(True)
        self.view.setMaximumBlockCount(3000)
        self.view.setStyleSheet(
            "QPlainTextEdit{background:#0b1220;color:#e2e8f0;border:none;"
            "font-family:'Consolas','Courier New',monospace;font-size:12px;}")
        self.view.setFixedHeight(150)
        root.addWidget(self.view)

        self.log_sig.connect(self._append)

        # มอนิเตอร์ระบบ (CPU/RAM/DISK) ทุก 2 วิ
        self._gpu_cache = "…"
        self._sys_timer = QTimer(self)
        self._sys_timer.timeout.connect(self._update_sys)
        self._sys_timer.start(2000)
        self._update_sys()
        # GPU เช็คใน background ทุก 5 วิ (กัน UI ค้างจาก subprocess)
        self._gpu_timer = QTimer(self)
        self._gpu_timer.timeout.connect(self._start_gpu)
        self._gpu_timer.start(5000)
        self._start_gpu()
        # มอนิเตอร์บริการ (เน็ต/FlowAccount/monday/แชร์ลิงก์) ทุก 30 วิ
        self._health_timer = QTimer(self)
        self._health_timer.timeout.connect(self._start_health)
        self._health_timer.start(30000)
        self._start_health()

    def _update_sys(self):
        if psutil is None:
            self.stat_sys.setText("CPU/RAM/DISK: ต้องติดตั้ง psutil")
            return
        try:
            cpu = psutil.cpu_percent()
            ram = psutil.virtual_memory().percent
            disk = psutil.disk_usage(os.path.splitdrive(sys.executable)[0] + "\\").percent
        except Exception:
            cpu = ram = disk = 0
        self.stat_sys.setText(
            f"CPU {cpu:.0f}% · RAM {ram:.0f}% · DISK {disk:.0f}% · GPU {self._gpu_cache}")

    def _start_gpu(self):
        w = getattr(self, "_gpu_worker", None)
        if w is not None and w.isRunning():
            return
        self._gpu_worker = GpuWorker(self)
        self._gpu_worker.result.connect(self._on_gpu)
        self._gpu_worker.start()

    def _on_gpu(self, val):
        self._gpu_cache = val

    def _start_health(self):
        w = getattr(self, "_health_worker", None)
        if w is not None and w.isRunning():
            return
        self._health_worker = HealthWorker(self)
        self._health_worker.result.connect(self._on_health)
        self._health_worker.start()

    def _on_health(self, res: dict):
        for k, v in res.items():
            if k in self._svc:
                # ไม่ทับสถานะ 'limit' ของ monday ที่ตั้งจากการ sync จริง
                if k == "monday" and self._svc.get("monday") == "limit" and v == "up":
                    continue
                self._svc[k] = v
        self._render_services()

    def set_service(self, name, state):
        """ตั้งสถานะบริการเอง (เช่น monday='limit' ตอน sync ติดลิมิต) — thread-safe ผ่าน log_sig ไม่ได้ ใช้ตรง"""
        if name in self._svc:
            self._svc[name] = state
            self._render_services()

    def _render_services(self):
        label = {"net": "เน็ต", "flowaccount": "FlowAccount",
                 "monday": "monday", "share": "แชร์ลิงก์"}
        color = {"up": "#4ade80", "down": "#f87171",
                 "limit": "#fbbf24", "?": "#64748b"}
        txt = {"up": "", "down": " ล่ม", "limit": " ลิมิต", "?": ""}
        parts = []
        for k in ("net", "flowaccount", "monday", "share"):
            st = self._svc.get(k, "?")
            parts.append(
                f"<span style='color:{color[st]}'>●</span>"
                f"<span style='color:#cbd5e1'> {label[k]}{txt[st]}</span>")
        self.stat_net.setText("&nbsp;&nbsp;".join(parts))

    def toggle(self):
        self._collapsed = not self._collapsed
        self.view.setVisible(not self._collapsed)
        self.btn_toggle.setText("▶ Dev Console" if self._collapsed else "▼ Dev Console")

    def log(self, msg, level="info"):
        # thread-safe (เรียกจาก worker ได้)
        self.log_sig.emit(str(msg), level)

    def _append(self, msg, level):
        colors = {"info": "#e2e8f0", "ok": "#4ade80",
                  "warn": "#fbbf24", "err": "#f87171"}
        c = colors.get(level, "#e2e8f0")
        ts = datetime.now().strftime("%H:%M:%S")
        safe = (msg.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))
        self.view.appendHtml(
            f"<span style='color:#475569'>{ts}</span> "
            f"<span style='color:{c}'>{safe}</span>")


def dev_log(msg, level="info"):
    """ส่งข้อความเข้า Dev Console (ถ้าเปิดอยู่)"""
    c = DevConsole._instance
    if c is not None:
        try:
            c.log(msg, level)
        except Exception:
            pass


class _StreamTee:
    """ดักจับ stdout/stderr ส่งเข้า Dev Console ด้วย (ยังพิมพ์ออกตามปกติ)"""
    def __init__(self, orig, level):
        self.orig = orig
        self.level = level

    def write(self, s):
        try:
            if self.orig:
                self.orig.write(s)
        except Exception:
            pass
        s2 = s.strip()
        if s2:
            dev_log(s2, self.level)

    def flush(self):
        try:
            if self.orig:
                self.orig.flush()
        except Exception:
            pass


# ──────────────────── Sensitive Manager Dialog ────────────────────

class SensitiveManagerDialog(QDialog):
    """ตั้งรหัสผ่าน + จัดการรายชื่อ sensitive (เฉพาะ admin/dev)"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("🔒 จัดการข้อมูลลับ (Sensitive Data)")
        self.setMinimumWidth(540)
        self.setModal(True)
        lay = QVBoxLayout(self)
        lay.setSpacing(14)
        lay.setContentsMargins(20, 18, 20, 18)

        # ── ส่วนรหัสผ่าน ──
        grp_pw = QGroupBox("รหัสผ่านสำหรับดูยอดเงิน")
        gp = QFormLayout(grp_pw)
        gp.setSpacing(8)
        self.ed_pw1 = QLineEdit(); self.ed_pw1.setEchoMode(QLineEdit.EchoMode.Password)
        self.ed_pw2 = QLineEdit(); self.ed_pw2.setEchoMode(QLineEdit.EchoMode.Password)
        self.ed_pw1.setPlaceholderText("รหัสผ่านใหม่")
        self.ed_pw2.setPlaceholderText("ยืนยันรหัสผ่านใหม่")
        gp.addRow("รหัสผ่านใหม่:", self.ed_pw1)
        gp.addRow("ยืนยัน:", self.ed_pw2)
        data = _sensitive_load()
        has_pw = bool(data.get("password_hash"))
        self.lbl_pw_status = QLabel("🔒 ตั้งรหัสไว้แล้ว" if has_pw else "⚠️ ยังไม่ได้ตั้งรหัส")
        self.lbl_pw_status.setStyleSheet(
            f"color:{'#16a34a' if has_pw else '#dc2626'};font-weight:600;font-size:12px;")
        gp.addRow("สถานะ:", self.lbl_pw_status)
        pw_btn_row = QHBoxLayout()
        self.btn_set_pw = QPushButton("💾 บันทึกรหัสผ่าน")
        self.btn_set_pw.setStyleSheet(
            "QPushButton{background:#2563eb;color:white;border:none;border-radius:5px;"
            "padding:5px 16px;font-weight:600;}"
            "QPushButton:hover{background:#1d4ed8;}")
        self.btn_set_pw.clicked.connect(self._save_password)
        self.btn_reset_pw = QPushButton("🔄 รีเซ็ตรหัสผ่าน")
        self.btn_reset_pw.setStyleSheet(
            "QPushButton{background:#fee2e2;color:#dc2626;border:1px solid #dc2626;"
            "border-radius:5px;padding:5px 16px;font-weight:600;}"
            "QPushButton:hover{background:#fecaca;}")
        self.btn_reset_pw.clicked.connect(self._reset_password)
        pw_btn_row.addWidget(self.btn_set_pw)
        pw_btn_row.addWidget(self.btn_reset_pw)
        pw_btn_row.addStretch()
        gp.addRow("", pw_btn_row)
        lay.addWidget(grp_pw)

        # ── ส่วนรายชื่อ ──
        grp_names = QGroupBox("รายชื่อที่ต้องซ่อนยอดเงิน")
        gn = QVBoxLayout(grp_names)
        gn.setSpacing(6)

        hint = QLabel("ชื่อผู้รับเงิน/Vendor ที่ตรงกันจะแสดงยอดเป็น ●●●●● ในหน้าตัดบิล (จับคู่สลิป)")
        hint.setStyleSheet("color:#64748b;font-size:11px;")
        hint.setWordWrap(True)
        gn.addWidget(hint)

        self.lst = QListWidget()
        self.lst.setAlternatingRowColors(True)
        self.lst.setWordWrap(False)   # ชื่อแสดงบรรทัดเดียว ไม่ตัดบรรทัด
        self.lst.setStyleSheet("font-size:13px;")
        _seen = set()
        for n in (data.get("names") or []):
            nm = re.sub(r"\s+", " ", n).strip()   # รวมช่องว่าง/บรรทัดใหม่ → บรรทัดเดียว
            if nm and nm.lower() not in _seen:
                self.lst.addItem(nm)
                _seen.add(nm.lower())
        gn.addWidget(self.lst, 1)

        bar_add = QHBoxLayout()
        self.ed_name = QLineEdit()
        self.ed_name.setPlaceholderText("พิมพ์ชื่อแล้วกด เพิ่ม หรือกด Enter")
        self.ed_name.returnPressed.connect(self._add_name)
        bar_add.addWidget(self.ed_name, 1)
        btn_add = QPushButton("➕ เพิ่ม")
        btn_add.setStyleSheet(
            "QPushButton{background:#16a34a;color:white;border:none;border-radius:5px;padding:5px 12px;}"
            "QPushButton:hover{background:#15803d;}")
        btn_add.clicked.connect(self._add_name)
        bar_add.addWidget(btn_add)
        gn.addLayout(bar_add)

        bar_act = QHBoxLayout()
        btn_del = QPushButton("🗑️ ลบที่เลือก")
        btn_del.setStyleSheet(
            "QPushButton{background:#dc2626;color:white;border:none;border-radius:5px;padding:5px 12px;}"
            "QPushButton:hover{background:#b91c1c;}")
        btn_del.clicked.connect(self._del_selected)
        btn_import = QPushButton("📂 Import จากไฟล์ (Excel/.csv/.txt)")
        btn_import.setStyleSheet(
            "QPushButton{background:#f1f5f9;border:1px solid #cbd5e1;border-radius:5px;"
            "padding:5px 12px;color:#334155;}"
            "QPushButton:hover{background:#e2e8f0;}")
        btn_import.clicked.connect(self._import_file)
        btn_clear = QPushButton("🧹 ล้างทั้งหมด")
        btn_clear.setStyleSheet(
            "QPushButton{background:#f1f5f9;border:1px solid #cbd5e1;border-radius:5px;"
            "padding:5px 12px;color:#64748b;}"
            "QPushButton:hover{background:#fee2e2;color:#dc2626;border-color:#dc2626;}")
        btn_clear.clicked.connect(self._clear_names)
        bar_act.addWidget(btn_del)
        bar_act.addWidget(btn_import)
        bar_act.addStretch()
        bar_act.addWidget(btn_clear)
        gn.addLayout(bar_act)

        self.lbl_count = QLabel(f"{self.lst.count()} รายชื่อ")
        self.lbl_count.setStyleSheet("color:#64748b;font-size:11px;")
        gn.addWidget(self.lbl_count)
        lay.addWidget(grp_names, 1)

        # ── ปุ่มล่าง ──
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_save = QPushButton("💾 บันทึกรายชื่อ")
        btn_save.setStyleSheet(
            "QPushButton{background:#16a34a;color:white;border:none;border-radius:6px;"
            "padding:6px 20px;font-weight:700;font-size:13px;}"
            "QPushButton:hover{background:#15803d;}")
        btn_save.clicked.connect(self._save_names)
        btn_close = QPushButton("ปิด")
        btn_close.setStyleSheet(
            "QPushButton{background:#f1f5f9;border:1px solid #cbd5e1;"
            "border-radius:6px;padding:6px 20px;color:#334155;}"
            "QPushButton:hover{background:#e2e8f0;}")
        btn_close.clicked.connect(self.accept)
        btn_row.addWidget(btn_save)
        btn_row.addWidget(btn_close)
        lay.addLayout(btn_row)

    def _save_password(self):
        pw1 = self.ed_pw1.text().strip()
        pw2 = self.ed_pw2.text().strip()
        if not pw1:
            QMessageBox.warning(self, "ข้อผิดพลาด", "กรุณาใส่รหัสผ่าน")
            return
        if pw1 != pw2:
            QMessageBox.warning(self, "ข้อผิดพลาด", "รหัสผ่านทั้งสองไม่ตรงกัน")
            return
        if len(pw1) < 4:
            QMessageBox.warning(self, "ข้อผิดพลาด", "รหัสผ่านต้องมีอย่างน้อย 4 ตัวอักษร")
            return
        data = _sensitive_load()
        data["password_hash"] = _sensitive_hash(pw1)
        _sensitive_save(data)
        self.ed_pw1.clear(); self.ed_pw2.clear()
        self.lbl_pw_status.setText("🔒 ตั้งรหัสไว้แล้ว")
        self.lbl_pw_status.setStyleSheet("color:#16a34a;font-weight:600;font-size:12px;")
        QMessageBox.information(self, "บันทึกแล้ว", "ตั้งรหัสผ่านใหม่เรียบร้อย")

    def _reset_password(self):
        data = _sensitive_load()
        if not data.get("password_hash"):
            QMessageBox.information(self, "รีเซ็ตรหัสผ่าน", "ยังไม่ได้ตั้งรหัสผ่านไว้")
            return
        ok = QMessageBox.question(
            self, "ยืนยันรีเซ็ตรหัสผ่าน",
            "ต้องการลบรหัสผ่านปัจจุบันใช่ไหม?\n"
            "หลังรีเซ็ต ยอดเงินที่ซ่อนไว้จะดูไม่ได้จนกว่าจะตั้งรหัสใหม่",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No)
        if ok != QMessageBox.StandardButton.Yes:
            return
        data["password_hash"] = ""
        _sensitive_save(data)
        self.ed_pw1.clear(); self.ed_pw2.clear()
        self.lbl_pw_status.setText("⚠️ ยังไม่ได้ตั้งรหัส")
        self.lbl_pw_status.setStyleSheet("color:#dc2626;font-weight:600;font-size:12px;")
        QMessageBox.information(self, "รีเซ็ตแล้ว",
            "ลบรหัสผ่านเรียบร้อย — กรุณาตั้งรหัสใหม่เพื่อใช้งานการดูยอดเงิน")

    def _add_name(self):
        name = re.sub(r"\s+", " ", self.ed_name.text()).strip()
        if not name:
            return
        # ตรวจซ้ำ
        existing = [self.lst.item(i).text().lower() for i in range(self.lst.count())]
        if name.lower() in existing:
            self.ed_name.clear()
            return
        self.lst.addItem(name)
        self.ed_name.clear()
        self._update_count()

    def _del_selected(self):
        for item in self.lst.selectedItems():
            self.lst.takeItem(self.lst.row(item))
        self._update_count()

    def _import_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "เลือกไฟล์รายชื่อ", "",
            "รายชื่อ (*.xlsx *.csv *.txt);;Excel (*.xlsx);;"
            "Text/CSV (*.txt *.csv);;All Files (*)")
        if not path:
            return
        try:
            if path.lower().endswith(".xlsx"):
                names = self._names_from_xlsx(path)   # Excel HR export
            else:
                with open(path, encoding="utf-8-sig") as f:
                    names = [re.sub(r"\s+", " ", l.split(",")[0].strip().strip('"')).strip()
                             for l in f if l.strip()]
        except Exception as e:
            QMessageBox.critical(self, "อ่านไฟล์ไม่ได้", str(e))
            return
        if not names:
            QMessageBox.warning(self, "ไม่พบรายชื่อ",
                "อ่านไฟล์ได้ แต่หาคอลัมน์ชื่อ/นามสกุลไม่เจอ\n"
                "(ไฟล์ Excel ควรมีหัวคอลัมน์ 'ชื่อ' และ 'นามสกุล')")
            return
        existing = {self.lst.item(i).text().lower() for i in range(self.lst.count())}
        added = 0
        for name in names:
            if name and name.lower() not in existing:
                self.lst.addItem(name)
                existing.add(name.lower())
                added += 1
        self._update_count()
        QMessageBox.information(self, "Import สำเร็จ",
            f"เพิ่มรายชื่อใหม่ {added} รายการ (ข้ามที่ซ้ำ)\n"
            f"จากทั้งหมดในไฟล์ {len(names)} ชื่อ")

    def _names_from_xlsx(self, path):
        """ดึงชื่อจากไฟล์ Excel (รายชื่อพนักงานจาก HR) — หาคอลัมน์ 'ชื่อ'/'นามสกุล'
        (ไทย+อังกฤษ) จากหัวตารางอัตโนมัติ → คืนชื่อเต็ม (ไม่เอาชื่อเล่น เพราะสั้น/โหล
        เสี่ยงบังเอิญตรงผู้ขายอื่น = ซ่อนยอดผิดคน). รองรับหลายชีต + ไฟล์อนาคตรูปแบบเดิม"""
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        out = []
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            # หาแถวหัวตาราง (มีทั้ง 'ชื่อ' และ 'นามสกุล')
            hdr = None
            for row in rows[:15]:
                cells = [str(c).strip() if c is not None else "" for c in row]
                if "ชื่อ" in cells and "นามสกุล" in cells:
                    hdr = cells
                    start = rows.index(row) + 1
                    break
            if hdr is None:
                continue

            def ci(name):
                return hdr.index(name) if name in hdr else None
            c_thf, c_thl = ci("ชื่อ"), ci("นามสกุล")
            c_enf, c_enl = ci("ชื่อ (EN)"), ci("นามสกุล (EN)")
            for row in rows[start:]:
                def g(idx):
                    if idx is None or idx >= len(row) or row[idx] is None:
                        return ""
                    return str(row[idx]).strip()
                th = re.sub(r"\s+", " ", (g(c_thf) + " " + g(c_thl))).strip()
                en = re.sub(r"\s+", " ", (g(c_enf) + " " + g(c_enl))).strip()
                if len(th) >= 2:
                    out.append(th)
                if len(en) >= 3:
                    out.append(en)
        return out

    def _clear_names(self):
        if QMessageBox.question(self, "ยืนยัน", "ล้างรายชื่อทั้งหมด?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                ) == QMessageBox.StandardButton.Yes:
            self.lst.clear()
            self._update_count()

    def _save_names(self):
        names = [re.sub(r"\s+", " ", self.lst.item(i).text()).strip()
                 for i in range(self.lst.count())]
        names = [n for n in names if n]
        data = _sensitive_load()
        data["names"] = names
        _sensitive_save(data)
        # refresh หน้าตัดบิลทันที (ไม่ต้องปิด/เปิดใหม่)
        self._refresh_slip_tab()
        QMessageBox.information(self, "บันทึกแล้ว",
            f"บันทึกรายชื่อ {len(names)} รายการเรียบร้อย\n"
            "ยอดเงินของรายชื่อเหล่านี้จะถูกซ่อนในหน้าตัดบิล (จับคู่สลิป)")

    def _refresh_slip_tab(self):
        """สั่งหน้าตัดบิลของ MainWindow ให้ render ใหม่เพื่อใช้ mask ล่าสุด"""
        try:
            w = self.parent()
            while w is not None and not hasattr(w, "slip_tab"):
                w = w.parent()
            if w is not None and getattr(w, "slip_tab", None) and w.slip_tab._results:
                w.slip_tab._revealed_slip_ids.clear()
                w.slip_tab._render_results()
        except Exception:
            pass

    def _update_count(self):
        self.lbl_count.setText(f"{self.lst.count()} รายชื่อ")


# ──────────────────── Main Window ────────────────────

APP_VERSION = "3.5.1"

# ──────────────────── Auto-Update (GitHub Releases) ────────────────────
# repo ที่เก็บ release (เปลี่ยนได้ผ่าน kcash_config.json คีย์ "update_repo")
UPDATE_REPO_DEFAULT = "dmatmk01-eng/kcash-queue"


def _update_repo() -> str:
    try:
        return (load_config().get("update_repo") or UPDATE_REPO_DEFAULT).strip()
    except Exception:
        return UPDATE_REPO_DEFAULT


def _ver_tuple(v: str):
    """แปลงสตริงเวอร์ชัน เช่น 'v1.10.2' → (1,10,2) เพื่อเทียบมาก/น้อยแบบตัวเลข"""
    nums = re.findall(r"\d+", str(v or ""))
    return tuple(int(n) for n in nums) if nums else (0,)


class UpdateCheckWorker(QThread):
    """เช็คเวอร์ชันล่าสุดจาก GitHub Releases — background, เงียบ ไม่บล็อก UI"""
    found = pyqtSignal(str, str, str)   # version, notes, download_url
    none  = pyqtSignal()
    error = pyqtSignal(str)

    def run(self):
        import urllib.request
        import json as _json
        try:
            repo = _update_repo()
            # ข้ามเฉพาะตอนยังไม่ได้ตั้ง repo (placeholder) — ไม่ใช่ repo จริง
            if not repo or repo.strip().lower() in ("", "owner/repo"):
                self.none.emit(); return
            url = f"https://api.github.com/repos/{repo}/releases/latest"
            headers = {"User-Agent": "KCash-Updater", "Accept": "application/vnd.github+json"}
            tok = ""
            try:
                tok = (load_config().get("update_token") or "").strip()
            except Exception:
                pass
            if tok:
                headers["Authorization"] = f"Bearer {tok}"
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=15) as r:
                data = _json.loads(r.read().decode("utf-8"))
            tag   = (data.get("tag_name") or "").lstrip("vV")
            notes = data.get("body") or ""
            dl = None
            for a in data.get("assets", []):
                if (a.get("name") or "").lower().endswith(".zip"):
                    dl = a.get("browser_download_url"); break
            if not tag or not dl:
                self.none.emit(); return
            if _ver_tuple(tag) > _ver_tuple(APP_VERSION):
                self.found.emit(tag, notes, dl)
            else:
                self.none.emit()
        except Exception as exc:
            self.error.emit(str(exc))


class UpdateDownloadWorker(QThread):
    """ดาวน์โหลดไฟล์อัปเดต (zip) พร้อมรายงาน % — background"""
    progress = pyqtSignal(int, int)   # done_bytes, total_bytes
    done     = pyqtSignal(str)        # path ไฟล์ zip ที่โหลดเสร็จ
    error    = pyqtSignal(str)

    def __init__(self, url, parent=None):
        super().__init__(parent)
        self.url = url

    def run(self):
        import urllib.request
        import tempfile
        try:
            headers = {"User-Agent": "KCash-Updater"}
            tok = ""
            try:
                tok = (load_config().get("update_token") or "").strip()
            except Exception:
                pass
            if tok:
                headers["Authorization"] = f"Bearer {tok}"
            req = urllib.request.Request(self.url, headers=headers)
            tmp = os.path.join(tempfile.gettempdir(), "kcash_update.zip")
            with urllib.request.urlopen(req, timeout=60) as r:
                total = int(r.headers.get("Content-Length") or 0)
                done = 0
                with open(tmp, "wb") as f:
                    while True:
                        chunk = r.read(131072)
                        if not chunk:
                            break
                        f.write(chunk)
                        done += len(chunk)
                        self.progress.emit(done, total)
            self.done.emit(tmp)
        except Exception as exc:
            self.error.emit(str(exc))

# สไตล์โหมดมืด — ครอบส่วนหลักของหน้าจอ (ปุ่มสีเฉพาะที่ตั้ง inline จะคงสีเดิม)
DARK_QSS = """
QMainWindow, QWidget { background-color: #202124; color: #e8eaed; }
QLabel { color: #e8eaed; background: transparent; }
QTableWidget { background-color: #2a2b2e; color: #e8eaed; gridline-color: #3c4043;
               border: 1px solid #3c4043; }
QTableWidget::item:selected { background: #2563eb; color: white; }
QHeaderView::section { background-color: #303134; color: #e8eaed; border: none;
                       border-right: 1px solid #3c4043; padding: 5px; }
QTreeWidget { background-color: #2a2b2e; color: #e8eaed; border: 1px solid #3c4043; }
QTreeWidget::item:selected { background: #2563eb; color: white; }
QLineEdit, QComboBox, QSpinBox, QDateEdit, QPlainTextEdit, QTextEdit, QTextBrowser {
    background-color: #2a2b2e; color: #e8eaed; border: 1px solid #5f6368;
    border-radius: 4px; padding: 2px 4px; }
QComboBox QAbstractItemView { background: #2a2b2e; color: #e8eaed;
    selection-background-color: #2563eb; }
QMenuBar { background-color: #2a2b2e; color: #e8eaed; }
QMenuBar::item:selected { background: #3c4043; }
QMenu { background-color: #2a2b2e; color: #e8eaed; border: 1px solid #5f6368; }
QMenu::item:selected { background: #2563eb; }
QTabWidget::pane { border: 1px solid #3c4043; background: #202124; }
QTabBar::tab { background: #2a2b2e; color: #bdc1c6; padding: 6px 12px; }
QTabBar::tab:selected { background: #202124; color: white; border-bottom: 2px solid #16a34a; }
QStatusBar { background: #2a2b2e; color: #bdc1c6; }
QScrollBar:vertical, QScrollBar:horizontal { background: #2a2b2e; }
QAbstractScrollArea, QTableView { background-color: #2a2b2e; color: #e8eaed; }
QTableCornerButton::section { background: #303134; }
QToolTip { background-color: #2a2b2e; color: #e8eaed; border: 1px solid #5f6368; }
"""

CHANGELOG = [
    {
        "version": "1.0",
        "date": "26/06/2569",
        "title": "เวอร์ชันแรก — อัปเดตใหญ่ระบบจับคู่สลิป + คิวจ่ายเงิน",
        "items": [
            "ระบบจับคู่สลิป: เรียนรู้เลขบัญชีอัตโนมัติ (ยิ่งใช้ยิ่งแม่น)",
            "จับคู่จากเลขบิล/เลขอ้างอิงที่พิมพ์ในสลิป (แม่นสุด)",
            "จับยอดรายการย่อย + ยอดหลังหัก ณ ที่จ่าย",
            "แยกหน้า 'ไม่เจอคู่' ออกจากหน้าหลัก (โชว์เฉพาะเจอคู่)",
            "เพิ่มคอลัมน์ 'รูป' ดูรูปสลิปในโปรแกรมได้ + ปุ่ม 'ดูใน FlowAccount'",
            "ตอนตัดบิล/แนบ มีตัวเลือก 'เฉพาะเจอคู่ / ทั้งหมด'",
            "คิวจ่ายเงิน: เปลี่ยนสถานะจากเซลล์ได้ (จ่ายแล้ว=ลิงก์ FlowAccount จริง)",
            "คลิกเลขเอกสาร = คัดลอกทันที, ค้นหาหลายงาน + ปุ่มค้นหา",
            "ปุ่มไม่อนุมัติ (ซ่อน ย้อนได้) + หน้าดูรายการไม่อนุมัติ (อนุมัติกลับได้)",
            "ปุ่ม 🗑️ ลบ (คอลัมน์สุดท้าย) = ลบเอกสาร EXP/PO/GR ใน FlowAccount จริง (EXP คืนสถานะ PO เป็น 'อนุมัติ')",
            "คอลัมน์ลิงก์แชร์ / ลิงก์แก้ไข, แนบไฟล์(รูป/PDF)เข้า FlowAccount",
            "ตารางคิวจ่าย(หลายวัน): คอลัมน์ชัดเจน, เช็คบ็อกซ์หัววัน, หมายเหตุ, ลิงก์แชร์, ค้นหาในคิว",
            "บังคับบันทึกคิวก่อนปิด, ดีดคิวออกเมื่อพ้น 3 วัน",
            "หน้าประวัติการจัดคิว (ดู/ค้นหา/ลบได้)",
            "Statement Matching: จัดหน้าใหม่ให้ดูง่าย",
            "Export PDF/Excel/Link: เพิ่มคอลัมน์ลิงก์แก้ไขใบ",
            "popup โหลดมีปุ่มปิด, popup เตือนทุกครั้งก่อนเปลี่ยนสถานะที่กระทบ FlowAccount",
            "ยกเลิกใบสั่งซื้อ(PO)จากคิว เหลือ ค่าใช้จ่าย + ใบรับสินค้า(RI) ตามมติประชุม",
            "ซ่อนเอกสารที่แปลงเป็นค่าใช้จ่าย(EXP)แล้ว — กันโชว์ซ้ำ/สถานะดูไม่ตรง (EXP โชว์สถานะจริง)",
            "โหมดมืด (Dark Mode) สลับที่เมนู 'มุมมอง'",
            "เรียงตารางคลิกหัวคอลัมน์ (ยอดเงินเรียงตัวเลขถูก)",
            "ตัดบิล: เพิ่มคอลัมน์แบรนด์ + ปุ่ม Export Excel/PDF + จับคู่รอบ 2 ตามยอด + บันทึกรูปตอนแนบ",
        ],
    },
    {
        "version": "1.5",
        "date": "30/06/2569",
        "title": "ซ่อนยอดเงิน (Sensitive) + อัปเดตอัตโนมัติ + ปรับหน้าตัดบิล",
        "items": [
            "ซ่อนยอดเงินรายชื่อที่กำหนด — กดไอคอนล็อก/ตา ใส่รหัสผ่านเพื่อดู (เฉพาะหน้าตัดบิล)",
            "เมนู 'จัดการผู้ใช้ → จัดการข้อมูลลับ' — ตั้งรหัสผ่าน + นำเข้ารายชื่อ .txt/.csv",
            "อัปเดตอัตโนมัติผ่าน GitHub Releases — โปรแกรมเช็คและดาวน์โหลดเองได้",
            "เมนู 'ช่วยเหลือ → ตรวจหาอัปเดต' สำหรับเช็คด้วยตัวเอง",
            "หน้าตัดบิล: แท็บ 'ทั้งหมด' ย้ายเป็นแท็บแรก + มี popup โหลดตอนนำเข้า PDF",
            "แก้การ Sort ยอดเงินในคิวจ่ายเงินให้เรียงตัวเลขถูกต้อง",
        ],
    },
    {
        "version": "1.8",
        "date": "30/06/2569",
        "title": "บังคับอัปเดตอัตโนมัติ + ปรับขนาดคอลัมน์อิสระ + ปรับปรุงข้อมูลลับ",
        "items": [
            "บังคับอัปเดตอัตโนมัติ — เมื่อมีเวอร์ชันใหม่ โปรแกรมจะแจ้งเตือนและอัปเดตให้ทันที (ไม่ต้องส่งไฟล์เอง)",
            "ตอนอัปเดต เก็บข้อมูลผู้ใช้/รหัสผ่าน/ตั้งค่าทั้งหมดไว้ — ไม่ต้องสมัครหรือตั้งค่าใหม่",
            "ทุกตาราง (คิวจ่ายเงิน/ตัดบิล/Statement) ลากปรับขนาดคอลัมน์ได้อิสระทุกช่อง",
            "ข้อมูลลับ: จับคู่ชื่อยืดหยุ่นขึ้น (ไม่ติดคำนำหน้า/ช่องว่าง) + กดบันทึกแล้วซ่อนทันที",
            "เพิ่มปุ่ม 'รีเซ็ตรหัสผ่าน' ในหน้าจัดการข้อมูลลับ",
            "รายชื่อข้อมูลลับแสดงบรรทัดเดียว ไม่ตัดบรรทัด",
        ],
    },
    {
        "version": "1.9",
        "date": "01/07/2569",
        "title": "เปิดมาโชว์ข้อมูลทั้งหมดก่อน (ไม่กรองวันที่อัตโนมัติ)",
        "items": [
            "เปิดโปรแกรม/ดึงข้อมูล → โชว์รายการทั้งหมดทันที เรียงล่าสุด→เก่าสุด (ไม่กรองวันที่)",
            "ตัวกรองวันที่จะทำงานเมื่อผู้ใช้เลือกวันเอง หรือกดปุ่ม 'เดือนนี้' เท่านั้น",
            "เพิ่มปุ่ม '🗓️ ทุกวันที่' — ยกเลิกกรองวันที่ กลับไปดูทั้งหมด",
            "แก้ปัญหาต้นเดือนแล้วตารางว่าง (เพราะเดิมกรองเป็นวันนี้วันเดียวอัตโนมัติ)",
        ],
    },
    {
        "version": "2.0",
        "date": "01/07/2569",
        "title": "แก้ระบบอัปเดตอัตโนมัติให้ทำงานจริง",
        "items": [
            "แก้บั๊กที่ทำให้โปรแกรมไม่เคยเช็คอัปเดตจริง (เด้งว่า 'ล่าสุดแล้ว' ตลอด)",
            "ตั้งแต่เวอร์ชันนี้เป็นต้นไป จะเช็ค + บังคับอัปเดตอัตโนมัติจาก GitHub ได้จริง",
        ],
    },
    {
        "version": "2.1.1",
        "date": "01/07/2569",
        "title": "ทดสอบระบบอัปเดตอัตโนมัติ",
        "items": [
            "เวอร์ชันทดสอบ — ยืนยันว่าระบบบังคับอัปเดตอัตโนมัติจาก GitHub ทำงานครบวงจร",
            "เก็บข้อมูลผู้ใช้/รหัสผ่าน/ตั้งค่าไว้ครบตอนอัปเดต",
        ],
    },
    {
        "version": "2.1.2",
        "date": "01/07/2569",
        "title": "แก้ตัวติดตั้งอัปเดต (robocopy เขียนทับไฟล์ไม่สำเร็จ)",
        "items": [
            "แก้บั๊กที่ทำให้อัปเดตแล้วไฟล์ไม่ถูกเขียนทับจริง (ยังเป็นเวอร์ชันเดิม)",
            "สคริปต์อัปเดตเปิดหน้าต่างจริง + รอโปรแกรมปิดสนิทก่อนก๊อปไฟล์ + เขียน log ไว้ตรวจสอบ",
            "ตั้งแต่เวอร์ชันนี้ อัปเดตอัตโนมัติเขียนทับไฟล์ได้สำเร็จจริง",
        ],
    },
    {
        "version": "2.1.3",
        "date": "01/07/2569",
        "title": "ทดสอบอัปเดตอัตโนมัติ (ตัวติดตั้งที่แก้แล้ว)",
        "items": [
            "เวอร์ชันทดสอบ — ยืนยันว่าอัปเดตจาก v2.1.2 เขียนทับไฟล์เป็นเวอร์ชันใหม่ได้จริง",
            "เก็บข้อมูลผู้ใช้/รหัสผ่าน/ตั้งค่าครบตอนอัปเดต",
        ],
    },
    {
        "version": "2.2.0",
        "date": "01/07/2569",
        "title": "จับคู่สลิปแม่นขึ้น — เลิกเดา ลดจับผิดใบ",
        "items": [
            "เลิกใช้ 'รหัส/ชื่อโปรเจกต์' มาจับคู่ (ใช้ร่วมหลายบิล ทำให้จับผิดใบ) — แก้เคสยอดเท่าแต่คนละบิล",
            "จับ 'เขียว (บันทึกออโต้ได้)' เฉพาะเมื่อมีกุญแจเฉพาะจริง: เลขบิล/เลข PO ในสลิป หรือบัญชีที่ผูกผู้ขายรายเดียว",
            "ยอดเท่ากันหลายใบ + ชื่อแยกไม่ออก → ไม่เดา ขึ้น 'ต้องยืนยัน' ให้เลือกเอง",
            "กันบัญชี 'คนกลาง' (1 บัญชีจ่ายหลายผู้ขาย) ไม่ให้จับเขียวอัตโนมัติ",
            "ผลรวม: สิ่งที่ขึ้นเขียว = เชื่อถือได้ ส่วนที่ไม่ชัวร์ให้คนยืนยันก่อน ไม่บันทึกผิดเอง",
        ],
    },
    {
        "version": "2.3.0",
        "date": "01/07/2569",
        "title": "ตัดบิลจากรูปภาพได้ (อ่านแท็ก ai#EXP อัตโนมัติ)",
        "items": [
            "รองรับสลิปเป็น 'รูปภาพ' (PNG/JPG) ไม่ใช่แค่ PDF — เลือกได้หลายรูปพร้อมกัน",
            "ปุ่มเปลี่ยนเป็น 'เลือกไฟล์สลิป PDF/รูปภาพ'",
            "อ่านแท็ก 'ai#EXP<เลข>' ที่บัญชีเขียนในรูปด้วย OCR (Windows) → จับคู่เลขเอกสารตรงเป๊ะอัตโนมัติ",
            "อ่านยอดเงิน/วันที่จากรูปให้ด้วย + ใช้รูปเป็นหลักฐานแนบเข้า FlowAccount",
            "ถ้า OCR อ่านไม่ได้/ไม่เจอเลข → ขึ้น 'ต้องยืนยัน' ให้จับคู่เอง (ไม่เดา)",
        ],
    },
    {
        "version": "2.3.1",
        "date": "01/07/2569",
        "title": "ตัดบิลรูปภาพ: ใช้เฉพาะเลข ai#EXP จับคู่ (กันข้อความอื่นกวน)",
        "items": [
            "แก้ปัญหารูปที่มีข้อความเยอะ (amount/transfer/เลขธุรกรรม) ทำให้จับผิด/ไม่เจอ",
            "ตอนนี้ใช้ 'เฉพาะเลข ai#EXP' เป็นกุญแจจับคู่ ตัดข้อความอื่นทิ้ง",
            "ทนต่อ OCR อ่านเพี้ยนมากขึ้น (ช่องว่างแทรก / O↔0 / I↔1)",
        ],
    },
    {
        "version": "2.3.2",
        "date": "01/07/2569",
        "title": "ข้อมูลลับ: Import รายชื่อจากไฟล์ Excel (.xlsx) ได้",
        "items": [
            "หน้าจัดการข้อมูลลับ Import ไฟล์ Excel (.xlsx) ได้แล้ว (นอกจาก .csv/.txt)",
            "อ่านคอลัมน์ 'ชื่อ'/'นามสกุล' (ไทย+อังกฤษ) จากหัวตารางอัตโนมัติ",
            "รองรับไฟล์รายชื่อพนักงานจากระบบ HR — อัปเดตรายชื่อใหม่แล้ว import ทับได้เลย",
            "ข้ามชื่อเล่น (สั้น/โหล เสี่ยงซ่อนยอดผิดคน) — ใช้ชื่อเต็มเท่านั้น",
        ],
    },
    {
        "version": "2.3.3",
        "date": "04/07/2569",
        "title": "จับเลข EXP ในสลิปแม่นขึ้น — แก้กรณีพิมพ์เพี้ยน (EXPO=EXP0)",
        "items": [
            "บัญชีเขียนเลข EXP ในรายละเอียดสลิป (PDF) → จับคู่เลขเอกสารตรงเป๊ะ = เขียว",
            "แก้กรณีพิมพ์ตัวอักษรแทนเลข เช่น 'EXPO38917' (โอ) = 'EXP038917' (ศูนย์) ให้จับเจอ",
            "ใช้เลข EXP เป็นตัวหลัก (สำคัญกว่าเลขใบวางบิลที่ใช้ร่วมหลายบิล)",
        ],
    },
    {
        "version": "3.5",
        "date": "04/07/2569",
        "title": "เวอร์ชันเสถียร — รวมทุกฟีเจอร์/การแก้ไขล่าสุด",
        "items": [
            "อัปเดตอัตโนมัติผ่าน GitHub (บังคับอัปเดต เก็บข้อมูลผู้ใช้/ตั้งค่าครบ)",
            "จับคู่สลิปแม่นขึ้น — จับด้วยเลข EXP/บิล/PO เป็นหลัก ไม่เดามั่ว",
            "ตัดบิลจากรูปภาพได้ (อ่านแท็ก ai#EXP ด้วย OCR)",
            "ซ่อนยอดเงิน (Sensitive) + Import รายชื่อพนักงานจาก Excel",
            "ปรับขนาดคอลัมน์อิสระ + เปิดมาโชว์ข้อมูลทั้งหมดก่อน",
        ],
    },
    {
        "version": "3.5.1",
        "date": "04/07/2569",
        "title": "ตัดบิล+แนบ: ตั้งวันที่เอกสาร = วันที่แนบ + เครดิต 0",
        "items": [
            "ตอน 'ตัดบิล + แนบเข้า FlowAccount' → ตั้งวันที่เอกสารและวันครบกำหนด = วันที่แนบ (วันนี้)",
            "ตั้งเครดิต (วัน) = 0",
            "คงยอด/รายการ/VAT เดิมไว้ครบ — เปลี่ยนแค่วันที่/เครดิต",
            "ถ้าแก้วันที่ไม่สำเร็จ → แจ้งเตือน (ไม่กระทบการแนบ/บันทึกจ่าย)",
        ],
    },
]


class ChangelogDialog(QDialog):
    """หน้าต่างแสดงเวอร์ชัน + รายการที่เพิ่ม/แก้ไข (changelog)"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"KCash Queue System — เวอร์ชัน {APP_VERSION}")
        self.resize(600, 560)
        lay = QVBoxLayout(self)

        head = QLabel(f"KCash Queue System  <span style='color:#16a34a'>v{APP_VERSION}</span>")
        head.setTextFormat(Qt.TextFormat.RichText)
        head.setStyleSheet("font-size:20px;font-weight:800;color:#0f172a;")
        lay.addWidget(head)
        sub = QLabel("ประวัติการอัปเดต / สิ่งที่เพิ่มและแก้ไข")
        sub.setStyleSheet("color:#64748b;font-size:12px;")
        lay.addWidget(sub)

        from PyQt6.QtWidgets import QScrollArea, QTextBrowser
        tb = QTextBrowser()
        tb.setOpenExternalLinks(False)
        parts = []
        for rel in CHANGELOG:
            parts.append(
                f"<div style='margin-top:10px'>"
                f"<span style='background:#dcfce7;color:#15803d;font-weight:700;"
                f"padding:2px 10px;border-radius:10px;'>v{rel['version']}</span> "
                f"<span style='color:#64748b'>&nbsp;{rel['date']}</span><br>"
                f"<b style='color:#0f172a'>{rel.get('title','')}</b></div>")
            parts.append("<ul style='margin-top:6px;line-height:1.6'>")
            for it in rel["items"]:
                parts.append(f"<li>{it}</li>")
            parts.append("</ul>")
        tb.setHtml("".join(parts))
        tb.setStyleSheet("QTextBrowser{border:1px solid #e2e8f0;border-radius:6px;"
                         "font-size:13px;padding:6px;}")
        lay.addWidget(tb, 1)

        btn = QPushButton("ปิด")
        btn.setStyleSheet("QPushButton{padding:6px 18px;border:1px solid #cbd5e1;"
                          "border-radius:5px;background:white;}")
        btn.clicked.connect(self.accept)
        row = QHBoxLayout(); row.addStretch(); row.addWidget(btn)
        lay.addLayout(row)


class ThemeSwitch(QWidget):
    """ปุ่มสลับธีมแบบ pill (สว่าง↔มืด) มีไอคอนพระอาทิตย์/พระจันทร์ + knob สีน้ำเงินเลื่อนได้"""
    toggled = pyqtSignal(bool)   # True = โหมดมืด

    def __init__(self, dark=False, parent=None):
        super().__init__(parent)
        self._dark = bool(dark)
        self.setFixedSize(62, 28)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setToolTip("สลับโหมดสว่าง / มืด")

    # ── API ให้เข้ากับ QAction เดิม ──
    def isChecked(self):
        return self._dark

    def setChecked(self, on, emit=False):
        on = bool(on)
        if on == self._dark:
            return
        self._dark = on
        self.update()
        if emit:
            self.toggled.emit(self._dark)

    def mousePressEvent(self, e):
        if e.button() == Qt.MouseButton.LeftButton:
            self._dark = not self._dark
            self.update()
            self.toggled.emit(self._dark)
        super().mousePressEvent(e)

    def paintEvent(self, e):
        from PyQt6.QtGui import QPainter, QColor, QPen, QPainterPath
        from PyQt6.QtCore import QPointF, QRectF
        p = QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        r = QRectF(self.rect()).adjusted(1, 1, -1, -1)
        rad = r.height() / 2
        dark = self._dark
        # ราง
        track  = QColor("#16181d") if dark else QColor("#ffffff")
        border = QColor("#3c4043") if dark else QColor("#cbd5e1")
        p.setPen(QPen(border, 1)); p.setBrush(track)
        p.drawRoundedRect(r, rad, rad)
        # ตำแหน่ง knob ซ้าย(สว่าง)/ขวา(มืด)
        d  = r.height() - 6
        cy = r.center().y()
        lx = r.left() + 3 + d / 2
        rx = r.right() - 3 - d / 2
        # knob สีน้ำเงิน
        p.setPen(Qt.PenStyle.NoPen); p.setBrush(QColor("#2563eb"))
        p.drawEllipse(QPointF(rx if dark else lx, cy), d / 2, d / 2)
        # ไอคอน: พระอาทิตย์(ซ้าย) / พระจันทร์(ขวา) — ขาวเมื่ออยู่ใต้ knob, เทาเมื่อปิด
        muted = QColor("#5f6368") if dark else QColor("#94a3b8")
        self._sun(p,  lx, cy, QColor("#ffffff") if not dark else muted)
        self._moon(p, rx, cy, QColor("#ffffff") if dark else muted)
        p.end()

    def _sun(self, p, cx, cy, col):
        from PyQt6.QtGui import QPen
        from PyQt6.QtCore import QPointF
        import math
        p.setPen(Qt.PenStyle.NoPen); p.setBrush(col)
        p.drawEllipse(QPointF(cx, cy), 3.0, 3.0)
        p.setPen(QPen(col, 1.4, cap=Qt.PenCapStyle.RoundCap))
        for i in range(8):
            a = math.radians(i * 45)
            dx, dy = math.cos(a), math.sin(a)
            p.drawLine(QPointF(cx + dx * 5.0, cy + dy * 5.0),
                       QPointF(cx + dx * 6.6, cy + dy * 6.6))

    def _moon(self, p, cx, cy, col):
        from PyQt6.QtGui import QPainterPath
        from PyQt6.QtCore import QPointF
        full = QPainterPath(); full.addEllipse(QPointF(cx, cy), 5.4, 5.4)
        cut  = QPainterPath(); cut.addEllipse(QPointF(cx + 2.6, cy - 1.6), 5.0, 5.0)
        p.setPen(Qt.PenStyle.NoPen); p.setBrush(col)
        p.drawPath(full.subtracted(cut))


class MainWindow(QMainWindow):
    def __init__(self, current_user=None):
        super().__init__()
        self.current_user = current_user or {"username": "", "fullname": "", "nickname": "", "role": "user"}
        self._logout_requested = False
        role = self.current_user.get("role", "user")
        title = f"KCash Queue System  v{APP_VERSION}"
        if role == "dev":
            title += "  —  Dev Mode"
        self.setWindowTitle(title)
        self.setMinimumSize(1100, 650)
        # ตั้งไอคอนหน้าต่าง
        icon_path = _resource_path("icon/icon.ico")
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
        self._build_ui()
        self._check_config()

    def _build_ui(self):
        self.setStyleSheet(f"""
            QMainWindow, QWidget {{ background: {C_BG}; font-family: 'Segoe UI', 'Noto Sans Thai'; }}
            QTabWidget::pane {{ border: 1px solid {C_BORDER}; background: {C_SURFACE}; }}
            QTabBar::tab {{
                padding: 8px 20px; font-size: 13px;
                border: 1px solid {C_BORDER}; border-bottom: none;
                border-radius: 4px 4px 0 0; margin-right: 2px;
                background: #f1f5f9;
            }}
            QTabBar::tab:selected {{ background: {C_SURFACE}; font-weight: 600; color: {C_PRIMARY}; }}
            QGroupBox {{ border: 1px solid {C_BORDER}; border-radius: 6px; margin-top: 8px; padding: 8px; }}
            QGroupBox::title {{ subcontrol-origin: margin; left: 10px; }}
            QLineEdit, QComboBox, QDateEdit {{
                border: 1px solid {C_BORDER}; border-radius: 4px; padding: 5px 8px; font-size: 13px;
                background: white;
            }}
            QLineEdit:focus, QComboBox:focus, QDateEdit:focus {{
                border-color: {C_PRIMARY};
            }}
            QPushButton {{
                border: 1px solid {C_BORDER}; border-radius: 4px;
                padding: 5px 12px; font-size: 13px; background: white;
            }}
            QPushButton:hover {{ background: #f1f5f9; }}
            QHeaderView::section {{
                background: #f1f5f9; padding: 5px 8px;
                border: none; border-right: 1px solid {C_BORDER}; border-bottom: 1px solid {C_BORDER};
                font-size: 12px; font-weight: 600; color: {C_MUTED};
            }}
            QStatusBar {{ background: #f1f5f9; color: {C_MUTED}; font-size: 11px; }}
        """)

        # Menubar
        role = self.current_user.get("role", "user")
        menubar = self.menuBar()
        file_menu = menubar.addMenu("ไฟล์")
        # ⚙️ ตั้งค่า — เฉพาะ dev เท่านั้น (admin/user มองไม่เห็นและแก้ไม่ได้)
        if role == "dev":
            act_settings = QAction("⚙️ ตั้งค่า", self)
            act_settings.triggered.connect(self._open_settings)
            file_menu.addAction(act_settings)
            file_menu.addSeparator()
        act_clear_links = QAction("🧹 ล้างลิงก์ Export เก่าทั้งหมด", self)
        act_clear_links.triggered.connect(self._clear_export_links)
        file_menu.addAction(act_clear_links)
        file_menu.addSeparator()
        act_quit = QAction("ออกจากโปรแกรม", self)
        act_quit.triggered.connect(self.close)
        file_menu.addAction(act_quit)

        # เมนูจัดการผู้ใช้ (เห็นเฉพาะ admin/dev)
        if role in ("admin", "dev"):
            user_menu = menubar.addMenu("จัดการผู้ใช้")
            act_users = QAction("👥 จัดการผู้ใช้ (สมัคร/แก้ไข/ลบ)", self)
            act_users.triggered.connect(self._open_user_manager)
            user_menu.addAction(act_users)
            user_menu.addSeparator()
            act_sensitive = QAction("🔒 จัดการข้อมูลลับ (ซ่อนยอดเงิน)", self)
            act_sensitive.triggered.connect(self._open_sensitive_manager)
            user_menu.addAction(act_sensitive)

        # หมายเหตุ: ปุ่มสลับ Dark Mode ย้ายไปไว้มุมขวาบน (ThemeSwitch) แล้ว

        help_menu = menubar.addMenu("ช่วยเหลือ")
        act_about = QAction("เกี่ยวกับ KCash", self)
        act_about.triggered.connect(self._about)
        help_menu.addAction(act_about)
        act_update = QAction("🔄 ตรวจหาอัปเดต", self)
        act_update.triggered.connect(lambda: self._check_for_updates(manual=True))
        help_menu.addAction(act_update)

        # มุมขวาบนของ menubar: badge Dev Mode (เฉพาะ dev) + ปุ่มออกจากระบบ
        corner = QWidget()
        ch = QHBoxLayout(corner)
        ch.setContentsMargins(0, 0, 8, 0)
        ch.setSpacing(8)
        # ── ปุ่มสลับธีม (สว่าง/มืด) แบบ switch — อยู่หน้าสุด ──
        try:
            _dm = bool(load_config().get("dark_mode", False))
        except Exception:
            _dm = False
        self.theme_switch = ThemeSwitch(dark=_dm)
        self.theme_switch.toggled.connect(self._toggle_dark)
        ch.addWidget(self.theme_switch)
        # ป้ายเวอร์ชัน — กดดู changelog ได้
        btn_ver = QPushButton(f"v{APP_VERSION}")
        btn_ver.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_ver.setToolTip("กดดูว่ามีอะไรเพิ่ม/แก้ไขในเวอร์ชันนี้")
        btn_ver.setStyleSheet(
            "QPushButton{color:#15803d;background:#dcfce7;border:1px solid #16a34a;"
            "border-radius:10px;padding:2px 12px;font-weight:700;font-size:12px;}"
            "QPushButton:hover{background:#bbf7d0;}")
        btn_ver.clicked.connect(lambda: ChangelogDialog(self).exec())
        ch.addWidget(btn_ver)
        if role == "dev":
            badge = QLabel("●  Dev Mode")
            badge.setStyleSheet(
                "color:white;background:#dc2626;border-radius:10px;"
                "padding:2px 12px;font-weight:700;font-size:12px;")
            ch.addWidget(badge)
        btn_logout = QPushButton("⏻ ออกจากระบบ")
        btn_logout.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_logout.setStyleSheet(
            "QPushButton{color:white;background:#dc2626;border:none;border-radius:10px;"
            "padding:3px 14px;font-weight:700;font-size:12px;}"
            "QPushButton:hover{background:#b91c1c;}")
        btn_logout.clicked.connect(self._logout)
        ch.addWidget(btn_logout)
        menubar.setCornerWidget(corner, Qt.Corner.TopRightCorner)

        # Tabs (ใบสั่งซื้อ + ใบรับสินค้า รวมเข้าหน้าคิวจ่ายเงินแล้ว)
        self.tabs      = QTabWidget()
        self.queue_tab = QueueTab()
        self.stmt_tab  = StatementTab()
        self.slip_tab  = SlipMatchTab()
        self.log_tab   = LogTab()

        self.tabs.addTab(self.queue_tab, "📋  คิวจ่ายเงิน")
        self.tabs.addTab(self.stmt_tab,  "🏦  Statement Matching")
        self.tabs.addTab(self.slip_tab,  "✂️  ตัดบิล (จับคู่สลิป)")
        self.tabs.addTab(self.log_tab,   "📜  Log")

        # Dev Console — เฉพาะ dev mode (พับเก็บได้)
        self.dev_console = None
        if role == "dev":
            container = QWidget()
            cv = QVBoxLayout(container)
            cv.setContentsMargins(0, 0, 0, 0)
            cv.setSpacing(0)
            cv.addWidget(self.tabs, 1)
            self.dev_console = DevConsole()
            cv.addWidget(self.dev_console)
            self.setCentralWidget(container)
        else:
            self.setCentralWidget(self.tabs)

        # Status bar
        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self.status.showMessage("ยินดีต้อนรับสู่ KCash Queue System")
        # ใช้ธีมตามที่บันทึกไว้ (โหมดมืด/สว่าง)
        if self.theme_switch.isChecked():
            self._apply_theme(True)
        if self.dev_console:
            # ดักจับ stdout/stderr + exception ทั้งหมดเข้า console
            sys.stdout = _StreamTee(sys.__stdout__, "info")
            sys.stderr = _StreamTee(sys.__stderr__, "err")
            _orig_hook = sys.excepthook
            def _hook(et, ev, tb, _o=_orig_hook):
                import traceback as _tb
                dev_log("".join(_tb.format_exception(et, ev, tb)), "err")
                _o(et, ev, tb)
            sys.excepthook = _hook
            dev_log(f"เข้าสู่ Dev Mode — ผู้ใช้ {self.current_user.get('nickname','')}", "ok")
            dev_log("Dev Console พร้อมใช้งาน (กดหัวข้อเพื่อพับเก็บ)", "info")

        # บันทึก Log การเปิดโปรแกรม + เดือนปีปัจจุบัน (requirement ข้อ 6)
        now = datetime.now()
        activity_log.log("เปิดโปรแกรม",
                         f"เข้าใช้งานเดือน {now.month}/{now.year + 543} (ปัจจุบัน)")

        # Wire signals
        self.queue_tab.status_message.connect(self.status.showMessage)
        self.stmt_tab.matched_updated.connect(self.queue_tab.set_matched)
        self.slip_tab.status_message.connect(self.status.showMessage)
        # ส่งสถานะทุกแท็บเข้า Dev Console ด้วย
        if self.dev_console:
            for tab in (self.queue_tab, self.slip_tab):
                tab.status_message.connect(lambda m: dev_log(m, "info"))
        self.queue_tab.company_changed.connect(self._on_company_switched)
        self.tabs.currentChanged.connect(self._on_tab_changed)

        # monday.com auto-sync timer — ทุก 5 นาที (มี cache+hash → กิน quota น้อย)
        self._monday_busy = False        # flag ป้องกัน sync ซ้อน
        self._monday_timer = QTimer(self)
        self._monday_timer.timeout.connect(self._check_monday_schedule)
        self._monday_timer.start(5 * 60_000)   # ทุก 5 นาที

        # ตรวจหาอัปเดตอัตโนมัติตอนเปิดโปรแกรม (เงียบ ๆ เบื้องหลัง หลัง UI พร้อม)
        QTimer.singleShot(2500, lambda: self._check_for_updates(manual=False))

    def _on_tab_changed(self, idx: int):
        # 0=คิวจ่ายเงิน 1=Statement 2=ตัดบิล 3=Log
        if idx == 1:
            self.stmt_tab.set_expenses(self.queue_tab.get_expenses())
        if idx == 2:
            self.slip_tab.set_expenses(self.queue_tab.get_expenses())
        if idx == 3:
            self.log_tab.reload()

    def _on_company_switched(self):
        """สลับบริษัท → ล้างข้อมูลแคชของแท็บอื่น ให้ดึงใหม่ของบริษัทนั้น"""
        self.slip_tab.set_expenses([])

    def _check_monday_schedule(self):
        """Auto-sync monday.com ทุก 5 นาที — เงียบ ๆ ไม่มี popup"""
        cfg = load_config()
        if not cfg.get("monday_auto_sync"):                    return
        if not cfg.get("monday_api_token"):                    return
        if not cfg.get("monday_board_id"):                     return
        if self._monday_busy:                                  return  # sync อยู่ ไม่ทำซ้อน

        expenses = self.queue_tab.get_expenses()
        if not expenses:
            return   # ยังไม่ได้ดึงข้อมูล ไม่ต้อง sync

        # เติมแบรนด์ก่อนส่ง
        from brand_assignments import load_assignments as _la
        assignments = _la()
        unpaid = []
        for e in expenses:
            if _status(e) == "paid":
                continue
            ec = dict(e)
            ec["_brand_name"] = _brand_name(e, assignments)
            unpaid.append(ec)

        if not unpaid:
            return

        self._monday_busy = True
        now = datetime.now().strftime("%H:%M")
        self.status.showMessage(f"⏰ {now} — Auto-sync → monday.com ({len(unpaid)} รายการ)...")

        worker = MondaySyncWorker(cfg, unpaid)
        worker.done.connect(self._on_auto_sync_done)
        worker.error.connect(self._on_auto_sync_error)
        worker.start()
        self._monday_bg = worker  # keep reference

    def _on_auto_sync_done(self, r):
        self._monday_busy = False
        c = r.get("created", 0)
        u = r.get("updated", 0)
        s = r.get("skipped", 0)
        now = datetime.now().strftime("%H:%M")
        if c == 0 and u == 0:
            self.status.showMessage(f"✅ {now} — Auto-sync: ไม่มีรายการเปลี่ยน (ข้าม {s})")
        else:
            self.status.showMessage(
                f"✅ {now} — Auto-sync เสร็จ: สร้าง {c}, อัพเดต {u}, ข้าม {s}")
        # Smart Auto-Refresh — ตรวจ error → refresh cache ถ้าจำเป็น
        self.queue_tab._maybe_smart_refresh(r.get("errors", []))

    def _on_auto_sync_error(self, msg):
        self._monday_busy = False
        now = datetime.now().strftime("%H:%M")
        # ตัดข้อความให้สั้นเข้า status bar
        short = msg.split("\n")[0][:80]
        self.status.showMessage(f"⚠️ {now} — Auto-sync ผิดพลาด: {short}")

    # ──────────── Auto-Update ────────────
    def _check_for_updates(self, manual=False):
        """ตรวจหาอัปเดตจาก GitHub Releases. manual=True → กดจากเมนู (โชว์ผลทุกกรณี)"""
        # อัปเดตได้เฉพาะตอนรันจาก .exe (frozen) — ตอน dev รันซอร์สไม่ต้อง
        if not getattr(sys, "frozen", False):
            if manual:
                QMessageBox.information(self, "ตรวจหาอัปเดต",
                    "โหมดพัฒนา (รันจากซอร์ส) ไม่ต้องอัปเดต — "
                    "ระบบอัปเดตอัตโนมัติทำงานเฉพาะตอนรันจากไฟล์ติดตั้ง (.exe)")
            return
        if getattr(self, "_upd_busy", False):
            return
        self._upd_busy = True
        self._upd_manual = manual
        if manual:
            self.status.showMessage("🔄 กำลังตรวจหาอัปเดต...")
        self._upd_check = UpdateCheckWorker(self)
        self._upd_check.found.connect(self._on_update_found)
        self._upd_check.none.connect(self._on_update_none)
        self._upd_check.error.connect(self._on_update_error)
        self._upd_check.start()

    def _on_update_none(self):
        self._upd_busy = False
        if getattr(self, "_upd_manual", False):
            QMessageBox.information(self, "ตรวจหาอัปเดต",
                f"คุณใช้เวอร์ชันล่าสุดอยู่แล้ว (v{APP_VERSION})")
        else:
            self.status.showMessage(f"✅ ใช้เวอร์ชันล่าสุดอยู่แล้ว (v{APP_VERSION})")

    def _on_update_error(self, msg):
        self._upd_busy = False
        if getattr(self, "_upd_manual", False):
            QMessageBox.warning(self, "ตรวจหาอัปเดตไม่สำเร็จ",
                f"เชื่อมต่อเซิร์ฟเวอร์อัปเดตไม่ได้\n\n{msg}")
        else:
            self.status.showMessage("⚠️ ตรวจหาอัปเดตไม่สำเร็จ (เน็ต/เซิร์ฟเวอร์)")

    def _on_update_found(self, version, notes, url):
        self._upd_busy = False
        note_txt = (notes or "").strip()
        if len(note_txt) > 600:
            note_txt = note_txt[:600] + "..."
        # บังคับอัปเดต — แจ้งเตือนแล้วอัปเดตให้ทันที (ปุ่มเดียว ปิดหน้าต่างไม่ได้)
        box = QMessageBox(self)
        box.setWindowTitle("จำเป็นต้องอัปเดต")
        box.setIcon(QMessageBox.Icon.Information)
        box.setText(
            f"มีเวอร์ชันใหม่ <b>v{version}</b> (ปัจจุบัน v{APP_VERSION})<br>"
            "ต้องอัปเดตเป็นเวอร์ชันล่าสุดก่อนใช้งาน<br>"
            "<span style='color:#16a34a'>ข้อมูลผู้ใช้/รหัสผ่าน/ตั้งค่าทั้งหมดจะถูกเก็บไว้</span>")
        if note_txt:
            box.setInformativeText("สิ่งที่เปลี่ยน:\n" + note_txt)
        box.setStandardButtons(QMessageBox.StandardButton.NoButton)
        box.addButton("⬇️ อัปเดตเดี๋ยวนี้", QMessageBox.ButtonRole.AcceptRole)
        # ถอดปุ่ม X เพื่อบังคับให้กดอัปเดต (กันปิดเลี่ยง)
        box.setWindowFlag(Qt.WindowType.WindowCloseButtonHint, False)
        box.setWindowFlag(Qt.WindowType.WindowContextHelpButtonHint, False)
        box.exec()
        # บังคับ: ไม่ว่าจะปิดด้วยวิธีใด → เริ่มดาวน์โหลด/อัปเดตเสมอ
        self._start_update_download(version, url)

    def _start_update_download(self, version, url):
        dlg = LoadingDialog(f"กำลังดาวน์โหลดเวอร์ชัน v{version}...", self)
        dlg.set_total(100)

        def _on_prog(done, total):
            if total > 0:
                dlg.set_progress(done, total,
                    f"ดาวน์โหลด {done // (1024*1024)} / {total // (1024*1024)} MB")
            else:
                dlg.set_indeterminate(f"ดาวน์โหลด {done // (1024*1024)} MB...")

        self._upd_dl = UpdateDownloadWorker(url, self)
        self._upd_dl.progress.connect(_on_prog)
        self._upd_dl.done.connect(lambda p: (dlg.accept(), self._apply_update(p)))
        self._upd_dl.error.connect(lambda m: (dlg.reject(),
            QMessageBox.warning(self, "ดาวน์โหลดไม่สำเร็จ", m)))
        self._upd_dl.start()
        dlg.exec()

    def _apply_update(self, zip_path):
        """แตกไฟล์ zip → เขียนสคริปต์อัปเดต → ปิดโปรแกรม → ก๊อปทับ → เปิดใหม่"""
        import zipfile
        import tempfile
        import shutil
        try:
            app_dir = os.path.dirname(sys.executable)   # โฟลเดอร์ที่ติดตั้งอยู่
            stage = os.path.join(tempfile.gettempdir(), "kcash_update_stage")
            if os.path.isdir(stage):
                shutil.rmtree(stage, ignore_errors=True)
            os.makedirs(stage, exist_ok=True)
            with zipfile.ZipFile(zip_path) as z:
                z.extractall(stage)
            # หาโฟลเดอร์ที่มีไฟล์ .exe (เผื่อ zip มีโฟลเดอร์ครอบ)
            new_dir = stage
            exe_name = os.path.basename(sys.executable)
            if not os.path.isfile(os.path.join(stage, exe_name)):
                for root, _dirs, files in os.walk(stage):
                    if exe_name in files:
                        new_dir = root
                        break
            if not os.path.isfile(os.path.join(new_dir, exe_name)):
                QMessageBox.warning(self, "อัปเดตไม่สำเร็จ",
                    "ไฟล์อัปเดตไม่ถูกต้อง (ไม่พบไฟล์โปรแกรมข้างใน)")
                return
        except Exception as exc:
            QMessageBox.warning(self, "อัปเดตไม่สำเร็จ", f"แตกไฟล์ไม่สำเร็จ:\n{exc}")
            return

        # ไฟล์ข้อมูลผู้ใช้ที่ต้องไม่เขียนทับ (robocopy /XF)
        keep = ("kcash_config.json kcash_users.dat kcash_remember.dat "
                "kcash_sensitive.json account_memory.json kcash_remarks.json "
                "queue_plan.json rejected.json paid_pending.json activity_log.json "
                "queue_log.json share_links.json")
        bat = os.path.join(tempfile.gettempdir(), "kcash_update.bat")
        log = os.path.join(tempfile.gettempdir(), "kcash_update_log.txt")
        exe_full = os.path.join(app_dir, exe_name)
        # ใช้ ping แทน timeout (timeout ต้องมี console/stdin ไม่งั้น error)
        # เปิดใน console จริง (CREATE_NEW_CONSOLE) เพื่อให้ robocopy/ping ทำงานปกติ
        bat_src = (
            "@echo off\r\n"
            "chcp 65001 >nul\r\n"
            "title KCash Update - do not close\r\n"
            f'echo === KCash update start %date% %time% === > "{log}"\r\n'
            "echo Updating KCash Queue System... please wait (do not close this window)\r\n"
            ":waitloop\r\n"
            f'tasklist /FI "IMAGENAME eq {exe_name}" 2>nul | find /I "{exe_name}" >nul\r\n'
            "if not errorlevel 1 (\r\n"
            "  ping -n 2 127.0.0.1 >nul\r\n"
            "  goto waitloop\r\n"
            ")\r\n"
            "ping -n 3 127.0.0.1 >nul\r\n"          # เผื่อ handle ไฟล์ยังไม่ปลดล็อกทันที
            f'echo copying %date% %time% >> "{log}"\r\n'
            f'robocopy "{new_dir}" "{app_dir}" /E /R:5 /W:1 /XF {keep} >> "{log}" 2>&1\r\n'
            f'echo robocopy exit=%errorlevel% >> "{log}"\r\n'
            f'start "" "{exe_full}"\r\n'
            f'echo relaunched %date% %time% >> "{log}"\r\n'
            f'rmdir /S /Q "{stage}" >nul 2>&1\r\n'
            'del "%~f0" >nul 2>&1\r\n'
        )
        try:
            # เขียน bat เป็น ANSI/mbcs — cmd อ่าน path ที่มีอักขระไทย/พิเศษได้ถูก
            with open(bat, "w", encoding="mbcs", errors="replace") as f:
                f.write(bat_src)
        except Exception:
            with open(bat, "w", encoding="utf-8") as f:
                f.write(bat_src)

        QMessageBox.information(self, "พร้อมอัปเดต",
            "โปรแกรมจะปิดตัวเองและอัปเดต จากนั้นจะเปิดขึ้นมาใหม่อัตโนมัติ\n"
            "(ใช้เวลาสักครู่ — กรุณาอย่าปิดหน้าต่างดำที่เด้งขึ้นมา)")
        try:
            # CREATE_NEW_CONSOLE = 0x10 → มีหน้าต่าง console จริง ให้คำสั่งทำงานครบ
            subprocess.Popen(["cmd", "/c", bat], creationflags=0x00000010)
        except Exception:
            os.startfile(bat)   # fallback สุดท้าย
        # ปิดโปรแกรมเพื่อให้สคริปต์ก๊อปทับได้
        QApplication.quit()
        os._exit(0)

    def _check_config(self):
        cfg = load_config()
        if not cfg.get("flowaccount_api_key"):
            self.status.showMessage("⚠️  ยังไม่ได้ตั้งค่า API Key  —  ไปที่เมนู ไฟล์ → ตั้งค่า")
        else:
            QTimer.singleShot(300, self.queue_tab.fetch_expenses)

    def _open_settings(self):
        if self.current_user.get("role", "user") != "dev":
            QMessageBox.warning(self, "ไม่มีสิทธิ์",
                "การตั้งค่าแก้ไขได้เฉพาะ Dev เท่านั้น")
            return
        dlg = SettingsDialog(self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.status.showMessage("บันทึกการตั้งค่าแล้ว — กำลังดึงข้อมูลใหม่...")
            self.queue_tab.fetch_expenses()

    def _clear_export_links(self):
        """ล้าง gist ลิงก์ Export เก่าทั้งหมด (ลิงก์เก่าจะใช้ไม่ได้)"""
        cfg = load_config()
        tok = cfg.get("github_token", "").strip()
        if not tok:
            QMessageBox.information(self, "ล้างลิงก์",
                "ยังไม่ได้ตั้งค่า GitHub Token จึงไม่มีลิงก์ถาวรให้ล้าง\n"
                "(ลิงก์ชั่วคราว 72 ชม. จะหมดอายุเอง)")
            return
        if QMessageBox.warning(self, "⚠️ ล้างลิงก์ Export เก่าทั้งหมด",
                "⚠️ ลิงก์ดู (Export Link) เก่าทั้งหมดจะใช้งานไม่ได้\n"
                "ใครที่ถือลิงก์เก่าอยู่จะเปิดไม่ขึ้นอีกต่อไป\n\n"
                "ยืนยันล้างทั้งหมด?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.Cancel,
                QMessageBox.StandardButton.Cancel) != QMessageBox.StandardButton.Yes:
            return
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        try:
            n = delete_all_kcash_gists(tok)
        except Exception as e:
            QApplication.restoreOverrideCursor()
            QMessageBox.critical(self, "❌ ล้างไม่สำเร็จ", str(e))
            return
        QApplication.restoreOverrideCursor()
        activity_log.log("ล้างลิงก์ Export", f"{n} ลิงก์")
        dev_log(f"ล้างลิงก์ Export เก่า {n} ลิงก์", "ok")
        QMessageBox.information(self, "✅ ล้างเสร็จแล้ว",
            f"ล้างลิงก์ Export เก่า {n} ลิงก์แล้ว\nลิงก์เก่าทั้งหมดใช้งานไม่ได้แล้ว")

    def _open_sensitive_manager(self):
        if self.current_user.get("role", "user") not in ("admin", "dev"):
            QMessageBox.warning(self, "ไม่มีสิทธิ์", "คุณไม่มีสิทธิ์เข้าส่วนนี้")
            return
        dlg = SensitiveManagerDialog(self)
        dlg.exec()
        # หลังปิด dialog → refresh หน้าตัดบิลให้แสดง mask ที่อัปเดตแล้ว
        try:
            if getattr(self, "slip_tab", None) and self.slip_tab._results:
                self.slip_tab._revealed_slip_ids.clear()   # reset unlock ทั้งหมด
                self.slip_tab._render_results()
        except Exception:
            pass

    def _open_user_manager(self):
        # user ทั่วไปห้ามเข้าเด็ดขาด
        if self.current_user.get("role", "user") not in ("admin", "dev"):
            QMessageBox.warning(self, "ไม่มีสิทธิ์", "คุณไม่มีสิทธิ์เข้าจัดการผู้ใช้")
            return
        dlg = UserManagerDialog(self.current_user, self)
        dlg.exec()

    def _logout(self):
        if QMessageBox.question(self, "ออกจากระบบ", "ต้องการออกจากระบบ?") != \
                QMessageBox.StandardButton.Yes:
            return
        activity_log.log("ออกจากระบบ", self.current_user.get("fullname", ""))
        self._logout_requested = True
        self.close()

    def _apply_theme(self, dark: bool):
        """ใส่/ถอด โหมดมืดทั้งแอป (palette + Fusion + QSS ครอบ viewport/พื้นหลัง)"""
        app = QApplication.instance()
        if not app:
            return
        if not hasattr(self, "_orig_style"):
            self._orig_style = app.style().objectName()
            self._orig_palette = QApplication.palette()
        if dark:
            app.setStyle("Fusion")
            p = QPalette()
            bg, base, txt = QColor("#202124"), QColor("#2a2b2e"), QColor("#e8eaed")
            p.setColor(QPalette.ColorRole.Window, bg)
            p.setColor(QPalette.ColorRole.WindowText, txt)
            p.setColor(QPalette.ColorRole.Base, base)
            p.setColor(QPalette.ColorRole.AlternateBase, QColor("#303134"))
            p.setColor(QPalette.ColorRole.Text, txt)
            p.setColor(QPalette.ColorRole.Button, QColor("#303134"))
            p.setColor(QPalette.ColorRole.ButtonText, txt)
            p.setColor(QPalette.ColorRole.ToolTipBase, base)
            p.setColor(QPalette.ColorRole.ToolTipText, txt)
            p.setColor(QPalette.ColorRole.PlaceholderText, QColor("#9aa0a6"))
            p.setColor(QPalette.ColorRole.Highlight, QColor("#2563eb"))
            p.setColor(QPalette.ColorRole.HighlightedText, QColor("white"))
            app.setPalette(p)
            app.setStyleSheet(DARK_QSS)
        else:
            try:
                app.setStyle(self._orig_style or "windowsvista")
            except Exception:
                pass
            app.setPalette(self._orig_palette)
            app.setStyleSheet("")

    def _toggle_dark(self, on: bool):
        """สลับธีมแบบสด — ไม่ต้องรีโปรแกรม"""
        _THEME["dark"] = bool(on)
        _apply_row_colors(on)
        self._apply_theme(on)   # palette + Fusion + DARK_QSS / คืนค่าเดิม
        # ใส่สไตล์ใหม่ให้ทุก widget ที่เคยตั้ง style ไว้ (เรียกด้วยต้นฉบับสีอ่อน → แปลงตามธีม)
        for w in QApplication.allWidgets():
            ss = getattr(w, "_orig_ss", None)
            if ss:
                w.setStyleSheet(ss)
        # รีเฟรชตารางให้สีพื้นแถวอัปเดตตามธีม
        for tab in (getattr(self, "queue_tab", None), getattr(self, "po_tab", None)):
            try:
                if tab is not None:
                    tab._apply_filter()
            except Exception:
                pass
        try:
            if getattr(self, "slip_tab", None) and self.slip_tab._results:
                self.slip_tab._render_results()
        except Exception:
            pass
        try:
            cfg = load_config()
            cfg["dark_mode"] = bool(on)
            save_config(cfg)
        except Exception:
            pass
        self.status.showMessage("🌙 โหมดมืด: เปิด" if on else "☀️ โหมดสว่าง")

    def _about(self):
        QMessageBox.about(self, "KCash Queue System",
                          "KCash Queue System\n\n"
                          "ระบบจัดการคิวจ่ายเงิน เชื่อมต่อ FlowAccount API\n\n"
                          "ฟีเจอร์:\n"
                          "• ดึง Expense จาก FlowAccount\n"
                          "• Download CSV สำหรับโอนเงินธนาคาร\n"
                          "• Mark จ่ายแล้ว + สร้างข้อความ Line\n"
                          "• จับคู่ Bank Statement กับ Expense\n\n"
                          "─────────────────────\n"
                          "ติดต่อ Dev:\n"
                          "นายพชร รัชนาทสกุล (คุณมายด์)\n"
                          "โทร 092-703-2121\n\n"
                          "พัฒนาด้วยภาษา Python (PyQt6)\n"
                          "ร่วมกับ Claude Code")


# ──────────────────── Entry point ────────────────────

def _apply_light_palette(app):
    """
    บังคับชุดสี (palette) ให้เป็นธีมสว่างทั้งหมด
    ป้องกันปัญหาตัวหนังสือจาง/มองไม่เห็นเมื่อผู้ใช้ตั้ง Windows เป็นธีมมืด
    """
    C_TEXT     = "#0f172a"   # ตัวหนังสือหลัก (เกือบดำ)
    C_DISABLED = "#94a3b8"   # ตัวหนังสือ disabled
    pal = QPalette()
    R = QPalette.ColorRole
    G = QPalette.ColorGroup

    def setall(role, color):
        c = QColor(color)
        pal.setColor(G.Active,   role, c)
        pal.setColor(G.Inactive, role, c)
        pal.setColor(G.Disabled, role, c)

    setall(R.Window,          C_BG)
    setall(R.WindowText,      C_TEXT)
    setall(R.Base,            C_SURFACE)
    setall(R.AlternateBase,   C_BG)
    setall(R.Text,            C_TEXT)
    setall(R.Button,          C_SURFACE)
    setall(R.ButtonText,      C_TEXT)
    setall(R.ToolTipBase,     C_SURFACE)
    setall(R.ToolTipText,     C_TEXT)
    setall(R.PlaceholderText, C_MUTED)
    setall(R.BrightText,      "#ffffff")
    setall(R.Highlight,       C_PRIMARY)
    setall(R.HighlightedText, "#ffffff")
    setall(R.Link,            C_PRIMARY)

    # override สีตัวหนังสือสำหรับ widget ที่ถูก disable
    pal.setColor(G.Disabled, R.Text,       QColor(C_DISABLED))
    pal.setColor(G.Disabled, R.WindowText, QColor(C_DISABLED))
    pal.setColor(G.Disabled, R.ButtonText, QColor(C_DISABLED))

    app.setPalette(pal)


def main():
    app = QApplication(sys.argv)
    app.setApplicationName("KCash Queue System")
    app.setStyle("Fusion")

    # ตั้งไอคอนระดับ application (taskbar + dialog ทั้งหมด)
    icon_path = _resource_path("icon/icon.ico")
    if os.path.exists(icon_path):
        app.setWindowIcon(QIcon(icon_path))

    # บังคับใช้ธีมสว่างเสมอ ไม่ว่า Windows จะตั้งเป็นธีมมืดหรือสว่าง
    # (แก้ปัญหาตัวหนังสือจางมองไม่เห็นเมื่อ Windows เป็นธีมมืด)
    _apply_light_palette(app)

    # ── Login (requirement ข้อ 8) — วน loop เพื่อรองรับ logout ──
    users.ensure_seed()
    while True:
        login = LoginDialog()
        if os.path.exists(icon_path):
            login.setWindowIcon(QIcon(icon_path))
        if login.exec() != QDialog.DialogCode.Accepted or not login.user:
            sys.exit(0)
        user = login.user   # {username, fullname, nickname, role}
        activity_log.set_session_user(user.get("nickname") or user.get("username"))
        activity_log.log("เข้าสู่ระบบ", f"{user.get('fullname','')} ({user.get('role','')})")

        win = MainWindow(current_user=user)
        win.show()
        app.exec()
        if not getattr(win, "_logout_requested", False):
            break   # ปิดโปรแกรมจริง
        # ถ้า logout → วนกลับไปหน้า login ใหม่
        activity_log.set_session_user(None)
    sys.exit(0)


if __name__ == "__main__":
    main()
